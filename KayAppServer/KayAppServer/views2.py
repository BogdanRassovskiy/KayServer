from django.shortcuts import render
import sqlite3
from django.http import HttpResponse
import requests as rq;
import os;
import traceback;
import openpyxl;
import logging;
import threading;
import math
import random
import json;
import hasher;
from PIL import Image;
import numpy as np;
import Stats;
from datetime import datetime;
import dolConv;
from qrGen import get_qrs;
import reWriteExcel as reW; 


URL=Stats.URL;
KAY=Stats.KAY;
h=hasher.h;
MYTGID="104932971";
VAC_LIST=["А","Б","В","Г","Д","Е","Ё","Ж","З","И","Й","К","Л","М","Н","О","П","Р","С","Т","У","Ф","Х","Ц","Ч","Ш","Щ","Ъ","Ы","Ь","Э","Ю","Я"];
VAC={"А":"A","Б":"B","В":"V","Г":"G","Д":"D","Е":"E","Ё":"YO","Ж":"J","З":"Z","И":"I","Й":"Y","К":"K","Л":"L","М":"M","Н":"N","О":"O",
"П":"P","Р":"R","С":"S","Т":"T","У":"U","Ф":"F","Х":"X","Ц":"C","Ч":"CH","Ш":"SH","Щ":"SH","Ъ":"","Ы":"I","Ь":"","Э":"E","Ю":"YU","Я":"YA",
"0":"0","1":"1","2":"2","3":"3","4":"4","5":"5","6":"6","7":"7","8":"8","9":"9"};
#url_for_bot="https://api.telegram.org/bot"+token+"/sendmessage?chat_id="+chat_id+"&text="txt;
'''
var updates

getMoney
getDistricts
getNak
getOrders
getDolgs
getPrices
getMyLocation
getMarketList
getProductList
getDefaultMarket
'''
#webVer
def web_enter(request):
    try:
        file=open("web/enter.html","r");
        page=file.read();
        file.close();
        page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
        page=page.replace("_URL_",URL);
        page=page.replace("_ERR_","");
        file=open("webBuf/enter.html","w");
        file.write(page);
        file.close();
        return render(request,"webBuf/enter.html");
    except Exception as e:
        logger(e);
def web_enter(request):
    try:
        file=open("web/enter.html","r");
        page=file.read();
        file.close();
        page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
        page=page.replace("_URL_",URL);
        page=page.replace("_ERR_","");
        file=open("webBuf/enter.html","w");
        file.write(page);
        file.close();
        return render(request,"webBuf/enter.html");
    except Exception as e:
        logger(e);
def privacy(request):
    try:
        return render(request,"web/privacy.html");
    except Exception as e:
        logger(e);
        
def check_web_enter(request):
    name=request.GET["name"];
    inn=request.GET["inn"];
    adres=request.GET["adres"];
    phone=request.GET["phone"];
    district=request.GET["district"];
    login=request.GET["login"];
    pswd=request.GET["pswd"];
    r=get_html(URL+"enter/?login={0}&pswd={1}".format(login,pswd)).text;
    print(r);
    err="";
    if "err=1" in r:
        err=r.replace("err=1,,text=","");
    else:
        r=get_html(URL+"create_new_market/?name={0}&inn={1}&adres={2}&phone={3}&district={4}&lon=0.0&lat=0.0&".format(name,inn,adres,phone,district)).text;
    if "err=1" not in r:
        "Магазин создан";
    file=open("web/enter.html","r");
    page=file.read();
    file.close();
    page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
    page=page.replace("_URL_",URL);
    page=page.replace("_ERR_",err);
    file=open("webBuf/enter.html","w");
    file.write(page);
    file.close();
    return render(request,"webBuf/enter.html");

#api
def check_version(request):
    versionNow=Stats.VERSION;
    return HttpResponse(versionNow, content_type='application/json')
def get_languages(request):
    conn=sqlite3.connect('strings.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT lang FROM 'languages'");
    lang=cort_to_list(cursor.fetchall());
    cursor.execute("SELECT code FROM 'languages'");
    code=cort_to_list(cursor.fetchall());
    conn.close();
    langs=[];
    for i in range(len(code)):
        langs.append([lang[i],code[i]]);
    langs=arrayToString(langs);
    send="err=0,,langs={0}".format(langs);
    return HttpResponse(send, content_type='application/json');
def get_strings(request,var="network",lang="_"):
    try:
        if lang=="_":
            lang=request.GET['lang'];
        conn=sqlite3.connect('strings.sqlite');
        cursor=conn.cursor();
        cursor.execute("SELECT id FROM '"+lang+"'");
        id_=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT text FROM '"+lang+"'");
        text=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT id FROM 'ru'");
        id_2=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT text FROM 'ru'");
        text2=cort_to_list(cursor.fetchall());
        id_=id_+id_2;
        text=text+text2;
        words=[];
        for i in range(len(id_)):
            words.append([id_[i],text[i],lang]);
        words=arrayToString(words);
        send="err=0,,words={0}".format(words);
    except Exception as e:
        logger(e);
        send="err=1,,text=ServerError";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_photo_html(request):
    pass
def add_mass_market(request):
    try:
        pswd=request.GET['pswd'];
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT val FROM 'const' WHERE key='pswd'");
        realPswd=cort_to_list(cursor.fetchall())[0];
        conn.close();
        print(h(pswd));
        print(realPswd);
        if h(pswd)==realPswd:
            data=request.GET['data'];
            markets=stringToArray(data);
            print(markets)
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            conn1=sqlite3.connect("markets.sqlite");
            cursor1=conn1.cursor();
            cursor.execute("SELECT login FROM 'users'");
            logins=cort_to_list(cursor.fetchall());
            cursor1.execute("SELECT phone FROM 'users'");
            phones=cort_to_list(cursor1.fetchall());
            i=0;
            for mark in markets:
                login=mark[0];
                pswd=mark[0];
                session="_";
                merchName=mark[1];
                user_type="market";
                lon="0.0";
                lat="0.0";
                adres=mark[2];
                inn=mark[3];
                phone=mark[4];
                varified="0";
                district=mark[5];
                lon=mark[6];
                lat=mark[7];
                phone=mark[8];

                clPhone=phone.replace("998","");
                if login not in logins:
                    if phone not in phones and clPhone not in phones:
                        i+=1;
                        print("OK");
                        cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                            (login,pswd,session,merchName,user_type,lon,lat,"0","ru"));
                        cursor1.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?))",
                            (login,adres,inn,phone,varified,district,phone2,));
                    else:
                        print("have phone");
                else:
                    cursor1.execute("SELECT inn FROM 'users'WHERE login=(?)",(login,));
                    In=cort_to_list(cursor1.fetchall());
                    cursor.execute("SELECT lon FROM 'users'WHERE login=(?)",(login,));
                    Lon=cort_to_list(cursor.fetchall());
                    if In=="_":
                        cursor1.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                        print("replaced inn > "+inn);
                    if Lon=="0.0" or Lon=="0.01":
                        cursor.execute("UPDATE users SET lon=(?)WHERE login=(?)",(lon,login,));
                        cursor.execute("UPDATE users SET lat=(?)WHERE login=(?)",(lat,login,));
                        print("replaced location > "+lon+","+lat);
                    if In!="_" and Lon!="0.0":
                        print("NOPE");
            conn.commit();
            conn.close();
            conn1.commit();
            conn1.close();
            i=str(i);
            z=str(len(markets)-int(i));
            send="err=0,,text=Загружено {0} магазинов. Дубликатов: {1}".format(i,z);
        else:
            send="err=1,,text=Доступ ограничен";
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
def api_create_order(request,var='network'):
    merchName=request.GET['merchName'];
    conn=sqlite3.connect("basic.sqlite");
    cursor=conn.cursor();
    cursor.execute("SELECT level FROM 'levels' WHERE merchName=(?)",(merchName,));
    level=cort_to_list(cursor.fetchall())[0];
    conn.close();
    if "a2|" in level:
        print("ok")
        return send_now_self_order(request);
    else:
        send="err=1,,text=Доступ ограничен";
        return HttpResponse(send, content_type='application/json')
def send_now_self_order(request,var='network'):
    try:
        d=datetime.now();
        key=request.GET['key'];
        merchName=request.GET['merchName'];
        try:
            payForm=request.GET['payForm'];
        except:
            payForm="nal";
        api_key=getConst(merchName,"api_key");
        if key==api_key:
            orderData=request.GET['orderData'].replace("|",":").replace("^",";");
            print(orderData)
            Data=stringToArrayData(orderData);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            for i in range(len(Data)):
                D=Data[i][0];
                cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(D,));
                prodName=cort_to_list(cursor.fetchall())[0];
            conn.close();
            orderName=request.GET['orderName'];
            if orderName=="timeUser":
                orderAdres=request.GET['orderAdres'];
                orderPhone=request.GET['orderPhone'];
                orderDistrict=request.GET['orderDistrict'];
                orderLon=request.GET['orderLon'];
                orderLat=request.GET['orderLat'];
                orderPhone2=request.GET['orderPhone2'];
                orderName=createTimeUser(orderAdres,orderPhone,merchName,orderDistrict,orderPhone2);
            else:
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT adres FROM 'users' WHERE login=(?)",(orderName,));
                orderAdres=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone FROM 'users' WHERE login=(?)",(orderName,));
                orderPhone=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT district FROM 'users' WHERE login=(?)",(orderName,));
                orderDistrict=cort_to_list(cursor.fetchall())[0];
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT lon FROM 'users' WHERE login=(?)",(orderName,));
                orderLon=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT lat FROM 'users' WHERE login=(?)",(orderName,));
                orderLat=cort_to_list(cursor.fetchall())[0];
                conn.close();

            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            Last(merchName,"new");
            last_index=Last(merchName);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("INSERT INTO 'order' VALUES((?),(?),(?),(?),(?),(?),(?),(?))",
                (orderData,orderName,"new",last_index,date,"new","",payForm));
            conn.commit();
            conn.close();

            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
            driver=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
            admin=cort_to_list(cursor.fetchall());
            conn.close();
            for d in driver:
                async2(d,"getOrders");
            for a in admin:
                async2(a,"getAdminOrders");
            send="err=0,,text=OK,,last_index={0},,user_name={1}".format(last_index,orderName);
        else:
            send="err=1,,text=Ключ не верный";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_app(request):
    file_location = '/home/tom/AndroidStudioProjects/KayKay/app/build/outputs/apk/debug/app-debug.apk'
    with open(file_location, 'rb') as f:
       file_data = f.read()
    response = HttpResponse(file_data, content_type='application/apk')
    response['Content-Disposition'] = 'attachment; filename="Kay-Space.apk"'
    return response
def get_photo(request):
    try:
        link=request.GET['link'];
        qrs=os.listdir("qrMarkets");
        marketImg=os.listdir("marketImg");
        file_location = 'qrMarkets/'+link;
        if link in qrs:
            pass;
        elif link in marketImg:
            file_location = 'marketImg/'+link;
        else:
            file_location = 'img/'+link;
        with open(file_location, 'rb') as f:
           file_data = f.read()
        response = HttpResponse(file_data, content_type='image/jpeg')
        response['Content-Disposition'] = 'attachment; filename="'+link+'.jpg"'
        return response;
    except Exception as e:
        logger(e);
def get_api_categories(request,var='network'):
    try:
        cats=[];
        merchName=request.GET['merchName'];
        apikey=request.GET['apikey'];
        api_key=getConst(merchName,"api_key");
        if api_key==apikey:
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT cat_id FROM categories WHERE work='1'");
            cat_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT id FROM categories WHERE work='1'");
            prod_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name FROM categories WHERE work='1'");
            name=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(name)):
                img=Img(merchName,"cat",prod_id[i]);
                cats.append([cat_id[i],prod_id[i],name[i],merchName,img]);
            cats=arrayToString(cats);
            send=cats
        else:
            send="err=1,,text=Ключ апи не действительный";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_api_products(request,var='network'):
    try:
        products=[];
        merchName=request.GET['merchName'];
        apikey=request.GET['apikey'];
        api_key=getConst(merchName,"api_key");
        if api_key==apikey:
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT cat_id FROM products WHERE work='1'");
            cat_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT id FROM products WHERE work='1'");
            prod_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name FROM products WHERE work='1'");
            name=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT rev FROM products WHERE work='1'");
            rev=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(name)):
                price=getProductPrice(merchName,"USER_",prod_id[i]);
                #if price!="0":
                img=Img(merchName,"prod",prod_id[i]);
                products.append([cat_id[i],prod_id[i],name[i],rev[i],price,merchName,img]);
            products=arrayToString(products);
            send=products
        else:
            send="err=1,,text=Ключ апи не действительный";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#api
def enter(request):
    try:
        login=request.GET['login'];
        pswd=request.GET['pswd'];
        conn=sqlite3.connect('basic.sqlite');
        cursor=conn.cursor();
        cursor.execute("SELECT login FROM 'users'");
        logins=cort_to_list(cursor.fetchall());
        print(login+"<<<<<<<<<<");
        if login in logins:
            cursor.execute("SELECT pswd FROM 'users' WHERE login=(?)",(login,));
            real=cort_to_list(cursor.fetchall())[0];
            if h(pswd)==real:
                merchName=Merch(login,0);
                user=User(login,0);
                adLevel="0";
                if user=="admin":
                    adLevel=ADLevel(merchName,login);
                if user=="driver":
                    cursor.execute("SELECT varified FROM 'levels' WHERE merchName=(?)",(merchName,));
                    varified=cort_to_list(cursor.fetchall())[0];
                    adLevel="0";
                conn.close();
                if user=="market":
                    conn=sqlite3.connect('markets.sqlite');
                    cursor=conn.cursor();
                    conn.close();
                session=genSession(login);
                send="err=0,,user={0},,session={1},,level={2}".format(user,session,adLevel);
            else:
                send="err=1,,text=Пароль не верный, проверьте написание";
        else:
            conn.close()
            send="err=1,,text=Такого пользователя не существует,проверьте написание";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
def check_mark_inn(request):
    try:
        inn=request.GET['inn'];
        conn=sqlite3.connect("markets.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT inn FROM 'users'");
        inns=cort_to_list(cursor.fetchall());
        if inn in inns or inn+"+" in inns:
            try:
                cursor.execute("SELECT varified FROM 'users' WHERE inn=(?)",(inn,));
                varified=cort_to_list(cursor.fetchall())[0];
                print(varified)
                if varified=="0":
                    cursor.execute("SELECT login FROM 'users' WHERE inn=(?)",(inn,));
                    login=cort_to_list(cursor.fetchall())[0];
                    send=login;
                    print(login)
                elif varified=="1":
                    send="NOT";
                elif varified=="2":
                    send="NOT";

            except:
                send="NOT"
            
        else:
            send="OK";
        conn.close();
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
def send_var_pswd(request):
    try:
        session1="c75de8c1b7c3ae5252091267a736a9bf57001d80e82668b3cb3cd09e2f6a43cb";
        session=request.GET['session'];
        if session==session1:
            login=request.GET['login'];
            pswd=request.GET['pswd'];
            pswd=h(pswd);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(pswd,login,));
            conn.commit();
            conn.close();

            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET varified='1'WHERE login=(?)",(login,));
            conn.commit();
            conn.close();
            send="OK";
        else:
            send="ACCESS DENIED";
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
def send_self_create_market(request):
    try:
        name=request.GET['name'];
        pswd=request.GET['pswd'];
        inn=request.GET['inn'];
        adres=request.GET['adres'];
        phone=request.GET['phone'];
        phone2=request.GET['phone2'];
        district=request.GET['district'];
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT login FROM 'users'");
        logins=cort_to_list(cursor.fetchall());
        conn.close();

        conn=sqlite3.connect("markets.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT adres FROM 'users'");
        adress=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT inn FROM 'users'");
        inns=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT phone FROM 'users'");
        phones=cort_to_list(cursor.fetchall());
        conn.close();
        if name in logins:
            send="Магазин с таким названием уже существует";
        elif inn in inns:
            send="Магазин с таким ИНН уже существует";
        elif adres in adress:
            send="Магазин с таким адресом уже существует, уточните адрес";
        elif phone in phones:
            send="Магазин с таким номером телефона уже существует";
        else:
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?))",
                (name,adres,inn,phone,"0",district,phone2));
            conn.commit();
            conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                    (name,h(pswd),"_","self","market","0.0","0.0","0","ru",));
            cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
            drivers=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
            admins=cort_to_list(cursor.fetchall());
            conn.commit();
            conn.close();
            #for d in drivers:
            #    async2(d,"getMarketList");
            #for a in admins:
            #    async2(a,"getMarketList");
            send="OK";
            txt="Зарегестрирован магазин🧩"
            txt=txt+"\n\nИмя: "+name
            txt=txt+"\nИНН: "+inn
            txt=txt+"\nАдрес: "+adres
            txt=txt+"\nТелефон: "+phone
            txt=txt+"\nРайон: "+district
            messageInChannel("Kay-Kay",txt);    
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
def send_self_create_prodavac(request):
    try:
        name=request.GET['name'];
        inn=request.GET['inn'];
        adres=request.GET['adres'];
        phone=request.GET['phone'];
        txt="📡Новая заявка от производителя.\nНаименование: "+name;
        txt=txt+"\nИНН: "+inn;
        txt=txt+"\nЮр.Адрес: "+adres;
        txt=txt+"\nТелефон: "+phone;
        messageInChannel("Kay-Kay",txt);    
        send="OK";
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
def check_order_status(request):
    try:
        merchName=request.GET['merchName'];
        last_index=request.GET['last_index'];
        print(last_index)
        print(merchName)
        lasts=stringToArray(last_index);
        conn=sqlite3.connect(mPath(merchName,"orders"));
        cursor=conn.cursor();
        cursor.execute("SELECT last_index FROM 'order'");
        orders=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT last_index FROM 'history'");
        history=cort_to_list(cursor.fetchall());
        ans="";
        for l in lasts:
            last_index=l[0];
            var="";
            if last_index in orders:
                var=last_index+"|1|^";
            elif last_index in history:
                var=last_index+"|2|^";
            else:
                var=last_index+"|3|^";
            ans+=var;
        conn.close();
        send=ans;
    except Exception as e:
        logger(e);
        send=e;
    return HttpResponse(send, content_type='application/json')
#root
def get_roots(request,var='network'):
    try:
        roots=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=Merch(session);
                login=Login(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'WHERE user_type='root'");
                logins=cort_to_list(cursor.fetchall());
                for l in logins:
                    cursor.execute("SELECT level FROM 'roots'WHERE login=(?)",(l,));
                    level=cort_to_list(cursor.fetchall())[0];
                    roots.append([l,level.replace("|",":")]);
                conn.close();
                roots=arrayToString(roots);
                send="err=0,,roots={0}".format(roots);
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_var_levels(request,var='network'):
    try:
        varLevels=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT var FROM 'percent'");
                var=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT cur FROM 'percent'");
                cur=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT per FROM 'percent'");
                per=cort_to_list(cursor.fetchall());
                conn.close();
                for i in range(len(var)):
                    varLevels.append([var[i],cur[i],per[i]]);
                varLevels=arrayToString(varLevels);
                send="err=0,,varLevels={0}".format(varLevels);
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_updates_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            updates=cort_to_list(cursor.fetchall());
            if updates[0]=="0":
                send="err=0,,update_stat=0";
            else:
                send="err=0,,update_stat=1,,updates={0}".format(updates[0]);
            cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            conn.commit();
            conn.close();
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def uber_root_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            #
            updates+="<<getDistricts>>"+get_districts(request,'local')
            updates+="<<getMarketRootList>>"+get_market_root_list(request,'local')
            updates+="<<getMerchants>>"+get_merchants(request,'local');
            updates+="<<getVarLevels>>"+get_var_levels(request,'local');
            updates+="<<getRoots>>"+get_roots(request,'local');
            updates+="<<getCislo>>"+get_cislo(request,'local');
            #
            send=updates
            send="err>>0"+send;
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
#root do
def send_market_var(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                v=request.GET['v'];
                if v=="remMarket":
                    conn=sqlite3.connect("markets.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    conn.commit();
                    conn.close();
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    conn.commit();
                    conn.close();
                    merchs=os.listdir("merchants");
                    for m in merchs:
                        conn=sqlite3.connect(mPath(m,"orders"));
                        cursor=conn.cursor();
                        cursor.execute("DELETE FROM 'buy_markets' WHERE login = (?)",(login,));
                        cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(login,));
                        conn.commit();
                        conn.close();
                        conn=sqlite3.connect(mPath(m,"inCar"));
                        cursor=conn.cursor();
                        cursor.execute("DELETE FROM 'buy_markets' WHERE login = (?)",(login,));
                        cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(login,));
                        conn.commit();
                        conn.close();
                if v=="remInn":
                    conn=sqlite3.connect("markets.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("UPDATE users SET inn='_'WHERE login=(?)",(login,));
                    conn.commit();
                    conn.close();
                if v=="acceptMarket":
                    conn=sqlite3.connect("markets.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("SELECT inn FROM 'users' WHERE login=(?)",(login,));
                    inn=cort_to_list(cursor.fetchall())[0];
                    inn=inn.replace("+","");
                    cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                    conn.commit();
                    conn.close();
                

                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_new_merch_pswd(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                pswd=request.GET['pswd'];
                
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT level FROM 'admins'");
                levels=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT login FROM 'admins'");
                admins=cort_to_list(cursor.fetchall());
                conn.close();
                for i in range(len(levels)):
                    if "b1|" in levels[i]:
                        maxAdmin=admins[i];
                        break;
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),maxAdmin,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_new_acces(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                acces=request.GET['acces'];
                acces=acces.replace(";","^").replace(":","|")
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                if "a3|"in acces:
                    acces=acces.replace("a3|","");
                    cursor.execute("UPDATE levels SET varified='1'WHERE merchName=(?)",(merchName,));
                else:
                    cursor.execute("UPDATE levels SET varified='0'WHERE merchName=(?)",(merchName,));
                cursor.execute("UPDATE levels SET level=(?)WHERE merchName=(?)",(acces,merchName,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_new_root_pswd(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                pswd=request.GET['pswd'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),login,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_admin_pswd(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                login=request.GET['login'];
                pswd=request.GET['pswd'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),login,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не admin";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_new_root_level(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                level=request.GET['level'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE 'roots' SET level=(?)WHERE login=(?)",(level,login,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_create_new_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                pswd=request.GET['pswd'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'WHERE user_type='root'");
                logins=cort_to_list(cursor.fetchall());
                if login in logins:
                    send="err=0,,text=Пользователь с таким логином существует";
                else:
                    cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        (login,h(pswd),"_","self","root","0.0","0.0","0","ru",));
                    cursor.execute("INSERT INTO 'roots' VALUES((?),(?))",(login,"",))
                    send="err=0,,text=OK";
                conn.commit();
                conn.close();
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_new_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                if login!="root":
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    cursor.execute("DELETE FROM 'roots' WHERE login = (?)",(login,));
                    conn.commit();
                    conn.close();
                    send="err=0,,text=OK";
                else:
                    send="err=1,,text=NOT";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_district(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'districts' WHERE district = (?)",(name,));
                cursor.execute("SELECT login FROM 'users'");
                logins=cort_to_list(cursor.fetchall());
                updates="getDistricts";
                for i in range(len(logins)):
                    cursor.execute("SELECT updates FROM 'users' WHERE login=(?)",(login[i]));
                    oldUpdates=cort_to_list(cursor.fetchall())[0];
                    if updates not in oldUpdates:
                        update=oldUpdates+updates+"|^";
                    else:
                        update=oldUpdates;
                    cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)",(update,login[i],));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_create_new_district(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT district FROM districts");
                districts=cort_to_list(cursor.fetchall());
                if name in districts:
                    send="err=0,,text=Район с таким именем уже существует";
                else:
                    cursor.execute("INSERT INTO 'districts' VALUES((?))",(name,));
                    cursor.execute("SELECT updates FROM 'users'");
                    updates=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT login FROM 'users'");
                    logins=cort_to_list(cursor.fetchall());
                    for i in range(len(updates)):
                        if "getDistricts|^" not in updates[i]:
                            upd=updates[i]+"getDistricts|^";
                            cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(upd,logins[i],));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_location(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET lon='0.0'WHERE login=(?)",(name,));
                cursor.execute("UPDATE users SET lat='0.0'WHERE login=(?)",(name,));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root" or User(session)=="admin":
                name=request.GET['name'];
                print(name+"<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'users' WHERE login = (?)",(name,));
                conn.commit();
                conn.close();
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'users' WHERE login = (?)",(name,));
                conn.commit();
                conn.close();
                merchs=os.listdir("merchants");
                for m in merchs:
                    conn=sqlite3.connect(mPath(m,"orders"));
                    cursor=conn.cursor();
                    cursor.execute("DELETE FROM 'buy_markets' WHERE name = (?)",(name,));
                    cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(name,));
                    conn.commit();
                    conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_save_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root"or User(session)=="admin":
                login=request.GET['login'];
                adres=request.GET['adres'];
                inn=request.GET['inn'];
                phone=request.GET['phone'];
                varified=request.GET['varified'];
                pswd=request.GET['pswd'];
                phone2=request.GET['phone2'];
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET adres=(?)WHERE login=(?)",(adres,login,));
                cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                cursor.execute("UPDATE users SET phone=(?)WHERE login=(?)",(phone,login,));
                cursor.execute("UPDATE users SET varified=(?)WHERE login=(?)",(varified,login,));
                cursor.execute("UPDATE users SET phone2=(?)WHERE login=(?)",(phone2,login,));
                conn.commit();
                conn.close();
                if pswd!="null":
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),login,));
                    conn.commit();
                    conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_get_this_history(request,var='network'):
    try:
        his=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=request.GET['merchName'];
            howMany=request.GET['howMany'].lower();
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT data FROM 'history' ORDER BY last_index DESC");
            datas=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT price FROM 'history' ORDER BY last_index DESC");
            price=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT type FROM 'history' ORDER BY last_index DESC");
            _type=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'history' ORDER BY last_index DESC");
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT driver FROM 'history' ORDER BY last_index DESC");
            driver=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'history' ORDER BY last_index DESC");
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT district FROM 'history' ORDER BY last_index DESC");
            district=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT last_index FROM 'history' ORDER BY last_index DESC");
            last_index=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT visible FROM 'history' ORDER BY last_index DESC");
            visible=cort_to_list(cursor.fetchall());
            for i in range(len(datas)):
                if ";" in datas[i] and ":" in datas[i]:
                    d=stringToArrayData(datas[i]);
                    dat="";
                    for z in range(len(d)):
                        prod_id=d[z][0];
                        how=d[z][1];
                        cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                        name=cort_to_list(cursor.fetchall())[0];
                        dat=name+" * "+how+"\n";
                        cursor.execute("UPDATE history SET data=(?)WHERE last_index=(?)",(dat,last_index[i]));
                else:
                    dat=datas[i];
                try:
                    cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(getter[i],));
                    buy_type=cort_to_list(cursor.fetchall())[0];
                except:
                    buy_type="_"
                if howMany in dat.lower() or howMany in last_index[i].lower() or howMany in driver[i].lower() or howMany in getter[i].lower() or howMany in date[i].lower():
                    his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                        visible[i],buy_type]);
            conn.commit();
            conn.close();
            his=arrayToString(his);
            send="err=0,,his={0}".format(his);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_merchant(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'");
                logins=cort_to_list(cursor.fetchall());
                conn.close();
                merchs=os.listdir("merchants");
                merchAdmin=merchName+"Admin";
                if merchName in merchs:
                    send="err=0,,text=Производитель с таким именем уже существует";
                elif merchAdmin in logins:
                    send="err=0,,text=Выберете другое имя";
                else:
                    os.system("cp -r merchBuf/buf merchants/"+merchName);
                    conn=sqlite3.connect(mPath(merchName,"orders"));
                    cursor=conn.cursor();
                    role="Главный администратор";
                    cursor.execute("INSERT INTO 'admins' VALUES((?),(?),(?))",(merchAdmin,"b1|a1|a2|a3|a4|a5|a6|a7|a8|a9|a10|a11|a12|a13|",role));
                    conn.commit();
                    conn.close();
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("INSERT INTO 'levels' VALUES((?),(?),(?))",(merchName,"","1"));
                    cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        (merchAdmin,h(merchAdmin),"_",merchName,"admin","0.0","0.0","0","ru",));
                    cursor.execute("SELECT login FROM 'users' WHERE user_type='market'");
                    markets=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT login FROM 'users' WHERE user_type='root'");
                    roots=cort_to_list(cursor.fetchall());
                    upd="getMerchants";
                    for m in markets:
                        cursor.execute("SELECT login FROM 'users' WHERE login=(?)",(m,));
                        updates=cort_to_list(cursor.fetchall())[0];
                        if upd not in updates:
                            updates=updates+upd+"|^";
                            cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(updates,m,));
                    for r in roots:
                        cursor.execute("SELECT login FROM 'users' WHERE login=(?)",(r,));
                        updates=cort_to_list(cursor.fetchall())[0];
                        if upd not in updates:
                            updates=updates+upd+"|^";
                            cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(updates,r,));
                    conn.commit();
                    conn.close();
                    send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_photo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            user=User(session)
            if user=="root" or user=="admin":
                market=request.GET['market'];
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET varified='1'WHERE login=(?)",(market,));
                conn.commit();
                conn.close();
                try:
                    os.remove("marketImg/"+market+"guvPhoto.jpg");
                    os.remove("marketImg/"+market+"pasPhoto.jpg");
                except:
                    pass;
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_accept_photo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            user=User(session)
            if user=="root" or user=="admin":
                market=request.GET['market'];
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE users SET varified='2'WHERE login=(?)",(market,));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_var_level_settings(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                level=request.GET['level'];
                currency=request.GET['currency'];
                percent=request.GET['percent'];
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE percent SET cur=(?)WHERE var=(?)",(currency,level,));
                cursor.execute("UPDATE percent SET per=(?)WHERE var=(?)",(percent,level,));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def get_cislo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT inn FROM 'users'");
                inn=cort_to_list(cursor.fetchall());
                w=0;
                wOut=0;
                for i in inn:
                    if i=="_":
                        wOut+=1;
                    else:
                        w+=1;
                conn.commit();
                conn.close();
                send="err=0,,cislo={0}".format(str(w)+"|"+str(wOut)+"|^");
            else:
                send="err=1,,text=Вы не root";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 

#market
def uber_market_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            #
            updates+="<<getMerchants>>"+get_merchants(request,'local');
            #updates+="<<getMarketCategories>>"+get_market_categories(request,'local');
            #updates+="<<getMarketProducts>>"+get_market_products(request,'local');
            updates+="<<getIMarket>>"+get_i_market(request,'local');
            updates+="<<getMarketNews>>"+get_market_news(request,'local');
            updates+="<<getDistricts>>"+get_districts(request,'local')
            updates+="<<getMarketList>>"+get_market_list(request,'local')
            #
            send=updates#.replace("err=0,,")
            send="err>>0"+send;
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
def get_merchants(request,var='network'):
    try:
        merchants=[];
        session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            for m in merchs:
                rev=getConst(m,"merchRev");
                img=getConst(m,"merchImg");
                site=getConst(m,"site");
                img="{0}get_photo/?link={1}".format(URL,img);
                cursor.execute("SELECT level FROM levels WHERE merchName=(?)",(m,));
                level=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT varified FROM levels WHERE merchName=(?)",(m,));
                varified=cort_to_list(cursor.fetchall())[0];
                if varified=="1":
                    level+="a3|";
                level=level.replace("^",";").replace("|",":")
                if m!="mchj_tashkent" or Login(session)=="W1W1W1_CHILANZAR":
                    if m!="Solod-Expo" or Login(session)=="W1W1W1_CHILANZAR":
                        generalLang=getConst(m,"generalLang");
                        secondLang=getConst(m,"secondLang");
                        merchants.append([m,rev,img,level,site,generalLang,secondLang]);
                print(site);
            conn.close();
            merchants=arrayToString(merchants);
            send="err=0,,merchants={0}".format(merchants);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_categories(request,var='network'):
    try:
        cats=[];
        session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            for merchName in merchs:
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT cat_id FROM categories WHERE work='1'");
                cat_id=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT id FROM categories WHERE work='1'");
                prod_id=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT name FROM categories WHERE work='1'");
                name=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT name2 FROM categories WHERE work='1'");
                name2=cort_to_list(cursor.fetchall());
                conn.close();
                for i in range(len(name)):
                    img=Img(merchName,"cat",prod_id[i]);
                    cats.append([cat_id[i],prod_id[i],name[i],merchName,img,name2[i]]);
            cats=arrayToString(cats);
            send="err=0,,market_categories={0}".format(cats);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_products(request,var='network'):
    try:
        products=[];
        session=request.GET['session'];
        if checkSession(session):
            market=Login(session);
            merchs=os.listdir("merchants");
            for merchName in merchs:
                print(merchName,"merchName");
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT cat_id FROM products WHERE work='1'");
                cat_id=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT id FROM products WHERE work='1'");
                prod_id=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT name FROM products WHERE work='1'");
                name=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT rev FROM products WHERE work='1'");
                rev=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT box FROM products WHERE work='1'");
                box=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT form FROM products WHERE work='1'");
                form=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT visible FROM products WHERE work='1'");
                visible=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT ost FROM products WHERE work='1'");
                ost=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT name2 FROM products WHERE work='1'");
                name2=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT rev2 FROM products WHERE work='1'");
                rev2=cort_to_list(cursor.fetchall());
                conn.close();
                for i in range(len(name)):
                    price=getProductPrice(merchName,market,prod_id[i]);
                    if price=="":
                        price="1";
                    if prod_id[i]=="1":
                        print(price+"price<<<<<<<")
                    if price!="0":
                        if visible[i]!="0":
                            img=Img(merchName,"prod",prod_id[i]);
                            #
                            #
                            if merchName!="Solod-Expo" or market=="W1W1W1_CHILANZAR":
                                products.append([cat_id[i],prod_id[i],name[i],rev[i],price,merchName,img,box[i],form[i],ost[i],name2[i],rev2[i]]);
            products=arrayToString(products);
            send="err=0,,market_products={0}".format(products);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_i_market(request,var='network'):
    try:
        info=[];
        session=request.GET['session'];
        if checkSession(session):
            market=Login(session);
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT adres FROM users WHERE login=(?)",(market,));
            adres=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT inn FROM users WHERE login=(?)",(market,));
            inn=cort_to_list(cursor.fetchall())[0].replace("+","");
            cursor.execute("SELECT phone FROM users WHERE login=(?)",(market,));
            phone=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT district FROM users WHERE login=(?)",(market,));
            district=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT varified FROM users WHERE login=(?)",(market,));
            varified=cort_to_list(cursor.fetchall())[0];
            checkCoinSettings(market);
            cursor.execute("SELECT tg_id FROM coinInfo WHERE login=(?)",(market,));
            tg_id=cort_to_list(cursor.fetchall())[0];

            marks=os.listdir("qrMarkets");
            if market not in marks:
                get_qrs(market,"qrMarkets/"+market);
            conn.close();
            info.append([adres,inn,phone,district,tg_id,varified]);
            info=arrayToString(info);
            send="err=0,,market_info={0}".format(info);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_news(request,var='network'):
    try:
        orders=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            merchs=os.listdir("merchants");
            for merchName in merchs:
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT data FROM 'order' WHERE getter=(?)",(login,));
                data=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT getter FROM 'order' WHERE getter=(?)",(login,));
                getter=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT owner FROM 'order' WHERE getter=(?)",(login,));
                owner=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT last_index FROM 'order' WHERE getter=(?)",(login,));
                last_index=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT date FROM 'order' WHERE getter=(?)",(login,));
                date=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT get_type FROM 'order' WHERE getter=(?)",(login,));
                get_type=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT price FROM 'order' WHERE getter=(?)",(login,));
                priceD=cort_to_list(cursor.fetchall());
                conn.close();
                for i in range(len(data)):
                    price=0;
                    d=stringToArrayData(data[i]);
                    for z in range(len(d)):
                        prod_id=d[z][0];
                        how=d[z][1];
                        pr=getProductPrice(merchName,login,prod_id);
                        price=price+int(pr)*int(how);
                    if get_type[i]=="dolg":
                        price=int(priceD[i]);
                    orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],owner[i],str(price),merchName]);
            orders=arrayToString(orders);
            send="err=0,,market_news={0}".format(orders);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send;
def get_market_history(request,var='network',user="all"):
    try:
        his=[];
        session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            login=Login(session);
            for merchName in merchs:
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT data FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                datas=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT price FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                price=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT type FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                _type=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT getter FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                getter=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT driver FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                driver=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT date FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                date=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT district FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                district=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT last_index FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                last_index=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT visible FROM 'history' WHERE getter=(?) ORDER BY last_index DESC ",(login,));
                visible=cort_to_list(cursor.fetchall());
                for i in range(len(datas)):
                    if ";" in datas[i] and ":" in datas[i]:
                        d=stringToArrayData(datas[i]);
                        dat="";
                        for z in range(len(d)):
                            prod_id=d[z][0];
                            how=d[z][1];
                            cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                            name=cort_to_list(cursor.fetchall())[0];
                            dat=name+" * "+how+"\n";
                            cursor.execute("UPDATE history SET data=(?)WHERE last_index=(?)",(dat,last_index[i]));
                    else:
                        dat=datas[i];
                    try:
                        cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(getter[i],));
                        buy_type=cort_to_list(cursor.fetchall())[0];
                    except:
                        buy_type="_"
                    his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                        visible[i],buy_type]);
                conn.commit();
                conn.close();
            his=arrayToString(his);
            send="err=0,,his={0}".format(his);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_updates_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            updates=cort_to_list(cursor.fetchall());
            if updates[0]=="0":
                send="err=0,,update_stat=0";
            else:
                send="err=0,,update_stat=1,,updates={0}".format(updates[0]);
            cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            conn.commit();
            conn.close();
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#market do
def send_market_order(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                market=Login(session);
                orderData=request.GET['data'].replace("|",":").replace("^",";");
                merchName=request.GET['merchName'];
                try:
                    payForm=request.GET['payForm'];
                except Exception as e:
                    logger(e);
                    payForm="nal";
                Data=stringToArrayData(orderData);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                dataTxt="";
                for i in range(len(Data)):
                    D=Data[i][0];
                    cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(D,));
                    prodName=cort_to_list(cursor.fetchall())[0];
                    try:
                        dataTxt=dataTxt+prodName+" x "+Data[i][1]+"\n";
                    except:
                        pass;
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                Last(merchName,"new");
                last_index=Last(merchName);
                cursor.execute("INSERT INTO 'order' VALUES((?),(?),(?),(?),(?),(?),(?),(?))",
                    (orderData,market,"new",last_index,date,"new","",payForm));
                cursor.execute("SELECT login FROM 'admins'");
                admins=cort_to_list(cursor.fetchall());
                conn.commit();
                conn.close();
                payFormTxt="Наличные";
                if payForm=="nal":
                    payFormTxt="Наличные";
                elif payForm=="term":
                    payFormTxt="Терминал";
                elif payForm=="per":
                    payFormTxt="Перечисление";
                payFormTxt="Форма оплаты: "+payFormTxt;
                txt="Новый заказ🎈№{0}\nЗаказчик: {1}\n\n{2}".format(last_index,market,dataTxt+"\n"+payFormTxt);
                messageInChannel(merchName,txt);
                send="err=0,,text=OK";
                for a in admins:
                    async2(a,"getAdminOrders");
                    #async2(a,"newOrder",merchName);
                    addNot(a,"title","text","marketOrder");
                async2(market,"getMarketNews");
                #if merchName =="Kay-Kay":
                #giveCoin(market,"ice","1");
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_market_rem_order(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                last_index=request.GET['last_index'];
                merchName=request.GET['merchName'];
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'order' WHERE last_index = (?)",(last_index,));
                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                (merchName,));
                admins=cort_to_list(cursor.fetchall());
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getAllPrices");
                conn.close();
                send="err=0 text=OK";
                txt="❌Заказ №"+last_index+" был удален заказчиком";
                messageInChannel(merchName,txt);
            else:
                send="err=1,,text=Вы не магазин";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_self_market_data(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                market=Login(session);
                newName=request.GET['newName'];
                newInn=request.GET['newInn'];
                newAdres=request.GET['newAdres'];
                newPhone=request.GET['newPhone'];
                newDistrict=request.GET['newDistrict'];
                newTgId=request.GET['newTgId'];
                newTgPswd=request.GET['newTgPswd'];
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT inn FROM 'users' WHERE login=(?)",(market,));
                oldInn=cort_to_list(cursor.fetchall())[0];
                V=False;
                if oldInn!=newInn:
                    cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(newInn,market,));
                    cursor.execute("UPDATE users SET varified='100'WHERE login=(?)",(market,));
                    V=True;
                cursor.execute("UPDATE users SET adres=(?)WHERE login=(?)",(newAdres,market,));
                cursor.execute("UPDATE users SET phone=(?)WHERE login=(?)",(newPhone,market,));
                cursor.execute("UPDATE users SET district=(?)WHERE login=(?)",(newDistrict,market,));
                cursor.execute("UPDATE coinInfo SET tg_id=(?)WHERE login=(?)",(newTgId,market,));
                if newTgPswd !="":
                    cursor.execute("UPDATE coinInfo SET pswd=(?)WHERE login=(?)",(h(newTgPswd),market,));
                conn.commit();
                conn.close();

                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                if newName!="":
                    cursor.execute("UPDATE users SET pswd=(?)WHERE session=(?)",(h(newName),session,));
                if V:
                    cursor.execute("UPDATE users SET session='_'WHERE session=(?)",(session,));
                cursor.execute("SELECT login FROM 'users'");
                us=cort_to_list(cursor.fetchall());
                conn.commit();
                conn.close();
                async2(market,"getMarketList");
                send="err=0 text=OK";
                if V:
                    txt="📕Пользователь {0} сменил себе ИНН и ждет варификации".format(market);
                    messageInChannel("Kay-Kay",txt);
            else:
                send="err=1,,text=Вы не магазин";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def check_send_tg_settings(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            tg_id=request.GET['tg_id'];
            tg_pswd=request.GET['tg_pswd'];
            if User(session)=="market":
                get_html(KAY+"/check_tg_data/?tg_id={0}&tg_pswd={1}".format(tg_id,tg_pswd,));
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не магазин";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
#admin
def uber_admin_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT id FROM 'products'");
            prod_id=arrayToString([cort_to_list(cursor.fetchall())]);
            conn.close();
            #
            updates+="<<getDistricts>>"+get_districts(request,'local')
            updates+="<<getHistory>>"+get_history(request,'local')
            updates+="<<getDrivers>>"+get_drivers(request,'local')
            updates+="<<getAdminOrders>>"+get_admin_orders(request,'local')
            #updates+="<<getProducts>>"+get_product_list(request,'local')
            #updates+="<<getCategories>>"+get_cat_list(request,'local')
            #updates+="<<getMarketList>>"+get_market_list(request,'local')
            updates+="<<getDefaultMarket>>"+get_default_market(request,'local')
            updates+="<<getAllPrices>>"+get_prices(request,prod_id,'local')
            updates+="<<getAllLogins>>"+get_all_users(request,'local')
            updates+="<<getAdmins>>"+get_admins(request,'local')
            #
            send=updates#.replace("err=0,,")
            send="err>>0"+send;
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
def get_history(request,var='network',user="all"):
    try:
        his=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT data FROM 'history' ORDER BY last_index DESC");
            datas=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT price FROM 'history' ORDER BY last_index DESC");
            price=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT type FROM 'history' ORDER BY last_index DESC");
            _type=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'history' ORDER BY last_index DESC");
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT driver FROM 'history' ORDER BY last_index DESC");
            driver=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'history' ORDER BY last_index DESC");
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT district FROM 'history' ORDER BY last_index DESC");
            district=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT last_index FROM 'history' ORDER BY last_index DESC");
            last_index=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT visible FROM 'history' ORDER BY last_index DESC");
            visible=cort_to_list(cursor.fetchall());
            for i in range(len(datas)):
                if ";" in datas[i] and ":" in datas[i]:
                    d=stringToArrayData(datas[i]);
                    dat="";
                    for z in range(len(d)):
                        prod_id=d[z][0];
                        how=d[z][1];
                        cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                        name=cort_to_list(cursor.fetchall())[0];
                        dat=name+" * "+how+"\n";
                        cursor.execute("UPDATE history SET data=(?)WHERE last_index=(?)",(dat,last_index[i]));
                else:
                    dat=datas[i];
                try:
                    cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(getter[i],));
                    buy_type=cort_to_list(cursor.fetchall())[0];
                except:
                    buy_type="_"
                his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                    visible[i],buy_type]);
            conn.commit();
            conn.close();
            his=arrayToString(his);
            send="err=0,,his={0}".format(his);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_drivers(request,var='network'):
    try:
        drivers=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName =(?)",(merchName,));
            logins=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT lon FROM 'users' WHERE user_type='driver' AND merchName =(?)",(merchName,));
            lon=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT lat FROM 'users' WHERE user_type='driver' AND merchName =(?)",(merchName,));
            lat=cort_to_list(cursor.fetchall());
            conn.close();
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            conn1=sqlite3.connect(mPath(merchName,"orders"));
            cursor1=conn1.cursor();
            for i in range(len(logins)):
                cursor.execute("SELECT name FROM 'naks' WHERE owner=(?)",(logins[i],));
                name=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT prod_id FROM 'naks' WHERE owner=(?)",(logins[i],));
                prod_id=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)",(logins[i],));
                free=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT reserve FROM 'naks' WHERE owner=(?)",(logins[i],));
                reserve=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT sell FROM 'naks' WHERE owner=(?)",(logins[i],));
                sell=cort_to_list(cursor.fetchall());
                nak="";
                for z in range(len(name)):
                    nak=nak+name[z]+":"+prod_id[z]+":"+free[z]+":"+reserve[z]+":"+sell[z]+":;";
                print(logins[i]+"<<<<")
                cursor.execute("SELECT cash FROM 'cashInCar' WHERE driver=(?)",(logins[i],));
                cash=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT term FROM 'cashInCar' WHERE driver=(?)",(logins[i],));
                term=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT per FROM 'cashInCar' WHERE driver=(?)",(logins[i],));
                per=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT on_day FROM 'cashInCar' WHERE driver=(?)",(logins[i],));
                on_day=cort_to_list(cursor.fetchall())[0];
                nakNum="";
                #nakNum
                cursor.execute("SELECT num FROM 'nakNum' WHERE owner=(?)",(logins[i],));
                num=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT date FROM 'nakNum' WHERE owner=(?)",(logins[i],));
                date=cort_to_list(cursor.fetchall());

                try:
                    cursor1.execute("SELECT district FROM 'drivers' WHERE login=(?)",(logins[i],));
                    diss=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT name FROM 'drivers' WHERE login=(?)",(logins[i],));
                    nakName=cort_to_list(cursor1.fetchall())[0];
                except:
                    cursor1.execute("INSERT INTO drivers VALUES((?),(?),(?))",(logins[i],":;",""))
                    conn1.commit();
                    diss=":;"
                    nakName=""

                nakNum="";
                for z in range(len(num)):
                    nakNum+="Накладная №"+num[z]+". Дата: "+date[z]+" ";
                drivers.append([logins[i],lon[i],lat[i],nak,cash,term,per,on_day,nakNum,diss,nakName]);
            drivers=arrayToString(drivers);
            send="err=0,,drivers={0}".format(drivers);
            conn.close();
            conn1.close();
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_admin_orders(request,var='network'):
    try:
        orders=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT data FROM 'order'");
            data=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'order'");
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT owner FROM 'order'");
            owner=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT last_index FROM 'order'");
            last_index=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'order'");
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT get_type FROM 'order'");
            get_type=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT price FROM 'order'");
            price=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT payForm FROM 'order'");
            payForm=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],price[i],owner[i],payForm[i]]);
            orders=arrayToString(orders);
            send="err=0,,orders={0}".format(orders);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_all_users(request,var='network'):
    try:
        logins=[];
        session=request.GET['session'];
        if True:
            #merchName=Merch(session);
            #login=Login(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            login=cort_to_list(cursor.fetchall());
            conn.close();
            for l in login:
                logins.append([l]);
            logins=arrayToString(logins);
            send="err=0,,logins={0}".format(logins);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        print(send+"<<<<")
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def get_updates_admin(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            updates=cort_to_list(cursor.fetchall());
            if updates[0]=="0":
                send="err=0,,update_stat=0";
            else:
                send="err=0,,update_stat=1,,updates={0}".format(updates[0]);
            cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            conn.commit();
            conn.close();
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_admins(request,var='network'):
    try:
        admins=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                login=Login(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'WHERE user_type='admin'AND merchName=(?)",(merchName,));
                logins=cort_to_list(cursor.fetchall());
                conn.close();
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                for l in logins:
                    cursor.execute("SELECT level FROM 'admins'WHERE login=(?)",(l,));
                    level=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT role FROM 'admins'WHERE login=(?)",(l,));
                    role=cort_to_list(cursor.fetchall())[0];
                    if role==None:
                        role="";
                    admins.append([l,level.replace("|","_"),role]);
                conn.close();
                admins=arrayToString(admins);
                send="err=0,,admins={0}".format(admins);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_box_settings(request,var='network'):
    try:
        session=request.GET['session'];
        box_settings=request.GET['box_settings'];
        box_settings=stringToArray(box_settings)[0];
        print(box_settings)
        prod_id=box_settings[0];
        boxes=box_settings[1];
        form=box_settings[2];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'products' SET form=(?)WHERE id=(?)",(form,prod_id,));
                cursor.execute("UPDATE 'products' SET 'box'=(?)WHERE id=(?)",(boxes,prod_id,));
                conn.commit();
                conn.close();
                print("will save")
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'WHERE user_type='driver'AND merchName=(?)",(merchName,));
                drivers=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT login FROM 'users'WHERE user_type='admin'AND merchName=(?)",(merchName,));
                admins=cort_to_list(cursor.fetchall());
                for d in drivers:
                    async2(d,"getProductList");
                for a in admins:
                    async2(a,"getProductList");
                conn.close()
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_unvarified_markets(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=ADLevel(myMerchName,login);
            if "max98" in level or login=="admin":
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE varified='100'");
                login=cort_to_list(cursor.fetchall());
                
                for i in range(len(login)):
                    log=login[i];
                
                    cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                    adres=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                    inn=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                    phone=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT tg_id FROM 'coinInfo' WHERE login=(?) ",(log,));
                    tg_id=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?) ",(log,));
                    pswd=cort_to_list(cursor.fetchall())[0];
                    if len(pswd)==0:
                        tg_pswd="0";
                    else:
                        tg_pswd="1";
                    markets.append([log,adres,inn,phone,'100',tg_id,tg_pswd]);
                conn.close();
                markets=arrayToString(markets);
                send="err=0,,markets={0}".format(markets);
            else:
                send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_excel_naks(request,var='network'):
    try:
        naks=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT id FROM 'nakSpace'");
                id_=cort_to_list(cursor.fetchall());
                for ID in id_:
                    cursor.execute("SELECT nak FROM 'nakSpace'WHERE id=(?)",(ID,));
                    nak=cort_to_list(cursor.fetchall())[0];
                    nak=nak.replace("|","a").replace("^","b");
                    cursor.execute("SELECT date FROM 'nakSpace'WHERE id=(?)",(ID,));
                    date=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT nakNum FROM 'nakSpace'WHERE id=(?)",(ID,));
                    nakNum=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT owner FROM 'nakSpace'WHERE id=(?)",(ID,));
                    owner=cort_to_list(cursor.fetchall())[0];
                    naks.append([ID,nak,date,nakNum,owner]);
                conn.close();
                naks=arrayToString(naks);
                send="err=0,,naks={0}".format(naks);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#admin do
def send_hist_var(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            last_index=request.GET['last_index'];
            vis=request.GET['var'];
            merchName=Merch(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            visOld="";
            if vis=="new" or vis=="old":
                cursor.execute("SELECT visible FROM 'history' WHERE last_index=(?)",(last_index,));
                visOld=cort_to_list(cursor.fetchall())[0];
            if visOld==vis:
                pass;
            else:
                if vis=="old" or vis=="new":
                    cursor.execute("UPDATE 'history' SET visible=(?)WHERE last_index=(?)",(vis,last_index,));
                elif vis=="clearAll":
                    cursor.execute("UPDATE 'history' SET visible='old'");
                elif vis=="choiseAllSell":
                    cursor.execute("UPDATE 'history' SET visible='new' WHERE type='sell'");
                elif vis=="choiseAllOrder":
                    cursor.execute("UPDATE 'history' SET visible='new' WHERE type='ord'");

                conn.commit();
            conn.close();
            send="err=0 text=OK";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_defaults(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                default_buyer=request.GET['default_buyer'];
                defDolgType=request.GET['defDolgType'];
                defDolgVal=request.GET['defDolgVal'];
                defDolgMax=request.GET['defDolgMax'];
                tg_id=request.GET['tg_id'];
                tg_token=request.GET['tg_token'].replace("|","");
                nots=request.GET['nots'];
                excelName=request.GET['excelName'];
                excelHow=request.GET['excelHow'];
                excelStartWrite=request.GET['excelStartWrite'];
                excelType=request.GET['excelType'];
                excelCat=request.GET['excelCat'];
                excelPriceName=request.GET['excelPriceName'];
                excelPrice=request.GET['excelPrice'];
                excelStartWritePrice=request.GET['excelStartWritePrice'];
                excelSheet=request.GET['excelSheet'];
                excelPriceSheet=request.GET['excelPriceSheet'];
                excelNakPlace=request.GET['excelNakPlace'];
                ostName=request.GET['ostName'];
                ostHowOst=request.GET['ostHowOst'];
                ostStart=request.GET['ostStart'];
                ostSheet=request.GET['ostSheet'];
                upRange=request.GET['upRange'];
                generalLang=request.GET['generalLang'];
                secondLang=request.GET['secondLang'];
                excelDriverName=request.GET['excelDriverName'];
                perInn=request.GET['perInn'];
                perOst=request.GET['perOst'];
                perStartRead=request.GET['perStartRead'];
                perSheet=request.GET['perSheet'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(default_buyer,"default_buyer",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(defDolgType,"defDolgType",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(defDolgVal,"defDolgVal",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(defDolgMax,"defDolgMax",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(tg_id,"tg_id",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(tg_token,"tg_token",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(nots,"nots",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelName,"excelName",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelHow,"excelHow",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelStartWrite,"excelStartWrite",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelType,"excelType",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelCat,"excelCat",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelPriceName,"excelPriceName",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelPrice,"excelPrice",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelStartWritePrice,"excelStartWritePrice",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelSheet,"excelNakSheet",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelPriceSheet,"excelProdSheet",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelNakPlace,"excelNakPlace",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(ostName,"ostName",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(ostHowOst,"ostHowOst",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(ostStart,"ostStart",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(ostSheet,"ostSheet",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(upRange,"upRange",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(generalLang,"generalLang",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(secondLang,"secondLang",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(excelDriverName,"exNakDriverName",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(perInn,"exPerInn",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(perOst,"exPerSum",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(perStartRead,"exPerStart",));
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(perSheet,"exPerSheet",));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Сессия истекла";
        else:
            send="err=1,,text=Вы не админ";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_site(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                site=request.GET['site'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(site,"site",));
                conn.commit();
                conn.close();
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Сессия истекла";
        else:
            send="err=1,,text=Вы не админ";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_prod_settings(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                #sendInfo=request.GET['sendInfo'];
                prod_id=request.GET['prod_id'];
                prod_name=request.GET['name'];
                cat_id=request.GET['cat_id'];
                rev=request.GET['rev'].replace("itsMySpace","\n");
                visible=request.GET['visible'];
                ost=request.GET['ost'];
                try:
                    name2=request.GET['name2'];
                    rev2=request.GET['rev2'];
                except:
                    name2=prod_name;
                    rev2=rev;
                print(rev)
                merchName=Merch(session);
                
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'products' SET name=(?) WHERE id=(?)",(prod_name,prod_id,));
                cursor.execute("UPDATE 'products' SET cat_id=(?) WHERE id=(?)",(cat_id,prod_id,));
                cursor.execute("UPDATE 'products' SET rev=(?) WHERE id=(?)",(rev,prod_id,));
                cursor.execute("UPDATE 'products' SET ost=(?) WHERE id=(?)",(ost,prod_id,));
                if ost=="0":
                    cursor.execute("UPDATE 'products' SET visible='0' WHERE id=(?)",(prod_id,));
                else:
                    cursor.execute("UPDATE 'products' SET visible='1' WHERE id=(?)",(prod_id,));
                cursor.execute("UPDATE 'products' SET visible=(?) WHERE id=(?)",(visible,prod_id,));
                
                cursor.execute("UPDATE 'products' SET name2=(?) WHERE id=(?)",(name2,prod_id,));
                cursor.execute("UPDATE 'products' SET rev2=(?) WHERE id=(?)",(rev2,prod_id,));
                print(rev);
                
                conn.commit();
                conn.close();
                #txt="Изменение в продуктах:\nПродукт:{0}\nЦены:{1}".format(prod_name,pr);
                txt="Изменение в продуктах:\nПродукт:{0}".format(prod_name);
                async2(Login(session),"getProducts");
                makeAct(merchName,Login(session),txt);
                send="err=0,, text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_cho_chosen_nak(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                driver=request.GET['driver'];
                NAK=request.GET['nak'];
                merchName=Merch(session);
                NAK=stringToArray(NAK)[0];
                conn1=sqlite3.connect(mPath(merchName,"orders"));
                cursor1=conn1.cursor();
                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                for NK in NAK:
                    cursor1.execute("SELECT nak FROM 'nakSpace' WHERE id=(?)",(NK,));
                    nak=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT nakNum FROM 'nakSpace' WHERE id=(?)",(NK,));
                    nakNum=cort_to_list(cursor1.fetchall())[0];
                    nak=stringToArray(nak);
                    for l in nak:
                        ids=l[0];
                        excelHowE=l[1];
                        cursor1.execute("SELECT name FROM 'products' WHERE id=(?)",(ids,));
                        excelNameE=cort_to_list(cursor1.fetchall())[0];

                        cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                        freeOld=cort_to_list(cursor.fetchall());
                        if len(freeOld)==0:
                            cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                                (excelNameE,ids,excelHowE,driver,));
                        else:
                            freeOld=freeOld[0];
                            newFree=str(int(excelHowE)+int(freeOld));
                            cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                                (newFree,driver,ids,));
                        ostMinus(merchName,ids,excelHowE);
                    cursor1.execute("DELETE FROM 'nakSpace'WHERE id=(?)",(NK,));
                    cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                        (driver,nakNum,date));
                conn1.commit();
                conn1.close();
                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
                admins=cort_to_list(cursor.fetchall());
                conn.commit();
                conn.close();
                for a in admins:
                    async2(a,"getDrivers");
                    async2(a,"getExcelNaks");
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def send_edit_prices(request,var="network"):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                prod_id=request.GET['prod_id'];
                type_=request.GET['type'];
                price=request.GET['price'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'prices' SET price=(?) WHERE id=(?)AND name=(?)",(price,prod_id,type_,));
                conn.commit();
                cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                prod_name=cort_to_list(cursor.fetchall())[0];
                conn.close();
                txt="Изменение в ценах:\nПродукт:{0}\nЦена:{1}".format(prod_name,price);
                makeAct(merchName,Login(session),txt);
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                driverName=request.GET['driverName'];
                driverPswd=request.GET['driverPswd'];
                driverVar=request.GET['driverVar'];
                driverOld=request.GET['driverOld'];
                login=Login(session);
                merchName=Merch(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'");
                login=cort_to_list(cursor.fetchall());
                if driverName in login and driverName!=driverOld:
                    send="err=1 text=Логин существует";
                else:
                    if driverVar=="new":
                        cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        (driverName,h(driverPswd),"_",merchName,"driver","0.01","0.01","0","ru",));
                        conn1=sqlite3.connect(mPath(merchName,"inCar"));
                        cursor1=conn1.cursor();
                        cursor1.execute("INSERT INTO cashInCar VALUES((?),(?),(?),(?),(?))",
                        (driverName,"0","0","0","0",));
                        conn1.commit();
                        conn1.close();
                        conn1=sqlite3.connect(mPath(merchName,"orders"));
                        cursor1=conn1.cursor();
                        cursor1.execute("INSERT INTO drivers VALUES((?),(?),(?))",(driverName,":;",""))
                        conn1.commit();
                        conn1.close();
                    else:
                        driverNakName=request.GET['driverNakName'];
                        cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(driverName,driverOld,));
                        if len(driverPswd)!=0:
                            cursor.execute("UPDATE 'users' SET pswd=(?)WHERE login=(?)",(h(driverPswd),driverName,));
                        conn1=sqlite3.connect(mPath(merchName,"inCar"));
                        cursor1=conn1.cursor();
                        cursor1.execute("UPDATE 'cashInCar' SET driver=(?)WHERE driver=(?)",(driverName,driverOld,));
                        cursor1.execute("UPDATE 'naks' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        cursor1.execute("UPDATE 'nakNum' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        conn1.commit();
                        conn1.close();
                        conn1=sqlite3.connect(mPath(merchName,"orders"));
                        cursor1=conn1.cursor();
                        cursor1.execute("UPDATE 'order' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        cursor1.execute("UPDATE 'drivers' SET login=(?)WHERE login=(?)",(driverName,driverOld));
                        cursor1.execute("UPDATE 'drivers' SET name=(?)WHERE login=(?)",(driverNakName,driverName));
                        cursor.execute("UPDATE 'users' SET session='_'WHERE login=(?)",(driverName,));
                        conn1.commit();
                        conn1.close();
                    send="err=0 text=OK";
                conn.commit();
                conn.close();
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                driverName=request.GET['driverName'];
                login=Login(session);
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT getter FROM 'order' WHERE get_type='dolg' AND owner=(?)",(driverName,));
                dolgs=len(cort_to_list(cursor.fetchall()));
                if dolgs==0:
                    cursor.execute("UPDATE 'order' SET get_type='new'WHERE owner=(?)",(driverName,));
                    cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)",(driverName,));
                    cursor.execute("DELETE FROM 'drivers' WHERE login = (?)",(driverName,));
                    conn.commit();
                    conn.close();
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(driverName,));
                    conn.commit();
                    conn.close();
                    conn=sqlite3.connect(mPath(merchName,"inCar"));
                    cursor=conn.cursor();
                    cursor.execute("SELECT prod_id FROM 'naks'");
                    ids=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT free FROM 'naks'");
                    free=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT reserve FROM 'naks'");
                    reserve=cort_to_list(cursor.fetchall());
                    for i in range(len(ids)):
                        prod_id=ids[i];
                        how_many=int(free[i])+int(reserve[i]);
                        ostPlus(merchName,prod_id,how_many);
                    cursor.execute("DELETE FROM 'cashInCar' WHERE driver = (?)",(driverName,));
                    cursor.execute("DELETE FROM 'nakNum' WHERE owner = (?)",(driverName,));
                    cursor.execute("DELETE FROM 'naks' WHERE owner = (?)",(driverName,));
                    conn.commit();
                    conn.close();
                    send="err=0 text=OK";
                    txt="Удален водитель {0}".format(driverName);
                    makeAct(merchName,Login(session),txt);
                    async2(Login(session),"getDrivers");
                    async2(Login(session),"getAdminOrders");
                else:
                    send="err=1 text=Остался долг";
                    conn.commit();
                    conn.close();
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_order(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                last_index=request.GET['last_index'];
                login=Login(session);
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT owner FROM 'order' WHERE last_index=(?)",(last_index,));
                driverName=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT getter FROM 'order' WHERE last_index=(?)",(last_index,));
                getter=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT data FROM 'order' WHERE last_index=(?)",(last_index,));
                data=cort_to_list(cursor.fetchall())[0];
                cursor.execute("UPDATE 'order' SET get_type='new'WHERE last_index=(?)",(last_index,));
                cursor.execute("UPDATE 'order' SET owner='new'WHERE last_index=(?)",(last_index,));
                conn.commit();
                conn.close();
                data=stringToArrayData(data);
                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                for i in range(len(data)):
                    prod_id=data[i][0];
                    how_many=data[i][1];
                    cursor.execute("SELECT reserve FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prod_id,));
                    reserveOld=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prod_id,));
                    freeOld=cort_to_list(cursor.fetchall())[0];
                    reserveNew=str(int(reserveOld)-int(how_many));
                    freeNew=str(int(freeOld)+int(how_many));
                    cursor.execute("UPDATE 'naks' SET reserve=(?)WHERE owner=(?)AND prod_id=(?)",(reserveNew,driverName,prod_id,));
                    cursor.execute("UPDATE 'naks' SET free=(?)WHERE owner=(?)AND prod_id=(?)",(freeNew,driverName,prod_id,));
                conn.commit();
                conn.close();
                async2(driverName,"getNak");
                async2(driverName,"getOrders");
                async2(driverName,"getPrices");
                async2(Login(session),"getDrivers");
                addNot(getter,merchName,last_index,"remMarketOrder");

                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_give_him_order(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                last_index=request.GET['last_index'];
                driverName=request.GET['driverName'];
                login=Login(session);
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT data FROM 'order' WHERE last_index=(?)",(last_index,));
                data=stringToArrayData(cort_to_list(cursor.fetchall())[0]);
                conn.close();

                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                
                for i in range(len(data)):
                    prodId=data[i][0];
                    how_many=data[i][1];
                    prod=prodById(merchName,prodId);
                    try:
                        cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prodId,));
                        free=cort_to_list(cursor.fetchall())[0];
                        cursor.execute("SELECT reserve FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prodId,));
                        reserve=cort_to_list(cursor.fetchall())[0];
                    except:
                        send="err=1,,text=Не хватает "+prod+"(есть 0. Нужно "+how_many+")";
                        return HttpResponse(send, content_type='application/json');
                    if int(how_many)>int(free):
                        conn.close();
                        send="err=1,,text=Не хватает "+prod+"(есть "+free+". Нужно "+how_many+")";
                        async2(Login(session),"getDrivers");
                        conn.close();
                        return HttpResponse(send, content_type='application/json');
                    else:
                        newFree=int(free)-int(how_many);
                        newReserve=int(reserve)+int(how_many);
                        cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?) AND prod_id=(?)",
                            (newFree,driverName,prodId,));
                        cursor.execute("UPDATE 'naks' SET reserve=(?) WHERE owner=(?) AND prod_id=(?)",
                            (newReserve,driverName,prodId,));
                        #nakNum
                        cursor.execute("SELECT num FROM 'nakNum' WHERE owner=(?)",(driverName,));
                        num=cort_to_list(cursor.fetchall());
                        cursor.execute("SELECT date FROM 'nakNum' WHERE owner=(?)",(driverName,));
                        date=cort_to_list(cursor.fetchall());
                        nakNum="";
                        for i in range(len(num)):
                            nakNum+="Накладная №"+num[i]+"\nДата: "+date[i]+"\n";
                txt="Заказ№{2} отдан водителю: {0}\nЗадействованы накладные:\n{1}".format(driverName,nakNum,last_index);
                messageInChannel(merchName,txt);
                conn.commit();
                conn.close();

                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'order' SET owner=(?) WHERE last_index=(?)",(driverName,last_index,));
                cursor.execute("UPDATE 'order' SET get_type=(?) WHERE last_index=(?)",("ord",last_index,));
                conn.commit();
                conn.close();

                #async2(driverName,"newOrder");
                addNot(driverName,"title",last_index,"driverOrder");

                async2(driverName,"getNak");
                async2(driverName,"getOrders");
                async2(driverName,"getPrices");
                async2(login,"getDrivers");
                txt="Заказ №{0} отдан водителю {1}".format(last_index,driverName);
                makeAct(merchName,login,txt);
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def refresh_gen_link(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                code=generate();
                newLink=URL+"excel_space/?code="+code;
                newLink=newLink.replace(":","}");
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE const SET val=(?) WHERE key=(?)",(newLink,"generalLink",));
                conn.commit();
                conn.close();
                send="err=0,,text={0}".format(newLink);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def send_remove_wrong_order(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                last_index=request.GET['last_index'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT data FROM 'order' WHERE last_index=(?)",(last_index,));
                data=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT getter FROM 'order' WHERE last_index=(?)",(last_index,));
                getter=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT price FROM 'order' WHERE last_index=(?)",(last_index,));
                price=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT price FROM 'order' WHERE last_index=(?)",(last_index,));
                price=cort_to_list(cursor.fetchall())[0];

                cursor.execute("DELETE FROM 'order' WHERE last_index = (?)",(last_index,));
                conn.commit();
                conn.close();
                if "USER_" in getter:
                    removeUser(merchName,getter);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                (merchName,));
                admins=cort_to_list(cursor.fetchall());
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getAllPrices");
                conn.close();
                txt="Удалён заказ№{0}".format(last_index);
                makeAct(merchName,Login(session),txt);

                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                cursor.execute("INSERT INTO history VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                    (data.replace("|",":").replace("^",";"),str(price),"ord",getter,"null",date,"district",last_index,"new"));
                conn.commit();
                conn.close();

                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_market_settings(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                marketName=request.GET['marketName'];
                clientPriceType=request.GET['clientPriceType'];
                dolgType=request.GET['dolgType'];
                dolgVal=request.GET['dolgVal'];
                wallet=request.GET['wallet'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE buy_markets SET type=(?) WHERE name=(?)",
                    (clientPriceType,marketName));
                cursor.execute("UPDATE buy_markets SET dolgType=(?) WHERE name=(?)",
                    (dolgType,marketName));
                cursor.execute("UPDATE buy_markets SET dolgVal=(?) WHERE name=(?)",
                    (dolgVal,marketName));
                cursor.execute("UPDATE buy_markets SET wallet=(?) WHERE name=(?)",
                    (wallet,marketName));

                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                    (merchName,));
                admins=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                    (merchName,));
                logins=cort_to_list(cursor.fetchall());
                conn.close();
                for a in admins:
                    async2(a,"getMarketList");
                for l in logins:
                    async2(l,"getMarketList");
                txt="Изменены настройка магазина:\n"+marketName+"\n"+dolgType+"\n"+dolgVal;
                makeAct(merchName,Login(session),txt);
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_new_price_type_name(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                oldPriceName=request.GET['oldPriceName'];
                priceName=request.GET['priceName'];
                merchName=Merch(session);

                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE buy_markets SET type=(?) WHERE type=(?)",
                    (priceName,oldPriceName));
                cursor.execute("UPDATE buyers_type SET name=(?) WHERE name=(?)",
                    (priceName,oldPriceName));
                cursor.execute("UPDATE const SET val=(?) WHERE val=(?) AND key='default_buyer'",
                    (priceName,oldPriceName));
                cursor.execute("UPDATE prices SET name=(?) WHERE name=(?)",
                    (priceName,oldPriceName));
                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                    (merchName,));
                admins=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                    (merchName,));
                logins=cort_to_list(cursor.fetchall());
                conn.close();
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getProducts");
                    async2(a,"getAllPrices");
                    async2(a,"getMarketList");
                for l in logins:
                    async2(l,"getNak");
                    async2(l,"getOrders");
                    async2(l,"getPrices");
                    async2(l,"getProductList");
                    async2(l,"getDefaultMarket");
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_rem_price_type_name(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                oldPriceName=request.GET['oldPriceName'];
                if oldPriceName =="user":
                    send="err=1,,text=Нельзя";
                    return HttpResponse(send, content_type='application/json')
                else:
                    merchName=Merch(session);
                    conn=sqlite3.connect(mPath(merchName,"orders"));
                    cursor=conn.cursor();
                    cursor.execute("UPDATE buy_markets SET type='user' WHERE type=(?)",
                        (oldPriceName,));
                    cursor.execute("DELETE FROM 'buyers_type' WHERE name = (?)",
                        (oldPriceName,));
                    cursor.execute("UPDATE 'const' SET val='user' WHERE key='default_buyer' AND val=(?)",
                        (oldPriceName,));
                    cursor.execute("DELETE FROM 'prices' WHERE name = (?)",
                        (oldPriceName,));
                    conn.commit();
                    conn.close();
                    txt="Удален тип цены:\n"+oldPriceName;
                    makeAct(merchName,Login(session),txt);
                    conn=sqlite3.connect("basic.sqlite");
                    cursor=conn.cursor();
                    cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                        (merchName,));
                    admins=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                        (merchName,));
                    logins=cort_to_list(cursor.fetchall());
                    conn.close();
                    for a in admins:
                        async2(a,"getAdminOrders");
                        async2(a,"getProducts");
                        async2(a,"getAllPrices");
                    for l in logins:
                        async2(l,"getNak");
                        async2(l,"getOrders");
                        async2(l,"getPrices");
                        async2(l,"getProductList");
                        async2(l,"getDefaultMarket");
                    send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_create_price_type(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                oldPriceName=request.GET['oldPriceName'];
                if oldPriceName=="user":
                    send="err=1,,text=Нельзя";
                    return HttpResponse(send, content_type='application/json')
                else:
                    merchName=Merch(session);
                    conn=sqlite3.connect(mPath(merchName,"orders"));
                    cursor=conn.cursor();
                    cursor.execute("SELECT name FROM 'buyers_type'");
                    myPriceTypes=cort_to_list(cursor.fetchall());
                    if oldPriceName not in myPriceTypes:
                        cursor.execute("INSERT INTO buyers_type VALUES((?))",(oldPriceName,));
                        cursor.execute("SELECT id FROM 'prices' WHERE name='user'");
                        id_=cort_to_list(cursor.fetchall());
                        cursor.execute("SELECT price FROM 'prices' WHERE name='user'");
                        price=cort_to_list(cursor.fetchall());
                        for i in range(len(id_)):
                            cursor.execute("INSERT INTO 'prices' VALUES((?),(?),(?))",
                                (id_[i],oldPriceName,price[i]));
                        conn.commit();
                        conn.close();

                        conn=sqlite3.connect("basic.sqlite");
                        cursor=conn.cursor();
                        cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                            (merchName,));
                        admins=cort_to_list(cursor.fetchall());
                        cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                            (merchName,));
                        logins=cort_to_list(cursor.fetchall());
                        conn.close();
                        for a in admins:
                            async2(a,"getAdminOrders");
                            async2(a,"getProducts");
                            async2(a,"getAllPrices");
                            async2(a,"getDefaultMarket");
                        for l in logins:
                            async2(l,"getNak");
                            async2(l,"getOrders");
                            async2(l,"getPrices");
                            async2(l,"getProductList");
                            async2(l,"getDefaultMarket");
                        send="err=0,,text=OK";
                    else:
                        send="err=0,,text=Такой тип цены уже существует";

            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_req_for_excel(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                exType=request.GET['exType'];
                merchName=Merch(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT code FROM 'exLink'");
                codes=cort_to_list(cursor.fetchall());
                code=generate();
                while code in codes:
                    code=generate();
                cursor.execute("INSERT INTO 'exLink' VALUES((?),(?),(?))",(code,exType,merchName));
                conn.commit();
                conn.close();
                link=URL+"get_excel/?code="+code;
                send="err=0,,link={0}".format(link);
                txt="Создана ссылка на отправку Excel:\nТип ссылки: {0}\nКод: {1}".format(exType,code);
                makeAct(merchName,Login(session),txt);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def get_excel(request):
    try:
        code=request.GET['code'];
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT code FROM 'exLink'");
        codes=cort_to_list(cursor.fetchall());
        if code in codes:
            cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
            _type=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
            merchName=cort_to_list(cursor.fetchall())[0];
            if _type=="prices" or _type=="ost" or _type=="per":
                html="html/prices.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL)
                htmlR=htmlR.replace("_VAL_",code)
                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                conn.close();
            else:
                html="html/nak.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL);
                htmlR=htmlR.replace("_VAL_",code);
                cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='driver'",(merchName,));
                logins=cort_to_list(cursor.fetchall());
                data="";
                for l in logins:
                    data+='\n<option value="{0}">{0}</option>'.format(l);
                print(data)
                htmlR=htmlR.replace("_OPTIONS_",data);
                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                conn.close();
            return render(request,htmlL);
        else:
            conn.close();
            send="err=1,,text=Кода не существует";
            return render(request,send);
    except Exception as e:
        logger(e);
        send="err=1,,text=Кода не существует";
        return render(request,errHtml("Недействительная ссылка"));
        #return HttpResponse(send, content_type='application/json')
def excel_space(request):
    try:
        merchs=os.listdir("merchants");
        code=request.GET['code'];
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                html="html/nakSpace.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL);
                htmlR=htmlR.replace("_VAL_",code);

                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                print(m);
                return render(request,htmlL);
        return render(request,errHtml("Недействительная ссылка"));
    except Exception as e:
        logger(e);
        send="err=1,,text=Кода не существует";
        return render(request,errHtml("Недействительная ссылка"));
        #return HttpResponse(send, content_type='application/json')
def save_excel_space(request):
    try:
        code=request.POST['code'];
        merchs=os.listdir("merchants");
        workNot=False;
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                workNot=True;
                merchName=m;
                break;
        if workNot:
            conn1=sqlite3.connect(mPath(merchName,"inCar"));
            cursor1=conn1.cursor();

            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT name FROM 'products'");
            prods=cort_to_list(cursor.fetchall());

            j=1;
            allFiles=request.FILES.getlist('nakFile');
            nakList=[];
            excelNakPlaceList=[];
            errorList=[];
            niceList=[];
            for f in allFiles:
                nakList=[];

                excelName=getConst(merchName,'excelName');
                excelHow=getConst(merchName,'excelHow');
                excelStartWrite=getConst(merchName,'excelStartWrite');
                excelNakSheet=getConst(merchName,'excelNakSheet');
                excelNakPlace=getConst(merchName,'excelNakPlace');

                exNakDriverName=getConst(merchName,'exNakDriverName');

                
                Path="merchants/{0}/xlsx/".format(merchName);
                #handle_uploaded_file(request.FILES['nakFile'],Path,code);
                try:
                    os.remove(Path+code+".xlsx");
                except:
                    pass;
                handle_uploaded_file(f,Path,code);
                error=False;
                try:
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                except:
                    reW.re(Path+code+".xlsx");
                    print("rewrite!<<<<<")
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                try:
                    sheet = wb[excelNakSheet];
                except Exception as e:
                    logger(e);
                    error=True;
                    #return render(request,errHtml("Файл: "+f.name+". \nНе найдена страница"));
                if not error:
                    i=int(excelStartWrite);
                    #nakList=[];
                    excelNakPlace = str(sheet[excelNakPlace].value);
                    exNakDriverName = str(sheet[exNakDriverName].value);
                    cursor.execute("SELECT login FROM 'drivers'WHERE name=(?)",(exNakDriverName,));
                    print(exNakDriverName);
                    try:
                        drivers=cort_to_list(cursor.fetchall())[0];
                    except:
                        drivers="null"
                    print(drivers);
                    cursor1.execute("SELECT num FROM 'nakNum'WHERE owner=(?)",(drivers,));
                    nums=cort_to_list(cursor1.fetchall());
                    if excelNakPlace=="None":
                        return render(request,errHtml("Номер накладной не найден."));
                    while True:
                        excelNameE = str(sheet[excelName+str(i)].value);
                        excelHowE = str(sheet[excelHow+str(i)].value);
                        if excelNameE=="None" or excelHowE=="None":
                            try:
                                os.remove("templates/{0}.html".format(code));
                            except:
                                pass;
                            print("Накладная №"+excelNakPlace);
                            for pp in nakList:
                                print(pp);
                            excelNakPlaceList.append(excelNakPlace);
                            break;
                        else:
                            try:
                                int(excelHowE)+1;
                            except:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Неккоректное количество. Строка "+str(i)+" Указанное колличество: \""+excelHowE+"\"";
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Неккоректное количество. Строка "+str(i)+" Указанное колличество: \""+excelHowE+"\""));
                            if excelNameE not in prods:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Продукта ("+excelNameE+") не существует. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Продукта ("+excelNameE+") не существует. Строка "+str(i)));
                            elif excelNameE=="None":
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Поле наименование пустое. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Поле наименование пустое. Строка "+str(i)));
                            elif excelHowE=="None":
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Поле количества пустое. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+"Поле количества пустое.Строка "+str(i)));

                            elif excelNakPlace in nums:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+". Накладная с таким номером уже есть у этого водителя и поэтому проигнорирована";
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+"Поле количества пустое.Строка "+str(i)));

                            else:
                                error=False;
                                if f.name not in niceList:
                                    niceList.append(f.name);
                                nakList.append(
                                    {"excelNameE":excelNameE,
                                    "excelHowE":excelHowE,
                                    "excelNum":excelNakPlace
                                    });
                        i+=1;
                else:
                    errorList.append("Файл: "+f.name+". => Не найдена страница");

                if not error:
                    writeExcelSpace(request,nakList,code,drivers,excelNakPlaceList);
                excelNakPlaceList.clear();
                #j+=1;
            conn.close();
            conn1.close();
            print(excelNakPlaceList)
            print("Общая")
            for pp in nakList:
                print(pp);
            print("Оконченно");
            drivers="null"
            txt="Ошибки произошли в файлах:";
            for t in errorList:
                txt+="<br>"+t;
            
            txt+="<br><br>Успешно записаны:";
            for t in niceList:
                txt+="<br>"+t;
            
            if len(errorList)==0:
                return render(request,doneHtml("Накладные записаны"));
            else:
                return render(request,errHtml(txt));

            #return render(request,doneHtml("Накладная записана"));
            
            #return writeExcelSpace(request,nakList,code,drivers,excelNakPlaceList);
        else:
            return render(request,errHtml("Недействительная ссылка"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def writeExcelSpace(request,exList,code,driver="null",nak="null"):
    try:
        merchs=os.listdir("merchants");
        workNot=False;
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                workNot=True;
                merchName=m;
                break;
        if workNot:
            if driver=="null":
                conn1=sqlite3.connect(mPath(merchName,"orders"));
                cursor1=conn1.cursor();
                NK="";
                for l in exList:
                    excelNameE=l["excelNameE"];
                    excelHowE=l["excelHowE"];
                    excelNum=l["excelNum"];
                    
                    cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                    ids=cort_to_list(cursor1.fetchall())[0];
                    NK=NK+ids+"|"+excelHowE+"|^";
                
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                cursor1.execute("SELECT id FROM 'nakSpace'");
                idsS=cort_to_list(cursor1.fetchall());
                id_=1;
                while str(id_) in idsS:
                    id_+=1;
                cursor1.execute("INSERT INTO 'nakSpace' VALUES((?),(?),(?),(?),(?))",
                        (str(id_),NK,date,excelNum,"_",));
                conn1.commit();
                conn1.close();
            else:
                conn1=sqlite3.connect(mPath(merchName,"orders"));
                cursor1=conn1.cursor();
                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                for l in exList:
                    excelNameE=l["excelNameE"];
                    excelHowE=l["excelHowE"];
                    cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                    ids=cort_to_list(cursor1.fetchall())[0];
                    cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                    freeOld=cort_to_list(cursor.fetchall());
                    if len(freeOld)==0:
                        cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                            (excelNameE,ids,excelHowE,driver,));
                    else:
                        freeOld=freeOld[0];
                        newFree=str(int(excelHowE)+int(freeOld));
                        cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                            (newFree,driver,ids,));
                    ostMinus(merchName,ids,excelHowE);
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                print("<<<<<<<<")
                for n in nak:
                    print(n+"<<<");
                    cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                        (driver,n,date));
                print("<<<<<<<<")
                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
                cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
                admins=cort_to_list(cursor.fetchall());
                conn.commit();
                conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
            admins=cort_to_list(cursor.fetchall());
            conn.commit();
            conn.close();
            for a in admins:
                async2(a,"getDrivers");
                async2(a,"getExcelNaks");
            return render(request,doneHtml("Накладная записана"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def get_chat_admins(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'WHERE user_type='admin'");
            login=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT merchName FROM 'users'WHERE user_type='admin'");
            merchName1=cort_to_list(cursor.fetchall());
            conn.close();
            sendList="";
            for i in range(len(login)):
                l=login[i];
                conn1=sqlite3.connect(mPath(merchName1[i],"orders"));
                cursor1=conn1.cursor();
                cursor1.execute("SELECT role FROM 'admins'WHERE login=(?)",(l,));
                role=cort_to_list(cursor1.fetchall())[0];
                conn1.close();
                if role==None:
                    role="";
                sendList=sendList+login[i]+"|"+merchName1[i]+"|"+role+"|^"
            send="err=0,,admins={0}".format(sendList);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_messages(request,var='network',Chat_id='null'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT chat_id FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            chat_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT message_id FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            message_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT text_or_photo FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            text_or_photo=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT sender FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            sender=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT wasRead FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            wasRead=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT text FROM 'messages'WHERE sender=(?)OR getter=(?)",(login,login,));
            text=cort_to_list(cursor.fetchall());
            conn.close();
            sendList="";
            for i in range(len(chat_id)):
                sendList=sendList+chat_id[i]+"|"+message_id[i]+"|"+text_or_photo[i]+"|"+sender[i]+"|"+getter[i]+"|"+date[i]+"|"+wasRead[i]+"|"+text[i]+"|^"
            send="err=0,,messages={0}".format(sendList);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    #print(send)
    if Chat_id!="null":
        send+=Chat_id;
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_readed(request,var='network'):
    try:
        session=request.GET['session'];
        readed=request.GET['readed'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            readed=stringToArray(readed);
            WR_iteration="0";
            for i in range(len(readed)):
                message_id=readed[i][0];
                cursor.execute("SELECT wasRead FROM 'messages'WHERE message_id=(?)",(message_id,));
                WR=cort_to_list(cursor.fetchall())[0];
                if WR=="0":
                    WR_iteration=str(int(WR_iteration)+1);
                    cursor.execute("UPDATE 'messages' SET wasRead='1' WHERE message_id=(?)",(message_id,));
                cursor.execute("SELECT sender FROM 'messages'WHERE message_id=(?)",(message_id,));
                sender=cort_to_list(cursor.fetchall())[0];
            conn.commit();
            conn.close();
            #if WR_iteration!="0":
            try:
                async2(sender,"getMessages");
            except:
                pass
            send="err=0,,text=OK";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_message_to_server(request,var='network'):
    try:
        session=request.GET['session'];
        chat_id=request.GET['chat_id'];
        sender=request.GET['sender'];
        getter=request.GET['getter'];
        text=request.GET['text'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT chat_id FROM messages");
            chat_ids=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT message_id FROM messages");
            message_ids=cort_to_list(cursor.fetchall());
            if "|^" in chat_id:
                getter=stringToArray(chat_id)[0][1];
                newChatId=1;
                while True:
                    if str(newChatId) in chat_ids:
                        newChatId+=1;
                    else:
                        break;
                chat_id=newChatId;
            newMessageId=1;
            while True:
                if str(newMessageId) in message_ids:
                    newMessageId+=1;
                else:
                    break;
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            cursor.execute("INSERT INTO 'messages' VALUES((?),(?),(?),(?),(?),(?),(?),(?))",
                (str(chat_id),str(newMessageId),"text",sender,getter,date,"0",text));
            conn.commit();
            conn.close();
            send=get_messages(request,'network',",,chat_id="+str(chat_id));
            async2(getter,"getMessages");
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def mass_page(request):
    html="templates/test.html";
    return render(request,html);
def test_mass(request):
    Path="html/"
    i=1
    for f in request.FILES.getlist('test_mass'):
        handle_uploaded_file(f,Path,"name"+str(i));
        i+=1;
    html="templates/test.html";
    return render(request,html);

def save_excel(request):
    try:
        code=request.POST['code'];
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT code FROM 'exLink'");
        codes=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
        try:
            _type=cort_to_list(cursor.fetchall())[0];
        except:
            return render(request,errHtml("Недействительная ссылка"));
        cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
        merchName=cort_to_list(cursor.fetchall())[0];
        cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
        drivers=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
        admins=cort_to_list(cursor.fetchall());
        conn.close();

        conn=sqlite3.connect(mPath(merchName,"orders"));
        cursor=conn.cursor();
        cursor.execute("SELECT name FROM 'buyers_type'");
        buyers_type=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM 'categories'");
        cats=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM 'products'");
        prods=cort_to_list(cursor.fetchall());
        conn.close();
        if _type=="per":
            exPerInn=getConst(merchName,'exPerInn');
            exPerSum=getConst(merchName,'exPerSum');
            exPerStart=getConst(merchName,'exPerStart');
            exPerSheet=getConst(merchName,'exPerSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[exPerSheet];
            except Exception as e:
                print(e);
                logger(e);
                return render(request,errHtml("Не найдена страница"));

            i=int(exPerStart);
            
            inns={};
            while True:
                inn = str(sheet[exPerInn+str(i)].value);
                summ = str(sheet[exPerSum+str(i)].value);
                if str(inn)=="None" and str(summ)=="None":
                    break;
                if str(inn)=="Итого"or str(inn)=="None":
                    break;
                if str(summ)=="None":
                    summ="0"
                inn=clearInn(inn);
                try:
                    inn=inn[-9:]
                except:
                    pass;
                if str(inn)=="":
                    inn="0"
                if inn in inns:
                    inns[inn]=str(int(inns[inn])+int(summ));
                else:
                    inns[inn]=summ;
                i+=1;
            print(inns)
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            for INN in inns:
                LGS=loginsByInn(INN);
                for l in LGS:
                    cursor.execute("UPDATE 'buy_markets' SET wallet=(?) WHERE name=(?)",(inns[INN],l,));
            conn.commit();
            conn.close();
            return render(request,doneHtml("Остатки по перечислению записаны"));
        if _type=="ost":
            ostName=getConst(merchName,'ostName');
            ostHowOst=getConst(merchName,'ostHowOst');
            ostStart=getConst(merchName,'ostStart');
            ostSheet=getConst(merchName,'ostSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[ostSheet];
            except Exception as e:
                print(e);
                logger(e);
                return render(request,errHtml("Не найдена страница"));

            i=int(ostStart);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            names=[];
            while True:
                name = str(sheet[ostName+str(i)].value);
                howOst = str(sheet[ostHowOst+str(i)].value);
                names.append(name);
                if name=="Итого":
                    break;
                if howOst=="None":
                    howOst="0";
                cursor.execute("SELECT id FROM 'products' WHERE name=(?)",(name,));
                if name=="None" and howOst=="None":
                    break;
                try:
                    prod_id=cort_to_list(cursor.fetchall())[0];
                except Exception as e:
                    logger(e);
                    return render(request,errHtml("Продукт не найден.\nСтрока "+str(i)+" ("+name+")"));
                howOst=howOst.replace(",000","");
                try:
                    int(howOst)+1;
                except:
                    return render(request,errHtml("Неккоректное количество остатка\nСтрока "+str(i)));

                cursor.execute("UPDATE 'products' SET ost=(?) WHERE id=(?)",(howOst,prod_id,));
                if howOst=="0":
                    cursor.execute("UPDATE 'products' SET visible='0' WHERE id=(?)",(prod_id,));
                else:
                    cursor.execute("UPDATE 'products' SET visible='1' WHERE id=(?)",(prod_id,));
                i+=1;
            cursor.execute("SELECT name FROM 'products'");
            allNames=cort_to_list(cursor.fetchall());
            for i in range(len(allNames)):
                if allNames[i] not in names:
                    cursor.execute("UPDATE 'products' SET ost='0' WHERE name=(?)",(allNames[i],));
                    cursor.execute("UPDATE 'products' SET ost='0' WHERE name=(?)",(allNames[i],));

            conn.commit();
            conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
            conn.commit();
            conn.close();
            for a in admins:
                async2(a,"getProducts");
            return render(request,doneHtml("Остатки записаны"));






        elif _type=="prices":
            excelType=getConst(merchName,'excelType');
            excelCat=getConst(merchName,'excelCat');
            excelPriceName=getConst(merchName,'excelPriceName');
            excelPrice=getConst(merchName,'excelPrice');
            excelStartWritePrice=getConst(merchName,'excelStartWritePrice');
            excelProdSheet=getConst(merchName,'excelProdSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[excelProdSheet];
            except Exception as e:
                print(e);
                logger(e);
                return render(request,errHtml("Не найдена страница"));
            i=int(excelStartWritePrice);
            priceList=[];
            while True:
                excelTypeE = str(sheet[excelType+str(i)].value);
                excelCatE = str(sheet[excelCat+str(i)].value);
                excelPriceNameE = str(sheet[excelPriceName+str(i)].value);
                excelPriceE = str(sheet[excelPrice+str(i)].value);
                if excelTypeE=="None" and excelCatE=="None" and excelPriceNameE=="None":
                    try:
                        os.remove("templates/{0}.html".format(code));
                    except:
                        pass;
                    for pp in priceList:
                        print(pp);
                    for a in admins:
                        async2(a,"getProducts");
                        async2(a,"getAllPrices");
                    for d in drivers:
                        async2(d,"getNak");
                        async2(d,"getPrices");
                    return writeExcel(request,priceList,code);
                    
                else:
                    try:
                        int(excelPriceE)+1;
                    except:
                        return render(request,errHtml("Неккоректная цена.\nСтрока "+str(i)));
                    if excelCatE not in cats:
                        return render(request,errHtml("Категории не существует.\nСтрока "+str(i)));
                    elif excelTypeE not in buyers_type:
                        return render(request,errHtml("Типа цены не существует.\nСтрока "+str(i)));
                    elif excelTypeE=="None":
                        return render(request,errHtml("Поле типа покупателя пустое.\nСтрока "+str(i)));
                    elif excelCatE=="None":
                        return render(request,errHtml("Поле категории пустое.\nСтрока "+str(i)));
                    elif excelPriceNameE=="None":
                        return render(request,errHtml("Поле типа цены пустое.\nСтрока "+str(i)));
                    elif excelPriceE=="None":
                        return render(request,errHtml("Поле цены пустое.\nСтрока "+str(i)));
                    else:
                        priceList.append(
                            {"excelTypeE":excelTypeE,
                            "excelCatE":excelCatE,
                            "excelPriceNameE":excelPriceNameE,
                            "excelPriceE":excelPriceE});
                i+=1;
            return writeExcel(request,priceList,code);

        else:
            j=1;
            allFiles=request.FILES.getlist('nakFile');
            nakList=[];
            excelNakPlaceList=[];
            for f in allFiles:

                excelName=getConst(merchName,'excelName');
                excelHow=getConst(merchName,'excelHow');
                excelStartWrite=getConst(merchName,'excelStartWrite');
                excelNakSheet=getConst(merchName,'excelNakSheet');
                excelNakPlace=getConst(merchName,'excelNakPlace');
                drivers=request.POST['drivers'];
                Path="merchants/{0}/xlsx/".format(merchName);
                #handle_uploaded_file(request.FILES['nakFile'],Path,code);
                try:
                    os.remove(Path+code+".xlsx");
                except:
                    pass;
                handle_uploaded_file(f,Path,code);
                try:
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                except:
                    reW.re(Path+code+".xlsx");
                    print("rewrite!<<<<<")
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                try:
                    sheet = wb[excelNakSheet];
                except Exception as e:
                    logger(e);
                    return render(request,errHtml("Не найдена страница"));
                i=int(excelStartWrite);
                #nakList=[];
                excelNakPlace = str(sheet[excelNakPlace].value);
                if excelNakPlace=="None":
                    return render(request,errHtml("Номер накладной не найден."));
                while True:
                    excelNameE = str(sheet[excelName+str(i)].value);
                    excelHowE = str(sheet[excelHow+str(i)].value);
                    if excelNameE=="None" or excelHowE=="None":
                        try:
                            os.remove("templates/{0}.html".format(code));
                        except:
                            pass;
                        print("Накладная №"+excelNakPlace);
                        for pp in nakList:
                            print(pp);
                        excelNakPlaceList.append(excelNakPlace);
                        break;

                        #if j==len(allFiles):
                        #    for a in admins:
                        #        async2(a,"getDrivers");
                        #    for d in drivers:
                        #        async2(d,"getNak");
                        #        async2(d,"getPrices");
                        #    print("Оконченно");
                        #    return writeExcel(request,nakList,code,drivers,excelNakPlace);
                        #else:
                        #    writeExcel(request,nakList,code,drivers,excelNakPlace);
                        #    break;
                    else:
                        try:
                            int(excelHowE)+1;
                        except:
                            return render(request,errHtml("Неккоректное количество.\nСтрока "+str(i)+"\nУказанное колличество: \""+excelHowE+"\""));
                        if excelNameE not in prods:
                            return render(request,errHtml("Продукта ("+excelNameE+") не существует.\nСтрока "+str(i)));
                        elif excelNameE=="None":
                            return render(request,errHtml("Поле наименование пустое.\nСтрока "+str(i)));
                        elif excelHowE=="None":
                            return render(request,errHtml("Поле количества пустое.\nСтрока "+str(i)));
                        else:
                            nakList.append(
                                {"excelNameE":excelNameE,
                                "excelHowE":excelHowE});
                    i+=1;
                j+=1;
            for a in admins:
                async2(a,"getDrivers");
            for d in drivers:
                async2(d,"getNak");
                async2(d,"getPrices");
            print(excelNakPlaceList)
            print("Общая")
            for pp in nakList:
                print(pp);
            print("Оконченно");
            return writeExcel(request,nakList,code,drivers,excelNakPlaceList);
            #return HttpResponse("OK", content_type='application/json');
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def errHtml(text):
    file=open("html/mistake.html","r");
    htmlR=file.read();
    file.close();
    htmlR=htmlR.replace("_TEXT_",text)
    file=open("html/err/resp.html","w");
    file.write(htmlR);
    file.close();
    return "html/err/resp.html";
def doneHtml(text):
    file=open("html/done.html","r");
    htmlR=file.read();
    file.close();
    htmlR=htmlR.replace("_TEXT_",text)
    file=open("html/err/"+text+".html","w");
    file.write(htmlR);
    file.close();
    return "html/err/"+text+".html";
def writeExcel(request,exList,code,driver="null",nak="null"):
    try:
        
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
        _type=cort_to_list(cursor.fetchall())[0];
        cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
        merchName=cort_to_list(cursor.fetchall())[0];
        conn.close();
        if _type=="prices":
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            for l in exList:
                excelTypeE=l["excelTypeE"];
                excelCatE=l["excelCatE"];
                excelPriceNameE=l["excelPriceNameE"];
                excelPriceE=l["excelPriceE"];
                cursor.execute("SELECT id FROM 'products' WHERE name=(?)",(excelPriceNameE,));
                ids=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT id FROM 'categories' WHERE name=(?)",(excelCatE,));
                cat_id=cort_to_list(cursor.fetchall())[0];
                if len(ids)==0:
                    ids=1;
                    cursor.execute("SELECT id FROM 'products'");
                    allId=cort_to_list(cursor.fetchall());
                    while str(ids) in allId:
                        ids+=1;
                    cursor.execute("INSERT INTO 'products' VALUES((?),(?),(?),'_','1','_','1','sh',(?),(?),(?),(?))",
                        (cat_id,str(ids),excelPriceNameE,"1","0",excelPriceNameE,"_"));
                    cursor.execute("SELECT name FROM 'buyers_type'");
                    bType=cort_to_list(cursor.fetchall());
                    for b in bType:
                        cursor.execute("INSERT INTO 'prices' VALUES((?),(?),(?))",
                            (str(ids),b,excelPriceE));
                else:
                    ids=ids[0];
                    cursor.execute("UPDATE 'products' SET work='1' WHERE id=(?)",(ids,));
                    cursor.execute("UPDATE 'products' SET cat_id=(?) WHERE id=(?)",(cat_id,ids,));
                    cursor.execute("UPDATE 'prices' SET price=(?) WHERE id=(?)AND name=(?)",
                        (excelPriceE,ids,excelTypeE,));
            conn.commit();
            conn.close();
            return render(request,doneHtml("Справочник записан"));
        else:
            conn1=sqlite3.connect(mPath(merchName,"orders"));
            cursor1=conn1.cursor();
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            for l in exList:
                excelNameE=l["excelNameE"];
                excelHowE=l["excelHowE"];
                cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                ids=cort_to_list(cursor1.fetchall())[0];
                cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                freeOld=cort_to_list(cursor.fetchall());
                if len(freeOld)==0:
                    cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                        (excelNameE,ids,excelHowE,driver,));
                else:
                    freeOld=freeOld[0];
                    newFree=str(int(excelHowE)+int(freeOld));
                    cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                        (newFree,driver,ids,));
                ostMinus(merchName,ids,excelHowE);
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            for n in nak:
                cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                    (driver,n,date));
            conn.commit();
            conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
            cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
            admins=cort_to_list(cursor.fetchall());
            conn.commit();
            conn.close();
            for a in admins:
                async2(a,"getDrivers");
            return render(request,doneHtml("Накладная записана"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def send_url_in_tg(request):
    try:
        session=request.GET['session'];
        url=request.GET['url'];
        messageInChannel(Merch(session),url)
        return HttpResponse("OK", content_type='application/json');
    except Exception as e:
        logger(e);
def send_new_admin(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                adminName=request.GET['adminName'];
                adminPswd=request.GET['adminPswd'];
                adminVar=request.GET['adminVar'];
                adminOld=request.GET['adminOld'];
                role=request.GET['role'];
                login=Login(session);
                merchName=Merch(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users'");
                login=cort_to_list(cursor.fetchall());
                if adminName in login and adminName!=adminOld:
                    send="err=1 text=Логин существует";
                else:
                    if adminVar=="new":
                        cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        (adminName,h(adminPswd),"_",merchName,"admin","0.01","0.01","0","ru",));
                        conn1=sqlite3.connect(mPath(merchName,"orders"));
                        cursor1=conn1.cursor();
                        cursor1.execute("INSERT INTO 'admins' VALUES((?),(?),(?))",(adminName,"",role,));
                        conn1.commit();
                        conn1.close();
                    elif adminVar=="old":
                        cursor.execute("UPDATE 'users' SET pswd=(?)WHERE login=(?)",(h(adminPswd),adminOld,));
                        cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(adminName,adminOld,));
                        conn1=sqlite3.connect(mPath(merchName,"orders"));
                        cursor1=conn1.cursor();
                        cursor1.execute("UPDATE 'admins'SET role=(?)WHERE login=(?)",(role,adminOld,));
                        cursor1.execute("UPDATE 'admins'SET login=(?)WHERE login=(?)",(adminName,adminOld,));
                        conn1.commit();
                        conn1.close();
                    send="err=0 text=OK";
                conn.commit();
                conn.close();
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send 
def send_remove_admin(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                adminName=request.GET['adminName'];
                merchName=Merch(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'users'WHERE login=(?)AND user_type='admin'",(adminName,));
                conn.commit();
                conn.close();
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("DELETE FROM 'admins'WHERE login=(?)",(adminName,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_acces_admin(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                adminName=request.GET['adminName'];
                adminAcces=request.GET['adminAcces'].replace("_","|");
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE admins SET level=(?) WHERE login=(?)",(adminAcces,adminName,));
                conn.commit();
                conn.close();
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_about_not(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                title=getConst(merchName,"notTitle")
                body=getConst(merchName,"notBody")
                conn.commit();
                conn.close();
                send="err=0 text={0}|{1}|^".format(title,body);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_product(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                prodName=request.GET['prodName'];
                prodPrice=request.GET['prodPrice'];
                try:
                    cat_id=request.GET['cat_id'];
                except:
                    cat_id="0";
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT name FROM 'products'");
                names=cort_to_list(cursor.fetchall());
                if prodName in names:
                    cursor.execute("SELECT work FROM 'products'WHERE name=(?)",(prodName,));
                    oldWork=cort_to_list(cursor.fetchall())[0];
                    if oldWork=="0":
                        cursor.execute("UPDATE products SET work='1'WHERE name=(?)",(prodName,));
                        cursor.execute("UPDATE products SET work='1'WHERE cat_id=(?)",(cat_id,));
                        send="err=0,,text=Продукт восстановлен";
                    else:
                        send="err=1,,text=Продукт уже существует";
                    conn.commit();
                    conn.close();
                    async3(Login(session),"getProducts");
                    async3(Login(session),"getAllPrices");
                else:
                    cursor.execute("SELECT id FROM 'products'");
                    ids=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT name FROM 'buyers_type'");
                    buyers_types=cort_to_list(cursor.fetchall());
                    id_=1;
                    while str(id_) in ids:
                        id_+=1;
                    cursor.execute("INSERT INTO 'products' VALUES((?),(?),(?),(?),(?),(?),'1','sh',(?),(?),(?),(?))",
                        (cat_id,str(id_),prodName,"_","1","_","1","0",prodName,"_"));
                    if prodPrice=="":
                        prodPrice="0";
                    for b in buyers_types:
                        cursor.execute("INSERT INTO 'prices' VALUES((?),(?),(?))",
                            (str(id_),b,prodPrice,));
                    conn.commit();
                    conn.close();
                    send="err=0 text=OK";
                    async3(Login(session),"getProducts");
                    async3(Login(session),"getAllPrices");
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_rem_product(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                prod_id=request.GET['prod_id'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                name=cort_to_list(cursor.fetchall())[0];
                cursor.execute("UPDATE products SET work='0'WHERE id=(?)",(prod_id,));
                conn.commit();
                conn.close();
                async2(Login(session),"getProducts");
                async2(Login(session),"getAllPrices");
                txt="Удален продукт: {0}".format(name);
                makeAct(merchName,Login(session),txt);
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_rem_cat(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                cat_id=request.GET['cat_id'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT name FROM 'categories' WHERE id=(?)",(cat_id,));
                name=cort_to_list(cursor.fetchall())[0];
                cursor.execute("UPDATE categories SET work='0'WHERE id=(?)",(cat_id,));
                conn.commit();
                conn.close();
                async2(Login(session),"getCategories");
                txt="Удалена категория: {0}".format(name);
                makeAct(merchName,Login(session),txt);
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_edit_cat(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                old_id=request.GET['old_id'];
                new_cat_name=request.GET['new_cat_name'];
                new_cat_name2=request.GET['new_cat_name2'];
                new_cat_id=request.GET['new_cat_id'];
                Var=request.GET['Var'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                if Var=="old":
                    cursor.execute("UPDATE categories SET name2=(?)WHERE id=(?)",(new_cat_name2,old_id,));
                    cursor.execute("UPDATE categories SET name=(?)WHERE id=(?)",(new_cat_name,old_id,));
                    cursor.execute("UPDATE categories SET cat_id=(?)WHERE id=(?)",(new_cat_id,old_id,));
                    send="err=0,,text=Сохранено";
                elif Var=="new":
                    cursor.execute("SELECT name FROM 'categories'");
                    names=cort_to_list(cursor.fetchall());
                    print(new_cat_name in names);
                    if new_cat_name in names:
                        cursor.execute("UPDATE categories SET work='1'WHERE name=(?)",(new_cat_name,));
                        cursor.execute("UPDATE categories SET cat_id=(?)WHERE name=(?)",(new_cat_id,new_cat_name,));
                        send="err=0,,text=Категория включена";
                    else:
                        cursor.execute("SELECT id FROM 'categories'");
                        ids=cort_to_list(cursor.fetchall());
                        id_=1;
                        while str(id_) in ids:
                            id_+=1;
                        cursor.execute("INSERT INTO categories VALUES((?),(?),(?),'1',NULL,(?))",
                            (new_cat_id,id_,new_cat_name,new_cat_name));
                        send="err=0 text=OK";
                conn.commit();
                conn.close();
                async2(Login(session),"getCategories");
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_edit_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                markName=request.GET['markName'];
                markAdres=request.GET['markAdres'];
                markInn=request.GET['markInn'].replace("}{","+").replace(" ","");
                markPhone=request.GET['markPhone'];
                markDistrict=request.GET['markDistrict'];
                markOld=request.GET['markOld'];
                markVarified=request.GET['markVarified'];
                merchName=Merch(session);
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                cursor.execute("UPDATE 'users' SET district=(?)WHERE login=(?)",(markDistrict,markOld,));
                cursor.execute("UPDATE 'users' SET varified=(?)WHERE login=(?)",(markVarified,markOld,));
                cursor.execute("UPDATE 'users' SET phone=(?)WHERE login=(?)",(markPhone,markOld,));
                cursor.execute("UPDATE 'users' SET inn=(?)WHERE login=(?)",(markInn,markOld,));
                cursor.execute("UPDATE 'users' SET adres=(?)WHERE login=(?)",(markAdres,markOld,));
                cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(markName,markOld,));
                conn.commit();
                conn.close();
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
                admins=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
                drivers=cort_to_list(cursor.fetchall());
                
                cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(markName,markOld,));
                conn.commit();
                conn.close();
                merchs=os.listdir("merchants");
                for m in merchs:
                    conn=sqlite3.connect(mPath(m,"orders"));
                    cursor=conn.cursor();
                    cursor.execute("UPDATE 'buy_markets' SET name=(?)WHERE name=(?)",(markName,markOld,));
                    cursor.execute("UPDATE 'order' SET getter=(?)WHERE getter=(?)",(markName,markOld,));
                    conn.commit();
                    conn.close();
                for a in admins:
                    async2(a,"getMarketList");
                for d in drivers:
                    async2(d,"getMarketList");
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_null_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                driver=request.GET['driverName'];
                merchName=Merch(session);
                login=Login(session);
                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                cursor.execute("SELECT prod_id FROM 'naks'");
                ids=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT free FROM 'naks'");
                free=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT reserve FROM 'naks'");
                reserve=cort_to_list(cursor.fetchall());
                for i in range(len(ids)):
                    prod_id=ids[i];
                    how_many=int(free[i])+int(reserve[i]);
                    ostPlus(merchName,prod_id,how_many);

                cursor.execute("UPDATE 'cashInCar' SET cash='0'WHERE driver=(?)",(driver,));
                cursor.execute("UPDATE 'cashInCar' SET term='0'WHERE driver=(?)",(driver,));
                cursor.execute("UPDATE 'cashInCar' SET per='0'WHERE driver=(?)",(driver,));
                cursor.execute("UPDATE 'cashInCar' SET on_day='0'WHERE driver=(?)",(driver,));
                cursor.execute("DELETE FROM 'nakNum' WHERE owner = (?)",(driver,));
                cursor.execute("DELETE FROM 'naks' WHERE owner = (?)",(driver,));

                conn.commit();
                conn.close();
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'order' SET get_type='new'WHERE owner=(?)AND get_type='ord'",(driver,));
                cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)AND get_type='ord'",(driver,));
                cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)AND get_type='new'",(driver,));
                conn.commit();
                conn.close();



                async2(driver,"getProducts");

                async2(login,"getProducts");
                async2(login,"getDrivers");
                async2(login,"getProducts");
                async2(login,"getAdminOrders");
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_districts_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                driver=request.GET['driverName'];
                data=request.GET['data'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE 'drivers' SET district=(?)WHERE login=(?)",(data,driver,));
                conn.commit();
                conn.close();
                async2(Login(session),"getDrivers");
                send="err=0,,text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_photo(request,var='network'):
    try:
        b=str(request.body[:600])
        for i in range(len(b)):
            if b[i]==";":
                if b[i+1]=="-":
                    b=b[i+2:];
                    break;
        for i in range(len(b)):
            if b[i]=="-":
                if b[i+1]==";":
                    b=b[:i];
                    break;
        print(b)
        if "guvPhoto" in b or "pasPhoto" in b:
            data=stringToArray(b);
            data=data[0];
            marketName=data[0];
            photoType=data[1];
            files=os.listdir("marketImg");
            nameFile=marketName+photoType+".jpg";
            path="marketImg/"+nameFile;
            #
            f=request.FILES[nameFile];
            with open("marketImg/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);
            #
            print("OK");
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET varified='100' WHERE login=(?)",(marketName,));
            conn.commit();
            conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'AND merchName='Kay-Kay'");
            admins=cort_to_list(cursor.fetchall());
            conn.close();
            for a in admins:
                l=ADLevel("Kay-Kay",a);
                print(a+"<<<<<<<<<<<<<<<<<,1")
                if "max98" in l:          
                    print(l+"<<<<<<<<<<<<<<<<<,2")
                    addNot(a,"title","text","varMarket");
            send="err=0,,text=OK";
            return HttpResponse(send, content_type='application/json');
        elif "chatSendPhoto" in b:
            data=stringToArray(b);
            data=data[0];
            chat_id=data[0];
            sender=data[1];
            getter=data[2];
            files=os.listdir("img");
            nameFile=generate()+".jpg";
            while nameFile in files:
                nameFile=generate()+".jpg";
            #
            f=request.FILES['image.jpg'];
            with open("img/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);
            #
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT chat_id FROM messages");
            chat_ids=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT message_id FROM messages");
            message_ids=cort_to_list(cursor.fetchall());
            if "|^" in chat_id:
                getter=stringToArray(chat_id)[0][1];
                newChatId=1;
                while True:
                    if str(newChatId) in chat_ids:
                        newChatId+=1;
                    else:
                        break;
                chat_id=newChatId;
            newMessageId=1;
            while True:
                if str(newMessageId) in message_ids:
                    newMessageId+=1;
                else:
                    break;
            
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            cursor.execute("INSERT INTO 'messages' VALUES((?),(?),(?),(?),(?),(?),(?),(?))",
                (str(chat_id),str(newMessageId),"photo",sender,getter,date,"0",nameFile));
            conn.commit();
            conn.close();
            send="err=0,,text=OK";
        else:
            data=stringToArray(b);
            data=data[0];
            photoType=data[0];
            merchName=data[1];
            files=os.listdir("img");
            nameFile=generate()+".jpg";
            while nameFile in files:
                nameFile=generate()+".jpg";
            path="merchants/"+merchName+"/imgBuf/"+nameFile;
            #
            f=request.FILES['image.jpg'];
            with open("img/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);
            #
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            if photoType=="merchPhoto":
                cursor.execute("UPDATE const SET val=(?) WHERE key='merchImg'",(nameFile,));
            elif photoType=="catPhoto":
                ID=data[2];
                cursor.execute("SELECT place FROM 'img' WHERE type='cat'AND _id=(?)",(ID,));
                olds=cort_to_list(cursor.fetchall());
                if len(olds)==0:
                    cursor.execute("INSERT INTO 'img' VALUES('cat',(?),(?))",(ID,nameFile,));
                else:
                    cursor.execute("UPDATE'img'SET place=(?) WHERE _id=(?) AND type='cat'",(nameFile,ID,));
            elif photoType=="prodPhoto":
                ID=data[2];
                cursor.execute("SELECT place FROM 'img' WHERE type='prod'AND _id=(?)",(ID,));
                olds=cort_to_list(cursor.fetchall());
                if len(olds)==0:
                    cursor.execute("INSERT INTO 'img' VALUES('prod',(?),(?))",(ID,nameFile,));
                else:
                    cursor.execute("UPDATE'img'SET place=(?) WHERE _id=(?) AND type='prod'",(nameFile,ID,));
            conn.commit();
            conn.close();
            send="err=0,,text=OK";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_merch_rev(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                newRev=request.GET['newRev'];
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("UPDATE const SET val=(?)WHERE key=(?)",(newRev,"merchRev"));
                conn.commit();
                conn.close();
                async2(Login(session),"getCategories");
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_hand_nak(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                driver=request.GET['driverName'];
                data=request.GET['data'];
                nak=request.GET['nakNum'];
                merchName=Merch(session);
                admin=Login(session);
                data=stringToArray(data);
                conn1=sqlite3.connect(mPath(merchName,"orders"));
                cursor1=conn1.cursor();
                conn=sqlite3.connect(mPath(merchName,"inCar"));
                cursor=conn.cursor();
                for i in range(len(data)):
                    ids=data[i][0]
                    excelHowE=data[i][1];

                    cursor1.execute("SELECT name FROM 'products' WHERE id=(?)",(ids,));
                    excelNameE=cort_to_list(cursor1.fetchall())[0];
                    cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                    freeOld=cort_to_list(cursor.fetchall());
                    if len(freeOld)==0:
                        cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                            (excelNameE,ids,excelHowE,driver,));
                    else:
                        freeOld=freeOld[0];
                        newFree=str(int(excelHowE)+int(freeOld));
                        cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                            (newFree,driver,ids,));
                    ostMinus(merchName,ids,excelHowE)
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                    (driver,nak,date));
                conn.commit();
                conn.close();
                conn1.commit();
                conn1.close();
                async2(admin,"getProducts");
                async2(admin,"getDrivers");
                async2(admin,"getProducts");
                txt="Создана накладная №{0}".format(nak);
                makeAct(merchName,Login(session),txt);
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_nulls(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                nulls=request.GET['nulls']+";";
                nulls=nulls.replace("|",":")
                merchName=Merch(session);
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("SELECT merchName FROM 'clearTime'");
                merchNames=cort_to_list(cursor.fetchall());
                if merchName in merchNames:
                    cursor.execute("UPDATE clearTime SET drivers=(?)WHERE merchName=(?)",(nulls,merchName,));
                else:
                    cursor.execute("INSERT INTO users VALUES((?),(?))",(merchName,nulls,));

                conn.commit();
                conn.close();
                async2(Login(session),"getDefaultMarket");
                send="err=0 text=OK";
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_act_req(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                acts=request.GET['acts'];
                try:
                    date=request.GET['date'];
                except:
                    date="";
                merchName=Merch(session);
                conn=sqlite3.connect(mPath(merchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT login FROM 'actHistory'");
                login=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT Date FROM 'actHistory'");
                Date=cort_to_list(cursor.fetchall());
                cursor.execute("SELECT Data FROM 'actHistory'");
                Data=cort_to_list(cursor.fetchall());
                act=[];
                for i in range(len(login)):
                    if date in Date[i]:
                        if acts in login[i] or acts in Data[i]:
                            act.append([login[i],Data[i],Date[i]]);
                acts=arrayToString(act);
                conn.close();
                send="err=0,,text=OK,,acts={0}".format(acts);
            else:
                send="err=1,,text=Вы не админ";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send

#driver
#driver do
def create_new_market(request,var='network'):
    try:
        session=request.GET['session'];
        merchName=Merch(session);
        login=Login(session);
        send="";
        if checkSession(session):
            name=request.GET['name'];
            inn=request.GET['inn'];
            adres=request.GET['adres'];
            phone=request.GET['phone'];
            phone2=request.GET['phone2'];
            district=request.GET['district'];
            lon=request.GET['lon'];
            lat=request.GET['lat'];
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            logins=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT district FROM 'districts'");
            districts=cort_to_list(cursor.fetchall());
            conn.close();

            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT inn FROM 'users'");
            inns=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT adres FROM 'users'");
            logins=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT phone FROM 'users'");
            phones=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT adres FROM 'users'");
            adress=cort_to_list(cursor.fetchall());
            if name in logins:
                send="err=1,,text=Такой Логин уже существует";
            if district not in districts:
                send="err=1,,text=Района не существует";
            if inn in inns:
                send="err=1,,text=Такой ИНН уже существует";
            if phone in phones:
                send="err=1,,text=Такой Телефон уже существует";
            if adres in adress:
                send="err=1,,text=Такой Адрес уже существует";
            if len(name)<6:
                send="err=1,,text=Логин слишком короткий";
            if len(phone)!=9:
                send="err=1,,text=Телефон слишком короткий";
            if len(inn)!=9 and len(inn)!=14:
                send="err=1,,text=инн слишком короткий";
            if len(adres)<6:
                send="err=1,,text=Адрес слишком короткий";
            float(lon)+1;
            float(lat)+1;
            if "err=1" not in send:
                cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?))",
                    (name,adres,inn,phone,"0",district,phone2));
                conn.commit();
            conn.close();

            if "err=1" not in send:
                conn=sqlite3.connect("basic.sqlite");
                cursor=conn.cursor();
                cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        (name,h(name),"_",merchName,"market",lon,lat,"0","ru",));
                cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
                logins=cort_to_list(cursor.fetchall());
                conn.commit();
                conn.close();
                checkMarket(merchName,name);
                send="err=0,,text=OK";
            else:
                for l in logins:
                    async2(l,"getMarketList");
            async2(login,"getMarketList");
            messageInChannel("Kay-Kay","Создан магазин🔐\n\nДобавил: {0}\nОт производителя:{3}\nМагазин: {1}\nИНН: {2}".format(Login(session),name,inn,merchName));
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
        async2(login,"getMarketList");
    return HttpResponse(send, content_type='application/json');
def get_adres_from_coords(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            lon=request.GET['lon'];
            lat=request.GET['lat'];
            adres=get_address_from_coords1(lon+","+lat);
            if "не установлен" in adres:
                adres=get_address_from_coords2(lon+","+lat);
            adres=adres.replace("Узбекистан","").replace(", "," ").replace(","," ")
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT district FROM 'districts'");
            districts=cort_to_list(cursor.fetchall());
            conn.close();
            district="null";
            for d in districts:
                dMin=d[:4]
                if dMin in adres:
                    district=d;
                    break;
            adresData=arrayToString([[adres,district]]);
            send="err=0,,adres={0}".format(adresData);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json');
def send_dolg(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            last_index=request.GET['last_index'];
            nal=request.GET['nal'];
            term=request.GET['term'];
            per=request.GET['per'];
            order_or_sell=request.GET['order_or_sell'];
            saldo=int(nal)+int(term)+int(per);
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT getter FROM 'order' WHERE last_index=(?) AND get_type='dolg'",(last_index,));
            getter=cort_to_list(cursor.fetchall())[0];
            cursor.execute("DELETE FROM 'order' WHERE last_index = (?)  AND get_type='dolg'",(last_index,));
            conn.commit();
            conn.close();
            checkMarket(merchName,getter);
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            cursor.execute("SELECT cash FROM 'cashInCar' WHERE driver=(?)",(login,));
            nal1=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT term FROM 'cashInCar' WHERE driver=(?)",(login,));
            term1=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT per FROM 'cashInCar' WHERE driver=(?)",(login,));
            per1=cort_to_list(cursor.fetchall())[0];
            cursor.execute("UPDATE cashInCar SET cash=(?) WHERE driver=(?)",(str(int(nal)+int(nal1)),login));
            cursor.execute("UPDATE cashInCar SET term=(?) WHERE driver=(?)",(str(int(term)+int(term1)),login));
            cursor.execute("UPDATE cashInCar SET per=(?) WHERE driver=(?)",(str(int(per)+int(per1)),login));
            conn.commit();
            conn.close();
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            dolgHave=getDolgInfo(merchName,getter)['dolgHave']
            if getDolgInfo(merchName,getter)['dolgType']=="sum":
                dolgHave=str(int(dolgHave)-saldo);
            else:
                dolgHave=str(int(dolgHave)-1);
            cursor.execute("UPDATE buy_markets SET dolgHave=(?) WHERE name=(?)",(dolgHave,getter));
            try:
                cursor.execute("SELECT price FROM 'history' WHERE last_index=(?)",(last_index,));
                historyPrice=cort_to_list(cursor.fetchall())[0];
                if historyPrice=="0":
                    cursor.execute("UPDATE 'history' SET price = (?) WHERE last_index=(?)",(str(saldo),last_index,));
            except:
                pass;
            conn.commit();
            conn.close();
            if "USER_" in getter:
                removeUser(merchName,getter);
            send="err=0,,text=OK";
            async2(login,"getDolgs");
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        async2(login,"getDolgs");
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
def send_order(request,var='network'):
    try:
        session=request.GET['session'];
        login=Login(session);
        if checkSession(session):
            last_index=request.GET['last_index'];
            d=request.GET['data'];
            nal=request.GET['nal'];
            term=request.GET['term'];
            per=request.GET['per'];
            order_or_sell=request.GET['order_or_sell'];
            orderType=order_or_sell;
            saldo=int(nal)+int(term)+int(per);
            merchName=Merch(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            if order_or_sell=="sell":
                Data=d;
                data=stringToArray(Data);
                getter=last_index;
                Last(merchName,"new");
                last_index=Last(merchName);
            else:
                cursor.execute("SELECT data FROM 'order' WHERE last_index=(?) AND get_type='ord'",(last_index,));
                Data=cort_to_list(cursor.fetchall())[0]
                data=stringToArrayData(Data);
                cursor.execute("SELECT getter FROM 'order' WHERE last_index=(?) AND get_type='ord'",(last_index,));
                getter=cort_to_list(cursor.fetchall())[0];
            checkMarket(merchName,getter);
            buyType=getBuyType(merchName,getter);
            realSaldo=0;
            d=datetime.now();
            try:
                date="{0}.{1}.{2} {3}:{4}".format(str(d.day),str(d.month),str(d.year),str(d.hour+5),str(d.minute));
            except:
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            district=District(getter);
            #inCar +1
            conn1=sqlite3.connect(mPath(merchName,"inCar"));
            cursor1=conn1.cursor();
            for i in range(len(data)):
                prod_id=data[i][0];
                how_many=data[i][1];
                cursor.execute("SELECT price FROM 'prices' WHERE name=(?) AND id=(?)",(buyType,prod_id,));
                price=cort_to_list(cursor.fetchall())[0];




                realSaldo=realSaldo+(int(price)*int(how_many));
                cursor1.execute("SELECT reserve FROM 'naks' WHERE owner=(?) AND prod_id=(?)",(login,prod_id,));
                reserve=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT free FROM 'naks' WHERE owner=(?) AND prod_id=(?)",(login,prod_id,));
                free=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT sell FROM 'naks' WHERE owner=(?) AND prod_id=(?)",(login,prod_id,));
                sell=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT cash FROM 'cashInCar' WHERE driver=(?)",(login,));
                nal1=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT term FROM 'cashInCar' WHERE driver=(?)",(login,));
                term1=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT per FROM 'cashInCar' WHERE driver=(?)",(login,));
                per1=cort_to_list(cursor1.fetchall())[0];
                

                if order_or_sell=="ord":
                    orderType="ord";
                    how_ost=reserve;
                    new_how_ost=str(int(how_ost)-int(how_many));
                    cursor1.execute("UPDATE naks SET reserve = (?) WHERE owner=(?) AND prod_id=(?)",(new_how_ost,login,prod_id,));
                    
                else:
                    orderType="sell";
                    how_ost=free;
                    new_how_ost=int(how_ost)-int(how_many);
                    cursor1.execute("UPDATE naks SET free = (?) WHERE owner=(?) AND prod_id=(?)",(new_how_ost,login,prod_id,));
                new_sell=str(int(sell)+int(how_many));
                cursor1.execute("UPDATE naks SET sell = (?) WHERE owner=(?) AND prod_id=(?)",(new_sell,login,prod_id,));
            cursor.execute("INSERT INTO history VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                    (Data.replace("|",":").replace("^",";"),str(saldo),orderType,getter,login,date,district,last_index,"new"));
            if realSaldo==saldo:
                cursor.execute("DELETE FROM 'order' WHERE last_index = (?)  AND get_type='ord'",(last_index,));
                if "USER_" in getter:
                    removeUser(merchName,getter);
            else:
                dolg=str(int(realSaldo)-int(saldo));
                if order_or_sell=="ord":
                    cursor.execute("UPDATE 'order' SET get_type = 'dolg' WHERE last_index=(?) AND get_type='ord'",(last_index,));
                    cursor.execute("UPDATE 'order' SET price=(?)WHERE last_index=(?)",(dolg,last_index,));
                else:
                    payForm="nal"
                    cursor.execute("INSERT INTO 'order' VALUES((?),(?),(?),(?),(?),(?),(?),(?))",
                    (Data.replace("|",":").replace("^",";"),getter,login,last_index,date,"dolg",dolg,payForm));

                dolgType=getDolgInfo(merchName,getter)['dolgType'];
                dolgHave=getDolgInfo(merchName,getter)['dolgHave'];
                if dolgType=="sum":
                    newHave=int(dolgHave)+int(dolg);
                else:
                    newHave=int(dolgHave)+1;
                cursor.execute("UPDATE buy_markets SET dolgHave=(?) WHERE name=(?)",(newHave,getter,));
            cursor1.execute("UPDATE cashInCar SET cash = (?) WHERE driver=(?)",(str(int(nal)+int(nal1)),login));
            cursor1.execute("UPDATE cashInCar SET term = (?) WHERE driver=(?)",(str(int(term)+int(term1)),login));
            cursor1.execute("UPDATE cashInCar SET per = (?) WHERE driver=(?)",(str(int(per)+int(per1)),login));
            conn.commit();
            conn.close();
            conn1.commit();
            conn1.close();
            if order_or_sell=="ord":
                giveCoinForOrder(getter,saldo);

            notification(merchName,"admin","Заказ выполнен",login+" выполнил заказ №"+last_index);
            send="err=0,,text=OK";
            async2(login,"getOrders");
            async2(login,"getDolgs");
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        async2(login,"getOrders");
        async2(login,"getDolgs");
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def ostMinus(merchName,prod_id,how_many):
    try:
        conn32=sqlite3.connect(mPath(merchName,"orders"));
        cursor32=conn32.cursor();
        cursor32.execute("SELECT ost FROM 'products' WHERE id=(?)",(prod_id,));
        ost1=cort_to_list(cursor32.fetchall())[0];
        ost1=str(int(ost1)-int(how_many));
        if ost1=="0":
            cursor32.execute("UPDATE products SET visible = '0' WHERE id=(?)",(prod_id,));
        cursor32.execute("UPDATE products SET ost = (?) WHERE id=(?)",(ost1,prod_id,));
        conn32.commit();
        conn32.close();
    except Exception as e:
        print("check log <<<<<<");
        logger(e);
def ostPlus(merchName,prod_id,how_many):
    try:
        how_many=str(how_many);
        conn32=sqlite3.connect(mPath(merchName,"orders"));
        cursor32=conn32.cursor();
        cursor32.execute("SELECT ost FROM 'products' WHERE id=(?)",(prod_id,));
        ost1=cort_to_list(cursor32.fetchall())[0];
        ost1=str(int(ost1)+int(how_many));
        if ost1=="0":
            cursor32.execute("UPDATE products SET visible = '0' WHERE id=(?)",(prod_id,));
        else:
            cursor32.execute("UPDATE products SET visible = '1' WHERE id=(?)",(prod_id,));
        cursor32.execute("UPDATE products SET ost = (?) WHERE id=(?)",(ost1,prod_id,));
        conn32.commit();
        conn32.close();
    except Exception as e:
        print("check log <<<<<<")
        logger(e);
def send_new_inn(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            new_inn=request.GET['new_inn'];
            getter=request.GET['getter'];
            int(new_inn)+1;
            login=Login(session);
            merchName=Merch(session);
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(new_inn,getter,));
            #cursor.execute("UPDATE users SET varified='0'WHERE login=(?)",(getter,));
            conn.commit();
            conn.close();
            send="err=0 text=OK";
            messageInChannel(merchName,"Добавлен ИНН🔐\n\nДобавил: {0}\nОт производителя:{3}\nМагазин: {1}\nИНН: {2}".format(login,getter,new_inn,merchName));
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_phone(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            new_phone=request.GET['new_phone'];
            getter=request.GET['getter'];
            int(new_phone)+1;
            login=Login(session);
            merchName=Merch(session);
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET phone=(?)WHERE login=(?)",(new_phone,getter,));
            conn.commit();
            conn.close();
            send="err=0 text=OK";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_location(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            lon=request.GET['new_lon'];
            lat=request.GET['new_lat'];
            getter=request.GET['getter'];
            login=Login(session);
            merchName=Merch(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET lon=(?)WHERE login=(?)",(lon,getter,));
            cursor.execute("UPDATE users SET lat=(?)WHERE login=(?)",(lat,getter,));
            conn.commit();
            conn.close();
            send="err=0 text=OK";
            messageInChannel(merchName,"Добавлена локация📍\n\nДобавил: {0}\nМагазин: {1}\nШирота: {2}\nДолгота: {3}".format(login,getter,lat,lon));
            locationInChannel(merchName,lon,lat);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#driver updates
def uber_driver_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            cursor.execute("SELECT prod_id FROM 'naks' WHERE owner = (?)",(login,));
            prod_id=arrayToString([cort_to_list(cursor.fetchall())]);
            conn.close();
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT lon FROM 'users' WHERE session=(?)",(session,));
            lon=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT lat FROM 'users' WHERE session=(?)",(session,));
            lat=cort_to_list(cursor.fetchall())[0];
            conn.close();
            #
            updates+="<<getMoney>>"+get_money(request,'local')
            updates+="<<getDistricts>>"+get_districts(request,'local')
            updates+="<<getNak>>"+get_nak(request,'local')
            updates+="<<getOrders>>"+get_orders(request,'local')
            updates+="<<getDolgs>>"+get_dolgs(request,'local')
            updates+="<<getPrices>>"+get_prices(request,prod_id,'local')
            #updates+="<<getLocation>>"+get_location(request,'local',utf-8');
            #updates+="<<getMarketList>>"+get_market_list(request,'local')
            updates+="<<getProductList>>"+get_product_list(request,'local')
            updates+="<<getDefaultMarket>>"+get_default_market(request,'local')
            #
            send=updates#.replace("err=0,,")
            send="err>>0"+send;
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    return HttpResponse(send, content_type='application/json')
def get_product_list(request,var='network'):
    try:
        products=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT cat_id FROM products WHERE work='1'");
            cat_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT id FROM products WHERE work='1'");
            prod_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name FROM products WHERE work='1'");
            name=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT rev FROM products WHERE work='1'");
            rev=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT box FROM products WHERE work='1'");
            box=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT form FROM products WHERE work='1'");
            form=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT visible FROM products WHERE work='1'");
            visible=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT ost FROM products WHERE work='1'");
            ost=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name2 FROM products WHERE work='1'");
            name2=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT rev2 FROM products WHERE work='1'");
            rev2=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(name)):
                img=Img(merchName,"prod",prod_id[i]);
                products.append([cat_id[i],prod_id[i],name[i],rev[i],img,merchName,box[i],form[i],visible[i],ost[i],name2[i],rev2[i]]);
            products=arrayToString(products);
            send="err=0,,products={0}".format(products);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_cat_list(request,var='network'):
    try:
        cats=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT cat_id FROM categories WHERE work='1'");
            cat_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT id FROM categories WHERE work='1'");
            prod_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name FROM categories WHERE work='1'");
            name=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name2 FROM categories WHERE work='1'");
            name2=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(name)):
                img=Img(merchName,"cat",prod_id[i]);
                cats.append([cat_id[i],prod_id[i],name[i],img,name2[i]]);
            cats=arrayToString(cats);
            send="err=0,,cats={0}".format(cats);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_default_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            merchName=Merch(session);
            default_buyer=getConst(merchName,"default_buyer");            
            defDolgType=getConst(merchName,"defDolgType");            
            defDolgVal=getConst(merchName,"defDolgVal");            
            defDolgMax=getConst(merchName,"defDolgMax");        
            tg_id=getConst(merchName,"tg_id");        
            tg_token=getConst(merchName,"tg_token");        
            nots=getConst(merchName,"nots");        
            excelName=getConst(merchName,"excelName");        
            excelHow=getConst(merchName,"excelHow");        
            excelStartWrite=getConst(merchName,"excelStartWrite");        
            excelType=getConst(merchName,"excelType");        
            excelCat=getConst(merchName,"excelCat");        
            excelPriceName=getConst(merchName,"excelPriceName");        
            excelPrice=getConst(merchName,"excelPrice");        
            excelStartWritePrice=getConst(merchName,"excelStartWritePrice");        
            excelNakSheet=getConst(merchName,"excelNakSheet");        
            excelProdSheet=getConst(merchName,"excelProdSheet");        
            excelNakPlace=getConst(merchName,"excelNakPlace");        
            api_key=getConst(merchName,"api_key");     
            merchImg=getConst(merchName,"merchImg");     
            merchRev=getConst(merchName,"merchRev");
            autoNull=getConst(merchName,"autoNull");
            site=getConst(merchName,"site");
            ostName=getConst(merchName,"ostName");
            ostHowOst=getConst(merchName,"ostHowOst");
            ostStart=getConst(merchName,"ostStart");
            ostSheet=getConst(merchName,"ostSheet");
            upRange=getConst(merchName,"upRange");
            generalLang=getConst(merchName,"generalLang");
            secondLang=getConst(merchName,"secondLang");
            generalLink=getConst(merchName,"generalLink");
            exNakDriverName=getConst(merchName,"exNakDriverName");

            exPerInn=getConst(merchName,"exPerInn");
            exPerSum=getConst(merchName,"exPerSum");
            exPerStart=getConst(merchName,"exPerStart");
            exPerSheet=getConst(merchName,"exPerSheet");
            merchImg="{0}get_photo/?link={1}".format(URL,merchImg);  
            defMarket=arrayToString([[default_buyer,defDolgType,defDolgVal,defDolgMax,
                tg_id,tg_token,nots,excelName,excelHow,excelStartWrite,excelType,
                excelCat,excelPriceName,excelPrice,excelStartWritePrice,excelNakSheet,
                excelProdSheet,excelNakPlace,api_key,merchName,merchImg,merchRev,autoNull,site
                ,ostName,ostHowOst,ostStart,ostSheet,upRange,generalLang,secondLang,generalLink,exNakDriverName,exPerInn,exPerSum,exPerStart,exPerSheet]]);
            send="err=0,,defMarket={0}".format(defMarket);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send)
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_list(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=Level(myMerchName);
            conn1=sqlite3.connect("basic.sqlite");
            cursor1=conn1.cursor();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            login=cort_to_list(cursor.fetchall());
            if False:#"a1|" not in level:
                log=[];
                cursor1.execute("SELECT login FROM 'users' WHERE merchName=(?) AND user_type='market'",(myMerchName,));
                lo=cort_to_list(cursor1.fetchall());
                for i in range(len(login)):
                    if login[i]in lo:
                        log.append(login[i]);
                login=log;
                conn.close();
            for i in range(len(login)):
                log=login[i];
                try:
                    cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    lon=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    lat=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    merchName=cort_to_list(cursor1.fetchall())[0];
                    lo=lon;
                    la=lat;
                    me=merchName;
                    cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                    adres=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                    inn=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                    phone=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                    varified=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                    district=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                    phone2=cort_to_list(cursor.fetchall())[0];
                    wallet=markWallet(myMerchName,log);
                    if str(phone2)=="None":
                        phone2="";
                    if User(session)!="root":
                        checkMarket(myMerchName,log);
                    dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                    dolgVal=getDolgInfo(myMerchName,log)['dolgVal']; 
                    dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                    dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                    buyType=getBuyType(myMerchName,log)
                    markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                except Exception as e:
                    print(log,"<<<<<");

            conn.close();
            conn1.close();
            markets=arrayToString(markets);
            send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_root_list(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=Level(myMerchName);
            conn1=sqlite3.connect("basic.sqlite");
            cursor1=conn1.cursor();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            login=cort_to_list(cursor.fetchall());
            if "a1|" not in level:
                log=[];
                cursor1.execute("SELECT login FROM 'users' WHERE merchName=(?) AND user_type='market'",(myMerchName,));
                lo=cort_to_list(cursor1.fetchall());
                for i in range(len(login)):
                    if login[i]in lo:
                        log.append(login[i]);
                login=log;
                conn.close();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            for i in range(len(login)):
                log=login[i];
                cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                varified=cort_to_list(cursor.fetchall())[0];
                if "100" in varified:
                    cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    lon=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    lat=cort_to_list(cursor1.fetchall())[0];
                    cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    merchName=cort_to_list(cursor1.fetchall())[0];
                    lo=lon;
                    la=lat;
                    me=merchName;
                    cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                    inn=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                    phone=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                    varified=cort_to_list(cursor.fetchall())[0];
                    cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                    district=cort_to_list(cursor.fetchall())[0];
                    if User(session)!="root":
                        checkMarket(myMerchName,log);
                    dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                    dolgVal=getDolgInfo(myMerchName,log)['dolgVal']; 
                    dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                    dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                    buyType=getBuyType(myMerchName,log)
                    markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax]);
            conn.close();
            conn1.close();
            markets=arrayToString(markets);
            send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_mikro_market(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            getMarkets=request.GET['getMarkets'];
            getMarkets=stringToArray(getMarkets)[0];
            myMerchName=Merch(session);
            level=Level(myMerchName);
            conn1=sqlite3.connect("basic.sqlite");
            cursor1=conn1.cursor();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            log=cort_to_list(cursor.fetchall());
            conn.close();
            login=[];
            for lo in log:
                if lo in getMarkets:
                    login.append(lo);
            for i in range(len(login)):
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                log=login[i];
                cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lon=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lat=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                merchName=cort_to_list(cursor1.fetchall())[0];
                lo=lon;
                la=lat;
                me=merchName;
                cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                adres=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                inn=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                phone=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                varified=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                district=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                phone2=cort_to_list(cursor.fetchall())[0];
                print(myMerchName,log,"<<<<<<<<<<<<<<<<<,")
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal']; 
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                conn.close();
            conn1.close();
            markets=arrayToString(markets);
            send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    print(send);
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_search_text(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            searchTxt=request.GET['searchTxt'].lower();
            myMerchName=Merch(session);
            level=Level(myMerchName);
            conn1=sqlite3.connect("basic.sqlite");
            cursor1=conn1.cursor();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT login FROM 'users'");
            log=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT adres FROM 'users'");
            adr=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT phone FROM 'users'");
            pho=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT inn FROM 'users'");
            inn=cort_to_list(cursor.fetchall());
            conn.close();
            login=[];
            for i in range(len(log)):
                if searchTxt in log[i].lower() or searchTxt in adr[i].lower() or searchTxt in pho[i].lower():
                    login.append(log[i]);
            for i in range(len(login)):
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                log=login[i];
                cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lon=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lat=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                merchName=cort_to_list(cursor1.fetchall())[0];
                lo=lon;
                la=lat;
                me=merchName;
                cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                adres=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                inn=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                phone=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                varified=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                district=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                phone2=cort_to_list(cursor.fetchall())[0];
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal']; 
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                conn.close();
            conn1.close();
            markets=arrayToString(markets);
            send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_search_loc(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            lon=request.GET['lon'];
            lat=request.GET['lat'];
            myMerchName=Merch(session);
            level=Level(myMerchName);
            conn1=sqlite3.connect("basic.sqlite");
            cursor1=conn1.cursor();
            conn=sqlite3.connect("markets.sqlite");
            cursor=conn.cursor();
            cursor1.execute("SELECT login FROM 'users'WHERE user_type='market'");
            log=cort_to_list(cursor1.fetchall());
            cursor1.execute("SELECT lon FROM 'users'WHERE user_type='market'");
            Lon=cort_to_list(cursor1.fetchall());
            cursor1.execute("SELECT lat FROM 'users'WHERE user_type='market'");
            Lat=cort_to_list(cursor1.fetchall());
            conn.close();
            login=[];
            print(len(log))
            for i in range(len(log)):
                res=get_distance(lon,lat,Lon[i],Lat[i]);
                if int(res)<3:
                    login.append(log[i]);
            for i in range(len(login)):
                conn=sqlite3.connect("markets.sqlite");
                cursor=conn.cursor();
                log=login[i];
                cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lon=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                lat=cort_to_list(cursor1.fetchall())[0];
                cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                merchName=cort_to_list(cursor1.fetchall())[0];
                lo=lon;
                la=lat;
                me=merchName;
                cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                adres=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                inn=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                phone=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                varified=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                district=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                phone2=cort_to_list(cursor.fetchall())[0];
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal']; 
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                conn.close();
            conn1.close();
            markets=arrayToString(markets);
            send="err=0,,markets={0}".format(markets);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def get_updates_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            get_location(request);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            updates=cort_to_list(cursor.fetchall());
            if updates[0]=="0":
                send="err=0,,update_stat=0";
            else:
                send="err=0,,update_stat=1,,updates={0}".format(updates[0]);
            cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            conn.commit();
            conn.close();
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        conn.close();
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if "new" in send:
        print(send+"<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_location(request,lon="l",lat="l",var='network'):
    try:
        session=request.GET['session'];
        if lon=="l":
            lon=request.GET['lon'];
            lat=request.GET['lat'];
        if checkSession(session):
            merchName=Merch(session);
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("UPDATE users SET lon=(?)WHERE session=(?)",(lon,session,));
            cursor.execute("UPDATE users SET lat=(?)WHERE session=(?)",(lat,session,));
            conn.commit();
            conn.close();
            send="err=0";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_prices(request,products=0,var='network'):
    try:
        session=request.GET['session'];
        if products==0:
            products=request.GET['products'];
        products=stringToArray(products)[0];
        stringList="";
        if checkSession(session):
            send="";
            merchName=Merch(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT val FROM 'const' WHERE key = 'default_buyer'");
            default_buyer=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT name FROM 'buyers_type'");
            buyers_type=cort_to_list(cursor.fetchall());
            for prod in products:
                for buy_type in buyers_type:
                    cursor.execute("SELECT name FROM 'prices' WHERE id=(?) AND name=(?)",(prod,buy_type));
                    name=cort_to_list(cursor.fetchall());
                    cursor.execute("SELECT price FROM 'prices' WHERE id=(?) AND name=(?)",(prod,buy_type));
                    price=cort_to_list(cursor.fetchall());
                    for i in range(len(name)):
                        stringList+=name[i]+"|"+price[i]+"|^";
                stringList=prod+"="+stringList;
                send+=stringList+",,"
                stringList="";
            conn.close();
            send=send+"err=0";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_dolgs(request,var='network'):
    try:
        dolgs=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT data FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            data=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT last_index FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            last_index=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT price FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            price=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT owner FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            owner=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT payForm FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            payForm=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                dolgs.append([data[i],getter[i],last_index[i],date[i],"dolg",price[i],owner[i],payForm[i]]);
            dolgs=arrayToString(dolgs);
            send="err=0,,dolgs={0}".format(dolgs);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_orders(request,var='network',order_type='ord'):
    try:
        orders=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT data FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            data=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT getter FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            getter=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT last_index FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            last_index=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            date=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT get_type FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            get_type=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT owner FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            owner=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT payForm FROM 'order' WHERE owner=(?) AND get_type=(?)",(login,order_type));
            payForm=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],owner[i],payForm[i]]);
            orders=arrayToString(orders);
            send="err=0,,orders={0}".format(orders);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_nak(request,var='network'):
    try:
        naks=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            cursor.execute("SELECT name FROM naks WHERE owner=(?)",(login,));
            name=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT prod_id FROM naks WHERE owner=(?)",(login,));
            prod_id=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT free FROM naks WHERE owner=(?)",(login,));
            free=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT reserve FROM naks WHERE owner=(?)",(login,));
            reserve=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT sell FROM naks WHERE owner=(?)",(login,));
            sell=cort_to_list(cursor.fetchall());
            conn.close();
            for i in range(len(name)):
                naks.append([name[i],prod_id[i],free[i],reserve[i],sell[i]]);
            naks=arrayToString(naks);
            send="err=0,,nak={0}".format(naks);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_districts(request,var='network'):
    try:
        session=request.GET['session'];
        if True:
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT district FROM 'districts'");
            dists=cort_to_list(cursor.fetchall());
            conn.close();
            dists=arrayToString([dists]);
            send="err=0,,districts={0}".format(dists);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_money(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"inCar"));
            cursor=conn.cursor();
            cursor.execute("SELECT cash FROM 'cashInCar' WHERE driver=(?)",(login,));
            cash=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT term FROM 'cashInCar' WHERE driver=(?)",(login,));
            term=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT per FROM 'cashInCar' WHERE driver=(?)",(login,));
            per=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT on_day FROM 'cashInCar' WHERE driver=(?)",(login,));
            on_day=cort_to_list(cursor.fetchall())[0];
            #nakNum
            cursor.execute("SELECT num FROM 'nakNum' WHERE owner=(?)",(login,));
            num=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'nakNum' WHERE owner=(?)",(login,));
            date=cort_to_list(cursor.fetchall());
            nakNum="";
            for i in range(len(num)):
                nakNum+="Накладная №"+num[i]+"\nДата: "+date[i]+"\n";
            conn.close();
            send="err=0,,moneyNal={0},,moneyTerm={1},,moneyPer={2},,moneyDay={3},,merchName={4},,nakNum={5}".format(cash,term,per,on_day,merchName,nakNum);
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def get_all_orders_sells(request,var='network',order_type='ord'):
    try:
        orders=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            conn=sqlite3.connect(mPath(merchName,"orders"));
            cursor=conn.cursor();
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            cursor.execute("SELECT date FROM 'history' WHERE driver=(?) AND type=(?)",(login,"ord",));
            date1=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM 'history' WHERE driver=(?) AND type=(?)",(login,"sell",));
            date2=cort_to_list(cursor.fetchall());
            orders=0;
            sells=0;
            for d in date1:
                if date in d:
                    orders+=1;
            for d in date2:
                if date in d:
                    sells+=1;
            conn.close();
            send="err=0,,ord={0},,sells={1}".format(str(orders),str(sells));
            print(send+"<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#functions
def clearInn(text):
    try:
        nums=["0","1","2","3","4","5","6","7","8","9"];
        clear="";
        i=len(text)-1;
        while i!=-1:
            if text[i]in nums:
                clear+=text[i];
            i-=1;
        i=len(clear)-1;
        clear2="";
        while i!=-1:
            clear2+=clear[i];
            i-=1;
        return clear2;
    except Exception as e:
        logger(e);

def loginsByInn(inn):
    conn=sqlite3.connect("markets.sqlite");
    cursor=conn.cursor();
    cursor.execute("SELECT login FROM 'users' WHERE inn=(?) ",(inn,));
    login=cort_to_list(cursor.fetchall());
    conn.commit();
    conn.close();
    return login

def markWallet(merchName,login):
    if "USER_" not in login:
        checkMarket(merchName,login)
        conn=sqlite3.connect(mPath(merchName,"orders"));
        cursor=conn.cursor();
        cursor.execute("SELECT wallet FROM 'buy_markets' WHERE name=(?) ",(login,));
        wallet=cort_to_list(cursor.fetchall())[0];
        if str(wallet)=="None":
            wallet="0";
            cursor.execute("UPDATE buy_markets SET wallet='0' WHERE name=(?)",(login,));
        conn.commit();
        conn.close();
    else:
        wallet="0";

    return wallet;
def not_was_seen(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            id_=request.GET['id_'];
            conn=sqlite3.connect('basic.sqlite');
            cursor=conn.cursor();
            cursor.execute("DELETE FROM 'nots' WHERE id = (?)",(id_,));
            conn.commit();
            conn.close();
            send="err=0,,text=OK";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def get_notifications(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            N=request.GET['nots'];
            login=Login(session);
            conn=sqlite3.connect('basic.sqlite');
            cursor=conn.cursor();
            cursor.execute("SELECT id FROM 'nots'WHERE owner=(?)",(login,));
            id_=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT title FROM 'nots'WHERE owner=(?)",(login,));
            title=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT text FROM 'nots'WHERE owner=(?)",(login,));
            text=cort_to_list(cursor.fetchall());
            cursor.execute("SELECT Type FROM 'nots'WHERE owner=(?)",(login,));
            Type=cort_to_list(cursor.fetchall());
            nots=[];
            for i in range(len(id_)):
                nots.append([id_[i],title[i],text[i],Type[i]]);
            nots=arrayToString(nots);
            conn.commit();
            conn.close();
            if int(N)<len(id_): 
                send="err=0,,nots={0}".format(nots);
            else:
                send="err=2,,nots=0";
        else:
            send="err=1,,text=Сессия истекла";
    except Exception as e:
        logger(e);
        send="err=1,,text=Ошибка сервера";
    if var=='network':
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def addNot(owner,title,text,Type):
    try:
        conn=sqlite3.connect('basic.sqlite');
        cursor=conn.cursor();
        cursor.execute("SELECT id FROM 'nots'");
        id_=cort_to_list(cursor.fetchall());
        cursor.execute("SELECT val FROM 'const'WHERE key='lastNot'");
        lastNot=cort_to_list(cursor.fetchall())[0];
        lastNot=str(int(lastNot)+1);
        cursor.execute("UPDATE const SET val=(?) WHERE key='lastNot'",(lastNot,));
        cursor.execute("INSERT INTO 'nots' VALUES((?),(?),(?),(?),(?))",
            (lastNot,owner,title,text,Type));
        conn.commit();
        conn.close();
    except Exception as e:
        logger(e);
def prodById(merchName,prodId):
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    try:
        cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prodId,));
        data=cort_to_list(cursor.fetchall())[0];
    except:
        data=prodId;
    conn.close();
    return prodId;
def get_distance(lon,lat,lon1,lat1):
    x1 = lon;
    y1 = lat;
    x2 = lon1;
    y2 = lat1;
    #+++++++++++++++++++++++++++++
    X = (float(x1)-float(x2))
    X = X * X
    Y = (float(y1)-float(y2))
    Y = Y * Y
    S0 = X + Y 
    S = math.sqrt(S0)
    S = round(float(S),5)
    km = S*100
    km = round(float(km),3)
    km = round(float(km),1)
    return km;

def giveCoinForOrder(login,price):
    try:
        if "USER_" not in login:
            conn=sqlite3.connect('markets.sqlite');
            cursor=conn.cursor();
            cursor.execute("SELECT varified FROM 'users' WHERE login=(?)",(login,));
            varified=cort_to_list(cursor.fetchall())[0];
            conn.close();
            if varified=="2":
                conn=sqlite3.connect('basic.sqlite');
                cursor=conn.cursor();
                cursor.execute("SELECT cur FROM 'percent' WHERE var=(?)",(varified,));
                cur=cort_to_list(cursor.fetchall())[0];
                cursor.execute("SELECT per FROM 'percent' WHERE var=(?)",(varified,));
                per=cort_to_list(cursor.fetchall())[0];
                conn.close();
                price=float(price)*(float(per)/100.0);
                print(price)
                #dol_cur=float(dolConv.get_dol());
                #price=str(float(price)/dol_cur);
                #print(price);
                if cur=="ice":
                    how_many=get_course("ice",price).text;
                else:
                    how_many=get_course("kay",price).text;
                print(">>>>>>>");
                print(login,cur,how_many);
                giveCoin(login,cur,how_many);
        else:
            pass;
    except Exception as e:
        logger(e);
def get_course(cur,price):
    return get_html(KAY+"/get_course/?cur={0}&price={1}".format(cur,price));
def giveCoin(login,kay_ice,how_many):
    try:
        conn=sqlite3.connect('markets.sqlite');
        cursor=conn.cursor();
        cursor.execute("SELECT tg_id FROM 'coinInfo' WHERE login=(?)",(login,));
        tg_id=cort_to_list(cursor.fetchall())[0];
        cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?)",(login,));
        tg_pswd=cort_to_list(cursor.fetchall())[0];
        conn.close();
        get_html(KAY+"/give_coin/?tg_id={0}&tg_pswd={1}&kay_ice={2}&how_many={3}".format(tg_id,tg_pswd,kay_ice,how_many,));
        print(KAY+"/give_coin/?tg_id={0}&tg_pswd={1}&kay_ice={2}&how_many={3}".format(tg_id,tg_pswd,kay_ice,how_many,))
    except Exception as e:
        print(e);
def createTimeUser(adres,phone,merchName,orderDistrict,phone2):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT login FROM 'users'");
    logins=cort_to_list(cursor.fetchall());
    i=1000;
    login="USER_"+str(i);
    while login in logins:
        login="USER_"+str(i);
        i+=1;
    cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
        (login,login,"_",merchName,"market","0.0","0.0","0","ru",));
    conn.commit();
    conn.close();
    conn=sqlite3.connect('markets.sqlite');
    cursor=conn.cursor();
    cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?))",
        (login,adres,"0",phone,"0",orderDistrict,phone2));
    conn.commit();
    conn.close();
    return login;
def removeUser(merchName,user):
    conn=sqlite3.connect("basic.sqlite");
    cursor=conn.cursor();
    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(user,));
    conn.commit();
    conn.close();
    conn=sqlite3.connect("markets.sqlite");
    cursor=conn.cursor();
    cursor.execute("DELETE FROM 'users' WHERE login = (?)",(user,));
    conn.commit();
    conn.close();
def handle_uploaded_file(f,Path,name):
    with open(Path+'/'+name+'.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk);
def async2(login,updates,merchName="0"):
    asy = threading.Thread(target=async3, args=(login,updates,merchName));
    asy.start()
def async3(login,updates,merchName):
    conn=sqlite3.connect("basic.sqlite");
    cursor=conn.cursor();
    if merchName=="0":
        cursor.execute("SELECT updates FROM 'users' WHERE login=(?)",(login,));
    else:
        cursor.execute("SELECT updates FROM 'users' WHERE login=(?)AND merchName=(?)",(login,merchName,));
    try:
        oldUpdates=cort_to_list(cursor.fetchall())[0];
    except:
        oldUpdates="";
    if oldUpdates=="0":
        update=updates+"|^";
    else:
        if updates not in oldUpdates:
            update=oldUpdates+updates+"|^";
        else:
            update=oldUpdates;
    if merchName=="0":
        cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)",(update,login,));
    else:
        cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)AND merchName=(?)",(update,login,merchName,));
    conn.commit();
    conn.close();
    send="OK";
    return send;
def notification(merchName,login,title,body):
    nots=getConst(merchName,"nots");
    if nots=="tg":
        messageInChannel(merchName,title+"\n"+body);
    elif nots=="app":
        conn=sqlite3.connect(mPath(merchName,"orders"));
        cursor=conn.cursor();
        cursor.execute("UPDATE 'const' SET val=(?)WHERE key='notTitle'",(title,));
        cursor.execute("UPDATE 'const' SET val=(?)WHERE key='notBody'",(body,));
        conn.commit();
        conn.close();
        async2(login,"sendAboutNot")
    else:
        print(">>>>");
        print(title);
        print(body);
        print(">>>>");   
def checkSession(session):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT session FROM 'users' WHERE session=(?)",(session,));
    length=len(cort_to_list(cursor.fetchall()));
    conn.close();
    if length==1:
        return True;
    else:
        return False;
def genSession(data):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT session FROM 'users'");
    sessions=cort_to_list(cursor.fetchall());
    session=generate();
    while session in sessions:
        session=generate();
    cursor.execute("UPDATE users SET session=(?)WHERE login=(?)",(session,data,));
    conn.commit();
    conn.close();
    return session;
def getBuyType(myMerchName,login):
    if "USER_" in login or myMerchName=="self":
        return "user";
    else:
        try:
            conn=sqlite3.connect(mPath(myMerchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(login,));
            Type=cort_to_list(cursor.fetchall())[0];
            conn.close();
            return Type;
        except:
            Def=getConst(myMerchName,"default_buyer");
            conn=sqlite3.connect(mPath(myMerchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("UPDATE buy_markets SET type=(?)WHERE name=(?)",(Def,login,));
            conn.commit();
            conn.close();
            return Def;
def getProductPrice(merchName,market,prod_id):
    buyType=getBuyType(merchName,market);
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    cursor.execute("SELECT price FROM 'prices' WHERE name=(?)AND id=(?)",(buyType,prod_id,));
    price=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return price;
def Img(merchName,cat_prod,prod_id):
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    try:
        cursor.execute("SELECT place FROM 'img' WHERE _id=(?)AND type=(?)",(prod_id,cat_prod,));
        img=cort_to_list(cursor.fetchall())[0];
    except:
        img=getConst(merchName,"merchImg");
    conn.close();
    ret="{0}get_photo/?link={1}".format(URL,img);
    return ret;
def getConst(myMerchName,key):
    try:
        if key=="autoNull":
            conn=sqlite3.connect("basic.sqlite");
            cursor=conn.cursor();
            cursor.execute("SELECT drivers FROM 'clearTime' WHERE merchName=(?)",(myMerchName,));
            val=cort_to_list(cursor.fetchall())[0];
            conn.close();
        else:
            conn=sqlite3.connect(mPath(myMerchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT val FROM 'const' WHERE key=(?)",(key,));
            val=cort_to_list(cursor.fetchall())[0];
            conn.close();
        return val;
    except Exception as e:
        print(e);
        return "null";
def Last(merchName,newLast="0"):
    newLast=str(newLast);
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    cursor.execute("SELECT val FROM 'const' WHERE key='last_index'");
    last_index=cort_to_list(cursor.fetchall())[0];
    if newLast=="0":
        conn.close();
        return last_index;
    else:
        newLast=str(int(last_index)+1);
        cursor.execute("UPDATE 'const' SET val=(?)WHERE key='last_index'",(newLast,));
        conn.commit();
        conn.close();
        return "OK";
def makeAct(merchName,login,data):
    d=datetime.now();
    date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    cursor.execute("INSERT INTO actHistory VALUES((?),(?),(?))",(login,data,date));
    conn.commit();
    conn.close();
def checkMarket(myMerchName,mark):
    if "USER_" in str(mark):
        return "";
    else:
        if myMerchName=="self":
            merchs=os.listdir("merchants");
            for myMerchName in merchs:
                conn=sqlite3.connect(mPath(myMerchName,"orders"));
                cursor=conn.cursor();
                cursor.execute("SELECT name FROM 'buy_markets' WHERE name=(?)",(mark,));
                markets=cort_to_list(cursor.fetchall());
                if len(markets)==0:
                    default_buyer=getConst(myMerchName,"default_buyer");
                    defDolgType=getConst(myMerchName,"defDolgType");
                    defDolgVal=getConst(myMerchName,"defDolgVal");
                    defDolgMax=getConst(myMerchName,"defDolgMax");
                    cursor.execute("INSERT INTO buy_markets VALUES((?),(?),(?),(?),(?),(?),(?))",
                                (mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0"));
                    conn.commit();
                    conn.close();
                else:
                    #conn.close();
                    pass;
        else:
            conn=sqlite3.connect(mPath(myMerchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT name FROM 'buy_markets' WHERE name=(?)",(mark,));
            markets=cort_to_list(cursor.fetchall());
            if len(markets)==0:
                default_buyer=getConst(myMerchName,"default_buyer");
                defDolgType=getConst(myMerchName,"defDolgType");
                defDolgVal=getConst(myMerchName,"defDolgVal");
                defDolgMax=getConst(myMerchName,"defDolgMax");
                cursor.execute("INSERT INTO buy_markets VALUES((?),(?),(?),(?),(?),(?),(?))",
                            (mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0"));
                conn.commit();
                conn.close();
            else:
                #conn.close();
                pass;
def checkCoinSettings(login):
    conn=sqlite3.connect("markets.sqlite");
    cursor=conn.cursor();
    cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?)",(login,));
    logins=cort_to_list(cursor.fetchall());
    if len(logins)==0:
        cursor.execute("INSERT INTO coinInfo VALUES((?),(?),(?))",(login,"",""));
        conn.commit();
    conn.close();
def getDolgInfo(myMerchName,login):
    try:
        if myMerchName!="self":
            if "USER_" in login:
                login="USER_";
            conn=sqlite3.connect(mPath(myMerchName,"orders"));
            cursor=conn.cursor();
            cursor.execute("SELECT dolgType FROM 'buy_markets' WHERE name=(?)",(login,));
            dolgType=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT dolgVal FROM 'buy_markets' WHERE name=(?)",(login,));
            dolgVal=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT dolgHave FROM 'buy_markets' WHERE name=(?)",(login,));
            dolgHave=cort_to_list(cursor.fetchall())[0];
            cursor.execute("SELECT dolgMax FROM 'buy_markets' WHERE name=(?)",(login,));
            dolgMax=cort_to_list(cursor.fetchall())[0];
            conn.close();
            vac={"dolgType":dolgType,"dolgVal":dolgVal,"dolgHave":dolgHave,"dolgMax":dolgMax};
            return vac;
        else:
            vac={"dolgType":"how","dolgVal":"0","dolgHave":"0","dolgMax":"0"};
            return vac;
    except:
        vac={"dolgType":"how","dolgVal":"0","dolgHave":"0","dolgMax":"0"};
        return vac;
def User(data,l=1):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    if l==1:
        cursor.execute("SELECT user_type FROM 'users' WHERE session=(?)",(data,));
        user_type=cort_to_list(cursor.fetchall())[0];
    else:
        cursor.execute("SELECT user_type FROM 'users' WHERE login=(?)",(data,));
        user_type=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return user_type;
def Login(data):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT login FROM 'users' WHERE session=(?)",(data,));
    login=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return login;
def Merch(data,l=1):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    if l==1:
        cursor.execute("SELECT merchName FROM 'users' WHERE session=(?)",(data,));
        merchName=cort_to_list(cursor.fetchall())[0];
    else:
        cursor.execute("SELECT merchName FROM 'users' WHERE login=(?)",(data,));
        merchName=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return merchName;
def Level(merchName):
    conn=sqlite3.connect('basic.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT level FROM 'levels' WHERE merchName=(?)",(merchName,));
    level=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return level;
def ADLevel(merchName,login):
    conn=sqlite3.connect(mPath(merchName,"orders"));
    cursor=conn.cursor();
    cursor.execute("SELECT level FROM 'admins' WHERE login=(?)",(login,));
    level=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return level;
def District(login):
    conn=sqlite3.connect('markets.sqlite');
    cursor=conn.cursor();
    cursor.execute("SELECT district FROM 'users' WHERE login=(?)",(login,));
    district=cort_to_list(cursor.fetchall())[0];
    conn.close();
    return district;
def Dolg(login,market):
    conn=sqlite3.connect(mPath(Merch(session),"orders"));
    cursor=conn.cursor();
    cursor.execute("SELECT data FROM order WHERE owner=(?)AND getter=(?)AND get_type='dolg'"
                ,(login,market,));
    dolgs=cort_to_list(cursor.fetchall());
    conn.close();
    if len(dolgs)==0:
        return True;
    else:
        return False; 
def mPath(merchName,base):
    return "merchants/"+merchName+"/"+base+'.sqlite';
def generate():
    gen=["q","w","e","r","t","y","u","i","o","p","a","s","d","f","g","h","j","k","l","z","x","c","v","b","n","m"]
    Gen=[]
    for i in gen:
        Gen.append(i.upper());
    nums=["0","1","2","3","4","5","6","7","8","9"];
    globList=gen+Gen+nums;

    word=""
    for i in range(32):
        symbol=globList[random.randint(0,len(globList)-1)];
        word+=str(symbol);
    return word;
def stringToArray(string):
    prods=[];
    prod_mass=[];
    tov=""
    for i in string:
        if i=="^":
            prod_mass.append(prods);
            prods=[];
        elif i =="|":
            prods.append(tov);
            tov="";
        else:
            tov=tov+i;
    return prod_mass;
def stringToMap(string):
    string=string+",,";
    Map={};
    a="";
    b="";
    var="a";
    try:
        for z in range(len(string)):
            i=string[z];
            if i=="=":
                var="b";
            elif string[z]=="," and string[z+1]==",":
                Map[a.replace(",","")]=b;
                var="a";
                a="";b="";
            else:
                if var=="a":
                    a=a+i;
                elif var=="b":
                    b=b+i;
    except:
        pass
    return Map;
def stringToArrayData(string):
    prods=[];
    prod_mass=[];
    tov=""
    for i in string:
        if i==";":
            prod_mass.append(prods);
            prods=[];
        elif i ==":":
            prods.append(tov);
            tov="";
        else:
            tov=tov+i;
    return prod_mass;
def cort_to_list(cort):
    list_=[];
    for i in range(len(cort)):
        list_.append(cort[i][0]);
    return list_;
def arrayToString(maxList):
    string="";
    for Min in maxList:
        for m in Min:
            string=string+str(m)+"|";
        string=string+"^";
    return string;
def messageInChannel(merchName,text):
    try:
        tg_token=getConst(merchName,"tg_token").replace(" ","");
        tg_id=getConst(merchName,"tg_id").replace(" ","");
        print(tg_token)
        print(tg_id)
        req = "https://api.telegram.org/bot"+tg_token+"/sendMessage?chat_id="+tg_id+"&text="+text;
        print(get_html(req).text);
    except Exception as e:
        logger(e);
def locationInChannel(merchName,lon,lat):
    try:
        conn=sqlite3.connect("basic.sqlite");
        cursor=conn.cursor();
        cursor.execute("SELECT token FROM 'telegram' WHERE merchName=(?)",(merchName,));
        token=cort_to_list(cursor.fetchall())[0];
        cursor.execute("SELECT chat_id FROM 'telegram' WHERE merchName=(?)",(merchName,));
        chat_id=cort_to_list(cursor.fetchall())[0];
        conn.close();
        req = "https://api.telegram.org/bot"+token+"/sendLocation?chat_id="+chat_id+"&latitude="+lat+"&longitude="+lon
        get_html(req);
    except Exception as e:
        logger(e);
def logger(e):
    try:
        logging.error(e);
        file=open("log/mylog.txt","w");
        file.write(traceback.format_exc());
        file.close();
        file=open("log/mylog.txt","r");
        text=file.read();
        file.close();
        req = "https://api.telegram.org/bot"+token+"/sendMessage?chat_id="+MYTGID+"&text="+text;
        get_html(req);
    except:
        print(e);
def get_html(url,params=None):
    HEADERS ={};
    r = rq.get(url,headers=HEADERS,params=params)
    return r
def get_address_from_coords1(coords):
    try:
        PARAMS = {
            "apikey":"dcc7de33-5acb-4746-9558-a2bfbccc8391",
            "format":"json",
            "lang":"ru_RU",
            "kind":"house",
            "geocode": coords
        }
        r = rq.get(url="https://geocode-maps.yandex.ru/1.x/", params=PARAMS)
        json_data = r.json()
        address_str = json_data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["metaDataProperty"]["GeocoderMetaData"]["AddressDetails"]["Country"]["AddressLine"]
        return address_str;
    except Exception as e:
        print(e);
        return "Адрес не установлен"
def get_address_from_coords2(coords):
    try:
        PARAMS = {
            "apikey":"dcc7de33-5acb-4746-9558-a2bfbccc8391",
            "format":"json",
            "lang":"ru_RU",
            "kind":"district",
            "geocode": coords
        }
        r = rq.get(url="https://geocode-maps.yandex.ru/1.x/", params=PARAMS)
        json_data = r.json()
        address_str2 = json_data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["metaDataProperty"]["GeocoderMetaData"]["AddressDetails"]["Country"]["AddressLine"]
        return address_str2;
    except Exception as e:
        print(e);
        return "Адрес не установлен"



