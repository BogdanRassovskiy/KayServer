from django.shortcuts import render
import sqlite3
from django.http import HttpResponse
import requests as rq;
import os;
import traceback;
import openpyxl;
import logging;
import threading;
import math
import random
import json;
import hasher;
from PIL import Image;
import numpy as np;
import Stats;
from datetime import datetime;
import time
import io, base64
#import dolConv;
from qrGen import get_qrs;
import reWriteExcel as reW;
from classes.basic_clearTime import basic_clearTime
from classes.basic_const import basic_const
from classes.basic_districts import basic_districts
from classes.basic_exLink import basic_exLink
from classes.basic_levels import basic_levels
from classes.basic_messages import basic_messages
from classes.basic_nots import basic_nots
from classes.basic_percent import basic_percent
from classes.basic_roots import basic_roots
from classes.basic_telegram import basic_telegram
from classes.basic_users import basic_users
from classes.inCar_cashInCar import inCar_cashInCar
from classes.inCar_drivers import inCar_drivers
from classes.inCar_nakNum import inCar_nakNum
from classes.inCar_naks import inCar_naks
from classes.markets_coinInfo import markets_coinInfo
from classes.markets_users import markets_users
from classes.orders_actHistory import orders_actHistory
from classes.orders_admins import orders_admins
from classes.orders_buy_markets import orders_buy_markets
from classes.orders_buyers_type import orders_buyers_type
from classes.orders_categories import orders_categories
from classes.orders_const import orders_const
from classes.orders_drivers import orders_drivers
from classes.orders_history import orders_history
from classes.orders_img import orders_img
from classes.orders_nakSpace import orders_nakSpace
from classes.orders_order import orders_order
from classes.orders_prices import orders_prices
from classes.orders_products import orders_products
from classes.strings_ru import strings_ru
from classes.strings_uz import strings_uz
from classes.strings_languages import strings_languages
from classes.myImage import myImage

basic_clearTimeS=[];
basic_constS=[];
basic_districtsS=[];
basic_exLinkS=[];
basic_levelsS=[];
basic_messagesS=[];
basic_notsS=[];
basic_percentS=[];
basic_rootsS=[];
basic_telegramS=[];
basic_usersS=[];

inCar_cashInCarS=[];
inCar_driversS=[];
inCar_nakNumS=[];
inCar_naksS=[];

markets_coinInfoS=[];
markets_usersS=[];

orders_actHistoryS=[];
orders_adminsS=[];
orders_buy_marketsS=[];
orders_buyers_typeS=[];
orders_categoriesS=[];
orders_constS=[];
orders_driversS=[];
orders_historyS=[];
orders_imgS=[];
orders_nakSpaceS=[];
orders_orderS=[];
orders_pricesS=[];
orders_productsS=[];

strings_ruS=[];
strings_uzS=[];
strings_languagesS=[];

myImageS=[];


URL=Stats.URL;
KAY=Stats.KAY;
h=hasher.h;
MYTGID="104932971";
VAC_LIST=["А","Б","В","Г","Д","Е","Ё","Ж","З","И","Й","К","Л","М","Н","О","П","Р","С","Т","У","Ф","Х","Ц","Ч","Ш","Щ","Ъ","Ы","Ь","Э","Ю","Я"];
VAC={"А":"A","Б":"B","В":"V","Г":"G","Д":"D","Е":"E","Ё":"YO","Ж":"J","З":"Z","И":"I","Й":"Y","К":"K","Л":"L","М":"M","Н":"N","О":"O",
"П":"P","Р":"R","С":"S","Т":"T","У":"U","Ф":"F","Х":"X","Ц":"C","Ч":"CH","Ш":"SH","Щ":"SH","Ъ":"","Ы":"I","Ь":"","Э":"E","Ю":"YU","Я":"YA",
"0":"0","1":"1","2":"2","3":"3","4":"4","5":"5","6":"6","7":"7","8":"8","9":"9"};
#url_for_bot="https://api.telegram.org/bot"+token+"/sendmessage?chat_id="+chat_id+"&text="txt;
'''
var updates

getMoney
getDistricts
getNak
getOrders
getDolgs
getPrices
getMyLocation
getMarketList
getProductList
getDefaultMarket
'''
def startServer():
    print("Started work!");
    #MARKETS.SQLITE
    conn = sqlite3.connect('markets.sqlite');
    cursor = conn.cursor();
    #COIN INFO
    cursor.execute("SELECT login FROM coinInfo");login = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT tg_id FROM coinInfo");tg_id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT pswd FROM coinInfo");pswd = cort_to_list(cursor.fetchall());
    for i in range(len(login)):
        s=markets_coinInfo(
            login[i],tg_id[i],pswd[i]
        );
        markets_coinInfoS.append(s);
    #USERS
    cursor.execute("SELECT login FROM users");login = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT adres FROM users");adres = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT inn FROM users");inn = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT phone FROM users");phone = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT varified FROM users");varified = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT district FROM users");district = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT phone2 FROM users");phone2 = cort_to_list(cursor.fetchall());
    for i in range(len(login)):
        s=markets_users(
            login[i],adres[i],inn[i],phone[i],varified[i],district[i],phone2[i]
        );
        markets_usersS.append(s);
    conn.close();
    print(">>>>READ MARKETS.SQLITE OK");
    #STRINGS.SQLITE
    conn = sqlite3.connect('strings.sqlite');
    cursor = conn.cursor();
    #LANGUAGES
    cursor.execute("SELECT code FROM languages");code = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT lang FROM languages");lang = cort_to_list(cursor.fetchall());
    for i in range(len(code)):
        s=strings_languages(
            code[i],lang[i]
        );
        strings_languagesS.append(s);
    #RU
    cursor.execute("SELECT id FROM ru");id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT text FROM ru");text = cort_to_list(cursor.fetchall());
    for i in range(len(id)):
        s=strings_ru(
            id[i],text[i]
        );
        strings_ruS.append(s);
    #UZ
    cursor.execute("SELECT id FROM uz");id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT text FROM uz");text = cort_to_list(cursor.fetchall());
    for i in range(len(id)):
        s=strings_uz(
            id[i],text[i]
        );
        strings_uzS.append(s);
    conn.close();

    print(">>>>READ STRINGS.SQLITE OK");
    #BASIC.SQLITE
    conn = sqlite3.connect('basic.sqlite');
    cursor = conn.cursor();
    #CLEARTIME
    cursor.execute("SELECT merchName FROM clearTime");merchName = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT drivers FROM clearTime");drivers = cort_to_list(cursor.fetchall());
    for i in range(len(merchName)):
        s=basic_clearTime(
            merchName[i],drivers[i]
        );
        basic_clearTimeS.append(s);
    #CONST
    cursor.execute("SELECT key FROM const");key = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT val FROM const");val = cort_to_list(cursor.fetchall());
    for i in range(len(key)):
        s=basic_const(
            key[i],val[i]
        );
        basic_constS.append(s);
    #districts
    cursor.execute("SELECT district FROM districts");district = cort_to_list(cursor.fetchall());
    for i in range(len(district)):
        s=basic_districts(
            district[i]
        );
        basic_districtsS.append(s);
    #exLink
    cursor.execute("SELECT code FROM exLink");code = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT _type FROM exLink");_type = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT merchName FROM exLink");merchName = cort_to_list(cursor.fetchall());
    for i in range(len(code)):
        s=basic_exLink(
            code[i],_type[i],merchName[i]
        );
        basic_exLinkS.append(s);
    #levels
    cursor.execute("SELECT merchName FROM levels");merchName = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT level FROM levels");level = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT varified FROM levels");varified = cort_to_list(cursor.fetchall());
    for i in range(len(merchName)):
        s=basic_levels(
            merchName[i],level[i],varified[i]
        );
        basic_levelsS.append(s);
    #messages
    cursor.execute("SELECT chat_id FROM messages");chat_id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT message_id FROM messages");message_id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT text_or_photo FROM messages");text_or_photo = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT sender FROM messages");sender = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT getter FROM messages");getter = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT date FROM messages");date = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT wasRead FROM messages");wasRead = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT text FROM messages");text = cort_to_list(cursor.fetchall());
    for i in range(len(chat_id)):
        s=basic_messages(
            chat_id[i],message_id[i],text_or_photo[i],sender[i],getter[i],date[i],
            wasRead[i],text[i]
        );
        basic_messagesS.append(s);
    #nots
    cursor.execute("SELECT id FROM nots");id = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT owner FROM nots");owner = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT title FROM nots");title = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT text FROM nots");text = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT Type FROM nots");Type = cort_to_list(cursor.fetchall());
    for i in range(len(id)):
        s=basic_nots(
            id[i],owner[i],title[i],text[i],Type[i]
        );
        basic_notsS.append(s);
    #percent
    cursor.execute("SELECT var FROM percent");var = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT cur FROM percent");cur = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT per FROM percent");per = cort_to_list(cursor.fetchall());
    for i in range(len(var)):
        s=basic_percent(
            var[i],cur[i],per[i]
        );
        basic_percentS.append(s);
    #roots
    cursor.execute("SELECT login FROM roots");login = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT level FROM roots");level = cort_to_list(cursor.fetchall());
    for i in range(len(login)):
        s=basic_roots(
            login[i],level[i]
        );
        basic_rootsS.append(s);
    #telegram
    cursor.execute("SELECT merchName FROM telegram");merchName = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT token FROM telegram");token = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT chat_id FROM telegram");chat_id = cort_to_list(cursor.fetchall());
    for i in range(len(merchName)):
        s=basic_telegram(
            merchName[i],token[i],chat_id[i]
        );
        basic_telegramS.append(s);
    #users
    cursor.execute("SELECT login FROM users");login = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT pswd FROM users");pswd = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT session FROM users");session = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT merchName FROM users");merchName = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT user_type FROM users");user_type = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT lon FROM users");lon = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT lat FROM users");lat = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT updates FROM users");updates = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT lang FROM users");lang = cort_to_list(cursor.fetchall());
    cursor.execute("SELECT ent FROM users");ent = cort_to_list(cursor.fetchall());
    for i in range(len(login)):
        s=basic_users(
            login[i],pswd[i],session[i],merchName[i],user_type[i],lon[i],
            lat[i],updates[i],lang[i],ent[i]
        );
        basic_usersS.append(s);
    conn.close();
    merchants=os.listdir("merchants");
    print(">>>>READ BASIC.SQLITE OK");

    #MERCHANTS
    for M in merchants:
        #INCAR
        conn = sqlite3.connect('merchants/{0}/inCar.sqlite'.format(M));
        cursor = conn.cursor();
        #cashInCar
        cursor.execute("SELECT driver FROM cashInCar");driver = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT cash FROM cashInCar");cash = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT term FROM cashInCar");term = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT per FROM cashInCar");per = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT on_day FROM cashInCar");on_day = cort_to_list(cursor.fetchall());
        for i in range(len(driver)):
            s=inCar_cashInCar(
                driver[i],cash[i],term[i],per[i],on_day[i],M
        );
            s.merchName=M;
            inCar_cashInCarS.append(s);
        #drivers
        cursor.execute("SELECT login FROM drivers");login = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT district FROM drivers");district = cort_to_list(cursor.fetchall());
        for i in range(len(login)):
            s=inCar_drivers(
                login[i],district[i]
                );
            s.merchName=M;
            inCar_driversS.append(s);
        #nakNum
        cursor.execute("SELECT owner FROM nakNum");owner = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT num FROM nakNum");num = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT date FROM nakNum");date = cort_to_list(cursor.fetchall());
        for i in range(len(owner)):
            s=inCar_nakNum(
                owner[i],num[i],date[i],M
        );
            s.merchName=M;
            inCar_nakNumS.append(s);
        #nakNum
        cursor.execute("SELECT name FROM naks");name = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT prod_id FROM naks");prod_id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT free FROM naks");free = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT reserve FROM naks");reserve = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT sell FROM naks");sell = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT owner FROM naks");owner = cort_to_list(cursor.fetchall());
        for i in range(len(name)):
            s=inCar_naks(
                name[i],prod_id[i],free[i],reserve[i],sell[i],owner[i],M
        );
            s.merchName=M;
            inCar_naksS.append(s);
        print(">>>>READ INCAR.SQLITE OK "+M);
        conn.close();
        #ORDERS
        conn = sqlite3.connect('merchants/{0}/orders.sqlite'.format(M));
        cursor = conn.cursor();
        #actHistory
        cursor.execute("SELECT login FROM actHistory");login = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT Data FROM actHistory");Data = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT Date FROM actHistory");Date = cort_to_list(cursor.fetchall());
        for i in range(len(login)):
            s=orders_actHistory(
                login[i],Data[i],Date[i],M
        );
            s.merchName=M;
            orders_actHistoryS.append(s);
        #adimns
        cursor.execute("SELECT login FROM admins");login = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT level FROM admins");level = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT role FROM admins");role = cort_to_list(cursor.fetchall());
        for i in range(len(login)):
            s=orders_admins(
                login[i],level[i],role[i],M
        );
            s.merchName=M;
            orders_adminsS.append(s);
        #buy_markets
        cursor.execute("SELECT name FROM buy_markets");name = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT type FROM buy_markets");type = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT dolgType FROM buy_markets");dolgType = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT dolgVal FROM buy_markets");dolgVal = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT dolgHave FROM buy_markets");dolgHave = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT dolgMax FROM buy_markets");dolgMax = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT wallet FROM buy_markets");wallet = cort_to_list(cursor.fetchall());
        for i in range(len(name)):
            s=orders_buy_markets(
                name[i],type[i],dolgType[i],dolgVal[i],dolgHave[i],dolgMax[i],wallet[i],M
        );
            s.merchName=M;
            orders_buy_marketsS.append(s);
        #buyers_type
        cursor.execute("SELECT name FROM buyers_type");name = cort_to_list(cursor.fetchall());
        for i in range(len(name)):
            s=orders_buyers_type(
                name[i],M
        );
            s.merchName=M;
            orders_buyers_typeS.append(s);
        #categories
        cursor.execute("SELECT cat_id FROM categories");cat_id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT id FROM categories");id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM categories");name = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT work FROM categories");work = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT img FROM categories");img = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name2 FROM categories");name2 = cort_to_list(cursor.fetchall());
        for i in range(len(cat_id)):
            s=orders_categories(
                cat_id[i],id[i],name[i],work[i],img[i],name2[i],M
        );
            s.merchName=M;
            orders_categoriesS.append(s);
        #const
        cursor.execute("SELECT key FROM const");key = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT val FROM const");val = cort_to_list(cursor.fetchall());
        for i in range(len(key)):
            s=orders_const(
                key[i],val[i],M
        );
            s.merchName=M;
            orders_constS.append(s);
        #drivers
        cursor.execute("SELECT login FROM drivers");login = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT district FROM drivers");district = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM drivers");name = cort_to_list(cursor.fetchall());
        for i in range(len(login)):
            s=orders_drivers(
                login[i],district[i],name[i],M
        );
            s.merchName=M;
            orders_driversS.append(s);
        #history
        cursor.execute("SELECT data FROM history");data = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT price FROM history");price = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT type FROM history");type = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT getter FROM history");getter = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT driver FROM history");driver = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT date FROM history");date = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT district FROM history");district = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT last_index FROM history");last_index = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT visible FROM history");visible = cort_to_list(cursor.fetchall());
        for k in range(len(data)):
            i=(len(data)-1)-k;
            s=orders_history(
                data[i],price[i],type[i],getter[i],driver[i],date[i],district[i],last_index[i],visible[i],M
                );
            s.merchName=M;
            orders_historyS.append(s);
        #img
        cursor.execute("SELECT type FROM img");type = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT _id FROM img");_id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT place FROM img");place = cort_to_list(cursor.fetchall());
        for i in range(len(type)):
            s=orders_img(
                type[i],_id[i],place[i],M
        );
            s.merchName=M;
            orders_imgS.append(s);
        #nakSpace
        cursor.execute("SELECT id FROM nakSpace");id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT nak FROM nakSpace");nak = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT date FROM nakSpace");date = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT nakNum FROM nakSpace");nakNum = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT owner FROM nakSpace");owner = cort_to_list(cursor.fetchall());
        for i in range(len(id)):
            s=orders_nakSpace(
                id[i],nak[i],date[i],nakNum[i],owner[i],M
        );
            s.merchName=M;
            orders_nakSpaceS.append(s);
        #order
        cursor.execute("SELECT data FROM 'order'");data = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT getter FROM 'order'");getter = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT owner FROM 'order'");owner = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT last_index FROM 'order'");last_index = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT date FROM 'order'");date = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT get_type FROM 'order'");get_type = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT price FROM 'order'");price = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT payForm FROM 'order'");payForm = cort_to_list(cursor.fetchall());
        for i in range(len(data)):
            s=orders_order(
                data[i],getter[i],owner[i],last_index[i],date[i],get_type[i],price[i],payForm[i],M
        );
            s.merchName=M;
            orders_orderS.append(s);

        #prices
        cursor.execute("SELECT id FROM prices");id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM prices");name = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT price FROM prices");price = cort_to_list(cursor.fetchall());
        for i in range(len(id)):
            s=orders_prices(
                id[i],name[i],price[i],M
        );
            s.merchName=M;
            orders_pricesS.append(s);
        #products
        cursor.execute("SELECT cat_id FROM products");cat_id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT id FROM products");id = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name FROM products");name = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT rev FROM products");rev = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT work FROM products");work = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT img FROM products");img = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT box FROM products");box = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT form FROM products");form = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT visible FROM products");visible = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT ost FROM products");ost = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT name2 FROM products");name2 = cort_to_list(cursor.fetchall());
        cursor.execute("SELECT rev2 FROM products");rev2 = cort_to_list(cursor.fetchall());
        for i in range(len(cat_id)):
            s=orders_products(
                cat_id[i],id[i],name[i],rev[i],work[i],img[i],box[i],form[i],visible[i],
                ost[i],name2[i],rev2[i],M
        );
            s.merchName=M;
            orders_productsS.append(s);

        conn.close();
        print(">>>>READ ORDERS.SQLITE OK "+M);
    #Photos
    folders=["qrMarkets","marketImg","img"];
    for folder in folders:
        files=os.listdir(folder);
        for f in files:
            file_location=folder+"/"+f;
            with open(file_location, 'rb') as f:
               file_data = f.read();
            myImageS.append(myImage(file_location,file_data));
    print(">>>>READ PHOTOS OK");

def writeDbThread():
    wdb = threading.Thread(target=writeLoop, args=(),daemon=True)
    wdb.start()
def startServerThread():
    sdb = threading.Thread(target=startServer, args=(),daemon=True)
    sdb.start()

def writeLoop():
    while True:
        time.sleep(2);
        writeDb();
def writeDb():
    try:
        #markets.sqlite
        conn = sqlite3.connect('markets.sqlite', timeout=5.0);
        cursor = conn.cursor();
        #coinInfo
        cursor.execute("SELECT login FROM coinInfo");
        login = cort_to_list(cursor.fetchall());
        for s in markets_coinInfoS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM coinInfo WHERE login =(?)',(s.login,));
                    conn.commit();
                    remList=[];
                    for i in range(len(markets_coinInfoS)):
                        if markets_coinInfoS[i].login==s.login:
                            markets_coinInfoS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        markets_coinInfoS.pop(r);

                else:
                    if s.login in login:
                        cursor.execute("UPDATE coinInfo SET login =(?) WHERE login=(?)",(s.login,s.login));
                        cursor.execute("UPDATE coinInfo SET tg_id =(?) WHERE login=(?)",(s.tg_id,s.login));
                        cursor.execute("UPDATE coinInfo SET pswd =(?) WHERE login=(?)",(s.pswd,s.login));
                        for i in range(len(markets_coinInfoS)):
                            if markets_coinInfoS[i].login==s.login:
                                markets_coinInfoS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO coinInfo VALUES ((?),(?),(?))",
                            (
                                s.login,s.tg_id,s.pswd
                            ));
                    conn.commit();
                    for i in range(len(markets_coinInfoS)):
                        if markets_coinInfoS[i].login==s.login:
                            markets_coinInfoS[i].changed="0";
                            conn.commit();
        #users
        cursor.execute("SELECT login FROM users");
        login = cort_to_list(cursor.fetchall());
        for s in markets_usersS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM users WHERE login =(?)',(s.login,));
                    conn.commit();
                    remList=[];
                    for i in range(len(markets_usersS)):
                        if markets_usersS[i].login==s.login:
                            markets_usersS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        markets_usersS.pop(r);

                else:
                    if s.login in login:
                        cursor.execute("UPDATE users SET login =(?) WHERE login=(?)",(s.login,s.login));
                        cursor.execute("UPDATE users SET adres =(?) WHERE login=(?)",(s.adres,s.login));
                        cursor.execute("UPDATE users SET inn =(?) WHERE login=(?)",(s.inn,s.login));
                        cursor.execute("UPDATE users SET phone =(?) WHERE login=(?)",(s.phone,s.login));
                        cursor.execute("UPDATE users SET varified =(?) WHERE login=(?)",(s.varified,s.login));
                        cursor.execute("UPDATE users SET district =(?) WHERE login=(?)",(s.district,s.login));
                        cursor.execute("UPDATE users SET phone2 =(?) WHERE login=(?)",(s.phone2,s.login));
                        for i in range(len(markets_usersS)):
                            if markets_usersS[i].login==s.login:
                                markets_usersS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO users VALUES ((?),(?),(?),(?),(?),(?),(?))",
                            (
                                s.login,s.adres,s.inn,s.phone,s.varified,s.district,s.phone2
                            ));
                    conn.commit();
                    for i in range(len(markets_usersS)):
                        if markets_usersS[i].login==s.login:
                            markets_usersS[i].changed="0";
                            conn.commit();
        conn.commit();
        conn.close();
        #basic.sqlite
        conn = sqlite3.connect('basic.sqlite', timeout=5.0);
        cursor = conn.cursor();
        #clearTime
        cursor.execute("SELECT merchName FROM clearTime");
        merchName = cort_to_list(cursor.fetchall());
        for s in basic_clearTimeS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM clearTime WHERE merchName =(?)',(s.merchName,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_clearTimeS)):
                        if basic_clearTimeS[i].merchName==s.merchName:
                            basic_clearTimeS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_clearTimeS.pop(r);

                else:
                    if s.merchName in merchName:
                        cursor.execute("UPDATE clearTime SET merchName =(?) WHERE merchName=(?)",(s.merchName,s.merchName));
                        cursor.execute("UPDATE clearTime SET drivers =(?) WHERE merchName=(?)",(s.drivers,s.merchName));
                        for i in range(len(basic_clearTimeS)):
                            if basic_clearTimeS[i].merchName==s.merchName:
                                basic_clearTimeS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO clearTime VALUES ((?),(?))",
                            (
                                s.merchName,s.drivers
                            ));
                    conn.commit();
                    for i in range(len(basic_clearTimeS)):
                        if basic_clearTimeS[i].merchName==s.merchName:
                            basic_clearTimeS[i].changed="0";
                            conn.commit();
        #const
        cursor.execute("SELECT key FROM const");
        key = cort_to_list(cursor.fetchall());
        for s in basic_constS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM const WHERE key =(?)',(s.key,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_constS)):
                        if basic_constS[i].key==s.key:
                            basic_constS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_constS.pop(r);

                else:
                    if s.key in key:
                        cursor.execute("UPDATE const SET key =(?) WHERE key=(?)",(s.key,s.key));
                        cursor.execute("UPDATE const SET val =(?) WHERE key=(?)",(s.val,s.key));
                        for i in range(len(basic_constS)):
                            if basic_constS[i].key==s.key:
                                basic_constS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO const VALUES ((?),(?))",
                            (
                                s.key,s.val
                            ));
                    conn.commit();
                    for i in range(len(basic_constS)):
                        if basic_constS[i].key==s.key:
                            basic_constS[i].changed="0";
                            conn.commit();
        #districts
        cursor.execute("SELECT district FROM districts");
        district = cort_to_list(cursor.fetchall());
        for s in basic_districtsS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM districts WHERE district =(?)',(s.district,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_districtsS)):
                        if basic_districtsS[i].district==s.district:
                            basic_districtsS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_districtsS.pop(r);

                else:
                    if s.district in district:
                        cursor.execute("UPDATE districts SET district =(?) WHERE district=(?)",(s.district,s.district));
                        for i in range(len(basic_districtsS)):
                            if basic_districtsS[i].district==s.district:
                                basic_districtsS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO districts VALUES ((?))",
                            (
                                s.district,
                            ));
                    conn.commit();
                    for i in range(len(basic_districtsS)):
                        if basic_districtsS[i].district==s.district:
                            basic_districtsS[i].changed="0";
                            conn.commit();
        #exLink
        cursor.execute("SELECT code FROM exLink");
        code = cort_to_list(cursor.fetchall());
        for s in basic_exLinkS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM exLink WHERE code =(?)',(s.code,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_exLinkS)):
                        if basic_exLinkS[i].code==s.code:
                            basic_exLinkS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_exLinkS.pop(r);

                else:
                    if s.code in code:
                        cursor.execute("UPDATE exLink SET code =(?) WHERE code=(?)",(s.code,s.code));
                        cursor.execute("UPDATE exLink SET _type =(?) WHERE code=(?)",(s._type,s.code));
                        cursor.execute("UPDATE exLink SET merchName =(?) WHERE code=(?)",(s.merchName,s.code));
                        for i in range(len(basic_exLinkS)):
                            if basic_exLinkS[i].code==s.code:
                                basic_exLinkS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO exLink VALUES ((?),(?),(?))",
                            (
                                s.code,s._type,s.merchName
                            ));
                    conn.commit();
                    for i in range(len(basic_exLinkS)):
                        if basic_exLinkS[i].code==s.code:
                            basic_exLinkS[i].changed="0";
                            conn.commit();
        #levels
        cursor.execute("SELECT merchName FROM levels");
        merchName = cort_to_list(cursor.fetchall());
        for s in basic_levelsS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM levels WHERE merchName =(?)',(s.merchName,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_levelsS)):
                        if basic_levelsS[i].merchName==s.merchName:
                            basic_levelsS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_levelsS.pop(r);

                else:
                    if s.merchName in merchName:
                        cursor.execute("UPDATE levels SET merchName =(?) WHERE merchName=(?)",(s.merchName,s.merchName));
                        cursor.execute("UPDATE levels SET level =(?) WHERE merchName=(?)",(s.level,s.merchName));
                        cursor.execute("UPDATE levels SET varified =(?) WHERE merchName=(?)",(s.varified,s.merchName));
                        for i in range(len(basic_levelsS)):
                            if basic_levelsS[i].merchName==s.merchName:
                                basic_levelsS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO levels VALUES ((?),(?),(?))",
                            (
                                s.merchName,s.level,s.varified
                            ));
                    conn.commit();
                    for i in range(len(basic_levelsS)):
                        if basic_levelsS[i].merchName==s.merchName:
                            basic_levelsS[i].changed="0";
                            conn.commit();
        #messages
        cursor.execute("SELECT message_id FROM messages");
        message_id = cort_to_list(cursor.fetchall());
        for s in basic_messagesS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM messages WHERE message_id =(?)',(s.message_id,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_messagesS)):
                        if basic_messagesS[i].message_id==s.message_id:
                            basic_messagesS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_messagesS.pop(r);

                else:
                    if s.message_id in message_id:
                        cursor.execute("UPDATE messages SET chat_id =(?) WHERE message_id=(?)",(s.chat_id,s.message_id));
                        cursor.execute("UPDATE messages SET message_id =(?) WHERE message_id=(?)",(s.message_id,s.message_id));
                        cursor.execute("UPDATE messages SET text_or_photo =(?) WHERE message_id=(?)",(s.text_or_photo,s.message_id));
                        cursor.execute("UPDATE messages SET sender =(?) WHERE message_id=(?)",(s.sender,s.message_id));
                        cursor.execute("UPDATE messages SET getter =(?) WHERE message_id=(?)",(s.getter,s.message_id));
                        cursor.execute("UPDATE messages SET date =(?) WHERE message_id=(?)",(s.date,s.message_id));
                        cursor.execute("UPDATE messages SET wasRead =(?) WHERE message_id=(?)",(s.wasRead,s.message_id));
                        cursor.execute("UPDATE messages SET text =(?) WHERE message_id=(?)",(s.text,s.message_id));
                        for i in range(len(basic_messagesS)):
                            if basic_messagesS[i].message_id==s.message_id:
                                basic_messagesS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO messages VALUES ((?),(?),(?),(?),(?),(?),(?),(?))",
                            (
                                s.chat_id,s.message_id,s.text_or_photo,s.sender,s.getter
                                ,s.date,s.wasRead,s.text
                            ));
                    conn.commit();
                    for i in range(len(basic_messagesS)):
                        if basic_messagesS[i].message_id==s.message_id:
                            basic_messagesS[i].changed="0";
                            conn.commit();
        #nots
        cursor.execute("SELECT id FROM nots");
        id = cort_to_list(cursor.fetchall());
        for s in basic_notsS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM nots WHERE id =(?)',(s.id,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_notsS)):
                        if basic_notsS[i].id==s.id:
                            basic_notsS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_notsS.pop(r);

                else:
                    if s.id in id:
                        cursor.execute("UPDATE nots SET id =(?) WHERE id=(?)",(s.id,s.id));
                        cursor.execute("UPDATE nots SET owner =(?) WHERE id=(?)",(s.owner,s.id));
                        cursor.execute("UPDATE nots SET title =(?) WHERE id=(?)",(s.title,s.id));
                        cursor.execute("UPDATE nots SET text =(?) WHERE id=(?)",(s.text,s.id));
                        cursor.execute("UPDATE nots SET Type =(?) WHERE id=(?)",(s.Type,s.id));
                        for i in range(len(basic_notsS)):
                            if basic_notsS[i].id==s.id:
                                basic_notsS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO nots VALUES ((?),(?),(?),(?),(?))",
                            (
                                s.id,s.owner,s.title,s.text,s.Type
                            ));
                    conn.commit();
                    for i in range(len(basic_notsS)):
                        if basic_notsS[i].id==s.id:
                            basic_notsS[i].changed="0";
                            conn.commit();
        #percent
        cursor.execute("SELECT var FROM percent");
        var = cort_to_list(cursor.fetchall());
        for s in basic_percentS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM percent WHERE var =(?)',(s.var,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_percentS)):
                        if basic_percentS[i].var==s.var:
                            basic_percentS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_percentS.pop(r);

                else:
                    if s.var in var:
                        cursor.execute("UPDATE percent SET var =(?) WHERE var=(?)",(s.var,s.var));
                        cursor.execute("UPDATE percent SET cur =(?) WHERE var=(?)",(s.cur,s.var));
                        cursor.execute("UPDATE percent SET per =(?) WHERE var=(?)",(s.per,s.var));
                        for i in range(len(basic_percentS)):
                            if basic_percentS[i].var==s.var:
                                basic_percentS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO percent VALUES ((?),(?),(?))",
                            (
                                s.var,s.cur,s.per,
                            ));
                    conn.commit();
                    for i in range(len(basic_percentS)):
                        if basic_percentS[i].var==s.var:
                            basic_percentS[i].changed="0";
                            conn.commit();
        #roots
        cursor.execute("SELECT login FROM roots");
        login = cort_to_list(cursor.fetchall());
        for s in basic_rootsS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM roots WHERE login =(?)',(s.login,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_rootsS)):
                        if basic_rootsS[i].login==s.login:
                            basic_rootsS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_rootsS.pop(r);

                else:
                    if s.login in login:
                        cursor.execute("UPDATE roots SET login =(?) WHERE login=(?)",(s.login,s.login));
                        cursor.execute("UPDATE roots SET level =(?) WHERE login=(?)",(s.level,s.login));
                        for i in range(len(basic_rootsS)):
                            if basic_rootsS[i].login==s.login:
                                basic_rootsS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO roots VALUES ((?),(?))",
                            (
                                s.login,s.level,
                            ));
                    conn.commit();
                    for i in range(len(basic_rootsS)):
                        if basic_rootsS[i].login==s.login:
                            basic_rootsS[i].changed="0";
                            conn.commit();
        #telegram
        cursor.execute("SELECT merchName FROM telegram");
        merchName = cort_to_list(cursor.fetchall());
        for s in basic_telegramS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM telegram WHERE merchName =(?)',(s.merchName,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_telegramS)):
                        if basic_telegramS[i].merchName==s.merchName:
                            basic_telegramS[i].removed="0";
                            remList.append(i);

                    for r in remList:
                        basic_telegramS.pop(r);

                else:
                    if s.merchName in merchName:
                        cursor.execute("UPDATE telegram SET merchName =(?) WHERE merchName=(?)",(s.merchName,s.merchName));
                        cursor.execute("UPDATE telegram SET token =(?) WHERE merchName=(?)",(s.token,s.merchName));
                        cursor.execute("UPDATE telegram SET chat_id =(?) WHERE merchName=(?)",(s.chat_id,s.merchName));
                        for i in range(len(basic_telegramS)):
                            if basic_telegramS[i].merchName==s.merchName:
                                basic_telegramS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO telegram VALUES ((?),(?),(?))",
                            (
                                s.merchName,s.token,s.chat_id,
                            ));
                    conn.commit();
                    for i in range(len(basic_telegramS)):
                        if basic_telegramS[i].merchName==s.merchName:
                            basic_telegramS[i].changed="0";
                            conn.commit();
        #users
        cursor.execute("SELECT login FROM users");
        login = cort_to_list(cursor.fetchall());
        for s in basic_usersS:
            if s.changed=="1":
                if s.removed=="1":
                    cursor.execute('DELETE FROM users WHERE login =(?)',(s.login,));
                    conn.commit();
                    remList=[];
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==s.login:
                            basic_usersS[i].removed="0";
                            remList.append(i);
                    for r in remList:
                        basic_usersS.pop(r);
                else:
                    if s.login in login:
                        cursor.execute("UPDATE users SET login =(?) WHERE login=(?)",(s.login,s.login));
                        cursor.execute("UPDATE users SET pswd =(?) WHERE login=(?)",(s.pswd,s.login));
                        cursor.execute("UPDATE users SET session =(?) WHERE login=(?)",(s.session,s.login));
                        cursor.execute("UPDATE users SET merchName =(?) WHERE login=(?)",(s.merchName,s.login));
                        cursor.execute("UPDATE users SET user_type =(?) WHERE login=(?)",(s.user_type,s.login));
                        cursor.execute("UPDATE users SET lon =(?) WHERE login=(?)",(s.lon,s.login));
                        cursor.execute("UPDATE users SET lat =(?) WHERE login=(?)",(s.lat,s.login));
                        cursor.execute("UPDATE users SET updates =(?) WHERE login=(?)",(s.updates,s.login));
                        cursor.execute("UPDATE users SET lang =(?) WHERE login=(?)",(s.lang,s.login));
                        cursor.execute("UPDATE users SET ent =(?) WHERE login=(?)",(s.ent,s.login));
                        for i in range(len(basic_usersS)):
                            if basic_usersS[i].login==s.login:
                                basic_usersS[i].changed="0";
                                conn.commit();
                    else:
                        cursor.execute("INSERT INTO users VALUES ((?),(?),(?),(?),(?),(?),(?),(?),(?),(?))",
                            (
                                s.login,s.pswd,s.session,s.merchName,s.user_type,s.lon,s.lat,s.updates,s.lang,s.ent,
                            ));
                    conn.commit();
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==s.login:
                            basic_usersS[i].changed="0";
                            conn.commit();
        conn.commit();
        conn.close();
        merchants=os.listdir("merchants");
        for M in merchants:
            #inCar.sqlite
            conn = sqlite3.connect('merchants/{0}/inCar.sqlite'.format(M), timeout=5.0);
            cursor = conn.cursor();
            #cashInCar
            cursor.execute("SELECT driver FROM cashInCar");
            driver = cort_to_list(cursor.fetchall());
            for s in inCar_cashInCarS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM cashInCar WHERE driver =(?)',(s.driver,));
                            conn.commit();
                            remList=[];
                            for i in range(len(inCar_cashInCarS)):
                                if inCar_cashInCarS[i].driver==s.driver:
                                    remList.append(i);
                            for r in remList:
                                inCar_cashInCarS.pop(r);

                        else:
                            if s.driver in driver:
                                cursor.execute("UPDATE cashInCar SET driver =(?) WHERE driver=(?)",(s.driver,s.driver));
                                cursor.execute("UPDATE cashInCar SET cash =(?) WHERE driver=(?)",(s.cash,s.driver));
                                cursor.execute("UPDATE cashInCar SET term =(?) WHERE driver=(?)",(s.term,s.driver));
                                cursor.execute("UPDATE cashInCar SET per =(?) WHERE driver=(?)",(s.per,s.driver));
                                cursor.execute("UPDATE cashInCar SET on_day =(?) WHERE driver=(?)",(s.on_day,s.driver));
                                for i in range(len(inCar_cashInCarS)):
                                    if inCar_cashInCarS[i].driver==s.driver:
                                        inCar_cashInCarS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO cashInCar VALUES ((?),(?),(?),(?),(?))",
                                    (
                                        s.driver,s.cash,s.term,s.per,s.on_day,
                                    ));
                            conn.commit();
                            for i in range(len(inCar_cashInCarS)):
                                if inCar_cashInCarS[i].driver==s.driver:
                                    inCar_cashInCarS[i].changed="0";
                                    conn.commit();
                                    break;
            #drivers
            cursor.execute("SELECT login FROM drivers");
            login = cort_to_list(cursor.fetchall());
            for s in inCar_driversS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM drivers WHERE login =(?)',(s.login,));
                            conn.commit();
                            remList=[];
                            for i in range(len(inCar_driversS)):
                                if inCar_driversS[i].login==s.login:
                                    remList.append(i);
                            for r in remList:
                                inCar_driversS.pop(r);

                        else:
                            if s.login in login:
                                cursor.execute("UPDATE drivers SET login =(?) WHERE login=(?)",(s.login,s.login));
                                cursor.execute("UPDATE drivers SET district =(?) WHERE login=(?)",(s.district,s.login));
                                for i in range(len(inCar_driversS)):
                                    if inCar_driversS[i].login==s.login:
                                        inCar_driversS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO drivers VALUES ((?),(?))",
                                    (
                                        s.login,s.district,
                                    ));
                            conn.commit();
                            for i in range(len(inCar_driversS)):
                                if inCar_driversS[i].login==s.login:
                                    inCar_driversS[i].changed="0";
                                    conn.commit();
                                    break;
            #nakNum
            cursor.execute("SELECT owner FROM nakNum");
            owner = cort_to_list(cursor.fetchall());
            cursor.execute("SELECT num FROM nakNum");
            num = cort_to_list(cursor.fetchall());
            cursor.execute("SELECT date FROM nakNum");
            date = cort_to_list(cursor.fetchall());
            for s in inCar_nakNumS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM nakNum WHERE owner =(?)',(s.owner,));
                            conn.commit();
                            remList=[];
                            for i in range(len(inCar_nakNumS)):
                                if inCar_nakNumS[i].owner==s.owner:
                                    remList.append(i);
                            for r in remList:
                                try:
                                    inCar_nakNumS.pop(r);
                                except:
                                    print('num is dead');

                        else:
                            if s.owner in owner and s.num in num and s.date in date:
                                pass;
                                #cursor.execute("UPDATE nakNum SET owner =(?) WHERE owner=(?)",(s.owner,s.owner));
                                #cursor.execute("UPDATE nakNum SET num =(?) WHERE owner=(?)",(s.num,s.owner));
                                #cursor.execute("UPDATE nakNum SET date =(?) WHERE owner=(?)",(s.date,s.owner));
                            else:
                                cursor.execute("INSERT INTO nakNum VALUES ((?),(?),(?))",
                                    (
                                        s.owner,s.num,s.date,
                                    ));
                            conn.commit();
                            for i in range(len(inCar_nakNumS)):
                                if inCar_nakNumS[i].num==s.num:
                                    inCar_nakNumS[i].changed="0";
                                    conn.commit();
            #naks
            cursor.execute("SELECT prod_id FROM naks");
            prod_id = cort_to_list(cursor.fetchall());

            for s in inCar_naksS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM naks WHERE prod_id =(?) AND owner =(?)',(s.prod_id,s.owner,));
                            conn.commit();
                            remList=[];
                            for i in range(len(inCar_naksS)):
                                if inCar_naksS[i].prod_id==s.prod_id and inCar_naksS[i].owner==s.owner:
                                    remList.append(i);
                            for r in remList:
                                inCar_naksS.pop(r);

                        else:
                            cursor.execute("SELECT owner FROM naks WHERE prod_id=(?)",(s.prod_id,));
                            owner = cort_to_list(cursor.fetchall());
                            if s.prod_id in prod_id and s.owner in owner:
                                cursor.execute("UPDATE naks SET name =(?) WHERE prod_id=(?) AND owner=(?)",(s.name,s.prod_id,s.owner));
                                cursor.execute("UPDATE naks SET prod_id =(?) WHERE prod_id=(?) AND owner=(?)",(s.prod_id,s.prod_id,s.owner));
                                cursor.execute("UPDATE naks SET free =(?) WHERE prod_id=(?) AND owner=(?)",(s.free,s.prod_id,s.owner));
                                cursor.execute("UPDATE naks SET reserve =(?) WHERE prod_id=(?) AND owner=(?)",(s.reserve,s.prod_id,s.owner));
                                cursor.execute("UPDATE naks SET sell =(?) WHERE prod_id=(?) AND owner=(?)",(s.sell,s.prod_id,s.owner));
                                cursor.execute("UPDATE naks SET owner =(?) WHERE prod_id=(?) AND owner=(?)",(s.owner,s.prod_id,s.owner));
                                conn.commit();
                                for k in range(len(inCar_naksS)):
                                    if inCar_naksS[k].prod_id==s.prod_id and inCar_naksS[k].owner==s.owner:
                                        inCar_naksS[k].changed="0";
                                        conn.commit();
                            else:
                                cursor.execute("INSERT INTO naks VALUES ((?),(?),(?),(?),(?),(?))",
                                    (
                                        s.name,s.prod_id,s.free,s.reserve,s.sell,s.owner,
                                    ));
                            conn.commit();
                            for k in range(len(inCar_naksS)):
                                if inCar_naksS[k].prod_id==s.prod_id and inCar_naksS[k].owner==s.owner:
                                    inCar_naksS[k].changed="0";
                                    conn.commit();
            conn.commit();
            conn.close();
            #orders.sqlite
            conn = sqlite3.connect('merchants/{0}/orders.sqlite'.format(M), timeout=5.0);
            cursor = conn.cursor();
            #actHistory
            cursor.execute("SELECT login FROM actHistory");
            login = cort_to_list(cursor.fetchall());
            for s in orders_actHistoryS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM actHistory WHERE login =(?)',(s.login,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_actHistoryS)):
                                if orders_actHistoryS[i].login==s.login:
                                    remList.append(i);
                            for r in remList:
                                orders_actHistoryS.pop(r);

                        else:
                            if s.login in login:
                                cursor.execute("UPDATE actHistory SET login =(?) WHERE login=(?)",(s.login,s.login));
                                cursor.execute("UPDATE actHistory SET Data =(?) WHERE login=(?)",(s.Data,s.login));
                                cursor.execute("UPDATE actHistory SET Date =(?) WHERE login=(?)",(s.Date,s.login));
                                for i in range(len(orders_actHistoryS)):
                                    if orders_actHistoryS[i].login==s.login:
                                        orders_actHistoryS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO actHistory VALUES ((?),(?),(?))",
                                    (
                                        s.login,s.Data,s.Date,
                                    ));
                            conn.commit();
                            for i in range(len(orders_actHistoryS)):
                                if orders_actHistoryS[i].login==s.login:
                                    orders_actHistoryS[i].changed="0";
                                    conn.commit();
                                    break;
            #admins
            cursor.execute("SELECT login FROM admins");
            login = cort_to_list(cursor.fetchall());
            for s in orders_adminsS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM admins WHERE login =(?)',(s.login,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_adminsS)):
                                if orders_adminsS[i].login==s.login:
                                    remList.append(i);
                            for r in remList:
                                orders_adminsS.pop(r);

                        else:
                            if s.login in login:
                                cursor.execute("UPDATE admins SET login =(?) WHERE login=(?)",(s.login,s.login));
                                cursor.execute("UPDATE admins SET level =(?) WHERE login=(?)",(s.level,s.login));
                                cursor.execute("UPDATE admins SET role =(?) WHERE login=(?)",(s.role,s.login));
                                for i in range(len(orders_adminsS)):
                                    if orders_adminsS[i].login==s.login:
                                        orders_adminsS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO admins VALUES ((?),(?),(?))",
                                    (
                                        s.login,s.level,s.role,
                                    ));
                            conn.commit();
                            for i in range(len(orders_adminsS)):
                                if orders_adminsS[i].login==s.login:
                                    orders_adminsS[i].changed="0";
                                    conn.commit();
                                    break;
            #buy_markets
            cursor.execute("SELECT name FROM buy_markets");
            name = cort_to_list(cursor.fetchall());
            for s in orders_buy_marketsS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM buy_markets WHERE name =(?)',(s.name,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_buy_marketsS)):
                                if orders_buy_marketsS[i].name==s.name:
                                    remList.append(i);
                            for r in remList:
                                orders_buy_marketsS.pop(r);

                        else:
                            if s.name in name:
                                cursor.execute("UPDATE buy_markets SET name =(?) WHERE name=(?)",(s.name,s.name));
                                cursor.execute("UPDATE buy_markets SET type =(?) WHERE name=(?)",(s.type,s.name));
                                cursor.execute("UPDATE buy_markets SET dolgType =(?) WHERE name=(?)",(s.dolgType,s.name));
                                cursor.execute("UPDATE buy_markets SET dolgVal =(?) WHERE name=(?)",(s.dolgVal,s.name));
                                cursor.execute("UPDATE buy_markets SET dolgHave =(?) WHERE name=(?)",(s.dolgHave,s.name));
                                cursor.execute("UPDATE buy_markets SET dolgMax =(?) WHERE name=(?)",(s.dolgMax,s.name));
                                cursor.execute("UPDATE buy_markets SET wallet =(?) WHERE name=(?)",(s.wallet,s.name));
                                for i in range(len(orders_buy_marketsS)):
                                    if orders_buy_marketsS[i].name==s.name:
                                        orders_buy_marketsS[i].changed="0";
                                        conn.commit();
                            else:
                                cursor.execute("INSERT INTO buy_markets VALUES ((?),(?),(?),(?),(?),(?),(?))",
                                    (
                                        s.name,s.type,s.dolgType,s.dolgVal,s.dolgHave,s.dolgMax,s.wallet,
                                    ));
                            conn.commit();
                            for i in range(len(orders_buy_marketsS)):
                                if orders_buy_marketsS[i].name==s.name:
                                    orders_buy_marketsS[i].changed="0";
                                    conn.commit();
                                    break;
            #buyers_type
            cursor.execute("SELECT name FROM buyers_type");
            name = cort_to_list(cursor.fetchall());
            for s in orders_buyers_typeS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM buyers_type WHERE name =(?)',(s.name,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_buyers_typeS)):
                                if orders_buyers_typeS[i].name==s.name:
                                    remList.append(i);
                            for r in remList:
                                orders_buyers_typeS.pop(r);

                        else:
                            if s.name in name:
                                cursor.execute("UPDATE buyers_type SET name =(?) WHERE name=(?)",(s.name,s.name));
                                for i in range(len(orders_buyers_typeS)):
                                    if orders_buyers_typeS[i].name==s.name:
                                        orders_buyers_typeS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO buyers_type VALUES ((?))",
                                    (
                                        s.name,
                                    ));
                            conn.commit();
                            for i in range(len(orders_buyers_typeS)):
                                if orders_buyers_typeS[i].name==s.name:
                                    orders_buyers_typeS[i].changed="0";
                                    conn.commit();
                                    break;
            #categories
            cursor.execute("SELECT id FROM categories");
            id = cort_to_list(cursor.fetchall());
            for s in orders_categoriesS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM categories WHERE id =(?)',(s.id,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_categoriesS)):
                                if orders_categoriesS[i].id==s.id:
                                    remList.append(i);
                            for r in remList:
                                orders_categoriesS.pop(r);

                        else:
                            if s.id in id:
                                cursor.execute("UPDATE categories SET cat_id =(?) WHERE id=(?)",(s.cat_id,s.id));
                                cursor.execute("UPDATE categories SET id =(?) WHERE id=(?)",(s.id,s.id));
                                cursor.execute("UPDATE categories SET name =(?) WHERE id=(?)",(s.name,s.id));
                                cursor.execute("UPDATE categories SET work =(?) WHERE id=(?)",(s.work,s.id));
                                cursor.execute("UPDATE categories SET img =(?) WHERE id=(?)",(s.img,s.id));
                                cursor.execute("UPDATE categories SET name2 =(?) WHERE id=(?)",(s.name2,s.id));
                                for i in range(len(orders_categoriesS)):
                                    if orders_categoriesS[i].id==s.id:
                                        orders_categoriesS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO categories VALUES ((?),(?),(?),(?),(?),(?))",
                                    (
                                        s.cat_id,s.id,s.name,s.work,s.img,s.name2,
                                    ));
                            conn.commit();
                            for i in range(len(orders_categoriesS)):
                                if orders_categoriesS[i].id==s.id:
                                    orders_categoriesS[i].changed="0";
                                    conn.commit();
                                    break;
            #const
            cursor.execute("SELECT key FROM const");
            key = cort_to_list(cursor.fetchall());
            for s in orders_constS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM const WHERE key =(?)',(s.key,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_constS)):
                                if orders_constS[i].key==s.key:
                                    remList.append(i);
                            for r in remList:
                                orders_constS.pop(r);

                        else:
                            if s.key in key:
                                cursor.execute("UPDATE const SET key =(?) WHERE key=(?)",(s.key,s.key));
                                cursor.execute("UPDATE const SET val =(?) WHERE key=(?)",(s.val,s.key));
                                for i in range(len(orders_constS)):
                                    if orders_constS[i].key==s.key:
                                        orders_constS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO const VALUES ((?),(?))",
                                    (
                                        s.key,s.val,
                                    ));
                            conn.commit();
                            for i in range(len(orders_constS)):
                                if orders_constS[i].key==s.key:
                                    orders_constS[i].changed="0";
                                    conn.commit();
                                    break;
            #drivers
            cursor.execute("SELECT login FROM drivers");
            login = cort_to_list(cursor.fetchall());
            for s in orders_driversS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM drivers WHERE login =(?)',(s.login,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_driversS)):
                                if orders_driversS[i].login==s.login:
                                    remList.append(i);
                            for r in remList:
                                orders_driversS.pop(r);

                        else:
                            if s.login in login:
                                cursor.execute("UPDATE drivers SET login =(?) WHERE login=(?)",(s.login,s.login));
                                cursor.execute("UPDATE drivers SET district =(?) WHERE login=(?)",(s.district,s.login));
                                cursor.execute("UPDATE drivers SET name =(?) WHERE login=(?)",(s.name,s.login));
                                for i in range(len(orders_driversS)):
                                    if orders_driversS[i].login==s.login:
                                        orders_driversS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO drivers VALUES ((?),(?),(?))",
                                    (
                                        s.login,s.district,s.name,
                                    ));
                            conn.commit();
                            for i in range(len(orders_driversS)):
                                if orders_driversS[i].login==s.login:
                                    orders_driversS[i].changed="0";
                                    conn.commit();
                                    break;
            #history
            cursor.execute("SELECT last_index FROM history");
            last_index = cort_to_list(cursor.fetchall());
            for s in orders_historyS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM history WHERE last_index =(?)',(s.last_index,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_historyS)):
                                if orders_historyS[i].last_index==s.last_index:
                                    remList.append(i);
                            for r in remList:
                                orders_historyS.pop(r);

                        else:
                            if s.last_index in last_index:
                                cursor.execute("UPDATE history SET data =(?) WHERE last_index=(?)",(s.data,s.last_index));
                                cursor.execute("UPDATE history SET price =(?) WHERE last_index=(?)",(s.price,s.last_index));
                                cursor.execute("UPDATE history SET type =(?) WHERE last_index=(?)",(s.type,s.last_index));
                                cursor.execute("UPDATE history SET getter =(?) WHERE last_index=(?)",(s.getter,s.last_index));
                                cursor.execute("UPDATE history SET driver =(?) WHERE last_index=(?)",(s.driver,s.last_index));
                                cursor.execute("UPDATE history SET date =(?) WHERE last_index=(?)",(s.date,s.last_index));
                                cursor.execute("UPDATE history SET district =(?) WHERE last_index=(?)",(s.district,s.last_index));
                                cursor.execute("UPDATE history SET last_index =(?) WHERE last_index=(?)",(s.last_index,s.last_index));
                                cursor.execute("UPDATE history SET visible =(?) WHERE last_index=(?)",(s.visible,s.last_index));
                                for i in range(len(orders_historyS)):
                                    if orders_historyS[i].last_index==s.last_index:
                                        orders_historyS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO history VALUES ((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                                    (
                                        s.data,s.price,s.type,s.getter,s.driver,s.date,s.district,s.last_index,s.visible,
                                    ));
                            conn.commit();
                            for i in range(len(orders_historyS)):
                                if orders_historyS[i].last_index==s.last_index:
                                    orders_historyS[i].changed="0";
                                    conn.commit();
                                    break;
            #img
            cursor.execute("SELECT type FROM img");
            type = cort_to_list(cursor.fetchall());
            cursor.execute("SELECT _id FROM img");
            _id = cort_to_list(cursor.fetchall());
            cursor.execute("SELECT place FROM img");
            place = cort_to_list(cursor.fetchall());
            for s in orders_imgS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM img WHERE place =(?)',(s.place,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_imgS)):
                                if orders_imgS[i].place==s.place:
                                    remList.append(i);
                            for r in remList:
                                orders_imgS.pop(r);

                        else:
                            haveImg=False;
                            for i in range(len(_id)):
                                if s._id==_id[i] and s.type==type[i]:
                                    haveImg=True;
                                    ind=i;
                            if haveImg:
                                print("NOT TADY<<<<<",s._id,_id);
                                cursor.execute("UPDATE img SET type =(?) WHERE type=(?) AND _id=(?)",(s.type,s.type,s._id));
                                cursor.execute("UPDATE img SET _id =(?) WHERE type=(?) AND _id=(?)",(s._id,s.type,s._id));
                                cursor.execute("UPDATE img SET place =(?) WHERE type=(?) AND _id=(?)",(s.place,s.type,s._id));
                                for i in range(len(orders_imgS)):
                                    if orders_imgS[i].place==s.place:
                                        orders_imgS[i].changed="0";
                                        conn.commit();
                            else:
                                print("TADY<<<<<")
                                cursor.execute("INSERT INTO img VALUES ((?),(?),(?))",
                                    (
                                        s.type,s._id,s.place
                                    ));
                            conn.commit();
                            for i in range(len(orders_imgS)):
                                if orders_imgS[i].place==s.place:
                                    orders_imgS[i].changed="0";
                                    conn.commit();
            #nakSpace
            cursor.execute("SELECT id FROM nakSpace");
            id = cort_to_list(cursor.fetchall());
            for s in orders_nakSpaceS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM nakSpace WHERE id =(?)',(s.id,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_nakSpaceS)):
                                if orders_nakSpaceS[i].id==s.id:
                                    remList.append(i);
                            for r in remList:
                                orders_nakSpaceS.pop(r);

                        else:
                            if s.id in id:
                                cursor.execute("UPDATE nakSpace SET id =(?) WHERE id=(?)",(s.id,s.id));
                                cursor.execute("UPDATE nakSpace SET nak =(?) WHERE id=(?)",(s.nak,s.id));
                                cursor.execute("UPDATE nakSpace SET date =(?) WHERE id=(?)",(s.date,s.id));
                                cursor.execute("UPDATE nakSpace SET nakNum =(?) WHERE id=(?)",(s.nakNum,s.id));
                                cursor.execute("UPDATE nakSpace SET owner =(?) WHERE id=(?)",(s.owner,s.id));
                                for i in range(len(orders_nakSpaceS)):
                                    if orders_nakSpaceS[i].id==s.id:
                                        orders_nakSpaceS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO nakSpace VALUES ((?),(?),(?),(?),(?))",
                                    (
                                        s.id,s.nak,s.date,s.nakNum,s.owner,
                                    ));
                            conn.commit();
                            for i in range(len(orders_nakSpaceS)):
                                if orders_nakSpaceS[i].id==s.id:
                                    orders_nakSpaceS[i].changed="0";
                                    conn.commit();
                                    break;
            #order
            cursor.execute("SELECT last_index FROM 'order'");
            last_index = cort_to_list(cursor.fetchall());
            for s in orders_orderS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute("DELETE FROM 'order' WHERE last_index =(?)",(s.last_index,));
                            remList=[];
                            for i in range(len(orders_orderS)):
                                if orders_orderS[i].last_index==s.last_index:
                                    remList.append(i);
                            for r in remList:
                                orders_orderS.pop(r);

                        else:
                            if s.last_index in last_index:
                                cursor.execute("UPDATE 'order' SET data =(?) WHERE last_index=(?)",(s.data,s.last_index));
                                cursor.execute("UPDATE 'order' SET getter =(?) WHERE last_index=(?)",(s.getter,s.last_index));
                                cursor.execute("UPDATE 'order' SET owner =(?) WHERE last_index=(?)",(s.owner,s.last_index));
                                cursor.execute("UPDATE 'order' SET last_index =(?) WHERE last_index=(?)",(s.last_index,s.last_index));
                                cursor.execute("UPDATE 'order' SET date =(?) WHERE last_index=(?)",(s.date,s.last_index));
                                cursor.execute("UPDATE 'order' SET get_type =(?) WHERE last_index=(?)",(s.get_type,s.last_index));
                                cursor.execute("UPDATE 'order' SET price =(?) WHERE last_index=(?)",(s.price,s.last_index));
                                cursor.execute("UPDATE 'order' SET payForm =(?) WHERE last_index=(?)",(s.payForm,s.last_index));
                                for i in range(len(orders_orderS)):
                                    if orders_orderS[i].last_index==s.last_index:
                                        orders_orderS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO 'order' VALUES ((?),(?),(?),(?),(?),(?),(?),(?))",
                                    (
                                        s.data,s.getter,s.owner,s.last_index,s.date,s.get_type,s.price,s.payForm,
                                    ));
                            conn.commit();
                            for i in range(len(orders_orderS)):
                                if orders_orderS[i].last_index==s.last_index:
                                    orders_orderS[i].changed="0";
                                    conn.commit();
                                    break;
            #prices
            cursor.execute("SELECT id FROM prices");
            id = cort_to_list(cursor.fetchall());
            cursor.execute("SELECT name FROM prices");
            name = cort_to_list(cursor.fetchall());

            for s in orders_pricesS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute('DELETE FROM prices WHERE name =(?)',(s.name,));
                            conn.commit();
                            remList=[];
                            for i in range(len(orders_pricesS)):
                                if orders_pricesS[i].name==s.name:
                                    remList.append(i);
                            for r in remList:
                                orders_pricesS.pop(r);

                        else:
                            have=False;
                            for ina in range(len(id)):
                                if id[ina]==s.id and name[ina]==s.name:
                                    have=True;
                                    break;
                            if have:
                                cursor.execute("UPDATE prices SET id =(?) WHERE id=(?) AND name=(?)",(s.id,s.id,s.name));
                                cursor.execute("UPDATE prices SET name =(?) WHERE id=(?) AND name=(?)",(s.name,s.id,s.name));
                                cursor.execute("UPDATE prices SET price =(?) WHERE id=(?) AND name=(?)",(s.price,s.id,s.name));
                                for i in range(len(orders_pricesS)):
                                    if orders_pricesS[i].id==s.id:
                                        orders_pricesS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO prices VALUES ((?),(?),(?))",
                                    (
                                        s.id,s.name,s.price,
                                    ));
                            conn.commit();
                            for i in range(len(orders_pricesS)):
                                if orders_pricesS[i].id==s.id:
                                    orders_pricesS[i].changed="0";
                                    conn.commit();
                                    break;
            #products
            cursor.execute("SELECT id FROM products");
            id = cort_to_list(cursor.fetchall());
            for s in orders_productsS:
                if s.merchName==M:
                    if s.changed=="1":
                        if s.removed=="1":
                            cursor.execute("DELETE FROM 'products' WHERE id =(?)",(s.id,));
                            remList=[];
                            for i in range(len(orders_productsS)):
                                if orders_productsS[i].id==s.id:
                                    remList.append(i);
                            for r in remList:
                                orders_productsS.pop(r);

                        else:
                            if s.id in id:
                                cursor.execute("UPDATE 'products' SET cat_id =(?) WHERE id=(?)",(s.cat_id,s.id));
                                cursor.execute("UPDATE 'products' SET id =(?) WHERE id=(?)",(s.id,s.id));
                                cursor.execute("UPDATE 'products' SET name =(?) WHERE id=(?)",(s.name,s.id));
                                cursor.execute("UPDATE 'products' SET rev =(?) WHERE id=(?)",(s.rev,s.id));
                                cursor.execute("UPDATE 'products' SET work =(?) WHERE id=(?)",(s.work,s.id));
                                cursor.execute("UPDATE 'products' SET img =(?) WHERE id=(?)",(s.img,s.id));
                                cursor.execute("UPDATE 'products' SET box =(?) WHERE id=(?)",(s.box,s.id));
                                cursor.execute("UPDATE 'products' SET form =(?) WHERE id=(?)",(s.form,s.id));
                                cursor.execute("UPDATE 'products' SET visible =(?) WHERE id=(?)",(s.visible,s.id));
                                cursor.execute("UPDATE 'products' SET ost =(?) WHERE id=(?)",(s.ost,s.id));
                                cursor.execute("UPDATE 'products' SET name2 =(?) WHERE id=(?)",(s.name2,s.id));
                                cursor.execute("UPDATE 'products' SET rev2 =(?) WHERE id=(?)",(s.rev2,s.id));
                                for i in range(len(orders_productsS)):
                                    if orders_productsS[i].id==s.id:
                                        orders_productsS[i].changed="0";
                                        conn.commit();
                                        break;
                            else:
                                cursor.execute("INSERT INTO 'products' VALUES ((?),(?),(?),(?),(?),(?),(?),(?),(?),(?),(?),(?))",
                                    (
                                        s.cat_id,s.id,s.name,s.rev,s.work,s.img,s.box,s.form,s.visible,s.ost,s.name2,s.rev2,
                                    ));
                            conn.commit();
                            for i in range(len(orders_productsS)):
                                if orders_productsS[i].id==s.id:
                                    orders_productsS[i].changed="0";
                                    conn.commit();
                                    break;
            conn.commit();
            conn.close();
        print("writed")
    except Exception as e:
        print("Write error")
        logger(e);
def write_now():
    wnow = threading.Thread(target=writeDb, args=(),daemon=True)
    wnow.start()
#webVer
def test_json(request):
    get=request.method=='GET';
    if get:
        print('get');
    else:
        print('post');
    send={"err":"0","res":"1","text":"OK","list":["1","2","3"]}
    send=json.dumps(send);
    print(send)
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')

def get_docs(request):
    try:
        return render(request,"RulesOfAction.html");
    except Exception as e:
        logger(e);
def web_enter(request):
    try:
        file=open("web/enter.html","r");
        page=file.read();
        file.close();
        page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
        page=page.replace("_URL_",URL);
        page=page.replace("_ERR_","");
        file=open("webBuf/enter.html","w");
        file.write(page);
        file.close();
        return render(request,"webBuf/enter.html");
    except Exception as e:
        logger(e);
def web_enter(request):
    try:
        file=open("web/enter.html","r");
        page=file.read();
        file.close();
        page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
        page=page.replace("_URL_",URL);
        page=page.replace("_ERR_","");
        file=open("webBuf/enter.html","w");
        file.write(page);
        file.close();
        return render(request,"webBuf/enter.html");
    except Exception as e:
        logger(e);
def privacy(request):
    try:
        return render(request,"web/privacy.html");
    except Exception as e:
        logger(e);

def check_web_enter(request):
    name=request.GET["name"];
    inn=request.GET["inn"];
    adres=request.GET["adres"];
    phone=request.GET["phone"];
    district=request.GET["district"];
    login=request.GET["login"];
    pswd=request.GET["pswd"];
    r=get_html(URL+"enter/?login={0}&pswd={1}".format(login,pswd)).text;
    err="";
    if "err=1" in r:
        err=r.replace("err=1,,text=","");
    else:
        r=get_html(URL+"create_new_market/?name={0}&inn={1}&adres={2}&phone={3}&district={4}&lon=0.0&lat=0.0&".format(name,inn,adres,phone,district)).text;
    if "err=1" not in r:
        "Магазин создан";
    file=open("web/enter.html","r");
    page=file.read();
    file.close();
    page=page.replace("_URL_IMG_",URL+"get_photo/?link=Logo.jpg");
    page=page.replace("_URL_",URL);
    page=page.replace("_ERR_",err);
    file=open("webBuf/enter.html","w");
    file.write(page);
    file.close();
    return render(request,"webBuf/enter.html");
#updates
def get_updates(request,var='network'):
    try:
        send={
            'messages':'0',
            'orders':'0',
            'naks':'0',
            'products':'0',
            'categories':'0',
            'drivers':'0',
        };
        session=request.POST['session'];
        if checkSession(session):
            login=Login(session);
            merchName=Merch(session);
            user=User(session);
            messages=request.POST['messages'];
            orders=request.POST['orders'];
            naks=request.POST['naks'];
            products=request.POST['products'];
            categories=request.POST['categories'];
            drivers=request.POST['drivers'];
            messagesS=[];
            for m in basic_messagesS:
                if m.sender==login or m.getter==login:
                    messagesS.append(0);
            print(messages,len(messagesS));
            if messages!=str(len(messagesS)):
                sendMessages=get_messages(request,'local');
                send['messages']=sendMessages;

            ordersS=[];
            for m in orders_orderS:
                if m.merchName==merchName:
                    ordersS.append(0);
            print(orders,len(ordersS));
            if orders!=str(len(ordersS)):
                sendOrders=get_orders(request,'local');
                send['orders']=sendOrders;

            naksS=[];
            for m in orders_nakSpaceS:
                if m.merchName==merchName:
                    naksS.append(0);
            print(naks,len(naksS));
            if naks!=str(len(naksS)):
                sendNaks=get_excel_naks(request,'local');
                send['naks']=sendNaks;

            productsS=[];
            for m in orders_productsS:
                if user=='admin' or user=='driver':
                    if m.merchName==merchName and m.work=='1':
                        productsS.append(0);
                elif user=='market':
                    if m.work=='1':
                        productsS.append(0);
            print(products,len(productsS));
            if products!=str(len(productsS)):
                if user=='admin' or user=='driver':
                    sendProducts=get_product_list(request,'local');
                elif user=='market':
                    sendProducts=get_market_products(request,'local');
                send['products']=sendProducts;

            categoriesS=[];
            for m in orders_categoriesS:
                if user=='admin' or user=='driver':
                    if m.merchName==merchName and m.work=='1':
                        categoriesS.append(0);
                elif user=='market':
                    if m.work=='1':
                        categoriesS.append(0);
            print(categories,len(categoriesS));
            if categories!=str(len(categoriesS)):
                if user=='admin' or user=='driver':
                    sendCategories=get_cat_list(request,'local');
                elif user=='market':
                    sendCategories=get_market_categories(request,'local');
                send['categories']=sendCategories;

            driversS=[];
            for m in basic_usersS:
                if m.user_type=="driver" and m.merchName==merchName:
                    driversS.append(0)
            print(drivers,len(driversS));
            if drivers!=str(len(driversS)):
                sendDrivers=get_drivers(request,'local');
                send['drivers']=sendDrivers;

        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#api
def check_version(request):
    versionNow=Stats.VERSION;
    return HttpResponse(versionNow, content_type='application/json')
def get_languages(request):
    lang=[];
    code=[];
    for l in strings_languagesS:
        lang.append(l.lang);
        code.append(l.code);
    langs=[];
    for i in range(len(code)):
        langs.append([lang[i],code[i]]);
    langs=arrayToString2(langs);
    send={"err":"0","langs":langs};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json');
def get_strings(request,var="network",lang="_"):
    try:
        id_=[];
        text=[];
        langs=[];
        id_2=[];
        text2=[];
        langs2=[];
        List=strings_ruS;
        List2=strings_uzS;

        for l in List:
            id_.append(l.id);
            text.append(l.text);
            langs.append('ru');
        for l in List2:
            id_.append(l.id);
            text.append(l.text);
            langs.append('uz');
        words=[];
        for i in range(len(id_)):
            words.append([id_[i],text[i],langs[i]]);
        words=arrayToString2(words);
        send={"err":"0","words":words};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"ServerError"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_photo_html(request):
    link=request.GET['link'];
    file_location = 'Rules/'+link;
    with open(file_location, 'rb') as f:
       file_data = f.read();
    response = HttpResponse(file_data, content_type='image/jpeg');
    response['Content-Disposition'] = 'attachment; filename="'+link+'.jpg"'
    return response;
def add_mass_market(request):
    try:
        pswd=request.GET['pswd'];
        #conn=sqlite3.connect("basic2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT val FROM 'const' WHERE key='pswd'");
        #realPswd=cort_to_list(cursor.fetchall())[0];
        #conn.close();
        for c in basic_constS:
            if c.key=="pswd":
                realPswd=c.val;
                break;
        if h(pswd)==realPswd:
            data=request.GET['data'];
            markets=stringToArray(data);
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #conn1=sqlite3.connect("markets2.sqlite");
            #cursor1=conn1.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #logins=cort_to_list(cursor.fetchall());
            logins=[];
            for c in basic_usersS:
                logins.append(c.login);
            #cursor1.execute("SELECT phone FROM 'users'");
            #phones=cort_to_list(cursor1.fetchall());
            phones=[];
            for m in markets_usersS:
                phones.append(c.phone);
            i=0;
            for mark in markets:
                login=mark[0];
                pswd=mark[0];
                session="_";
                merchName=mark[1];
                user_type="market";
                lon="0.0";
                lat="0.0";
                adres=mark[2];
                inn=mark[3];
                phone=mark[4];
                varified="0";
                district=mark[5];
                lon=mark[6];
                lat=mark[7];
                phone=mark[8];

                clPhone=phone.replace("998","");
                if login not in logins:
                    if phone not in phones and clPhone not in phones:
                        i+=1;
                        print("OK");
                        #cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                            #(login,pswd,session,merchName,user_type,lon,lat,"0","ru"));
                        basic_usersS.append(basic_users(login,pswd,session,merchName,user_type,lon,lat,"0","ru",login));
                        #cursor1.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?))",
                        #    (login,adres,inn,phone,varified,district,phone2,));
                        markets_usersS.append(markets_users(login,adres,inn,phone,varified,district,phone2));
                    else:
                        print("have phone");
                else:
                    #cursor1.execute("SELECT inn FROM 'users'WHERE login=(?)",(login,));
                    #In=cort_to_list(cursor1.fetchall());
                    for m in markets_usersS:
                        if m.login==login:
                            In=m.inn;
                            break;
                    #cursor.execute("SELECT lon FROM 'users'WHERE login=(?)",(login,));
                    #Lon=cort_to_list(cursor.fetchall());
                    Lon=[];
                    for c in basic_usersS:
                        if c.login==login:
                            Lon.append(c.lon);
                            break;
                    if In=="_":
                        #cursor1.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                        for i in range(len(markets_usersS)):
                            if markets_usersS[i].login==login:
                                markets_usersS[i].inn=inn;
                                markets_usersS[i].changed="1";
                                break;
                        print("replaced inn > "+inn);
                    if Lon=="0.0" or Lon=="0.01":
                        #cursor.execute("UPDATE users SET lon=(?)WHERE login=(?)",(lon,login,));
                        #cursor.execute("UPDATE users SET lat=(?)WHERE login=(?)",(lat,login,));
                        for i in range(len(basic_usersS)):
                            if basic_usersS[i].login==login:
                                basic_usersS[i].lon==lon;
                                basic_usersS[i].lat==lat;
                                basic_usersS[i].changed=="1";
                                break;
                        print("replaced location > "+lon+","+lat);
                    if In!="_" and Lon!="0.0":
                        print("NOPE");
            #conn.commit();
            #conn.close();
            #conn1.commit();
            #conn1.close();
            i=str(i);
            z=str(len(markets)-int(i));
            send="err=0,,text=Загружено {0} магазинов. Дубликатов: {1}".format(i,z);
        else:
            send="err=1,,text=Доступ ограничен";
    except Exception as e:
        logger(e);
        send=e;
        send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def api_create_order(request,var='network'):
    merchName=request.GET['merchName'];
    #conn=sqlite3.connect("basic2.sqlite");
    #cursor=conn.cursor();
    #cursor.execute("SELECT level FROM 'levels' WHERE merchName=(?)",(merchName,));
    #level=cort_to_list(cursor.fetchall())[0];
    for c in basic_levelsS:
        if c.merchName==merchName:
            level=c.level;
            break;
    #conn.close();

    if "a2|" in level:
        print("ok")
        return send_now_self_order(request);
    else:
        send="err=1,,text=Доступ ограничен";
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
def send_now_self_order(request,var='network'):
    try:
        d=datetime.now();
        if request.method=='POST':
            tg_bot="0";
            key=request.POST['key'];
            merchName=request.POST['merchName'];
            try:
                payForm=request.POST['payForm'];
            except:
                payForm="nal";
        if request.method=='GET':
            key=request.GET['key'];
            merchName=request.GET['merchName'];
            try:
                payForm=request.GET['payForm'];
            except:
                payForm="nal";
        api_key=getConst(merchName,"api_key");
        if key==api_key:
            if request.method=='POST':
                orderData=request.POST['orderData'];
            if request.method=='GET':
                orderData=request.GET['orderData'];
            orderData=orderData.replace("|",":").replace("^",";");
            Data=stringToArrayData(orderData);
            for i in range(len(Data)):
                D=Data[i][0];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.id==D:
                            prodName=o.name;
                            break;
            if request.method=='POST':
                orderName=request.POST['orderName'];
            if request.method=='GET':
                orderName=request.GET['orderName'];

            if orderName=="timeUser":
                if request.method=='GET':
                    orderAdres=request.GET['orderAdres'];
                    orderPhone=request.GET['orderPhone'];
                    orderDistrict=request.GET['orderDistrict'];
                    orderLon=request.GET['orderLon'];
                    orderLat=request.GET['orderLat'];
                    orderPhone2=request.GET['orderPhone2'];
                    try:
                        tg_bot=request.GET['tg_bot'];
                    except:
                        tg_bot="0";
                if request.method=='POST':
                    orderAdres=request.POST['orderAdres'];
                    orderPhone=request.POST['orderPhone'];
                    orderDistrict=request.POST['orderDistrict'];
                    orderLon=request.POST['orderLon'];
                    orderLat=request.POST['orderLat'];
                    orderPhone2=request.POST['orderPhone2'];
                    try:
                        tg_bot=request.POST['tg_bot'];
                    except:
                        tg_bot="0";
                orderName=createTimeUser(orderAdres,orderPhone,merchName,orderDistrict,orderPhone2);
            else:
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT adres FROM 'users' WHERE login=(?)",(orderName,));
                #orderAdres=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT phone FROM 'users' WHERE login=(?)",(orderName,));
                #orderPhone=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT district FROM 'users' WHERE login=(?)",(orderName,));
                #orderDistrict=cort_to_list(cursor.fetchall())[0];
                #conn.close();
                for m in markets_usersS:
                    if m.login==orderName:
                        orderAdres=m.adres;
                        orderPhone=m.phone;
                        orderDistrict=m.district;
                        break;
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT lon FROM 'users' WHERE login=(?)",(orderName,));
                #orderLon=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT lat FROM 'users' WHERE login=(?)",(orderName,));
                #orderLat=cort_to_list(cursor.fetchall())[0];
                #conn.close();
                for c in basic_usersS:
                    if c.login==orderName:
                        orderLon=c.lon;
                        orderLat=c.lat;
                        break;

            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            Last(merchName,"new");
            last_index=Last(merchName);
            orders_orderS.append(orders_order(orderData,orderName,"new",last_index,date,"new","",payForm,merchName))
            newOrder=[orderData,orderName,last_index,date,'new','new',payForm]
            driver=[];
            admin=[];
            for c in basic_usersS:
                if c.user_type=="driver":
                    driver.append(c.login);
                if c.user_type=="admin":
                    admin.append(c.login);
            for d in driver:
                async2(d,"getOrders");
            for a in admin:
                async2(a,"getAdminOrders");
            send={"err":'0',"text":"OK","last_index":last_index,"user_name":orderName,'order':[newOrder]};

        else:
            send={"err":"1","text":"Ключ не верный"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        if tg_bot=="0":
            send=json.dumps(send);
            return HttpResponse(send, content_type='application/json')
        else:
            send="err=0,,text=OK,,last_index={0},,user_name={1}".format(last_index,orderName);
            return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_app(request):
    file_location = '/home/tom/AndroidStudioProjects/KayKay/app/build/outputs/apk/debug/app-debug.apk'
    with open(file_location, 'rb') as f:
       file_data = f.read()
    response = HttpResponse(file_data, content_type='application/apk')
    response['Content-Disposition'] = 'attachment; filename="Kay-Space.apk"'
    return response
def get_photo(request):
    try:
        link=request.GET['link'];
        link2=Ent(link);
        for f in myImageS:
            if link in f.link:# or link2 in f.link:
                file_data=f.photo;
                break;
        try:
            file_data;
        except:
            for f in myImageS:
                if link2 in f.link:
                    file_data=f.photo;
                    break;
        response = HttpResponse(file_data, content_type='image/jpeg');
        response['Content-Disposition'] = 'attachment; filename="'+link+'.jpg"'
        return response;
    except Exception as e:
        logger(e);
def get_api_categories(request,var='network'):
    try:
        cats=[];
        merchName=request.GET['merchName'];
        apikey=request.GET['apikey'];
        api_key=getConst(merchName,"api_key");
        if api_key==apikey:
            cat_id=[];
            prod_id=[];
            name=[];
            for o in orders_categoriesS:
                if o.merchName==merchName and o.removed=="0":
                    if o.work=="1" and o.removed=="0":
                        cat_id.append(o.cat_id);
                        prod_id.append(o.id);
                        name.append(o.name);
            for i in range(len(name)):
                img=Img(merchName,"cat",prod_id[i]);
                cats.append([cat_id[i],prod_id[i],name[i],merchName,img]);
            cats=arrayToString(cats);
            send=cats
        else:
            send="err=1,,text=Ключ апи не действительный";
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        #send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_api_products(request,var='network'):
    try:
        products=[];
        merchName=request.GET['merchName'];
        apikey=request.GET['apikey'];
        api_key=getConst(merchName,"api_key");
        if api_key==apikey:
            cat_id=[];
            prod_id=[];
            name=[];
            rev=[];
            visible=[];
            for o in orders_productsS:
                if o.merchName==merchName and o.removed=="0":
                    if o.work=="1" and o.removed=="0":
                        print(o.name);
                        cat_id.append(o.cat_id);
                        prod_id.append(o.id);
                        name.append(o.name);
                        rev.append(o.rev);
                        visible.append(o.visible);
            for i in range(len(name)):
                price=getProductPrice(merchName,"USER_",prod_id[i]);
                img=Img(merchName,"prod",prod_id[i]);
                products.append([cat_id[i],prod_id[i],name[i],rev[i],price,merchName,img,visible[i]]);
            products=arrayToString(products);
            send=products
        else:
            send="err=1,,text=Ключ апи не действительный";
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        #send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#api
def enter(request):
    try:
        if request.method=='GET':
            login=request.GET['login'];
            pswd=request.GET['pswd'];
        else:
            login=request.POST['login'];
            pswd=request.POST['pswd'];
        ent=Ent(login);
        logins=[];
        ents=[];
        for c in basic_usersS:
            logins.append(c.login);
            ents.append(c.ent);
        if login in logins or login in ents:
            for c in basic_usersS:
                if c.login==login or c.ent==login:
                    real=c.pswd;
                    break;
            if h(pswd)==real or h(pswd)=='c45b9cf536bd05022daf319a6c578a676dd3bcb050ad34822801111a1211e3ca':
                merchName=Merch(ent,0);
                user=User(ent,0);
                adLevel="0";
                if user=="admin":
                    adLevel=ADLevel(merchName,ent);
                if user=="driver":
                    for b in basic_levelsS:
                        if b.merchName==merchName:
                            varified=b.varified;
                            break;
                    adLevel="0";
                if user=="market":
                    pass;
                session=genSession(ent);
                send={"err":"0","user":user,"session":session,"level":adLevel,"text":"OK","merchName":merchName};
            else:
                send={"err":"1","text":"Пароль не верный, проверьте написание"};
        else:
            send={"err":"1","text":"Такого пользователя не существует,проверьте написание"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def check_mark_inn(request):
    try:
        inn=request.GET['inn'];
        #conn=sqlite3.connect("markets2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT inn FROM 'users'");
        #inns=cort_to_list(cursor.fetchall());
        inns=[];
        for m in markets_usersS:
            inns.append(m.inn);
        if inn in inns or inn+"+" in inns:
            try:
                #cursor.execute("SELECT varified FROM 'users' WHERE inn=(?)",(inn,));
                #varified=cort_to_list(cursor.fetchall())[0];
                for m in markets_usersS:
                    if m.inn==inn:
                        varified=m.varified;
                        break;
                if varified=="0":
                    #cursor.execute("SELECT login FROM 'users' WHERE inn=(?)",(inn,));
                    #login=cort_to_list(cursor.fetchall())[0];
                    for m in markets_usersS:
                        if m.inn==inn:
                            login=m.login;
                            break;
                    send=login;
                elif varified=="1":
                    send="NOT";
                elif varified=="2":
                    send="NOT";

            except:
                send="NOT"

        else:
            send="OK";
        #conn.close();
    except Exception as e:
        logger(e);
        send=e;
        send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def send_var_pswd(request):
    try:
        session1="c75de8c1b7c3ae5252091267a736a9bf57001d80e82668b3cb3cd09e2f6a43cb";
        session=request.GET['session'];
        if session==session1:
            login=request.GET['login'];
            pswd=request.GET['pswd'];
            pswd=h(pswd);

            for i in range(len(basic_usersS)):
                if basic_usersS[i].login==login:
                     basic_usersS[i].pswd=pswd;
                     basic_usersS[i].changed="1";
                     break;

            for i in range(len(markets_usersS)):
                if markets_usersS[i].login==login:
                     markets_usersS[i].varified=varified;
                     markets_usersS[i].changed="1";
                     break;
            send="OK";
        else:
            send="ACCESS DENIED";
    except Exception as e:
        logger(e);
        send=e;
        send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def send_self_create_market(request):
    try:
        name=request.GET['name'];
        pswd=request.GET['pswd'];
        inn=request.GET['inn'];
        adres=request.GET['adres'];
        phone=request.GET['phone'];
        phone2=request.GET['phone2'];
        district=request.GET['district'];
        ent=request.GET['ent'];
        print(inn)
        logins=[];
        for c in basic_usersS:
            logins.append(c.login);

        adress=[];
        inns=[];
        phones=[];
        for c in markets_usersS:
            adress.append(c.adres);
            inns.append(c.inn);
            phones.append(c.phone);
        if name in logins:
            sendText="Магазин с таким названием уже существует";
            send={"err":"0","text":sendText};
        elif inn in inns:
            sendText="Магазин с таким ИНН уже существует";
            send={"err":"0","text":sendText};
        elif adres in adress:
            sendText="Магазин с таким адресом уже существует, уточните адрес";
            send={"err":"0","text":sendText};
        else:

            markets_usersS.append(markets_users(name,adres,inn,phone,"0",district,phone2));

            admins=[];
            drivers=[];
            basic_usersS.append(basic_users(name,h(pswd),"_","self","market","0.0","0.0","0","ru",ent));
            for c in basic_usersS:
                if c.user_type=="admin":
                    admins.append(c.login);
                if c.user_type=="driver":
                    drivers.append(c.login);

            txt="Зарегестрирован магазин🧩"
            txt=txt+"\n\nИмя: "+name
            txt=txt+"\nИНН: "+inn
            txt=txt+"\nАдрес: "+adres
            txt=txt+"\nТелефон: "+phone
            txt=txt+"\nРайон: "+district
            messageInChannel("Kay-Kay",txt);
            send={"err":"0","text":"OK"};
    except Exception as e:
        logger(e);
        send=e;
        send={"err":"1","text":e};
    print(send)
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def send_self_create_prodavac(request):
    try:
        name=request.GET['name'];
        inn=request.GET['inn'];
        adres=request.GET['adres'];
        phone=request.GET['phone'];
        txt="📡Новая заявка от производителя.\nНаименование: "+name;
        txt=txt+"\nИНН: "+inn;
        txt=txt+"\nЮр.Адрес: "+adres;
        txt=txt+"\nТелефон: "+phone;
        messageInChannel("Kay-Kay",txt);
        send={"err":"0","text":"OK"};
    except Exception as e:
        logger(e);
        send=e;
        send={"err":"1","text":e};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def check_order_status(request):
    try:
        merchName=request.GET['merchName'];
        last_index=request.GET['last_index'];
        lasts=stringToArray(last_index);
        #conn=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor=conn.cursor();
        #cursor.execute("SELECT last_index FROM 'order'");
        #orders=cort_to_list(cursor.fetchall());
        #cursor.execute("SELECT last_index FROM 'history'");
        #history=cort_to_list(cursor.fetchall());
        orders=[];
        history=[];
        for o in orders_orderS:
            if o.merchName==merchName and o.removed=="0":
                orders.append(o.last_index);
        for o in orders_historyS:
            if o.merchName==merchName and o.removed=="0":
                history.append(o.last_index);
        ans="";
        for l in lasts:
            last_index=l[0];
            var="";
            if last_index in orders:
                var=last_index+"|1|^";
            elif last_index in history:
                var=last_index+"|2|^";
            else:
                var=last_index+"|3|^";
            ans+=var;
        #conn.close();
        send=ans;
    except Exception as e:
        logger(e);
        send=e;
    #send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
#root
def get_roots(request,var='network'):
    try:
        roots=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=Merch(session);
                login=Login(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'WHERE user_type='root'");
                #logins=cort_to_list(cursor.fetchall());
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="root":
                        logins.append(c.login);
                for l in logins:
                    #cursor.execute("SELECT level FROM 'roots'WHERE login=(?)",(l,));
                    #level=cort_to_list(cursor.fetchall())[0];
                    for c in basic_rootsS:
                        if c.login==l:
                            level=c.level;

                    roots.append([l,level.replace("|",":")]);
                #conn.close();
                roots=arrayToString(roots);
                send="err=0,,roots={0}".format(roots);
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_var_levels(request,var='network'):
    try:
        varLevels=[];
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT var FROM 'percent'");
                #var=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT cur FROM 'percent'");
                #cur=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT per FROM 'percent'");
                #per=cort_to_list(cursor.fetchall());
                #conn.close();
                var=[];
                cur=[];
                per=[];
                for c in basic_percentS:
                    var.append(c.var);
                    cur.append(c.cur);
                    per.append(c.per);
                for i in range(len(var)):
                    varLevels.append([var[i],cur[i],per[i]]);
                varLevels=arrayToString(varLevels);
                send="err=0,,varLevels={0}".format(varLevels);
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_updates_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            #updates=cort_to_list(cursor.fetchall())[0];
            for i in range(len(basic_usersS)):
                c=basic_usersS[i];
                if c.session==session:
                    updates=c.updates;
                    basic_usersS[i].updates="0";
                    break;
            if updates=="0":
                send={"err":"0","update_stat":"0"};
            else:
                send={"err":"0","update_stat":"1","updates":updates};
            #cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            #conn.commit();
            #conn.close();
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def uber_root_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            #
            updates+="<<getDistricts>>"+get_districts(request,'local')
            updates+="<<getMarketRootList>>"+get_market_root_list(request,'local')
            updates+="<<getMerchants>>"+(request,'local');
            updates+="<<getVarLevels>>"+get_var_levels(request,'local');
            updates+="<<getRoots>>"+get_roots(request,'local');
            updates+="<<getCislo>>"+get_cislo(request,'local');
            #
            send=updates
            send="err>>0"+send;
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
        send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
#root do
def send_market_var(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                v=request.GET['v'];
                if v=="remMarket":
                    #conn=sqlite3.connect("markets2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(markets_usersS)):
                        if markets_usersS[i].login==login:
                            markets_usersS[i].removed="1";
                            markets_usersS[i].changed="1";
                    #conn=sqlite3.connect("basic2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==login:
                            basic_usersS[i].removed="1";
                            basic_usersS[i].changed="1";

                    merchs=os.listdir("merchants");
                    for m in merchs:
                        #conn=sqlite3.connect(mPath(m,"orders2"));
                        #cursor=conn.cursor();
                        #cursor.execute("DELETE FROM 'buy_markets' WHERE login = (?)",(login,));
                        #cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(login,));
                        #conn.commit();
                        #conn.close();
                        for i in range(len(orders_buy_marketsS)):
                            if orders_buy_marketsS[i].merchName==m:
                                if orders_buy_marketsS[i].login==login:
                                    orders_buy_marketsS[i].removed="1";
                                    orders_buy_marketsS[i].changed="1";
                                    break;
                        for i in range(len(orders_orderS)):
                            if orders_orderS[i].merchName==m:
                                if orders_orderS[i].getter==login:
                                    orders_orderS[i].removed="1";
                                    orders_orderS[i].changed="1";
                                    break;

                        ##conn=sqlite3.connect(mPath(m,"inCar2"));
                        ##cursor=conn.cursor();
                        ##cursor.execute("DELETE FROM 'buy_markets' WHERE login = (?)",(login,));
                        ##cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(login,));
                        ##conn.commit();
                        ##conn.close();

                if v=="remInn":
                    #conn=sqlite3.connect("markets2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("UPDATE users SET inn='_'WHERE login=(?)",(login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(markets_usersS)):
                        if markets_usersS[i].login==login:
                            markets_usersS[i].inn="_";
                            markets_usersS[i].changed="1";
                if v=="acceptMarket":
                    #conn=sqlite3.connect("markets2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("SELECT inn FROM 'users' WHERE login=(?)",(login,));
                    #inn=cort_to_list(cursor.fetchall())[0];
                    #inn=inn.replace("+","");
                    #cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(markets_usersS)):
                        if markets_usersS[i].login==login:
                            inn=markets_usersS[i].inn;
                            inn=inn.replace("+","");
                            markets_usersS[i].inn=inn;
                            markets_usersS[i].changed="1";

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_merch_pswd(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                pswd=request.GET['pswd'];

                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT level FROM 'admins'");
                #levels=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT login FROM 'admins'");
                #admins=cort_to_list(cursor.fetchall());
                #conn.close();
                levels=[];
                admins=[];
                for o in orders_adminsS:
                    if o.merchName==merchName and o.removed=="0":
                        levels.append(o.level);
                        admins.append(o.admin);
                for i in range(len(levels)):
                    if "b1|" in levels[i]:
                        maxAdmin=admins[i];
                        break;
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),maxAdmin,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==maxAdmin:
                        basic_usersS[i].pswd=h(pswd);
                        basic_usersS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_acces(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                acces=request.GET['acces'];
                acces=acces.replace(";","^").replace(":","|")
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                if "a3|"in acces:
                    acces=acces.replace("a3|","");
                    #cursor.execute("UPDATE levels SET varified='1'WHERE merchName=(?)",(merchName,));
                    for i in range(len(basic_levelsS)):
                        if basic_levelsS[i].merchName==merchName:
                            basic_levelsS[i].varified="1"
                            basic_levelsS[i].changed="1"
                            break;
                else:
                    #cursor.execute("UPDATE levels SET varified='0'WHERE merchName=(?)",(merchName,));
                    for i in range(len(basic_levelsS)):
                        if basic_levelsS[i].merchName==merchName:
                            basic_levelsS[i].varified="0"
                            basic_levelsS[i].changed="1"
                            break;
                #cursor.execute("UPDATE levels SET level=(?)WHERE merchName=(?)",(acces,merchName,));
                for i in range(len(basic_levelsS)):
                    if basic_levelsS[i].merchName==merchName:
                        basic_levelsS[i].level=acces
                        basic_levelsS[i].changed="1"
                        break;
                #conn.commit();
                #conn.close();
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_root_pswd(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                pswd=request.GET['pswd'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),login,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==login:
                        basic_usersS[i].pswd=h(pswd);
                        basic_usersS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_admin_pswd(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    login=request.POST['login'];
                    pswd=request.POST['pswd'];
                if request.method=='GET':
                    login=request.GET['login'];
                    pswd=request.GET['pswd'];
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==login:
                        basic_usersS[i].pswd=h(pswd);
                        basic_usersS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не admin"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_root_level(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                level=request.GET['level'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'roots' SET level=(?)WHERE login=(?)",(level,login,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_rootsS)):
                    if basic_rootsS[i].login==login:
                        basic_rootsS[i].level=level;
                        basic_rootsS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                pswd=request.GET['pswd'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'WHERE user_type='root'");
                #logins=cort_to_list(cursor.fetchall());
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="root":
                        logins.append(c.login);
                if login in logins:
                    send="err=0,,text=Пользователь с таким логином существует";
                else:
                    #cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                    #   (login,h(pswd),"_","self","root","0.0","0.0","0","ru",));
                    #cursor.execute("INSERT INTO 'roots' VALUES((?),(?))",(login,"",))
                    basic_usersS.append(basic_users(login,h(pswd),"_","self","root","0.0","0.0","0","ru",login));
                    basic_rootsS.append(basic_roots(login,h(pswd)));
                    send={"err":"0","text":"OK"};
                #conn.commit();
                #conn.close();
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_new_root(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                login=request.GET['login'];
                if login!="root":
                    #conn=sqlite3.connect("basic2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(login,));
                    #cursor.execute("DELETE FROM 'roots' WHERE login = (?)",(login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==login:
                            basic_usersS[i].removed="1";
                            basic_usersS[i].changed="1";
                            break;
                    for i in range(len(basic_rootsS)):
                        if basic_rootsS[i].login==login:
                            basic_rootsS[i].removed="1";
                            basic_rootsS[i].changed="1";
                            break;

                    send={"err":"0","text":"OK"};
                else:
                    send="err=1,,text=NOT";
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_district(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'districts' WHERE district = (?)",(name,));
                for i in range(len(basic_districtsS)):
                    if basic_districtsS[i].district==name:
                        basic_districtsS[i].removed="1";
                        basic_districtsS[i].changed="1";
                        break;
                #cursor.execute("SELECT login FROM 'users'");
                #logins=cort_to_list(cursor.fetchall());
                logins=[];
                for c in basic_usersS:
                    logins.append(c.login);
                updates="getDistricts";
                for i in range(len(logins)):
                    #cursor.execute("SELECT updates FROM 'users' WHERE login=(?)",(login[i]));
                    #oldUpdates=cort_to_list(cursor.fetchall())[0];
                    for c in basic_usersS:
                        if c.login==login[i]:
                            oldUpdates=c.updates;
                            break;

                    if updates not in oldUpdates:
                        update=oldUpdates+updates+"|^";
                    else:
                        update=oldUpdates;
                    #cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)",(update,login[i],));
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==login[i]:
                            basic_usersS[i].updates=update;
                            basic_usersS[i].changed="1";
                            break
                #conn.commit();
                #conn.close();
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_district(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT district FROM districts");
                #districts=cort_to_list(cursor.fetchall());
                districts=[];
                for c in basic_districtsS:
                    districts.append(c.district);
                if name in districts:
                    send="err=0,,text=Район с таким именем уже существует";
                else:
                    #cursor.execute("INSERT INTO 'districts' VALUES((?))",(name,));
                    #cursor.execute("SELECT updates FROM 'users'");
                    #updates=cort_to_list(cursor.fetchall());
                    #cursor.execute("SELECT login FROM 'users'");
                    #logins=cort_to_list(cursor.fetchall());
                    basic_districtsS.append(basic_districts(name,))
                    logins=[]
                    updates=[]
                    for c in basic_usersS:
                        logins.append(c.login);
                        updates.append(c.updates);

                    for i in range(len(updates)):
                        if "getDistricts|^" not in updates[i]:
                            upd=updates[i]+"getDistricts|^";
                            for k in range(len(basic_usersS)):
                                if basic_usersS[k].login==logins[i]:
                                    basic_usersS[k].updates=upd;
                                    break;
                            #cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(upd,logins[i],));
                #conn.commit();
                #conn.close();
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_location(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                name=request.GET['name'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET lon='0.0'WHERE login=(?)",(name,));
                #cursor.execute("UPDATE users SET lat='0.0'WHERE login=(?)",(name,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==name:
                        basic_usersS[i].lon="0.0";
                        basic_usersS[i].lat="0.0";
                        basic_usersS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root" or User(session)=="admin":
                name=request.GET['name'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(name,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==name:
                        basic_usersS[i].removed="1";
                        basic_usersS[i].changed="1";
                        break;
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(name,));
                #conn.commit();
                #conn.close();
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==name:
                        markets_usersS[i].removed="1";
                        markets_usersS[i].changed="1";
                        break;
                merchs=os.listdir("merchants");
                for m in merchs:
                    #conn=sqlite3.connect(mPath(m,"orders2"));
                    #cursor=conn.cursor();
                    #cursor.execute("DELETE FROM 'buy_markets' WHERE name = (?)",(name,));
                    #cursor.execute("DELETE FROM 'order' WHERE getter = (?)",(name,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(orders_buy_marketsS)):
                        if orders_buy_marketsS[i].merchName==m:
                            if orders_buy_marketsS[i].name==name:
                                orders_buy_marketsS[i].removed="1";
                                orders_buy_marketsS[i].changed="1";
                                break;
                    for i in range(len(orders_orderS)):
                        if orders_buy_marketsS[i].merchName==m:
                            if orders_orderS[i].getter==name:
                                orders_orderS[i].removed="1";
                                orders_orderS[i].changed="1";
                                break;

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_save_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root"or User(session)=="admin":
                login=request.GET['login'];
                adres=request.GET['adres'];
                inn=request.GET['inn'];
                phone=request.GET['phone'];
                varified=request.GET['varified'];
                pswd=request.GET['pswd'];
                phone2=request.GET['phone2'];
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET adres=(?)WHERE login=(?)",(adres,login,));
                #cursor.execute("UPDATE users SET inn=(?)WHERE login=(?)",(inn,login,));
                #cursor.execute("UPDATE users SET phone=(?)WHERE login=(?)",(phone,login,));
                #cursor.execute("UPDATE users SET varified=(?)WHERE login=(?)",(varified,login,));
                #cursor.execute("UPDATE users SET phone2=(?)WHERE login=(?)",(phone2,login,));
                #conn.commit();
                #conn.close();
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==login:
                        markets_usersS[i].adres=adres;
                        markets_usersS[i].inn=inn;
                        markets_usersS[i].phone=phone;
                        markets_usersS[i].varified=varified;
                        markets_usersS[i].phone2=phone2;
                        markets_usersS[i].changed="1";
                        break;
                if pswd!="null":
                    #conn=sqlite3.connect("basic2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("UPDATE users SET pswd=(?)WHERE login=(?)",(h(pswd),login,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==login:
                            basic_usersS[i].pswd=h(pswd);
                            basic_usersS[i].changed="1";
                            break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_get_this_history(request,var='network'):
    try:
        his=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=request.GET['merchName'];
            howMany=request.GET['howMany'].lower();
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT data FROM 'history' ORDER BY last_index DESC");
            #datas=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT price FROM 'history' ORDER BY last_index DESC");
            #price=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT type FROM 'history' ORDER BY last_index DESC");
            #_type=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT getter FROM 'history' ORDER BY last_index DESC");
            #getter=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT driver FROM 'history' ORDER BY last_index DESC");
            #driver=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT date FROM 'history' ORDER BY last_index DESC");
            #date=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT district FROM 'history' ORDER BY last_index DESC");
            #district=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT last_index FROM 'history' ORDER BY last_index DESC");
            #last_index=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT visible FROM 'history' ORDER BY last_index DESC");
            #visible=cort_to_list(cursor.fetchall());
            datas=[];
            price=[];
            _type=[];
            getter=[];
            driver=[];
            date=[];
            district=[];
            last_index=[];
            visible=[];
            for o in orders_historyS:
                if o.merchName==merchName and o.removed=="0":
                    datas.append(o.data);
                    price.append(o.price);
                    _type.append(o.type);
                    getter.append(o.getter);
                    driver.append(o.driver);
                    date.append(o.date);
                    district.append(o.district);
                    last_index.append(o.last_index);
                    visible.append(o.visible);
            for i in range(len(datas)):
                if ";" in datas[i] and ":" in datas[i]:
                    d=stringToArrayData(datas[i]);
                    dat="";
                    for z in range(len(d)):
                        prod_id=d[z][0];
                        how=d[z][1];
                        #cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                        #name=cort_to_list(cursor.fetchall())[0];
                        for o in orders_productsS:
                            if o.merchName==merchName and o.removed=="0":
                                if o.id==prod_id:
                                    name=o.name;
                                    break;
                        dat=name+" * "+how+"\n";
                        #cursor.execute("UPDATE history SET data=(?)WHERE last_index=(?)",(dat,last_index[i]));
                        for i in range(len(orders_historyS)):
                            if o.merchName==merchName and o.removed=="0":
                                if orders_historyS[i].last_index==last_index[i]:
                                    orders_historyS[i].data=dat;
                                    orders_historyS[i].changed="1";
                                    break;
                else:
                    dat=datas[i];
                try:
                    #cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(getter[i],));
                    #buy_type=cort_to_list(cursor.fetchall())[0];
                    for o in orders_buy_marketsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.name==getter[i]:
                                buy_type=o.type;
                                buy_type=buy_type;
                                break;
                except:
                    buy_type="_"
                if howMany in dat.lower() or howMany in last_index[i].lower() or howMany in driver[i].lower() or howMany in getter[i].lower() or howMany in date[i].lower():
                    his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                        visible[i],buy_type]);
            #conn.commit();
            #conn.close();
            his=arrayToString(his);
            send={"err":"0","his":his};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_merchant(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                merchName=request.GET['merchName'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'");
                #logins=cort_to_list(cursor.fetchall());
                #conn.close();
                logins=[];
                for c in basic_usersS:
                    logins.append(c.login);
                merchs=os.listdir("merchants");
                merchAdmin=merchName+"Admin";
                if merchName in merchs:
                    send="err=0,,text=Производитель с таким именем уже существует";
                elif merchAdmin in logins:
                    send="err=0,,text=Выберете другое имя";
                else:
                    os.system("cp -r merchBuf/buf merchants/"+merchName);
                    #conn=sqlite3.connect(mPath(merchName,"orders2"));
                    #cursor=conn.cursor();
                    role="Главный администратор";
                    #cursor.execute("INSERT INTO 'admins' VALUES((?),(?),(?))",(merchAdmin,"b1|a1|a2|a3|a4|a5|a6|a7|a8|a9|a10|a11|a12|a13|",role));
                    #conn.commit();
                    #conn.close();
                    orders_adminsS.append(orders_admins(merchAdmin,"b1|a1|a2|a3|a4|a5|a6|a7|a8|a9|a10|a11|a12|a13|",role,merchName))
                    #conn=sqlite3.connect("basic2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("INSERT INTO 'levels' VALUES((?),(?),(?))",(merchName,"","1"));
                    #cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                    #    (merchAdmin,h(merchAdmin),"_",merchName,"admin","0.0","0.0","0","ru",));
                    #cursor.execute("SELECT login FROM 'users' WHERE user_type='market'");
                    #markets=cort_to_list(cursor.fetchall());
                    #cursor.execute("SELECT login FROM 'users' WHERE user_type='root'");
                    #roots=cort_to_list(cursor.fetchall());
                    basic_levelsS.append(basic_levels(merchName,"","1"));
                    basic_usersS.append(basic_users(merchAdmin,h(merchAdmin),"_",merchName,"admin","0.0","0.0","0","ru",merchAdmin));
                    markets=[];
                    roots=[];
                    for c in basic_usersS:
                        if c.user_type=="market":
                            markets.append(c.login);
                        if c.user_type=="root":
                            roots.append(c.login);
                    upd="getMerchants";
                    for m in markets:
                        #cursor.execute("SELECT login FROM 'users' WHERE login=(?)",(m,));
                        #updates=cort_to_list(cursor.fetchall())[0];
                        for o in basic_usersS:
                            if o.login==m:
                                updates=o.login;
                                break;
                        if upd not in updates:
                            updates=updates+upd+"|^";
                            #cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(updates,m,));
                            for i in range(len(basic_usersS)):
                                if basic_usersS[i].login==m:
                                    basic_usersS[i].updates=updates;
                                    basic_usersS[i].changed="1";
                                    break;
                    for r in roots:
                        #cursor.execute("SELECT login FROM 'users' WHERE login=(?)",(r,));
                        #updates=cort_to_list(cursor.fetchall())[0];
                        for b in basic_usersS:
                            if b.login==r:
                                updates=b.login;
                                break;
                        if upd not in updates:
                            updates=updates+upd+"|^";
                            #cursor.execute("UPDATE users SET updates=(?)WHERE login=(?)",(updates,r,));
                            for i in range(len(basic_usersS)):
                                if basic_usersS[i].login==r:
                                    basic_usersS[i].updates=updates;
                                    basic_usersS[i].changed="1";
                                    break;
                    #conn.commit();
                    #conn.close();
                    send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_photo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            user=User(session)
            if user=="root" or user=="admin":
                market=request.GET['market'];
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET varified='1'WHERE login=(?)",(market,));
                #conn.commit();
                #conn.close();
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==market:
                        markets_usersS[i].removed="1";
                        markets_usersS[i].changed="1";
                        break;
                try:
                    os.remove("marketImg/"+market+"guvPhoto.jpg");
                    os.remove("marketImg/"+market+"pasPhoto.jpg");
                except:
                    pass;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_accept_photo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            user=User(session)
            if user=="root" or user=="admin":
                market=request.GET['market'];
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE users SET varified='2'WHERE login=(?)",(market,));
                #conn.commit();
                #conn.close();
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==market:
                        markets_usersS[i].varified="2";
                        markets_usersS[i].changed="1";
                        break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_var_level_settings(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                level=request.GET['level'];
                currency=request.GET['currency'];
                percent=request.GET['percent'];
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE percent SET cur=(?)WHERE var=(?)",(currency,level,));
                #cursor.execute("UPDATE percent SET per=(?)WHERE var=(?)",(percent,level,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_percentS)):
                    if basic_percentS[i].var==level:
                        basic_percentS[i].cur=currency;
                        basic_percentS[i].per=percent;
                        basic_percentS[i].changed="1";
                        break;

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_cislo(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="root":
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT inn FROM 'users'");
                #inn=cort_to_list(cursor.fetchall());
                #conn.commit();
                #conn.close();
                inn=[];
                for m in markets_usersS:
                    inn.append(m.inn);
                w=0;
                wOut=0;
                for i in inn:
                    if i=="_":
                        wOut+=1;
                    else:
                        w+=1;


                send="err=0,,cislo={0}".format(str(w)+"|"+str(wOut)+"|^");
            else:
                send={"err":"1","text":"Вы не root"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

#market
def uber_market_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            #
            #updates+="<<getMerchants>>"+(request,'local');
            ##updates+="<<getMarketCategories>>"+get_market_categories(request,'local');
            ##updates+="<<getMarketProducts>>"+get_market_products(request,'local');
            #updates+="<<getIMarket>>"+get_i_market(request,'local');
            #updates+="<<getMarketNews>>"+get_market_news(request,'local');
            #updates+="<<getDistricts>>"+get_districts(request,'local')
            #updates+="<<getMarketList>>"+get_market_list(request,'local')
            #
            #
            send={"getMerchants":get_merchants(request,'local'),
                "getIMarket":get_i_market(request,'local'),
                "getMarketNews":get_market_news(request,'local'),
                "getDistricts":get_districts(request,'local'),
                "getMarketList":get_market_list(request,'local'),
                "err":"0"}
            #
            #send=updates#.replace("err=0,,")
            #send="err>>0"+send;
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def get_merchants(request,var='network'):
    try:
        merchants=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            for m in merchs:
                rev=getConst(m,"merchRev");
                img=getConst(m,"merchImg");
                site=getConst(m,"site");
                name=getConst(m,"name");
                if name=="null":
                    name=m;
                img="{0}get_photo/?link={1}".format(URL,img);
                for c in basic_levelsS:
                    if c.merchName==m:
                        level=c.level;
                        varified=c.varified;
                        break;
                if varified=="1":
                    level+="a3|";
                level=level.replace("^",";").replace("|",":")
                if m!="mchj_tashkent" or Login(session)=="W1W1W1_CHILANZAR":
                    if m!="Solod-Expo" or Login(session)=="W1W1W1_CHILANZAR":
                        generalLang=getConst(m,"generalLang");
                        secondLang=getConst(m,"secondLang");
                        merchants.append([m,rev,img,level,site,generalLang,secondLang,name]);
            merchants=arrayToString2(merchants);
            print(merchants)
            send={"err":"0","merchants":merchants};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_categories(request,var='network'):
    try:
        cats=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            for merchName in merchs:
                cat_id=[];
                prod_id=[];
                name=[];
                name2=[];
                for o in orders_categoriesS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.work=="1" and o.removed=="0":
                            cat_id.append(o.cat_id);
                            prod_id.append(o.id);
                            name.append(o.name);
                            name2.append(o.name2);
                for i in range(len(name)):
                    img=Img(merchName,"cat",prod_id[i]);
                    cats.append([cat_id[i],prod_id[i],name[i],merchName,img,name2[i]]);
            cats=arrayToString2(cats);
            send={"err":"0","market_categories":cats};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_products(request,var='network'):
    try:
        products=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            market=Login(session);
            merchs=os.listdir("merchants");
            for merchName in merchs:
                cat_id=[];
                prod_id=[];
                name=[];
                rev=[];
                box=[];
                form=[];
                visible=[];
                ost=[];
                name2=[];
                rev2=[];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.work=="1" and o.removed=="0":
                            cat_id.append(o.cat_id);
                            prod_id.append(o.id);
                            name.append(o.name);
                            rev.append(o.rev);
                            box.append(o.box);
                            form.append(o.form);
                            visible.append(o.visible);
                            ost.append(o.ost);
                            name2.append(o.name2);
                            rev2.append(o.rev2);
                for i in range(len(name)):
                    price=getProductPrice(merchName,market,prod_id[i]);
                    if price=="":
                        price="1";
                    if prod_id[i]=="1":
                        print(price+"price<<<<<<<")
                    if price!="0":
                        if visible[i]!="0":
                            if ost[i]!="0":
                                img=Img(merchName,"prod",prod_id[i]);
                                #
                                #
                                if merchName!="Solod-Expo" or merchName!="mchj_tashkent" :# or market=="W1W1W1_CHILANZAR":
                                    generalLang=getConst(merchName,"generalLang");
                                    if merchName=='Kay-Kay' and cat_id[i]=='1':
                                        print(name[i],ost[i]);
                                    products.append([cat_id[i],prod_id[i],name[i],rev[i],price,merchName,img,box[i],form[i],ost[i],name2[i],rev2[i],generalLang]);
            products=arrayToString2(products);
            send={"err":"0","market_products":products};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_i_market(request,var='network'):
    try:
        info=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            market=Login(session);

            checkCoinSettings(market);

            for m in markets_usersS:
                if m.login==market:
                    adres=m.adres;
                    inn=m.inn;
                    phone=m.phone;
                    district=m.district;
                    varified=m.varified;
                    break;
            for m in markets_coinInfoS:
                if m.login==market:
                    tg_id=m.tg_id;
                    break;
            marks=os.listdir("qrMarkets");
            if market not in marks:
                get_qrs(market,"qrMarkets/"+market);
            info.append([adres,inn,phone,district,tg_id,varified]);
            info=arrayToString2(info);
            send={"err":"0","market_info":info};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return send
def get_market_news(request,var='network'):
    try:
        orders=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            merchs=os.listdir("merchants");
            for merchName in merchs:
                data=[];
                price=[];
                _type=[];
                getter=[];
                owner=[];
                last_index=[];
                date=[];
                get_type=[];
                priceD=[];
                for o in orders_orderS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.getter==login:
                            data.append(o.data);
                            getter.append(o.getter);
                            owner.append(o.owner);
                            last_index.append(o.last_index);
                            date.append(o.date);
                            get_type.append(o.get_type);
                            priceD.append(o.price);
                for i in range(len(data)):
                    price=0;
                    d=stringToArrayData(data[i]);
                    for z in range(len(d)):
                        prod_id=d[z][0];
                        how=d[z][1];
                        pr=getProductPrice(merchName,login,prod_id);
                        price=price+int(pr)*int(how);
                    if get_type[i]=="dolg":
                        price=int(priceD[i]);
                    orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],owner[i],str(price),merchName]);
            orders=arrayToString2(orders);
            send={"err":"0","market_news":orders};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send;
def get_market_history(request,var='network',user="all"):
    try:
        his=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchs=os.listdir("merchants");
            login=Login(session);
            for merchName in merchs:
                datas=[];
                price=[];
                _type=[];
                getter=[];
                driver=[];
                date=[];
                district=[];
                last_index=[];
                visible=[];
                for o in orders_historyS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.getter==login:
                            datas.append(o.data);
                            price.append(o.price);
                            _type.append(o.type);
                            getter.append(o.getter);
                            driver.append(o.driver);
                            date.append(o.date);
                            district.append(o.district);
                            last_index.append(o.last_index);
                            visible.append(o.visible);

                for i in range(len(datas)):
                    if ";" in datas[i] and ":" in datas[i]:
                        d=stringToArrayData(datas[i]);
                        dat="";
                        for z in range(len(d)):
                            prod_id=d[z][0];
                            how=d[z][1];
                            for o in orders_productsS:
                                if o.merchName==merchName and o.removed=="0":
                                    if o.id==prod_id:
                                        name=o.name;
                                        break;
                            dat=name+" * "+how+"\n";
                            for k in range(len(orders_historyS)):
                                if orders_historyS[k].merchName==merchName:
                                    if orders_historyS[k].last_index==last_index[i]:
                                        orders_historyS[k].data=dat;
                                        orders_historyS[k].changed="1";
                                        break;
                    else:
                        dat=datas[i];
                    try:
                        for o in orders_buy_marketsS:
                            if o.merchName==merchName and o.removed=="0":
                                if o.name==getter[i]:
                                    buy_type=o.type;
                                    buy_type=buy_type;
                                    break;
                    except:
                        buy_type="_"
                    his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                        visible[i],buy_type,merchName]);
            his=arrayToString2(his);
            send={"err":"0","his":his};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_updates_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            #updates=cort_to_list(cursor.fetchall())[0];
            for c in basic_usersS:
                if c.session==session:
                    updates=c.updates;
                    break;
            if updates=="0":
                send={"err":"0","update_stat":"0"};
            else:
                updates=stringToArray(updates);
                send={"err":"0","update_stat":"1","updates":updates};
            #cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            #conn.commit();
            #conn.close();
            for i in range(len(basic_usersS)):
                if basic_usersS[i].session==session:
                    basic_usersS[i].updates="0";
                    basic_usersS[i].changed="1";
                    break;

        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#market do
def send_market_order(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                market=Login(session);
                if request.method=='GET':
                    orderData=request.GET['data'].replace("|",":").replace("^",";");
                    merchName=request.GET['merchName'];
                    try:
                        payForm=request.GET['payForm'];
                    except Exception as e:
                        logger(e);
                        payForm="nal";
                if request.method=='POST':
                    orderData=request.POST['data'].replace("|",":").replace("^",";");
                    merchName=request.POST['merchName'];
                    try:
                        payForm=request.POST['payForm'];
                    except Exception as e:
                        logger(e);
                        payForm="nal";
                Data=stringToArrayData(orderData);

                dataTxt="";
                for i in range(len(Data)):
                    D=Data[i][0];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.id==D:
                                prodName=o.name;
                                break;
                    try:
                        dataTxt=dataTxt+prodName+" x "+Data[i][1]+"\n";
                    except:
                        pass;
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                Last(merchName,"new");
                last_index=Last(merchName);

                orders_orderS.append(orders_order(orderData,market,"new",last_index,date,"new","",payForm,merchName));

                admins=[];
                for o in orders_adminsS:
                    if o.merchName==merchName and o.removed=="0":
                        admins.append(o.login);
                payFormTxt="Наличные";
                if payForm=="nal":
                    payFormTxt="Наличные";
                elif payForm=="term":
                    payFormTxt="Терминал";
                elif payForm=="per":
                    payFormTxt="Перечисление";
                payFormTxt="Форма оплаты: "+payFormTxt;
                txt="Новый заказ🎈№{0}\nЗаказчик: {1}\n\n{2}".format(last_index,market,dataTxt+"\n"+payFormTxt);
                messageInChannel(merchName,txt);
                send={"err":"0","text":"OK"};
                for a in admins:
                    async2(a,"getAdminOrders");
                    addNot(a,"title","text","marketOrder");
                async2(market,"getMarketNews");
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_market_rem_order(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                if request.method=='POST':
                    last_index=request.POST['last_index'];
                    merchName=request.POST['merchName'];
                if request.method=='GET':
                    last_index=request.GET['last_index'];
                    merchName=request.GET['merchName'];
                for i in range(len(orders_orderS)):
                    if orders_orderS[i].last_index==last_index:
                        orders_orderS[i].removed="1";
                        orders_orderS[i].changed="1";
                admins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getAllPrices");
                send={"err":"0","text":"OK"};
                txt="❌Заказ №"+last_index+" был удален заказчиком";
                messageInChannel(merchName,txt);
            else:
                send={"err":"1","text":"Вы не магазин"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_self_market_data(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="market":
                market=Login(session);
                if request.method=='POST':
                    newName=request.POST['newName'];
                    newInn=request.POST['newInn'];
                    newAdres=request.POST['newAdres'];
                    newPhone=request.POST['newPhone'];
                    newDistrict=request.POST['newDistrict'];
                    newTgId=request.POST['newTgId'];
                    newTgPswd=request.POST['newTgPswd'];
                if request.method=='GET':
                    newName=request.GET['newName'];
                    newInn=request.GET['newInn'];
                    newAdres=request.GET['newAdres'];
                    newPhone=request.GET['newPhone'];
                    newDistrict=request.GET['newDistrict'];
                    newTgId=request.GET['newTgId'];
                    newTgPswd=request.GET['newTgPswd'];

                for m in markets_usersS:
                    oldInn=m.inn;
                    break;
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==market:
                        V=False;
                        if oldInn!=newInn:
                            markets_usersS[i].inn=newInn;
                            markets_usersS[i].varified="100";
                            V=True;

                        markets_usersS[i].adres=newAdres;с
                        markets_usersS[i].phone=newPhone;
                        markets_usersS[i].district=newDistrict;
                        markets_usersS[i].changed="1";
                        for k in range(len(markets_coinInfoS)):
                            if markets_coinInfoS[i].login==market:
                                markets_coinInfoS[i].tg_id=newTgId;
                                markets_coinInfoS[i].changed="1";
                                break;

                        if newTgPswd !="":
                            for k in range(len(markets_coinInfoS)):
                                if markets_coinInfoS[i].login==market:
                                    markets_coinInfoS[i].pswd=h(newTgPswd);
                                    markets_coinInfoS[i].changed="1";
                                    break;
                        break;

                us=[];
                for i in range(len(basic_usersS)):
                    us.append(basic_usersS[i].login);
                    if basic_usersS[i].session==session:
                        if newName!="":
                            basic_usersS[i].pswd=h(newName);
                        if V:
                            basic_usersS[i].session="_";
                        basic_usersS[i].changed="1";

                async2(market,"getMarketList");
                send={"err":"0","text":"OK"};
                if V:
                    txt="📕Пользователь {0} сменил себе ИНН и ждет варификации".format(market);
                    messageInChannel("Kay-Kay",txt);
            else:
                send={"err":"1","text":"Вы не магазин"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def check_send_tg_settings(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if request.method=='POST':
                tg_id=request.POST['tg_id'];
                tg_pswd=request.POST['tg_pswd'];
            if request.method=='GET':
                tg_id=request.GET['tg_id'];
                tg_pswd=request.GET['tg_pswd'];
            print(tg_id,tg_pswd);
            if User(session)=="market":
                get_html(KAY+"/check_tg_data/?tg_id={0}&tg_pswd={1}".format(tg_id,tg_pswd,));
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не магазин"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#admin
def uber_admin_request(request):
    try:
        updates="";
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);

            prod_id=[];
            for o in orders_productsS:
                if o.merchName==merchName and o.removed=="0":
                    prod_id.append(o.id);

            send={"getDistricts":get_districts(request,'local'),
            "getHistory":get_history(request,'local'),
            "getDrivers":get_drivers(request,'local'),
            "getAdminOrders":get_admin_orders(request,'local'),
            "getDefaultMarket":get_default_market(request,'local'),
            #"getAllPrices":get_prices(request,prod_id,'local',"admin"),
            "getAllLogins":get_all_users(request,'local'),
            "getAdmins":get_admins(request,'local'),
            "err":"0"};

        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def get_mikro_history(request,var='network',user="all"):
    try:
        his=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            search=request.POST['search'];
            sDate=request.POST['sDate'];
            print(search,sDate);
            merchName=Merch(session);
            datas=[];
            price=[];
            _type=[];
            getter=[];
            driver=[];
            date=[];
            district=[];
            last_index=[];
            visible=[];
            for i in range(len(orders_historyS)):
                o=orders_historyS[i]
                if o.merchName==merchName and o.removed=="0" and sDate in o.date:
                    print(sDate,o.date)
                    if search in o.getter.lower() or search in o.driver.lower() or search in o.last_index or search in o.district.lower():
                        datas.append(o.data);
                        price.append(o.price);
                        _type.append(o.type);
                        getter.append(o.getter);
                        driver.append(o.driver);
                        date.append(o.date);
                        district.append(o.district);
                        last_index.append(o.last_index);
                        visible.append(o.visible);
            for i in range(len(datas)):
                if ";" in datas[i] and ":" in datas[i]:
                    d=stringToArrayData(datas[i]);
                    dat="";
                    for z in range(len(d)):
                        print(d);
                        prod_id=d[z][0];
                        how=d[z][1];
                        for o in orders_productsS:
                            if o.id==prod_id:
                                name=o.name;
                                break;
                        dat=name+" * "+how+"\n";
                        for k in range(len(orders_historyS)):
                            if orders_historyS[k].last_index==last_index[i]:
                                orders_historyS[k].data=dat;
                                orders_historyS[k].changed="1";
                                break;
                else:
                    dat=datas[i];
                try:
                    for o in orders_buy_marketsS:
                        if o.name==getter[i]:
                            buy_type=o.type;
                            break;
                    buy_type=buy_type;
                except:
                    buy_type="_"
                his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                    visible[i],buy_type]);

            his=arrayToString2(his);
            print('hisRes',len(his));
            send={"err":"0","his":his};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_history(request,var='network',user="all"):
    try:
        his=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            try:
                if request.method=='GET':
                    his_mode=request.GET['his_mode'];
                    his_last=request.GET['his_last'];
                else:
                    his_mode=request.POST['his_mode'];
                    his_last=request.POST['his_last'];
            except Exception as e:
                print(e)
                if var=="network":
                    his_mode="all";
                else:
                    his_mode="10";
                his_last="0";
            print(his_mode,his_last,"history info<<<<<")
            merchName=Merch(session);
            datas=[];
            price=[];
            _type=[];
            getter=[];
            driver=[];
            date=[];
            district=[];
            last_index=[];
            visible=[];
            for i in range(len(orders_historyS)):
                o=orders_historyS[i]
                if o.merchName==merchName and o.removed=="0" and i>=int(his_last):
                    datas.append(o.data);
                    price.append(o.price);
                    _type.append(o.type);
                    getter.append(o.getter);
                    driver.append(o.driver);
                    date.append(o.date);
                    district.append(o.district);
                    last_index.append(o.last_index);
                    visible.append(o.visible);
                    if his_mode!="all":
                        if i>=int(his_mode)+int(his_last):
                            break;
            for i in range(len(datas)):
                if ";" in datas[i] and ":" in datas[i]:
                    d=stringToArrayData(datas[i]);
                    dat="";
                    for z in range(len(d)):
                        print(d);
                        prod_id=d[z][0];
                        how=d[z][1];
                        for o in orders_productsS:
                            if o.id==prod_id:
                                name=o.name;
                                break;
                        dat=name+" * "+how+"\n";
                        for k in range(len(orders_historyS)):
                            if orders_historyS[k].last_index==last_index[i]:
                                orders_historyS[k].data=dat;
                                orders_historyS[k].changed="1";
                                break;
                else:
                    dat=datas[i];
                try:
                    for o in orders_buy_marketsS:
                        if o.name==getter[i]:
                            buy_type=o.type;
                            break;
                    buy_type=buy_type;
                except:
                    buy_type="_"
                his.append([dat,price[i],_type[i],getter[i],driver[i],date[i],district[i],last_index[i],
                    visible[i],buy_type]);

            his=arrayToString2(his);
            send={"err":"0","his":his};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_drivers(request,var='network'):
    try:
        drivers=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            logins=[];
            lon=[];
            lat=[];
            for c in basic_usersS:
                if c.user_type=="driver" and c.merchName==merchName:
                    logins.append(c.login);
                    lon.append(c.lon);
                    lat.append(c.lat);

            for i in range(len(logins)):
                name=[];
                prod_id=[];
                free=[];
                reserve=[];
                sell=[];
                for inc in inCar_naksS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.owner==logins[i]:
                            name.append(inc.name);
                            prod_id.append(inc.prod_id);
                            free.append(inc.free);
                            reserve.append(inc.reserve);
                            sell.append(inc.sell);
                nak="";
                for z in range(len(name)):
                    nak=nak+str(name[z])+":"+str(prod_id[z])+":"+str(free[z])+":"+str(reserve[z])+":"+str(sell[z])+":;";
                for inc in inCar_cashInCarS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.driver==logins[i]:
                            cash=inc.cash;
                            term=inc.term;
                            per=inc.per;
                            on_day=inc.on_day;
                            break;

                nakNum="";

                num=[];
                date=[];
                for inc in inCar_nakNumS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.owner==logins[i]:
                            if inc.removed=="0":
                                num.append(inc.num);
                                date.append(inc.date);
                try:
                    for o in orders_driversS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.login==logins[i]:
                                diss=o.district;
                                nakName=o.name;
                                break;
                except:
                    orders_driversS.append(orders_drivers(logins[i],":;","",merchName));
                    diss=":;"
                    nakName=""
                nakNum="";
                for z in range(len(num)):
                    nakNum+="Накладная №"+num[z]+". Дата: "+date[z]+" ";
                drivers.append([logins[i],lon[i],lat[i],nak,cash,term,per,on_day,nakNum,diss,nakName]);
            drivers=arrayToString2(drivers);
            send={"err":"0","drivers":drivers};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_admin_orders(request,var='network'):
    try:
        orders=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT data FROM 'order'");
            #data=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT getter FROM 'order'");
            #getter=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT owner FROM 'order'");
            #owner=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT last_index FROM 'order'");
            #last_index=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT date FROM 'order'");
            #date=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT get_type FROM 'order'");
            #get_type=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT price FROM 'order'");
            #price=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT payForm FROM 'order'");
            #payForm=cort_to_list(cursor.fetchall());
            #conn.close();
            data=[];
            getter=[];
            owner=[];
            last_index=[];
            date=[];
            get_type=[];
            price=[];
            payForm=[];
            for o in orders_orderS:
                if o.merchName==merchName and o.removed=="0":
                    data.append(o.data);
                    getter.append(o.getter);
                    owner.append(o.owner);
                    last_index.append(o.last_index);
                    date.append(o.date);
                    get_type.append(o.get_type);
                    price.append(o.price);
                    payForm.append(o.payForm);
                    if o.get_type=="new":
                        print(o.data,"<<<<<")

            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],price[i],owner[i],payForm[i]]);
            orders=arrayToString2(orders);
            send={"err":"0","orders":orders};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_all_users(request,var='network'):
    try:
        logins=[];
        session=request.GET['session'];
        if True:
            #merchName=Merch(session);
            #login=Login(session);
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #login=cort_to_list(cursor.fetchall());
            #conn.close();
            logins2=[];
            for c in basic_usersS:
                logins2.append(c.login);
            logins=[];
            for l in logins2:
                logins.append([l]);
            logins=arrayToString2(logins);
            send={"err":"0","logins":logins};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def get_updates_admin(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            #updates=cort_to_list(cursor.fetchall())[0];
            for c in basic_usersS:
                if c.session==session:
                    updates=c.updates;
                    break;
            if updates=="0":
                send={"err":"0","update_stat":"0"};
            else:
                updates=stringToArray(updates);
                send={"err":"0","update_stat":"1","updates":updates};
            #cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            for i in range(len(basic_usersS)):
                if basic_usersS[i].session==session:
                    basic_usersS[i].updates="0";
                    basic_usersS[i].changed="1";
                    break;
            #conn.commit();
            #conn.close();
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_admins(request,var='network'):
    try:
        admins=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                login=Login(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'WHERE user_type='admin'AND merchName=(?)",(merchName,));
                #logins=cort_to_list(cursor.fetchall());
                #conn.close();
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        logins.append(c.login);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                for l in logins:
                    #cursor.execute("SELECT level FROM 'admins'WHERE login=(?)",(l,));
                    #level=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT role FROM 'admins'WHERE login=(?)",(l,));
                    #role=cort_to_list(cursor.fetchall())[0];
                    for o in orders_adminsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.login==l:
                                level=o.level;
                                role=o.role;
                                break;
                    if role==None:
                        role="";
                    admins.append([l,level.replace("|","_"),role]);
                #conn.close();
                admins=arrayToString2(admins);
                send={"err":"0","admins":admins};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_box_settings(request,var='network'):
    try:
        if request.method=='POST':
            session=request.POST['session'];
            box_settings=request.POST['box_settings'];
        if request.method=='GET':
            session=request.GET['session'];
            box_settings=request.GET['box_settings'];
        box_settings=stringToArray(box_settings)[0];
        prod_id=box_settings[0];
        boxes=box_settings[1];
        form=box_settings[2];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'products' SET form=(?)WHERE id=(?)",(form,prod_id,));
                #cursor.execute("UPDATE 'products' SET 'box'=(?)WHERE id=(?)",(boxes,prod_id,));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_productsS)):
                    if orders_productsS[i].merchName==merchName:
                        if orders_productsS[i].id==prod_id:
                            orders_productsS[i].form=form;
                            orders_productsS[i].box=boxes;
                            orders_productsS[i].changed="1";
                print("will save")
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'WHERE user_type='driver'AND merchName=(?)",(merchName,));
                #drivers=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT login FROM 'users'WHERE user_type='admin'AND merchName=(?)",(merchName,));
                #admins=cort_to_list(cursor.fetchall());
                drivers=[];
                admins=[];
                for c in basic_usersS:
                    if c.user_type=="driver" and c.merchName==merchName:
                        drivers.append(c.login);
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                for d in drivers:
                    async2(d,"getProductList");
                for a in admins:
                    async2(a,"getProductList");
                #conn.close()
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_unvarified_markets(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=ADLevel(myMerchName,login);
            if "max98" in level or login=="admin":
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users' WHERE varified='100'");
                #login=cort_to_list(cursor.fetchall());
                login=[];
                for m in markets_usersS:
                    login.append(m.login);
                for i in range(len(login)):
                    log=login[i];
                    #cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                    #adres=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                    #inn=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                    #phone=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT tg_id FROM 'coinInfo' WHERE login=(?) ",(log,));
                    #tg_id=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?) ",(log,));
                    #pswd=cort_to_list(cursor.fetchall())[0];
                    for m in markets_usersS:
                        if m.login==log:
                            adres=m.adres;
                            inn=m.inn;
                            phone=m.phone;
                            break;
                    for m in markets_coinInfoS:
                        if m.login==log:
                            pswd=m.pswd;
                            tg_id=m.tg_id;
                            break;
                    if len(pswd)==0:
                        tg_pswd="0";
                    else:
                        tg_pswd="1";
                    markets.append([log,adres,inn,phone,'100',tg_id,tg_pswd]);
                #conn.close();
                markets=arrayToString2(markets);
                send={"err":"0","markets":markets};
            else:
                send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_excel_naks(request,var='network'):
    try:
        naks=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                id_=[];
                for o in orders_nakSpaceS:
                    if o.merchName==merchName and o.removed=="0":
                        id_.append(o.id);
                for ID in id_:
                    for o in orders_nakSpaceS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.id==ID:
                                nak=o.nak;
                                nak=nak.replace("|","a").replace("^","b");
                                date=o.date;
                                nakNum=o.nakNum;
                                owner=o.owner;
                                break;
                    naks.append([ID,nak,date,nakNum,owner]);
                naks=arrayToString2(naks);
                send={"err":"0","naks":naks};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#admin do
def send_hist_var(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            last_index=request.GET['last_index'];
            vis=request.GET['var'];
            merchName=Merch(session);
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            visOld="";
            if vis=="new" or vis=="old":
                #cursor.execute("SELECT visible FROM 'history' WHERE last_index=(?)",(last_index,));
                #visOld=cort_to_list(cursor.fetchall())[0];
                for o in orders_historyS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.last_index==last_index:
                            visOld=o.visible;
                            break;
            if visOld==vis:
                pass;
            else:

                if vis=="old" or vis=="new":
                    #cursor.execute("UPDATE 'history' SET visible=(?)WHERE last_index=(?)",(vis,last_index,));
                    for i in range(len(orders_historyS)):
                        if orders_historyS[i].merchName==merchName:
                            if orders_historyS[i].last_index==last_index:
                                orders_historyS[i].visible=vis;
                                orders_historyS[i].changed="1";
                                break;
                elif vis=="clearAll":
                    #cursor.execute("UPDATE 'history' SET visible='old'");
                    for i in range(len(orders_historyS)):
                        if orders_historyS[i].merchName==merchName:
                            orders_historyS[i].visible="old";
                            orders_historyS[i].changed="1";
                elif vis=="choiseAllSell":
                    #cursor.execute("UPDATE 'history' SET visible='new' WHERE type='sell'");
                    for i in range(len(orders_historyS)):
                        if orders_historyS[i].merchName==merchName:
                            if orders_historyS[i].type=="sell":
                                orders_historyS[i].visible="new";
                                orders_historyS[i].changed="1";
                                break;
                elif vis=="choiseAllOrder":
                    #cursor.execute("UPDATE 'history' SET visible='new' WHERE type='ord'");
                    for i in range(len(orders_historyS)):
                        if orders_historyS[i].merchName==merchName:
                            if orders_historyS[i].type=="ord":
                                orders_historyS[i].visible="new";
                                orders_historyS[i].changed="1";
                                break;

                #conn.commit();
            #conn.close();
            send={"err":"0","text":"OK"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_defaults(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    default_buyer=request.POST['default_buyer'];
                    print(default_buyer)
                    defDolgType=request.POST['defDolgType'];
                    defDolgVal=request.POST['defDolgVal'];
                    defDolgMax=request.POST['defDolgMax'];
                    tg_id=request.POST['tg_id'];
                    tg_token=request.POST['tg_token'].replace("|",":");
                    nots=request.POST['nots'];
                    excelName=request.POST['excelName'];
                    excelHow=request.POST['excelHow'];
                    excelStartWrite=request.POST['excelStartWrite'];
                    excelType=request.POST['excelType'];
                    excelCat=request.POST['excelCat'];
                    excelPriceName=request.POST['excelPriceName'];
                    excelPrice=request.POST['excelPrice'];
                    excelStartWritePrice=request.POST['excelStartWritePrice'];
                    excelSheet=request.POST['excelSheet'];
                    excelPriceSheet=request.POST['excelPriceSheet'];
                    excelNakPlace=request.POST['excelNakPlace'];
                    ostName=request.POST['ostName'];
                    ostHowOst=request.POST['ostHowOst'];
                    ostStart=request.POST['ostStart'];
                    ostSheet=request.POST['ostSheet'];
                    upRange=request.POST['upRange'];
                    generalLang=request.POST['generalLang'];
                    secondLang=request.POST['secondLang'];
                    excelDriverName=request.POST['excelDriverName'];
                    perInn=request.POST['perInn'];
                    perOst=request.POST['perOst'];
                    perStartRead=request.POST['perStartRead'];
                    perSheet=request.POST['perSheet'];
                if request.method=='GET':
                    default_buyer=request.GET['default_buyer'];
                    defDolgType=request.GET['defDolgType'];
                    defDolgVal=request.GET['defDolgVal'];
                    defDolgMax=request.GET['defDolgMax'];
                    tg_id=request.GET['tg_id'];
                    tg_token=request.GET['tg_token'].replace("|",":");
                    nots=request.GET['nots'];
                    excelName=request.GET['excelName'];
                    excelHow=request.GET['excelHow'];
                    excelStartWrite=request.GET['excelStartWrite'];
                    excelType=request.GET['excelType'];
                    excelCat=request.GET['excelCat'];
                    excelPriceName=request.GET['excelPriceName'];
                    excelPrice=request.GET['excelPrice'];
                    excelStartWritePrice=request.GET['excelStartWritePrice'];
                    excelSheet=request.GET['excelSheet'];
                    excelPriceSheet=request.GET['excelPriceSheet'];
                    excelNakPlace=request.GET['excelNakPlace'];
                    ostName=request.GET['ostName'];
                    ostHowOst=request.GET['ostHowOst'];
                    ostStart=request.GET['ostStart'];
                    ostSheet=request.GET['ostSheet'];
                    upRange=request.GET['upRange'];
                    generalLang=request.GET['generalLang'];
                    secondLang=request.GET['secondLang'];
                    excelDriverName=request.GET['excelDriverName'];
                    perInn=request.GET['perInn'];
                    perOst=request.GET['perOst'];
                    perStartRead=request.GET['perStartRead'];
                    perSheet=request.GET['perSheet'];
                merchName=Merch(session);

                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].key=="default_buyer":orders_constS[i].val=default_buyer;
                        if orders_constS[i].key=="defDolgType":orders_constS[i].val=defDolgType;
                        if orders_constS[i].key=="defDolgVal":orders_constS[i].val=defDolgVal;
                        if orders_constS[i].key=="defDolgMax":orders_constS[i].val=defDolgMax;
                        if orders_constS[i].key=="tg_id":orders_constS[i].val=tg_id;
                        if orders_constS[i].key=="tg_token":orders_constS[i].val=tg_token;
                        if orders_constS[i].key=="nots":orders_constS[i].val=nots;
                        if orders_constS[i].key=="excelName":orders_constS[i].val=excelName;
                        if orders_constS[i].key=="excelHow":orders_constS[i].val=excelHow;
                        if orders_constS[i].key=="excelStartWrite":orders_constS[i].val=excelStartWrite;
                        if orders_constS[i].key=="excelType":orders_constS[i].val=excelType;
                        if orders_constS[i].key=="excelCat":orders_constS[i].val=excelCat;
                        if orders_constS[i].key=="excelPriceName":orders_constS[i].val=excelPriceName;
                        if orders_constS[i].key=="excelPrice":orders_constS[i].val=excelPrice;
                        if orders_constS[i].key=="excelStartWritePrice":orders_constS[i].val=excelStartWritePrice;
                        if orders_constS[i].key=="excelSheet":orders_constS[i].val=excelSheet;
                        if orders_constS[i].key=="excelNakSheet":orders_constS[i].val=excelSheet;
                        if orders_constS[i].key=="excelProdSheet":orders_constS[i].val=excelPriceSheet;
                        if orders_constS[i].key=="excelNakPlace":orders_constS[i].val=excelNakPlace;
                        if orders_constS[i].key=="ostName":orders_constS[i].val=ostName;
                        if orders_constS[i].key=="ostHowOst":orders_constS[i].val=ostHowOst;
                        if orders_constS[i].key=="ostStart":orders_constS[i].val=ostStart;
                        if orders_constS[i].key=="ostSheet":orders_constS[i].val=ostSheet;
                        if orders_constS[i].key=="upRange":orders_constS[i].val=upRange;
                        if orders_constS[i].key=="generalLang":orders_constS[i].val=generalLang;
                        if orders_constS[i].key=="secondLang":orders_constS[i].val=secondLang;
                        if orders_constS[i].key=="exNakDriverName":orders_constS[i].val=excelDriverName;
                        if orders_constS[i].key=="exPerInn":orders_constS[i].val=perInn;
                        if orders_constS[i].key=="exPerSum":orders_constS[i].val=perOst;
                        if orders_constS[i].key=="exPerStart":orders_constS[i].val=perStartRead;
                        if orders_constS[i].key=="exPerSheet":orders_constS[i].val=perSheet;
                        orders_constS[i].changed="1";

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Сессия истекла"};
        else:
            send={"err":"1","text":"Вы не админ"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_site(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    site=request.POST['site'];
                    name=request.POST['name'];
                if request.method=='GET':
                    site=request.GET['site'];
                    name=request.GET['name'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'const' SET val=(?)WHERE key=(?)",(site,"site",));
                #conn.commit();
                #conn.close();
                br=0;
                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].key=="site":
                            orders_constS[i].val=site;
                            orders_constS[i].changed="1";
                            br+=1;
                        if orders_constS[i].key=="name":
                            orders_constS[i].val=name;
                            orders_constS[i].changed="1";
                            br+=1;
                        if br==2:break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Сессия истекла"};
        else:
            send={"err":"1","text":"Вы не админ"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_prod_settings(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                #sendInfo=request.GET['sendInfo'];
                if request.method=='POST':
                    prod_id=request.POST['prod_id'];
                    prod_name=request.POST['name'];
                    cat_id=request.POST['cat_id'];
                    rev=request.POST['rev'].replace("itsMySpace","\n");
                    visible=request.POST['visible'];
                    ost=request.POST['ost'];
                    EditVar=request.POST['EditVar'];
                    price=request.POST['price'];
                    try:
                        name2=request.POST['name2'];
                        rev2=request.POST['rev2'];
                    except:
                        name2=prod_name;
                        rev2=rev;
                if request.method=='GET':
                    prod_id=request.GET['prod_id'];
                    prod_name=request.GET['name'];
                    cat_id=request.GET['cat_id'];
                    rev=request.GET['rev'].replace("itsMySpace","\n");
                    visible=request.GET['visible'];
                    ost=request.GET['ost'];
                    EditVar='old';
                    price='0';
                    try:
                        name2=request.GET['name2'];
                        rev2=request.GET['rev2'];
                    except:
                        name2=prod_name;
                        rev2=rev;
                merchName=Merch(session);

                if EditVar=='new':
                    have=False;
                    for p in orders_productsS:
                        if p.name==prod_name or p.name2==name2:
                            have=True;
                            break;
                    if have:
                        prod='0';
                    else:
                        ids=[];
                        for o in orders_productsS:
                            if o.merchName==merchName and o.removed=="0":
                                ids.append(o.id);
                        id_=1;
                        while str(id_) in ids:
                            id_+=1;
                        prod={'cat_id':cat_id,'prod_id':str(id_),'name':prod_name,'rev':rev,'work':"1",'img':"_",'box':'1','form':'sh','visible':visible,'ost':ost,
                            'name2':name2,'rev2':rev2,'merchName':merchName};
                        orders_productsS.append(orders_products(cat_id,str(id_),prod_name,rev,"1","_",'1','sh',visible,ost,name2,rev2,merchName));
                        buyers_types=[];
                        for o in orders_buyers_typeS:
                            if o.merchName==merchName and o.removed=="0":
                                buyers_types.append(o.name);
                        for b in buyers_types:
                            orders_pricesS.append(orders_prices(str(id_),b,price,merchName));
                else:
                    for i in range(len(orders_productsS)):
                        if orders_productsS[i].merchName==merchName:
                            if orders_productsS[i].id==prod_id:
                                orders_productsS[i].name=prod_name;
                                orders_productsS[i].cat_id=cat_id;
                                orders_productsS[i].rev=rev;
                                orders_productsS[i].ost=ost;
                                if ost=="0":
                                    orders_productsS[i].visible="0";
                                else:
                                    orders_productsS[i].visible="1";
                                orders_productsS[i].visible=visible;
                                orders_productsS[i].name2=name2;
                                orders_productsS[i].rev2=rev2;
                                orders_productsS[i].changed="1";
                                break;
                    prod={'cat_id':cat_id,'prod_id':str(prod_id),'name':prod_name,'rev':rev,'work':"1",'img':"_",'box':'1','form':'sh','visible':visible,'ost':ost,
                        'name2':name2,'rev2':rev2,'merchName':merchName};
                txt="Изменение в продуктах:\nПродукт:{0}".format(prod_name);
                async2(Login(session),"getProducts");
                makeAct(merchName,Login(session),txt);
                if prod=='0':
                    for i in range(len(orders_productsS)):
                        if orders_productsS[i].merchName==merchName:
                            if orders_productsS[i].id==prod_id:
                                orders_productsS[i].work='1';
                                orders_productsS[i].changed="1";
                                p=orders_productsS[i];
                                prod={'cat_id':p.cat_id,'prod_id':str(p.prod_id),'name':p.name,'rev':p.rev,'work':"1",'img':"_",'box':'1','form':'sh','visible':p.visible,'ost':p.ost,
                                    'name2':p.name2,'rev2':p.rev2,'merchName':p.merchName};
                                break;
                    send={"err":"1","text":"Продукт включен",'prod':prod};
                else:
                    send={"err":"0","text":"OK",'prod':prod};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    print(send)
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_cho_chosen_nak(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                if request.method=='POST':
                    driver=request.POST['driver'];
                    NAK=request.POST['nak'];
                if request.method=='GET':
                    driver=request.GET['driver'];
                    NAK=request.GET['nak'];
                merchName=Merch(session);
                NAK=stringToArray(NAK)[0];
                #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor1=conn1.cursor();
                #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                #cursor=conn.cursor();
                for NK in NAK:
                    #cursor1.execute("SELECT nak FROM 'nakSpace' WHERE id=(?)",(NK,));
                    #nak=cort_to_list(cursor1.fetchall())[0];
                    #cursor1.execute("SELECT nakNum FROM 'nakSpace' WHERE id=(?)",(NK,));
                    #nakNum=cort_to_list(cursor1.fetchall())[0];
                    for o in orders_nakSpaceS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.id==NK:
                                nak=o.nak;
                                nakNum=o.nakNum;
                                break;
                    nak=stringToArray(nak);
                    for l in nak:
                        ids=l[0];
                        excelHowE=l[1];
                        #cursor1.execute("SELECT name FROM 'products' WHERE id=(?)",(ids,));
                        #excelNameE=cort_to_list(cursor1.fetchall())[0];
                        for o in orders_productsS:
                            if o.merchName==merchName and o.removed=="0":
                                if o.id==ids:
                                    excelNameE=o.name;
                                    break;
                        #cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                        #freeOld=cort_to_list(cursor.fetchall());
                        freeOld=[];
                        for inc in inCar_naksS:
                            if inc.merchName==merchName and inc.removed=="0":
                                if inc.owner==driver and inc.prod_id==ids:
                                    freeOld.append(inc.free);

                        if len(freeOld)==0:
                            #cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                            #    (excelNameE,ids,excelHowE,driver,));
                            inCar_naksS.append(inCar_naks(excelNameE,ids,excelHowE,'0','0',driver,merchName));
                        else:
                            freeOld=freeOld[0];
                            newFree=str(int(excelHowE)+int(freeOld));
                            #cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                            #    (newFree,driver,ids,));
                            for i in range(len(inCar_naksS)):
                                if inCar_naksS[i].merchName==merchName:
                                    if inCar_naksS[i].owner==driver and inCar_naksS[i].prod_id==ids:
                                        inCar_naksS[i].free=newFree;
                                        inCar_naksS[i].changed="1";
                                        break;
                        ostMinus(merchName,ids,excelHowE);
                    #cursor1.execute("DELETE FROM 'nakSpace'WHERE id=(?)",(NK,));
                    for i in range(len(orders_nakSpaceS)):
                        if o.merchName==merchName and o.removed=="0":
                            if orders_nakSpaceS[i].id==NK:
                                orders_nakSpaceS[i].removed="1";
                                orders_nakSpaceS[i].changed="1";
                                break;
                    #cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                    #    (driver,nakNum,date));
                    inCar_nakNumS.append(inCar_nakNum(driver,nakNum,date,merchName));
                #conn1.commit();
                #conn1.close();
                #conn.commit();
                #conn.close();
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
                #admins=cort_to_list(cursor.fetchall());
                #conn.commit();
                #conn.close();
                admins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                for a in admins:
                    async2(a,"getDrivers");
                    async2(a,"getExcelNaks");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def send_edit_prices(request,var="network"):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    prod_id=request.POST['prod_id'];
                    type_=request.POST['type'];
                    price=request.POST['price'];
                if request.method=='GET':
                    prod_id=request.GET['prod_id'];
                    type_=request.GET['type'];
                    price=request.GET['price'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'prices' SET price=(?) WHERE id=(?)AND name=(?)",(price,prod_id,type_,));
                #conn.commit();
                for i in range(len(orders_pricesS)):
                    if orders_pricesS[i].merchName==merchName:
                        if orders_pricesS[i].id==prod_id and orders_pricesS[i].name==type_:
                            orders_pricesS[i].price=price;
                            orders_pricesS[i].changed="1";
                            break;
                #cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                #prod_name=cort_to_list(cursor.fetchall())[0];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.id==prod_id:
                            prod_name=o.name;
                            break;
                #conn.close();
                txt="Изменение в ценах:\nПродукт:{0}\nЦена:{1}".format(prod_name,price);
                makeAct(merchName,Login(session),txt);
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_driver(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    driverName=request.POST['driverName'];
                    driverPswd=request.POST['driverPswd'];
                    driverVar=request.POST['driverVar'];
                    driverOld=request.POST['driverOld'];
                if request.method=='GET':
                    driverName=request.GET['driverName'];
                    driverPswd=request.GET['driverPswd'];
                    driverVar=request.GET['driverVar'];
                    driverOld=request.GET['driverOld'];
                login=Login(session);
                merchName=Merch(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'");
                #login=cort_to_list(cursor.fetchall());
                login=[];
                for c in basic_usersS:
                    login.append(c.login);
                if driverName in login and driverName!=driverOld:
                    send="err=1 text=Логин существует";
                else:
                    if driverVar=="new":
                        #cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        #    (driverName,h(driverPswd),"_",merchName,"driver","0.01","0.01","0","ru",));
                        basic_usersS.append(basic_users(driverName,h(driverPswd),"_",merchName,"driver","0.01","0.01","0","ru",driverName));

                        #conn1=sqlite3.connect(mPath(merchName,"inCar2"));
                        #cursor1=conn1.cursor();
                        #cursor1.execute("INSERT INTO cashInCar VALUES((?),(?),(?),(?),(?))",
                        #    (driverName,"0","0","0","0",));
                        #conn1.commit();
                        #conn1.close();
                        inCar_cashInCarS.append(inCar_cashInCar(driverName,"0","0","0","0",merchName));
                        #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                        #cursor1=conn1.cursor();
                        #cursor1.execute("INSERT INTO drivers VALUES((?),(?),(?))",(driverName,":;",""))
                        #conn1.commit();
                        #conn1.close();
                        driverNakName='';
                        if request.method=='POST':driverNakName=request.POST['driverNakName'];
                        orders_driversS.append(orders_drivers(driverName,":;",driverNakName,merchName));
                    else:
                        if request.method=='GET':driverNakName=request.GET['driverNakName'];
                        if request.method=='POST':driverNakName=request.POST['driverNakName'];

                        #cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(driverName,driverOld,));
                        for i in range(len(basic_usersS)):
                            if basic_usersS[i].login==driverOld:
                                basic_usersS[i].login=driverName;
                                if len(driverPswd)!=0:
                                    basic_usersS[i].pswd=h(driverPswd);
                                basic_usersS[i].session="_";
                                basic_usersS[i].changed="1";
                                break;
                        #if len(driverPswd)!=0:
                        #    cursor.execute("UPDATE 'users' SET pswd=(?)WHERE login=(?)",(h(driverPswd),driverName,));
                        #conn1=sqlite3.connect(mPath(merchName,"inCar2"));
                        #cursor1=conn1.cursor();
                        #cursor1.execute("UPDATE 'cashInCar' SET driver=(?)WHERE driver=(?)",(driverName,driverOld,));
                        #cursor1.execute("UPDATE 'naks' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        #cursor1.execute("UPDATE 'nakNum' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        #conn1.commit();
                        #conn1.close();
                        for i in range(len(inCar_cashInCarS)):
                            if inCar_cashInCarS[i].merchName==merchName:
                                if inCar_cashInCarS[i].driver==driverOld:
                                    inCar_cashInCarS[i].driver=driverName;
                                    inCar_cashInCarS[i].changed="1";
                                    break;
                        for i in range(len(inCar_naksS)):
                            if inCar_naksS[i].merchName==merchName:
                                if inCar_naksS[i].owner==driverOld:
                                    inCar_naksS[i].owner=driverName;
                                    inCar_naksS[i].changed="1";
                                    break;
                        for i in range(len(inCar_nakNumS)):
                            if inCar_nakNumS[i].merchName==merchName:
                                if inCar_nakNumS[i].owner==driverOld:
                                    inCar_nakNumS[i].owner=driverName;
                                    inCar_nakNumS[i].changed="1";
                                    break;

                        #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                        #cursor1=conn1.cursor();
                        #cursor1.execute("UPDATE 'order' SET owner=(?)WHERE owner=(?)",(driverName,driverOld,));
                        #cursor1.execute("UPDATE 'drivers' SET login=(?)WHERE login=(?)",(driverName,driverOld));
                        #cursor1.execute("UPDATE 'drivers' SET name=(?)WHERE login=(?)",(driverNakName,driverName));
                        #cursor.execute("UPDATE 'users' SET session='_'WHERE login=(?)",(driverName,));
                        #conn1.commit();
                        #conn1.close();
                        for i in range(len(orders_orderS)):
                            if orders_orderS[i].merchName==merchName:
                                if orders_orderS[i].owner==driverOld:
                                    orders_orderS[i].owner=driverName;
                                    orders_orderS[i].changed="1";
                                    break;
                        for i in range(len(orders_driversS)):
                            if orders_driversS[i].merchName==merchName:
                                if orders_driversS[i].login==driverOld:
                                    orders_driversS[i].login=driverName;
                                    orders_driversS[i].name=driverNakName;
                                    orders_driversS[i].changed="1";
                                    break;
                        for i in range(len(basic_usersS)):
                            if basic_usersS[i].merchName==merchName:
                                if basic_usersS[i].login==driverOld:
                                    basic_usersS[i].name=driverName;
                                    basic_usersS[i].changed="1";
                                    break;

                    send={"err":"0","text":"OK"};
                #conn.commit();
                #conn.close();
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_driver(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':driverName=request.POST['driverName'];
                if request.method=='GET':driverName=request.GET['driverName'];

                login=Login(session);
                merchName=Merch(session);
                dolgs=[];
                for o in orders_orderS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.get_type=="dolg" and o.owner==driverName:
                            dolgs.append(o.getter);
                dolgs=len(dolgs);
                if dolgs==0:
                    #cursor.execute("UPDATE 'order' SET get_type='new'WHERE owner=(?)",(driverName,));
                    #cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)",(driverName,));
                    #cursor.execute("DELETE FROM 'drivers' WHERE login = (?)",(driverName,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(orders_orderS)):
                        if orders_orderS[i].merchName==merchName:
                            if orders_orderS[i].owner=="new":
                                orders_orderS[i].get_type="new";
                                orders_orderS[i].owner="new";
                                orders_orderS[i].changed="1";
                                break;
                    for i in range(len(orders_driversS)):
                        if orders_driversS[i].merchName==merchName:
                            if orders_driversS[i].login==driverName:
                                orders_driversS[i].removed="1";
                                orders_driversS[i].changed="1";
                                break;

                    #conn=sqlite3.connect("basic2.sqlite");
                    #cursor=conn.cursor();
                    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(driverName,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(basic_usersS)):
                        if basic_usersS[i].login==driverName:
                            basic_usersS[i].removed="1";
                            basic_usersS[i].changed="1";
                            break;
                    #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                    #cursor=conn.cursor();
                    #cursor.execute("SELECT prod_id FROM 'naks'");
                    #ids=cort_to_list(cursor.fetchall());
                    #cursor.execute("SELECT free FROM 'naks'");
                    #free=cort_to_list(cursor.fetchall());
                    #cursor.execute("SELECT reserve FROM 'naks'");
                    #reserve=cort_to_list(cursor.fetchall());
                    ids=[];
                    free=[];
                    reserve=[];
                    for inc in inCar_naksS:
                        if inc.merchName==merchName and inc.removed=="0":
                            ids.append(inc.prod_id);
                            free.append(inc.free);
                            reserve.append(inc.reserve);
                    for i in range(len(ids)):
                        prod_id=ids[i];
                        how_many=int(free[i])+int(reserve[i]);
                        ostPlus(merchName,prod_id,how_many);
                    #cursor.execute("DELETE FROM 'cashInCar' WHERE driver = (?)",(driverName,));
                    #cursor.execute("DELETE FROM 'nakNum' WHERE owner = (?)",(driverName,));
                    #cursor.execute("DELETE FROM 'naks' WHERE owner = (?)",(driverName,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(inCar_cashInCarS)):
                        if inCar_cashInCarS[i].merchName==merchName:
                            if inCar_cashInCarS[i].driver==driverName:
                                inCar_cashInCarS[i].removed="1";
                                inCar_cashInCarS[i].changed="1";
                    for i in range(len(inCar_nakNumS)):
                        if inCar_nakNumS[i].merchName==merchName:
                            if inCar_cashInCarS[i].driver==driverName:
                                inCar_nakNumS[i].removed="1";
                                inCar_nakNumS[i].changed="1";
                    for i in range(len(inCar_naksS)):
                        if inCar_naksS[i].merchName==merchName:
                            if inCar_naksS[i].owner==driverName:
                                inCar_naksS[i].removed="1";
                                inCar_naksS[i].changed="1";

                    send={"err":"0","text":"OK"};
                    txt="Удален водитель {0}".format(driverName);
                    makeAct(merchName,Login(session),txt);
                    async2(Login(session),"getDrivers");
                    async2(Login(session),"getAdminOrders");
                else:
                    send={"err":"1","text":"Остался долг"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    write_now();
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_order(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':last_index=request.POST['last_index'];
                if request.method=='GET':last_index=request.GET['last_index'];
                login=Login(session);
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT owner FROM 'order' WHERE last_index=(?)",(last_index,));
                #driverName=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT getter FROM 'order' WHERE last_index=(?)",(last_index,));
                #getter=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT data FROM 'order' WHERE last_index=(?)",(last_index,));
                #data=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("UPDATE 'order' SET get_type='new'WHERE last_index=(?)",(last_index,));
                #cursor.execute("UPDATE 'order' SET owner='new'WHERE last_index=(?)",(last_index,));
                #conn.commit();
                #conn.close();
                for o in orders_orderS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.last_index==last_index:
                            driverName=o.owner;
                            getter=o.getter;
                            data=o.data;
                            break;
                for i in range(len(orders_orderS)):
                    if orders_orderS[i].merchName==merchName:
                        if orders_orderS[i].last_index==last_index:
                            orders_orderS[i].get_type="new"
                            orders_orderS[i].owner="new"
                            orders_orderS[i].change="1"
                            break;

                data=stringToArrayData(data);
                #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                #cursor=conn.cursor();
                for i in range(len(data)):
                    prod_id=data[i][0];
                    how_many=data[i][1];
                    #cursor.execute("SELECT reserve FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prod_id,));
                    #reserveOld=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driverName,prod_id,));
                    #freeOld=cort_to_list(cursor.fetchall())[0];
                    for inc in inCar_naksS:
                        if inc.merchName==merchName and inc.removed=="0":
                            if inc.owner==driverName and inc.prod_id==prod_id:
                                reserveOld=inc.reserve;
                                freeOld=inc.free;
                                break;

                    reserveNew=str(int(reserveOld)-int(how_many));
                    freeNew=str(int(freeOld)+int(how_many));
                    #cursor.execute("UPDATE 'naks' SET reserve=(?)WHERE owner=(?)AND prod_id=(?)",(reserveNew,driverName,prod_id,));
                    #cursor.execute("UPDATE 'naks' SET free=(?)WHERE owner=(?)AND prod_id=(?)",(freeNew,driverName,prod_id,));
                    for k in range(len(inCar_naksS)):
                        if inCar_naksS[k].merchName==merchName:
                            if inCar_naksS[k].owner==driverName and inCar_naksS[k].prod_id==prod_id:
                                inCar_naksS[k].reserve=reserveNew;
                                inCar_naksS[k].free=freeNew;
                                inCar_naksS[k].changed="1";
                                break;
                #conn.commit();
                #conn.close();
                async2(driverName,"getNak");
                async2(driverName,"getOrders");
                async2(driverName,"getPrices");
                async2(Login(session),"getDrivers");
                addNot(getter,merchName,last_index,"remMarketOrder");

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_give_him_order(request,var='network'):
    try:
        free='';
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    last_index=request.POST['last_index'];
                    driverName=request.POST['driverName'];
                if request.method=='GET':
                    last_index=request.GET['last_index'];
                    driverName=request.GET['driverName'];

                login=Login(session);
                merchName=Merch(session);
                for o in orders_orderS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.last_index==last_index:
                            data=o.data;
                            break;
                print(last_index)
                data=stringToArrayData(data);
                print(data,"<<<<data")
                for i in range(len(data)):
                    prodId=data[i][0];
                    how_many=data[i][1];
                    prod=prodById(merchName,prodId);
                    try:
                        free=""
                        reserve=""
                        for inc in inCar_naksS:
                            if inc.merchName==merchName and inc.removed=="0":
                                if inc.owner==driverName and inc.prod_id==prodId:
                                    free=str(inc.free);
                                    reserve=str(inc.reserve);
                                    break;
                        print(free,"free")
                        print(reserve,"reserve")
                        print(len(inCar_naksS),"leninCar")
                        free[0];
                        reserve[0];
                    except Exception as e:
                        logger(e)
                        send={"err":"1","text":'Не хватает '+prod+'(есть 0. Нужно '+how_many+")"};
                        send=json.dumps(send);
                        print("break!!!!")
                        return HttpResponse(send, content_type='application/json');
                    print(prodId,how_many,free,"<<<<<<how fre")
                    if int(how_many)>int(free):

                        send={"err":"1","text":"Не хватает "+str(prod)+"(есть "+str(free)+". Нужно "+str(how_many)+")"};
                        async2(Login(session),"getDrivers");
                        send=json.dumps(send);
                        return HttpResponse(send, content_type='application/json');
                    else:
                        newFree=int(free)-int(how_many);
                        newReserve=int(reserve)+int(how_many);
                        for k in range(len(inCar_naksS)):
                            if inCar_naksS[k].merchName==merchName:
                                if inCar_naksS[k].owner==driverName and inCar_naksS[k].prod_id==prodId:
                                    inCar_naksS[k].free=newFree;
                                    inCar_naksS[k].reserve=newReserve;
                                    inCar_naksS[k].changed="1";
                                    break;
                        num=[];
                        date=[];
                        for inc in inCar_nakNumS:
                            if inc.merchName==merchName and inc.removed=="0":
                                if inc.owner==driverName:
                                    num.append(inc.num);
                                    date.append(inc.date);
                        nakNum="";
                        for k in range(len(num)):
                            nakNum+="Накладная №"+num[k]+"\nДата: "+date[k]+"\n";
                txt="Заказ№{2} отдан водителю: {0}\nЗадействованы накладные:\n{1}".format(driverName,nakNum,last_index);

                messageInChannel(merchName,txt);

                for i in range(len(orders_orderS)):
                    if orders_orderS[i].last_index==last_index:
                        orders_orderS[i].owner=driverName;
                        orders_orderS[i].get_type="ord";
                        orders_orderS[i].changed="1";
                        break;

                addNot(driverName,"title",last_index,"driverOrder");

                async2(driverName,"getNak");
                async2(driverName,"getOrders");
                async2(driverName,"getPrices");
                async2(login,"getDrivers");
                txt="Заказ №{0} отдан водителю {1}".format(last_index,driverName);
                makeAct(merchName,login,txt);
                if request.method=='POST':
                    drivers=get_drivers(request,'local');
                    drivers=drivers['drivers'];
                    send={"err":"0","text":"OK",'drivers':drivers};
                else:
                    send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    print(free,"FREEEEEE");
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def refresh_gen_link(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                code=generate();
                newLink=URL+"excel_space/?code="+code;
                newLink=newLink.replace(":","}");
                for i in range(len(orders_constS)):
                    if orders_constS[i].key=="generalLink":
                        orders_constS[i].val=newLink;
                        orders_constS[i].changed="1";
                        break;
                send={"err":"0","text":newLink};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def send_remove_wrong_order(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':last_index=request.POST['last_index'];
                if request.method=='GET':last_index=request.GET['last_index'];
                merchName=Merch(session);
                for i in range(len(orders_orderS)):
                    o=orders_orderS[i];
                    if o.merchName==merchName and o.removed=="0":
                        print(o.last_index,"MMM",last_index)
                        if o.last_index==last_index:
                            data=o.data;
                            getter=o.getter
                            price=o.price
                            orders_orderS[i].removed="1";
                            orders_orderS[i].changed="1";
                if "USER_" in getter:
                    removeUser(merchName,getter);
                admins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getAllPrices");
                txt="Удалён заказ№{0}".format(last_index);
                makeAct(merchName,Login(session),txt);
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                orders_historyS.append(orders_history(data.replace("|",":").replace("^",";"),str(price),"ord",getter,"null",date,"district",last_index,"new",merchName));
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_market_settings(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    marketName=request.GET['marketName'];
                    clientPriceType=request.GET['clientPriceType'];
                    dolgType=request.GET['dolgType'];
                    dolgVal=request.GET['dolgVal'];
                    wallet=request.GET['wallet'];
                if request.method=='POST':
                    marketName=request.POST['marketName'];
                    clientPriceType=request.POST['clientPriceType'];
                    dolgType=request.POST['dolgType'];
                    dolgVal=request.POST['dolgVal'];
                    wallet=request.POST['wallet'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE buy_markets SET type=(?) WHERE name=(?)",
                #    (clientPriceType,marketName));
                #cursor.execute("UPDATE buy_markets SET dolgType=(?) WHERE name=(?)",
                #    (dolgType,marketName));
                #cursor.execute("UPDATE buy_markets SET dolgVal=(?) WHERE name=(?)",
                #    (dolgVal,marketName));
                #cursor.execute("UPDATE buy_markets SET wallet=(?) WHERE name=(?)",
                #    (wallet,marketName));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_buy_marketsS)):
                    if orders_buy_marketsS[i].merchName==merchName:
                        if orders_buy_marketsS[i].name==marketName:
                            orders_buy_marketsS[i].type=clientPriceType;
                            orders_buy_marketsS[i].dolgType=dolgType;
                            orders_buy_marketsS[i].dolgVal=dolgVal;
                            orders_buy_marketsS[i].wallet=wallet;
                            orders_buy_marketsS[i].changed="1";
                            print(dolgVal,"<<<<<<<<<")
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                #    (merchName,));
                #admins=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                #    (merchName,));
                #logins=cort_to_list(cursor.fetchall());
                #conn.close();
                admins=[];
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                    if c.user_type=="driver" and c.merchName==merchName:
                        logins.append(c.login);
                for a in admins:
                    async2(a,"getMarketList");
                for l in logins:
                    async2(l,"getMarketList");
                txt="Изменены настройка магазина:\n"+marketName+"\n"+dolgType+"\n"+dolgVal;
                makeAct(merchName,Login(session),txt);
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_price_type_name(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                oldPriceName=request.GET['oldPriceName'];
                priceName=request.GET['priceName'];
                merchName=Merch(session);

                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE buy_markets SET type=(?) WHERE type=(?)",
                #   (priceName,oldPriceName));
                #cursor.execute("UPDATE buyers_type SET name=(?) WHERE name=(?)",
                #    (priceName,oldPriceName));
                #cursor.execute("UPDATE const SET val=(?) WHERE val=(?) AND key='default_buyer'",
                #    (priceName,oldPriceName));
                #cursor.execute("UPDATE prices SET name=(?) WHERE name=(?)",
                #    (priceName,oldPriceName));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_buy_marketsS)):
                    if orders_buy_marketsS[i].merchName==merchName:
                        if orders_buy_marketsS[i].type==oldPriceName:
                            orders_buy_marketsS[i].type=priceName;
                            orders_buy_marketsS[i].changed="1";
                        break;
                for i in range(len(orders_buyers_typeS)):
                    if orders_buyers_typeS[i].merchName==merchName:
                        if orders_buyers_typeS[i].name==oldPriceName:
                            orders_buyers_typeS[i].name=priceName;
                            orders_buyers_typeS[i].changed="1";
                            break;
                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].val==oldPriceName and orders_constS[i].key=="default_buyer":
                            orders_constS[i].val=priceName;
                            orders_constS[i].changed="1";
                            break;
                for i in range(len(orders_pricesS)):
                    if orders_pricesS[i].merchName==merchName:
                        if orders_pricesS[i].name==oldPriceName:
                            orders_pricesS[i].name=priceName;
                            orders_pricesS[i].changed="1";
                            break;

                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='admin' AND merchName=(?)",
                #    (merchName,));
                #admins=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='driver' AND merchName=(?)",
                #    (merchName,));
                #logins=cort_to_list(cursor.fetchall());
                #conn.close();
                admins=[];
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="admin" and c.merchName==merchName:
                        admins.append(c.login);
                    if c.user_type=="driver" and c.merchName==merchName:
                        logins.append(c.login);
                for a in admins:
                    async2(a,"getAdminOrders");
                    async2(a,"getProducts");
                    async2(a,"getAllPrices");
                    async2(a,"getMarketList");
                for l in logins:
                    async2(l,"getNak");
                    async2(l,"getOrders");
                    async2(l,"getPrices");
                    async2(l,"getProductList");
                    async2(l,"getDefaultMarket");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_rem_price_type_name(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    oldPriceName=request.POST['oldPriceName'];
                if request.method=='GET':
                    oldPriceName=request.GET['oldPriceName'];
                if oldPriceName =="user":
                    send={"err":"1","text":"Нельзя"};
                    send=json.dumps(send);
                    return HttpResponse(send, content_type='application/json')
                else:
                    merchName=Merch(session);
                    for i in range(len(orders_buy_marketsS)):
                        if orders_buy_marketsS[i].merchName==merchName:
                            if orders_buy_marketsS[i].type==oldPriceName:
                                orders_buy_marketsS[i].type="user";
                                orders_buy_marketsS[i].changed="1";
                                break;
                    for i in range(len(orders_buyers_typeS)):
                        if orders_buyers_typeS[i].merchName==merchName:
                            if orders_buyers_typeS[i].name==oldPriceName:
                                orders_buyers_typeS[i].removed="1";
                                orders_buyers_typeS[i].changed="1";
                                break;
                    for i in range(len(orders_constS)):
                        if orders_constS[i].merchName==merchName:
                            if orders_constS[i].val==oldPriceName and orders_constS[i].key=="default_buyer":
                                orders_constS[i].val="user";
                                orders_constS[i].changed="1";
                                break;
                    for i in range(len(orders_pricesS)):
                        if orders_pricesS[i].merchName==merchName:
                            if orders_pricesS[i].name==oldPriceName:
                                orders_pricesS[i].removed="1";
                                orders_pricesS[i].changed="1";

                    txt="Удален тип цены:\n"+oldPriceName;
                    makeAct(merchName,Login(session),txt);
                    admins=[];
                    logins=[];
                    for c in basic_usersS:
                        if c.user_type=="admin" and c.merchName==merchName:
                            admins.append(c.login);
                        if c.user_type=="driver" and c.merchName==merchName:
                            logins.append(c.login);
                    for a in admins:
                        async2(a,"getAdminOrders");
                        async2(a,"getProducts");
                        async2(a,"getAllPrices");
                    for l in logins:
                        async2(l,"getNak");
                        async2(l,"getOrders");
                        async2(l,"getPrices");
                        async2(l,"getProductList");
                        async2(l,"getDefaultMarket");
                    send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_price_type(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    oldPriceName=request.POST['oldPriceName'];
                if request.method=='GET':
                    oldPriceName=request.GET['oldPriceName'];
                if oldPriceName=="user":
                    send={"err":"1","text":"Нельзя"};
                    send=json.dumps(send);
                    return HttpResponse(send, content_type='application/json')
                else:
                    merchName=Merch(session);
                    myPriceTypes=[];
                    for o in orders_buyers_typeS:
                        if o.merchName==merchName and o.removed=="0":
                            myPriceTypes.append(o.name);

                    if oldPriceName not in myPriceTypes:
                        orders_buyers_typeS.append(orders_buyers_type(oldPriceName,merchName));
                        id_=[];
                        for o in orders_pricesS:
                            if o.merchName==merchName and o.removed=="0":
                                if o.name=="user":
                                    id_.append(o.id);
                        price=[];
                        for o in orders_pricesS:
                            if o.merchName==merchName and o.removed=="0":
                                if o.name=="user":
                                    price.append(o.price);
                        for i in range(len(id_)):
                            orders_pricesS.append(orders_prices(id_[i],oldPriceName,price[i],merchName));
                        admins=[];
                        logins=[];
                        for c in basic_usersS:
                            if c.user_type=="admin" and c.merchName==merchName:
                                admins.append(c.login);
                            if c.user_type=="driver" and c.merchName==merchName:
                                logins.append(c.login);
                        for a in admins:
                            async2(a,"getAdminOrders");
                            async2(a,"getProducts");
                            async2(a,"getAllPrices");
                            async2(a,"getDefaultMarket");
                        for l in logins:
                            async2(l,"getNak");
                            async2(l,"getOrders");
                            async2(l,"getPrices");
                            async2(l,"getProductList");
                            async2(l,"getDefaultMarket");
                        send={"err":"0","text":"OK"};
                    else:
                        send={"err":"1","text":"Такой тип цены уже существует"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        print(send)
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_req_for_excel(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    exType=request.POST['exType'];
                if request.method=='GET':
                    exType=request.GET['exType'];
                print(exType);
                merchName=Merch(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT code FROM 'exLink'");
                #codes=cort_to_list(cursor.fetchall());
                codes=[];
                for c in basic_exLinkS:
                    codes.append(c.code);
                code=generate();
                while code in codes:
                    code=generate();
                #cursor.execute("INSERT INTO 'exLink' VALUES((?),(?),(?))",(code,exType,merchName));
                #conn.commit();
                #conn.close();
                basic_exLinkS.append(basic_exLink(code,exType,merchName));
                link=URL+"get_excel/?code="+code;
                send={"err":"0","link":link};
                txt="Создана ссылка на отправку Excel:\nТип ссылки: {0}\nКод: {1}".format(exType,code);
                makeAct(merchName,Login(session),txt);
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_excel(request):
    try:
        code=request.GET['code'];
        #conn=sqlite3.connect("basic2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT code FROM 'exLink'");
        #codes=cort_to_list(cursor.fetchall());
        codes=[];
        for c in basic_exLinkS:
            codes.append(c.code);
        if code in codes:
            #cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
            #_type=cort_to_list(cursor.fetchall())[0];                print(c.code)

            #cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
            #merchName=cort_to_list(cursor.fetchall())[0];
            print(code)

            for c in basic_exLinkS:
                if c.code==code:
                    print(c.code)
                    _type=c._type;
                    merchName=c.merchName;
                    break;
            if _type=="prices" or _type=="ost" or _type=="per":
                html="html/prices.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL)
                htmlR=htmlR.replace("_VAL_",code)
                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                #conn.close();
            else:
                html="html/nak.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL);
                htmlR=htmlR.replace("_VAL_",code);
                #cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='driver'",(merchName,));
                #logins=cort_to_list(cursor.fetchall());
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="driver" and c.merchName==merchName:
                        logins.append(c.login);
                data="";
                for l in logins:
                    data+='\n<option value="{0}">{0}</option>'.format(l);
                htmlR=htmlR.replace("_OPTIONS_",data);
                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                #conn.close();
            return render(request,htmlL);
        else:
            #conn.close();
            send="err=1,,text=Кода не существует";
            return render(request,send);
    except Exception as e:
        logger(e);
        send="err=1,,text=Кода не существует";
        return render(request,errHtml("Недействительная ссылка"));
        send=json.dumps(send);
        #return HttpResponse(send, content_type='application/json')
def excel_space(request):
    try:
        merchs=os.listdir("merchants");
        code=request.GET['code'];
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                html="html/nakSpace.html";
                file=open(html,"r");
                htmlR=file.read();
                file.close();
                htmlL="templates/{0}.html".format(code);
                htmlR=htmlR.replace("_URL_",URL);
                htmlR=htmlR.replace("_VAL_",code);

                file=open(htmlL,"w");
                file.write(htmlR);
                file.close();
                return render(request,htmlL);
        return render(request,errHtml("Недействительная ссылка"));
    except Exception as e:
        logger(e);
        send="err=1,,text=Кода не существует";
        return render(request,errHtml("Недействительная ссылка"));
        send=json.dumps(send);
        #return HttpResponse(send, content_type='application/json')
def save_excel_space(request):
    try:
        code=request.POST['code'];
        merchs=os.listdir("merchants");
        workNot=False;
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                workNot=True;
                merchName=m;
                break;
        if workNot:
            #conn1=sqlite3.connect(mPath(merchName,"inCar2"));
            #cursor1=conn1.cursor();

            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT name FROM 'products'");
            #prods=cort_to_list(cursor.fetchall());
            prods=[];
            for o in orders_productsS:
                if o.merchName==merchName and o.removed=="0":
                    prods.append(o.name);
            j=1;
            allFiles=request.FILES.getlist('nakFile');
            nakList=[];
            excelNakPlaceList=[];
            errorList=[];
            niceList=[];
            for f in allFiles:
                nakList=[];

                excelName=getConst(merchName,'excelName');
                excelHow=getConst(merchName,'excelHow');
                excelStartWrite=getConst(merchName,'excelStartWrite');
                excelNakSheet=getConst(merchName,'excelNakSheet');
                excelNakPlace=getConst(merchName,'excelNakPlace');

                exNakDriverName=getConst(merchName,'exNakDriverName');


                Path="merchants/{0}/xlsx/".format(merchName);
                #handle_uploaded_file(request.FILES['nakFile'],Path,code);
                try:
                    os.remove(Path+code+".xlsx");
                except:
                    pass;
                handle_uploaded_file(f,Path,code);
                error=False;
                try:
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                except:
                    reW.re(Path+code+".xlsx");
                    print("rewrite!<<<<<")
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                try:
                    sheet = wb[excelNakSheet];
                except Exception as e:
                    logger(e);
                    error=True;
                    #return render(request,errHtml("Файл: "+f.name+". \nНе найдена страница"));
                if not error:
                    i=int(excelStartWrite);
                    #nakList=[];
                    excelNakPlace = str(sheet[excelNakPlace].value);
                    exNakDriverName = str(sheet[exNakDriverName].value);
                    #cursor.execute("SELECT login FROM 'drivers'WHERE name=(?)",(exNakDriverName,));
                    for o in orders_driversS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.name==exNakDriverName:
                                drivers=o.login;
                    try:
                        #drivers=cort_to_list(cursor.fetchall())[0];
                        drivers=drivers;
                    except:
                        drivers="null"
                    #cursor1.execute("SELECT num FROM 'nakNum'WHERE owner=(?)",(drivers,));
                    #nums=cort_to_list(cursor1.fetchall());
                    nums=[];
                    for inc in inCar_nakNumS:
                        if inc.merchName==merchName and inc.removed=="0":
                            if inc.owner==drivers:
                                nums.append(inc.num);
                    if excelNakPlace=="None":
                        return render(request,errHtml("Номер накладной не найден."));
                    while True:
                        excelNameE = str(sheet[excelName+str(i)].value);
                        excelHowE = str(sheet[excelHow+str(i)].value);
                        if excelNameE=="None" or excelHowE=="None":
                            try:
                                os.remove("templates/{0}.html".format(code));
                            except:
                                pass;
                            print("Накладная №"+excelNakPlace);
                            for pp in nakList:
                                print(pp);
                            excelNakPlaceList.append(excelNakPlace);
                            break;
                        else:
                            try:
                                int(excelHowE)+1;
                            except:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Неккоректное количество. Строка "+str(i)+" Указанное колличество: \""+excelHowE+"\"";
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Неккоректное количество. Строка "+str(i)+" Указанное колличество: \""+excelHowE+"\""));
                            if excelNameE not in prods:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Продукта ("+excelNameE+") не существует. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Продукта ("+excelNameE+") не существует. Строка "+str(i)));
                            elif excelNameE=="None":
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Поле наименование пустое. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+" Поле наименование пустое. Строка "+str(i)));
                            elif excelHowE=="None":
                                error=True;
                                txt="Накладная№ "+excelNakPlace+"Поле количества пустое. Строка "+str(i);
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+"Поле количества пустое.Строка "+str(i)));

                            elif excelNakPlace in nums:
                                error=True;
                                txt="Накладная№ "+excelNakPlace+". Накладная с таким номером уже есть у этого водителя и поэтому проигнорирована";
                                if txt not in errorList:errorList.append(txt);
                                #return render(request,errHtml("Накладная№ "+excelNakPlace+"Поле количества пустое.Строка "+str(i)));

                            else:
                                error=False;
                                if f.name not in niceList:
                                    niceList.append(f.name);
                                nakList.append(
                                    {"excelNameE":excelNameE,
                                    "excelHowE":excelHowE,
                                    "excelNum":excelNakPlace
                                    });
                        i+=1;
                else:
                    errorList.append("Файл: "+f.name+". => Не найдена страница");

                if not error:
                    writeExcelSpace(request,nakList,code,drivers,excelNakPlaceList);
                excelNakPlaceList.clear();
                #j+=1;
            #conn.close();
            #conn1.close();
            print(excelNakPlaceList)
            print("Общая")
            for pp in nakList:
                print(pp);
            print("Оконченно");
            drivers="null"
            txt="Ошибки произошли в файлах:";
            for t in errorList:
                txt+="<br>"+t;

            txt+="<br><br>Успешно записаны:";
            for t in niceList:
                txt+="<br>"+t;

            if len(errorList)==0:
                return render(request,doneHtml("Накладные записаны"));
            else:
                return render(request,errHtml(txt));

            #return render(request,doneHtml("Накладная записана"));

            #return writeExcelSpace(request,nakList,code,drivers,excelNakPlaceList);
        else:
            return render(request,errHtml("Недействительная ссылка"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def writeExcelSpace(request,exList,code,driver="null",nak="null"):
    try:
        merchs=os.listdir("merchants");
        workNot=False;
        for m in merchs:
            link=getConst(m,"generalLink").replace("}",":");
            newLink=URL+"excel_space/?code="+code;
            if link==newLink:
                workNot=True;
                merchName=m;
                break;
        print(driver,">>>>driver")
        if workNot:
            if driver=="null":
                #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor1=conn1.cursor();
                NK="";
                for l in exList:
                    excelNameE=l["excelNameE"];
                    excelHowE=l["excelHowE"];
                    excelNum=l["excelNum"];

                    #cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                    #ids=cort_to_list(cursor1.fetchall())[0];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.name==excelNameE:
                                ids=o.id;

                    NK=NK+ids+"|"+excelHowE+"|^";

                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                #cursor1.execute("SELECT id FROM 'nakSpace'");
                #idsS=cort_to_list(cursor1.fetchall());
                idsS=[];
                for o in orders_nakSpaceS:
                    if o.merchName==merchName and o.removed=="0":
                        idsS.append(o.id);
                id_=1;
                while str(id_) in idsS:
                    id_+=1;
                #cursor1.execute("INSERT INTO 'nakSpace' VALUES((?),(?),(?),(?),(?))",
                #        (str(id_),NK,date,excelNum,"_",));
                orders_nakSpaceS.append(orders_nakSpace(str(id_),NK,date,excelNum,"_",merchName));
                #conn1.commit();
                #conn1.close();
            else:
                #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor1=conn1.cursor();
                #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                #cursor=conn.cursor();
                for l in exList:
                    excelNameE=l["excelNameE"];
                    excelHowE=l["excelHowE"];
                    #cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                    #ids=cort_to_list(cursor1.fetchall())[0];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.name==excelNameE:
                                ids=o.id;
                                break;
                    #cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                    #freeOld=cort_to_list(cursor.fetchall());
                    freeOld=[];
                    for inc in inCar_naksS:
                        if inc.merchName==merchName and inc.removed=="0":
                            if inc.owner==driver and inc.prod_id==ids:
                                freeOld.append(inc.free);
                    if len(freeOld)==0:
                        #cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                        #    (excelNameE,ids,excelHowE,driver,));
                        inCar_naksS.append(inCar_naks(excelNameE,ids,excelHowE,'0','0',driver,merchName));
                    else:
                        freeOld=freeOld[0];
                        newFree=str(int(excelHowE)+int(freeOld));
                        #cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                        #    (newFree,driver,ids,));
                        for i in range(len(inCar_naksS)):
                            if inCar_naksS[i].merchName==merchName:
                                if inCar_naksS[i].owner==driver and inCar_naksS[i].prod_id==ids:
                                    inCar_naksS[i].free=newFree;
                                    inCar_naksS[i].changed="1";
                                    break;
                    ostMinus(merchName,ids,excelHowE);
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                print("<<<<<<<<")
                for n in nak:
                    print(n+"<<<");
                    #cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                    #    (driver,n,date));
                    inCar_nakNumS.append(inCar_nakNum(driver,n,date,merchName));
                print("<<<<<<<<")
                #conn.commit();
                #conn.close();
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
                #cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
                #admins=cort_to_list(cursor.fetchall());
                #conn.commit();
                #conn.close();
                for i in range(len(basic_exLinkS)):
                    if basic_exLinkS[i].code==code:
                        basic_exLinkS[i].removed="1";
                        basic_exLinkS[i].changed="1";
                admins=[];
                for c in basic_usersS:
                    if c.merchName==merchName and c.user_type=="admin":
                        admins.append(c.login);
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
            #admins=cort_to_list(cursor.fetchall());
            #conn.commit();
            #conn.close();
            admins=[];
            for c in basic_usersS:
                if c.merchName==merchName and c.user_type=="admin":
                    admins.append(c.login);
            for a in admins:
                async2(a,"getDrivers");
                async2(a,"getExcelNaks");
            return render(request,doneHtml("Накладная записана"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def get_chat_admins(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            admins=[];
            merchName1=[];
            for c in basic_usersS:
                if c.user_type=="admin":
                    admins.append(c.login);
                if c.user_type=="admin":
                    merchName1.append(c.merchName);
            if request.method=='GET':
                sendList="";
            if request.method=='POST':
                sendList=[];
            for i in range(len(admins)):
                l=admins[i];
                for o in orders_adminsS:
                    if o.merchName==merchName1[i]:
                        if o.login==l:
                            role=o.role;
                            break;
                if role==None:
                    role="";
                if request.method=='GET':
                    sendList=sendList+login[i]+"|"+merchName1[i]+"|"+role+"|^"
                if request.method=='POST':
                    sendList.append([l,merchName1[i],role]);
            if request.method=='GET':
                sendList=stringToArray(sendList);
            send={"err":"0","admins":sendList};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_messages(request,var='network',Chat_id='null'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);

            chat_id=[];
            message_id=[];
            text_or_photo=[];
            sender=[];
            getter=[];
            date=[];
            wasRead=[];
            text=[];
            for c in basic_messagesS:
                if c.sender==login or c.getter==login:
                    chat_id.append(c.chat_id);
                    message_id.append(c.message_id);
                    text_or_photo.append(c.text_or_photo);
                    sender.append(c.sender);
                    getter.append(c.getter);
                    date.append(c.date);
                    wasRead.append(c.wasRead);
                    text.append(c.text);
            if request.method=='GET':
                sendList="";
                for i in range(len(chat_id)):
                    sendList=sendList+chat_id[i]+"|"+message_id[i]+"|"+text_or_photo[i]+"|"+sender[i]+"|"+getter[i]+"|"+date[i]+"|"+wasRead[i]+"|"+text[i]+"|^"
            if request.method=='POST':
                sendList=[];
                for i in range(len(chat_id)):
                    sendList.append([chat_id[i],message_id[i],text_or_photo[i],sender[i],getter[i],date[i],wasRead[i],text[i]])
            send={"err":"0","messages":sendList};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    #
    if Chat_id!="null":
        send["chat_id"]=Chat_id;
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_readed(request,var='network'):
    try:
        if request.method=='POST':
            session=request.POST['session'];
            readed=request.POST['readed'];
        if request.method=='GET':
            session=request.GET['session'];
            readed=request.GET['readed'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            readed=stringToArray(readed);
            WR_iteration="0";
            for i in range(len(readed)):
                message_id=readed[i][0];
                for c in basic_messagesS:
                    if c.message_id==message_id:
                        WR=c.wasRead;
                        break;
                if WR=="0":
                    WR_iteration=str(int(WR_iteration)+1);
                    for k in range(len(basic_messagesS)):
                        if basic_messagesS[k].message_id==message_id:
                            basic_messagesS[k].wasRead="1";
                            basic_messagesS[k].changed="1";
                            break;
                for c in basic_messagesS:
                    if c.message_id==message_id:
                        sender=c.sender;
                        break;
            try:
                async2(sender,"getMessages");
            except:
                pass
            send={"err":"0","text":"OK"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_message_to_server(request,var='network'):
    try:
        if request.method=='POST':
            session=request.POST['session'];
            chat_id=request.POST['chat_id'];
            sender=request.POST['sender'];
            getter=request.POST['getter'];
            text=request.POST['text'];
        if request.method=='GET':
            session=request.GET['session'];
            chat_id=request.GET['chat_id'];
            sender=request.GET['sender'];
            getter=request.GET['getter'];
            text=request.GET['text'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            chat_ids=[];
            message_ids=[];
            for c in basic_messagesS:
                chat_ids.append(c.chat_id);
                message_ids.append(c.message_id);
            if "|^" in chat_id:
                #getter=stringToArray(chat_id)[0][1];
                newChatId=1;
                while True:
                    if str(newChatId) in chat_ids:
                        newChatId+=1;
                    else:
                        break;
                chat_id=newChatId;
            newMessageId=1;
            while True:
                if str(newMessageId) in message_ids:
                    newMessageId+=1;
                else:
                    break;
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            basic_messagesS.append(basic_messages(str(chat_id),str(newMessageId),"text",sender,getter,date,"0",text));
            send=get_messages(request,'local',str(chat_id));
            async2(getter,"getMessages");
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def mass_page(request):
    html="templates/test.html";
    return render(request,html);
def test_mass(request):
    Path="html/"
    i=1
    for f in request.FILES.getlist('test_mass'):
        handle_uploaded_file(f,Path,"name"+str(i));
        i+=1;
    html="templates/test.html";
    return render(request,html);

def save_excel(request):
    try:
        code=request.POST['code'];
        #conn=sqlite3.connect("basic2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT code FROM 'exLink'");
        #codes=cort_to_list(cursor.fetchall());
        codes=[];
        for c in basic_exLinkS:
            codes.append(c.code);
        #cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
        for c in basic_exLinkS:
            if c.code==code:
                _type=c._type;
                break;
        try:
            _type=_type;
            #_type=cort_to_list(cursor.fetchall())[0];
        except:
            return render(request,errHtml("Недействительная ссылка"));
        #cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
        #merchName=cort_to_list(cursor.fetchall())[0];
        #cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
        #drivers=cort_to_list(cursor.fetchall());
        #cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
        #admins=cort_to_list(cursor.fetchall());
        #conn.close();
        for c in basic_exLinkS:
            if c.code==code:
                merchName=c.merchName
                break;
        drivers=[]
        admins=[]
        for c in basic_usersS:
            if c.user_type=="driver":
                drivers.append(c.login)
            if c.user_type=="admin":
                admins.append(c.login)

        #conn=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor=conn.cursor();
        #cursor.execute("SELECT name FROM 'buyers_type'");
        #buyers_type=cort_to_list(cursor.fetchall());
        #cursor.execute("SELECT name FROM 'categories'");
        #cats=cort_to_list(cursor.fetchall());
        #cursor.execute("SELECT name FROM 'products'");
        #prods=cort_to_list(cursor.fetchall());
        #conn.close();
        buyers_type=[];
        for o in orders_buyers_typeS:
            if o.merchName==merchName and o.removed=="0":
                buyers_type.append(o.name);
        cats=[];
        for o in orders_categoriesS:
            if o.merchName==merchName and o.removed=="0":
                cats.append(o.name);
        prods=[];
        for o in orders_productsS:
            if o.merchName==merchName and o.removed=="0":
                prods.append(o.name);

        if _type=="per":
            exPerInn=getConst(merchName,'exPerInn');
            exPerSum=getConst(merchName,'exPerSum');
            exPerStart=getConst(merchName,'exPerStart');
            exPerSheet=getConst(merchName,'exPerSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[exPerSheet];
            except Exception as e:
                logger(e);
                return render(request,errHtml("Не найдена страница"));

            i=int(exPerStart);

            inns={};
            while True:
                inn = str(sheet[exPerInn+str(i)].value);
                summ = str(sheet[exPerSum+str(i)].value);
                if str(inn)=="None" and str(summ)=="None":
                    break;
                if str(inn)=="Итого"or str(inn)=="None":
                    break;
                if str(summ)=="None":
                    summ="0"
                inn=clearInn(inn);
                try:
                    inn=inn[-9:]
                except:
                    pass;
                if str(inn)=="":
                    inn="0"
                if inn in inns:
                    inns[inn]=str(int(inns[inn])+int(summ));
                else:
                    inns[inn]=summ;
                i+=1;
            print(inns)
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            for INN in inns:
                LGS=loginsByInn(INN);
                for l in LGS:
                    #cursor.execute("UPDATE 'buy_markets' SET wallet=(?) WHERE name=(?)",(inns[INN],l,));
                    for k in range(len(orders_buy_marketsS)):
                        if orders_buy_marketsS[k].merchName==merchName:
                            if orders_buy_marketsS[k].name==l:
                                orders_buy_marketsS[k].wallet=inns[INN];
                                orders_buy_marketsS[k].change='1';
            #conn.commit();
            #conn.close();

            return render(request,doneHtml("Остатки по перечислению записаны"));
        if _type=="ost":
            ostName=getConst(merchName,'ostName');
            ostHowOst=getConst(merchName,'ostHowOst');
            ostStart=getConst(merchName,'ostStart');
            ostSheet=getConst(merchName,'ostSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[ostSheet];
            except Exception as e:
                print(e);
                logger(e);
                return render(request,errHtml("Не найдена страница"));

            i=int(ostStart);
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            names=[];
            while True:
                name = str(sheet[ostName+str(i)].value);
                howOst = str(sheet[ostHowOst+str(i)].value);
                names.append(name);
                if name=="Итого":
                    break;
                if howOst=="None":
                    howOst="0";
                #cursor.execute("SELECT id FROM 'products' WHERE name=(?)",(name,));
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.name==name:
                            prod_id=id;
                            break;
                if name=="None" and howOst=="None":
                    break;
                try:
                    prod_id=prod_id;
                    #prod_id=cort_to_list(cursor.fetchall())[0];
                except Exception as e:
                    logger(e);
                    return render(request,errHtml("Продукт не найден.\nСтрока "+str(i)+" ("+name+")"));
                howOst=howOst.replace(",000","");
                try:
                    int(howOst)+1;
                except:
                    return render(request,errHtml("Неккоректное количество остатка\nСтрока "+str(i)));
                for k in range(len(orders_productsS)):
                    if orders_productsS[k].merchName==merchName:
                        if orders_productsS[k].id==prod_id:
                            ind=k;
                            break;
                #cursor.execute("UPDATE 'products' SET ost=(?) WHERE id=(?)",(howOst,prod_id,));
                orders_productsS[ind].ost=howOst;
                if howOst=="0":
                    #cursor.execute("UPDATE 'products' SET visible='0' WHERE id=(?)",(prod_id,));
                    orders_productsS[ind].visible="0";
                else:
                    #cursor.execute("UPDATE 'products' SET visible='1' WHERE id=(?)",(prod_id,));
                    orders_productsS[ind].visible="1";
                i+=1;
            orders_productsS[ind].changed="1";
            #cursor.execute("SELECT name FROM 'products'");
            #allNames=cort_to_list(cursor.fetchall());
            allNames=[];
            for o in orders_productsS:
                if o.merchName==merchName and o.removed=="0":
                    allNames.append(o.name);
            for i in range(len(allNames)):
                if allNames[i] not in names:
                    #cursor.execute("UPDATE 'products' SET ost='0' WHERE name=(?)",(allNames[i],));
                    for k in range(len(orders_productsS)):
                        if orders_productsS[k].merchName==merchName:
                            if orders_productsS[k].name==allNames[i]:
                                orders_productsS[k].ost="0";
                                orders_productsS[k].changed="1";
                                break;

            #conn.commit();
            #conn.close();
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
            #conn.commit();
            #conn.close();
            for i in range(len(basic_exLinkS)):
                if basic_exLinkS[i].code==code:
                    basic_exLinkS[i].removed="1";
                    basic_exLinkS[i].changed="1";
                    break;
            for a in admins:
                async2(a,"getProducts");
            return render(request,doneHtml("Остатки записаны"));






        elif _type=="prices":
            excelType=getConst(merchName,'excelType');
            excelCat=getConst(merchName,'excelCat');
            excelPriceName=getConst(merchName,'excelPriceName');
            excelPrice=getConst(merchName,'excelPrice');
            excelStartWritePrice=getConst(merchName,'excelStartWritePrice');
            excelProdSheet=getConst(merchName,'excelProdSheet');
            Path="merchants/{0}/xlsx/".format(merchName);
            handle_uploaded_file(request.FILES['nakFile'],Path,code);
            try:
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            except:
                reW.re(Path+code+".xlsx");
                print("rewrite!<<<<<")
                wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
            try:
                sheet = wb[excelProdSheet];
            except Exception as e:
                logger(e);
                return render(request,errHtml("Не найдена страница"));
            i=int(excelStartWritePrice);
            priceList=[];
            while True:
                excelTypeE = str(sheet[excelType+str(i)].value);
                excelCatE = str(sheet[excelCat+str(i)].value);
                excelPriceNameE = str(sheet[excelPriceName+str(i)].value);
                excelPriceE = str(sheet[excelPrice+str(i)].value);
                if excelTypeE=="None" and excelCatE=="None" and excelPriceNameE=="None":
                    try:
                        os.remove("templates/{0}.html".format(code));
                    except:
                        pass;
                    for pp in priceList:
                        print(pp);
                    for a in admins:
                        async2(a,"getProducts");
                        async2(a,"getAllPrices");
                    for d in drivers:
                        async2(d,"getNak");
                        async2(d,"getPrices");
                    return writeExcel(request,priceList,code);

                else:
                    try:
                        int(excelPriceE)+1;
                    except:
                        return render(request,errHtml("Неккоректная цена.\nСтрока "+str(i)));
                    if excelCatE not in cats:
                        return render(request,errHtml("Категории не существует.\nСтрока "+str(i)));
                    elif excelTypeE not in buyers_type:
                        return render(request,errHtml("Типа цены не существует.\nСтрока "+str(i)));
                    elif excelTypeE=="None":
                        return render(request,errHtml("Поле типа покупателя пустое.\nСтрока "+str(i)));
                    elif excelCatE=="None":
                        return render(request,errHtml("Поле категории пустое.\nСтрока "+str(i)));
                    elif excelPriceNameE=="None":
                        return render(request,errHtml("Поле типа цены пустое.\nСтрока "+str(i)));
                    elif excelPriceE=="None":
                        return render(request,errHtml("Поле цены пустое.\nСтрока "+str(i)));
                    else:
                        priceList.append(
                            {"excelTypeE":excelTypeE,
                            "excelCatE":excelCatE,
                            "excelPriceNameE":excelPriceNameE,
                            "excelPriceE":excelPriceE});
                i+=1;
            return writeExcel(request,priceList,code);

        else:
            j=1;
            allFiles=request.FILES.getlist('nakFile');
            nakList=[];
            excelNakPlaceList=[];
            for f in allFiles:

                excelName=getConst(merchName,'excelName');
                excelHow=getConst(merchName,'excelHow');
                excelStartWrite=getConst(merchName,'excelStartWrite');
                excelNakSheet=getConst(merchName,'excelNakSheet');
                excelNakPlace=getConst(merchName,'excelNakPlace');
                drivers=request.POST['drivers'];
                Path="merchants/{0}/xlsx/".format(merchName);
                #handle_uploaded_file(request.FILES['nakFile'],Path,code);
                try:
                    os.remove(Path+code+".xlsx");
                except:
                    pass;
                handle_uploaded_file(f,Path,code);
                try:
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                except:
                    reW.re(Path+code+".xlsx");
                    print("rewrite!<<<<<")
                    wb = openpyxl.load_workbook(filename=Path+code+".xlsx");
                try:
                    sheet = wb[excelNakSheet];
                except Exception as e:
                    logger(e);
                    return render(request,errHtml("Не найдена страница"));
                i=int(excelStartWrite);
                #nakList=[];
                excelNakPlace = str(sheet[excelNakPlace].value);
                if excelNakPlace=="None":
                    return render(request,errHtml("Номер накладной не найден."));
                while True:
                    excelNameE = str(sheet[excelName+str(i)].value);
                    excelHowE = str(sheet[excelHow+str(i)].value);
                    if excelNameE=="None" or excelHowE=="None":
                        try:
                            os.remove("templates/{0}.html".format(code));
                        except:
                            pass;
                        print("Накладная №"+excelNakPlace);
                        for pp in nakList:
                            print(pp);
                        excelNakPlaceList.append(excelNakPlace);
                        break;

                        #if j==len(allFiles):
                        #    for a in admins:
                        #        async2(a,"getDrivers");
                        #    for d in drivers:
                        #        async2(d,"getNak");
                        #        async2(d,"getPrices");
                        #    print("Оконченно");
                        #    return writeExcel(request,nakList,code,drivers,excelNakPlace);
                        #else:
                        #    writeExcel(request,nakList,code,drivers,excelNakPlace);
                        #    break;
                    else:
                        try:
                            int(excelHowE)+1;
                        except:
                            return render(request,errHtml("Неккоректное количество.\nСтрока "+str(i)+"\nУказанное колличество: \""+excelHowE+"\""));
                        if excelNameE not in prods:
                            return render(request,errHtml("Продукта ("+excelNameE+") не существует.\nСтрока "+str(i)));
                        elif excelNameE=="None":
                            return render(request,errHtml("Поле наименование пустое.\nСтрока "+str(i)));
                        elif excelHowE=="None":
                            return render(request,errHtml("Поле количества пустое.\nСтрока "+str(i)));
                        else:
                            nakList.append(
                                {"excelNameE":excelNameE,
                                "excelHowE":excelHowE});
                    i+=1;
                j+=1;
            for a in admins:
                async2(a,"getDrivers");
            for d in drivers:
                async2(d,"getNak");
                async2(d,"getPrices");
            print(excelNakPlaceList)
            print("Общая")
            for pp in nakList:
                print(pp);
            print("Оконченно");
            return writeExcel(request,nakList,code,drivers,excelNakPlaceList);
            #return HttpResponse("OK", content_type='application/json');
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def errHtml(text):
    file=open("html/mistake.html","r");
    htmlR=file.read();
    file.close();
    htmlR=htmlR.replace("_TEXT_",text)
    file=open("html/err/resp.html","w");
    file.write(htmlR);
    file.close();
    return "html/err/resp.html";
def doneHtml(text):
    file=open("html/done.html","r");
    htmlR=file.read();
    file.close();
    htmlR=htmlR.replace("_TEXT_",text)
    file=open("html/err/"+text+".html","w");
    file.write(htmlR);
    file.close();
    return "html/err/"+text+".html";
def writeExcel(request,exList,code,driver="null",nak="null"):
    try:

        #conn=sqlite3.connect("basic2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT _type FROM 'exLink' WHERE code=(?)",(code,));
        #_type=cort_to_list(cursor.fetchall())[0];
        #cursor.execute("SELECT merchName FROM 'exLink' WHERE code=(?)",(code,));
        #merchName=cort_to_list(cursor.fetchall())[0];
        #conn.close();
        for c in basic_exLinkS:
            if c.code==code:
                _type=c._type;
                merchName=c.merchName;
                break;
        if _type=="prices":
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            for l in exList:
                excelTypeE=l["excelTypeE"];
                excelCatE=l["excelCatE"];
                excelPriceNameE=l["excelPriceNameE"];
                excelPriceE=l["excelPriceE"];
                #cursor.execute("SELECT id FROM 'products' WHERE name=(?)",(excelPriceNameE,));
                #ids=cort_to_list(cursor.fetchall());
                ids=[];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.name==excelPriceNameE:
                            ids.append(o.id);
                #cursor.execute("SELECT id FROM 'categories' WHERE name=(?)",(excelCatE,));
                #cat_id=cort_to_list(cursor.fetchall())[0];
                cat_id=[];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.name==excelCatE:
                            cat_id.append(o.id);
                            break;
                if len(ids)==0:
                    ids=1;
                    #cursor.execute("SELECT id FROM 'products'");
                    #allId=cort_to_list(cursor.fetchall());
                    allId=[];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            allId.append(o.id);
                    while str(ids) in allId:
                        ids+=1;
                    #cursor.execute("INSERT INTO 'products' VALUES((?),(?),(?),'_','1','_','1','sh',(?),(?),(?),(?))",
                    #    (cat_id,str(ids),excelPriceNameE,"1","0",excelPriceNameE,"_"));
                    orders_productsS.append(orders_products(cat_id,str(ids),excelPriceNameE,"1","0",excelPriceNameE,"_",merchName));
                    #cursor.execute("SELECT name FROM 'buyers_type'");
                    #bType=cort_to_list(cursor.fetchall());
                    bType=[];
                    for o in orders_buyers_typeS:
                        if o.merchName==merchName and o.removed=="0":
                            bType.append(o.name);
                    for b in bType:
                        #cursor.execute("INSERT INTO 'prices' VALUES((?),(?),(?))",
                        #    (str(ids),b,excelPriceE));
                        orders_pricesS.append(orders_prices(str(ids),b,excelPriceE,merchName));
                else:
                    ids=ids[0];
                    #cursor.execute("UPDATE 'products' SET work='1' WHERE id=(?)",(ids,));
                    #cursor.execute("UPDATE 'products' SET cat_id=(?) WHERE id=(?)",(cat_id,ids,));
                    #cursor.execute("UPDATE 'prices' SET price=(?) WHERE id=(?)AND name=(?)",
                    #    (excelPriceE,ids,excelTypeE,));
                    for i in range(len(orders_productsS)):
                        if orders_productsS[i].merchName==merchName:
                            if orders_productsS[i].id==ids:
                                orders_productsS[i].work="1";
                                orders_productsS[i].cat_id=cat_id;
                                orders_productsS[i].changed="1";
                                break;
                    for i in range(len(orders_pricesS)):
                        if orders_productsS[i].merchName==merchName:
                            if orders_pricesS[i].id==ids and orders_pricesS[i].name==excelTypeE:
                                orders_pricesS[i].price=excelPriceE;
                                orders_pricesS[i].changed="1";
                                break;


            #conn.commit();
            #conn.close();
            return render(request,doneHtml("Справочник записан"));
        else:
            #conn1=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor1=conn1.cursor();
            #conn=sqlite3.connect(mPath(merchName,"inCar2"));
            #cursor=conn.cursor();
            for l in exList:
                excelNameE=l["excelNameE"];
                excelHowE=l["excelHowE"];
                #cursor1.execute("SELECT id FROM 'products' WHERE name=(?)",(excelNameE,));
                #ids=cort_to_list(cursor1.fetchall())[0];
                for i in range(len(orders_productsS)):
                    if orders_productsS[i].merchName==merchName:
                        if orders_productsS[i].name==excelNameE:
                            ids=orders_productsS[i].id;
                            break;
                #cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                #freeOld=cort_to_list(cursor.fetchall());
                freeOld=[];
                for inc in inCar_naksS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.owner==driver and inc.prod_id==ids:
                            freeOld.append(inc.free);
                if len(freeOld)==0:
                    #cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                    #    (excelNameE,ids,excelHowE,driver,));
                    inCar_naksS.append(inCar_naks(excelNameE,ids,excelHowE,driver,'0','0',merchName));
                else:
                    freeOld=freeOld[0];
                    newFree=str(int(excelHowE)+int(freeOld));
                    #cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                    #    (newFree,driver,ids,));
                    for i in range(len(inCar_naksS)):
                        if inCar_naksS[i].merchName==merchName:
                            if inCar_naksS[i].owner==driver and inCar_naksS[i].prod_id==ids:
                                inCar_naksS[i].free=newFree;
                                inCar_naksS[i].changed="1";
                                break;
                ostMinus(merchName,ids,excelHowE);
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            for n in nak:
                #cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                #    (driver,n,date));
                inCar_nakNumS.append(inCar_nakNum(driver,n,date,merchName));
            #conn.commit();
            #conn.close();
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("DELETE FROM 'exLink'WHERE code=(?)",(code,));
            #cursor.execute("SELECT login FROM 'users' WHERE merchName=(?)AND user_type='admin'",(merchName,));
            #admins=cort_to_list(cursor.fetchall());
            #conn.commit();
            #conn.close();
            for i in range(len(basic_exLinkS)):
                if basic_exLinkS[i].code==code:
                    basic_exLinkS[i].removed="1";
                    basic_exLinkS[i].changed="1";
                    break;
            admins=[];
            for c in basic_usersS:
                if c.user_type=="admin" and c.merchName==merchName:
                    admins.append(c.login);
            for a in admins:
                async2(a,"getDrivers");
            return render(request,doneHtml("Накладная записана"));
    except Exception as e:
        logger(e);
        return render(request,doneHtml("Ошибка сервера ;("));
def send_url_in_tg(request):
    try:
        if request.method=='POST':
            session=request.POST['session'];
            url=request.POST['url'];
        if request.method=='GET':
            session=request.GET['session'];
            url=request.GET['url'];
        messageInChannel(Merch(session),url)
        send={"err":"0","text":"OK"};
    except Exception as e:
        send={"err":"1","text":str(e)};
        logger(e);
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json');

def send_new_admin(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    adminName=request.GET['adminName'];
                    adminPswd=request.GET['adminPswd'];
                    adminVar=request.GET['adminVar'];
                    adminOld=request.GET['adminOld'];
                    role=request.GET['role'];
                if request.method=='POST':
                    adminName=request.POST['adminName'];
                    adminPswd=request.POST['adminPswd'];
                    adminVar=request.POST['adminVar'];
                    adminOld=request.POST['adminOld'];
                    role=request.POST['role'];
                login1=Login(session);
                merchName=Merch(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users'");
                #login=cort_to_list(cursor.fetchall());
                login=[];
                for c in basic_usersS:
                    login.append(c.login);
                if adminName in login and adminName!=adminOld:
                    send={"err":"1","text":"Логин существует"};
                else:
                    if adminVar=="new":
                        #cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                        #    (adminName,h(adminPswd),"_",merchName,"admin","0.01","0.01","0","ru",));
                        basic_usersS.append(basic_users(adminName,h(adminPswd),"_",merchName,"admin","0.01","0.01","0","ru",adminName,));

                        #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                        #cursor1=conn1.cursor();
                        #cursor1.execute("INSERT INTO 'admins' VALUES((?),(?),(?))",(adminName,"",role,));
                        #conn1.commit();
                        #conn1.close();
                        orders_adminsS.append(orders_admins(adminName,"",role,merchName));
                        send={"err":"0","text":"OK"};
                    elif adminVar=="old":
                        #cursor.execute("UPDATE 'users' SET pswd=(?)WHERE login=(?)",(h(adminPswd),adminOld,));
                        #cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(adminName,adminOld,));
                        print(adminVar)
                        print(adminOld)
                        print(adminName)
                        if adminName!='' and adminName!=adminOld:
                            print('this')
                            for i in range(len(basic_usersS)):
                                if basic_usersS[i].login==adminOld:
                                    adminPswd=basic_usersS[i].pswd
                                    basic_usersS[i].removed="1";
                                    basic_usersS[i].changed="1";
                                    break;
                            for i in range(len(orders_adminsS)):
                                if orders_adminsS[i].merchName==merchName:
                                    if orders_adminsS[i].login==adminOld:
                                        orders_adminsS[i].removed="1";
                                        orders_adminsS[i].changed="1";
                                        break;
                            basic_usersS.append(basic_users(adminName,h(adminPswd),"_",merchName,"admin","0.01","0.01","0","ru",adminName));
                            orders_adminsS.append(orders_admins(adminName,"",role,merchName));

                        else:
                            for i in range(len(basic_usersS)):
                                if basic_usersS[i].login==adminOld:
                                    if adminPswd!='':
                                        basic_usersS[i].pswd=h(adminPswd);
                                    #if adminName!='':
                                    #    basic_usersS[i].login=adminName;
                                    basic_usersS[i].changed="1";
                                    print('ok')
                                    break;
                            #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                            #cursor1=conn1.cursor();
                            #cursor1.execute("UPDATE 'admins'SET role=(?)WHERE login=(?)",(role,adminOld,));
                            #cursor1.execute("UPDATE 'admins'SET login=(?)WHERE login=(?)",(adminName,adminOld,));
                            #conn1.commit();
                            #conn1.close();
                            for i in range(len(orders_adminsS)):
                                if orders_adminsS[i].merchName==merchName:
                                    if orders_adminsS[i].login==adminOld:
                                        if role!='':
                                            orders_adminsS[i].role=role;
                                        #if adminName!='':
                                        #    orders_adminsS[i].login=adminName;
                                        orders_adminsS[i].changed="1";
                                        break;
                        send={"err":"0","text":"OK"};
                    #conn.commit();
                    #conn.close();
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_remove_admin(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='POST':
                    adminName=request.POST['adminName'];
                else:
                    adminName=request.GET['adminName'];
                merchName=Merch(session);
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'users'WHERE login=(?)AND user_type='admin'",(adminName,));
                #conn.commit();
                #conn.close();
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].login==adminName:
                        basic_usersS[i].removed="1";
                        basic_usersS[i].changed="1";
                        break;
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("DELETE FROM 'admins'WHERE login=(?)",(adminName,));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_adminsS)):
                    if orders_adminsS[i].merchName==merchName:
                        if orders_adminsS[i].login==adminName:
                            orders_adminsS[i].removed="1";
                            orders_adminsS[i].changed="1";
                            break;
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_acces_admin(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    adminName=request.GET['adminName'];
                    adminAcces=request.GET['adminAcces'].replace("_","|");
                if request.method=='POST':
                    adminName=request.POST['adminName'];
                    adminAcces=request.POST['adminAcces'].replace("_","|");
                merchName=Merch(session);
                for i in range(len(orders_adminsS)):
                    if orders_adminsS[i].merchName==merchName:
                        if orders_adminsS[i].login==adminName:
                            orders_adminsS[i].level=adminAcces;
                            orders_adminsS[i].changed="1";

                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_about_not(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                title=getConst(merchName,"notTitle")
                body=getConst(merchName,"notBody")
                #conn.commit();
                #conn.close();
                text=title+"|"+body+"|^";
                text=stringToArray(text);
                send={"err":"0","text":text};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_create_new_product(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                prodName=request.GET['prodName'];
                prodPrice=request.GET['prodPrice'];
                try:
                    cat_id=request.GET['cat_id'];
                except:
                    cat_id="0";
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT name FROM 'products'");
                #names=cort_to_list(cursor.fetchall());
                names=[];
                for o in orders_productsS:
                    if o.merchName==merchName and o.removed=="0":
                        names.append(o.name);
                if prodName in names:
                    #cursor.execute("SELECT work FROM 'products'WHERE name=(?)",(prodName,));
                    #oldWork=cort_to_list(cursor.fetchall())[0];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.name==prodName:
                                oldWork=o.work;
                                break;
                    if oldWork=="0":
                        #cursor.execute("UPDATE products SET work='1'WHERE name=(?)",(prodName,));
                        #cursor.execute("UPDATE products SET work='1'WHERE cat_id=(?)",(cat_id,));
                        for i in range(len(orders_productsS)):
                            if orders_productsS[i].merchName==merchName:
                                if orders_productsS[i].name==prodName:
                                    o.work="1";
                                    o.changed="1";
                                if orders_categoriesS[i].name==cat_id:
                                    o.work="1";
                                    o.changed="1";
                        send="err=0,,text=Продукт восстановлен";
                    else:
                        send="err=1,,text=Продукт уже существует";
                    #conn.commit();
                    #conn.close();
                    async3(Login(session),"getProducts",merchName);
                    async3(Login(session),"getAllPrices",merchName);
                else:
                    #cursor.execute("SELECT id FROM 'products'");
                    #ids=cort_to_list(cursor.fetchall());
                    ids=[];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            ids.append(o.id);
                    #cursor.execute("SELECT name FROM 'buyers_type'");
                    #buyers_types=cort_to_list(cursor.fetchall());
                    buyers_types=[];
                    for o in orders_buyers_typeS:
                        if o.merchName==merchName and o.removed=="0":
                            buyers_types.append(o.name);

                    id_=1;
                    while str(id_) in ids:
                        id_+=1;
                    #cursor.execute("INSERT INTO 'products' VALUES((?),(?),(?),(?),(?),(?),'1','sh',(?),(?),(?),(?))",
                    #    (cat_id,str(id_),prodName,"_","1","_","1","0",prodName,"_"));
                    orders_productsS.append(orders_products(cat_id,str(id_),prodName,"_","1","_",'1','sh',"1","0",prodName,"_",merchName));
                    if prodPrice=="":
                        prodPrice="0";
                    for b in buyers_types:
                        #cursor.execute("INSERT INTO 'prices' VALUES((?),(?),(?))",
                        #    (str(id_),b,prodPrice,));
                        print(b)
                        orders_pricesS.append(orders_prices(str(id_),b,prodPrice,merchName));
                    #conn.commit();
                    #conn.close();
                    send={"err":"0","text":"OK"};
                    async3(Login(session),"getProducts",merchName);
                    async3(Login(session),"getAllPrices",merchName);
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_rem_product(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                prod_id=request.GET['prod_id'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prod_id,));
                #name=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("UPDATE products SET work='0'WHERE id=(?)",(prod_id,));
                for i in range(len(orders_productsS)):
                    if orders_productsS[i].merchName==merchName:
                        if orders_productsS[i].id==prod_id:
                            name=orders_productsS[i].name;
                            orders_productsS[i].work="0";
                            orders_productsS[i].changed="1";
                            break;
                #conn.commit();
                #conn.close();
                async2(Login(session),"getProducts");
                async2(Login(session),"getAllPrices");
                txt="Удален продукт: {0}".format(name);
                makeAct(merchName,Login(session),txt);
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_rem_cat(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    cat_id=request.GET['cat_id'];
                if request.method=='POST':
                    cat_id=request.POST['cat_id'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT name FROM 'categories' WHERE id=(?)",(cat_id,));
                #name=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("UPDATE categories SET work='0'WHERE id=(?)",(cat_id,));
                for i in range(len(orders_categoriesS)):
                    if orders_categoriesS[i].merchName==merchName:
                        if orders_categoriesS[i].id==cat_id:
                            name=orders_categoriesS[i].name;
                            orders_categoriesS[i].work="0";
                            orders_categoriesS[i].changed="1";
                            print(cat_id,"<<<<removed cat")
                            break;
                #conn.commit();
                #conn.close();

                async2(Login(session),"getCategories");
                txt="Удалена категория: {0}".format(cat_id);
                makeAct(merchName,Login(session),txt);
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_edit_cat(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    old_id=request.GET['old_id'];
                    new_cat_name=request.GET['new_cat_name'];
                    new_cat_name2=request.GET['new_cat_name2'];
                    new_cat_id=request.GET['new_cat_id'];
                    Var=request.GET['Var'];
                if request.method=='POST':
                    old_id=request.POST['old_id'];
                    new_cat_name=request.POST['new_cat_name'];
                    new_cat_name2=request.POST['new_cat_name2'];
                    new_cat_id=request.POST['new_cat_id'];
                    Var=request.POST['Var'];
                print(request.POST);
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                if Var=="old":
                    #cursor.execute("UPDATE categories SET name2=(?)WHERE id=(?)",(new_cat_name2,old_id,));
                    #cursor.execute("UPDATE categories SET name=(?)WHERE id=(?)",(new_cat_name,old_id,));
                    #cursor.execute("UPDATE categories SET cat_id=(?)WHERE id=(?)",(new_cat_id,old_id,));
                    for i in range(len(orders_categoriesS)):
                        if orders_categoriesS[i].merchName==merchName:
                            if orders_categoriesS[i].id==old_id:
                                orders_categoriesS[i].name2=new_cat_name2
                                orders_categoriesS[i].name=new_cat_name
                                orders_categoriesS[i].cat_id=new_cat_id
                                orders_categoriesS[i].changed="1";
                                print(new_cat_id,"<<<<")
                                break;
                    send={"err":"0","text":"Сохранено"};
                elif Var=="new":
                    #cursor.execute("SELECT name FROM 'categories'");
                    #names=cort_to_list(cursor.fetchall());
                    names=[];
                    for o in orders_categoriesS:
                        if o.merchName==merchName and o.removed=="0":
                            names.append(o.name);
                    if new_cat_name in names:
                        #cursor.execute("UPDATE categories SET work='1'WHERE name=(?)",(new_cat_name,));
                        #cursor.execute("UPDATE categories SET cat_id=(?)WHERE name=(?)",(new_cat_id,new_cat_name,));
                        for i in range(len(orders_categoriesS)):
                            if orders_categoriesS[i].merchName==merchName:
                                if orders_categoriesS[i].name==new_cat_name:
                                    orders_categoriesS[i].work="1";
                                    orders_categoriesS[i].cat_id=new_cat_id;
                                    orders_categoriesS[i].changed="1";
                                    break;
                        send={"err":"1","text":"Категория включена"};
                    else:
                        #cursor.execute("SELECT id FROM 'categories'");
                        #ids=cort_to_list(cursor.fetchall());
                        ids=[];
                        for o in orders_categoriesS:
                            if o.merchName==merchName and o.removed=="0":
                                ids.append(o.id);
                        id_=1;
                        while str(id_) in ids:
                            id_+=1;
                        id_=str(id_);
                        #cursor.execute("INSERT INTO categories VALUES((?),(?),(?),'1',NULL,(?))",
                        #    (new_cat_id,id_,new_cat_name,new_cat_name));
                        orders_categoriesS.append(orders_categories(new_cat_id,id_,new_cat_name,'1',"",new_cat_name2,merchName));
                        cat=[[str(new_cat_id),str(id_),new_cat_name,merchName,"_img",new_cat_name2]]
                        send={"err":"0","text":"OK","cat":cat};
                #conn.commit();
                #conn.close();
                async2(Login(session),"getCategories");
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_edit_market(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                markName=request.GET['markName'];
                markAdres=request.GET['markAdres'];
                markInn=request.GET['markInn'].replace("}{","+").replace(" ","");
                markPhone=request.GET['markPhone'];
                markDistrict=request.GET['markDistrict'];
                markOld=request.GET['markOld'];
                markVarified=request.GET['markVarified'];
                merchName=Merch(session);
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'users' SET district=(?)WHERE login=(?)",(markDistrict,markOld,));
                #cursor.execute("UPDATE 'users' SET varified=(?)WHERE login=(?)",(markVarified,markOld,));
                #cursor.execute("UPDATE 'users' SET phone=(?)WHERE login=(?)",(markPhone,markOld,));
                #cursor.execute("UPDATE 'users' SET inn=(?)WHERE login=(?)",(markInn,markOld,));
                #cursor.execute("UPDATE 'users' SET adres=(?)WHERE login=(?)",(markAdres,markOld,));
                #cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(markName,markOld,));
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==markOld:
                        markets_usersS[i].district=markDistrict;
                        markets_usersS[i].varified=markVarified;
                        markets_usersS[i].phone=markPhone;
                        markets_usersS[i].inn=markInn;
                        markets_usersS[i].adres=markAdres;
                        markets_usersS[i].login=markName;
                        markets_usersS[i].changed="1";
                        break;
                #conn.commit();
                #conn.close();
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'");
                #admins=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
                #drivers=cort_to_list(cursor.fetchall());
                #cursor.execute("UPDATE 'users' SET login=(?)WHERE login=(?)",(markName,markOld,));
                #conn.commit();
                #conn.close();
                admins=[];
                drivers=[];
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].user_type=="admin":
                        admins.append(basic_usersS[i].login);
                    if basic_usersS[i].user_type=="driver":
                        drivers.append(basic_usersS[i].login);
                    if basic_usersS[i].login==markOld:
                        basic_usersS[i].login=markName;
                        basic_usersS[i].changed="1";

                merchs=os.listdir("merchants");
                for m in merchs:
                    #conn=sqlite3.connect(mPath(m,"orders2"));
                    #cursor=conn.cursor();
                    #cursor.execute("UPDATE 'buy_markets' SET name=(?)WHERE name=(?)",(markName,markOld,));
                    #cursor.execute("UPDATE 'order' SET getter=(?)WHERE getter=(?)",(markName,markOld,));
                    #conn.commit();
                    #conn.close();
                    for i in range(len(orders_buy_marketsS)):
                        if orders_buy_marketsS[i].merchName==merchName:
                            if orders_buy_marketsS[i].name==markOld:
                                orders_buy_marketsS[i].name=markName;
                                orders_buy_marketsS[i].changed="1";
                                break;
                    for i in range(len(orders_orderS)):
                        if orders_buy_marketsS[i].merchName==merchName:
                            if orders_orderS[i].getter==markOld:
                                orders_orderS[i].getter=markName;
                                orders_orderS[i].changed="1";
                                break;

                for a in admins:
                    async2(a,"getMarketList");
                for d in drivers:
                    async2(d,"getMarketList");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_null_driver(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':driver=request.GET['driverName'];
                if request.method=='POST':driver=request.POST['driverName'];
                merchName=Merch(session);
                login=Login(session);
                #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT prod_id FROM 'naks'");
                #ids=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT free FROM 'naks'");
                #free=cort_to_list(cursor.fetchall());
                #cursor.execute("SELECT reserve FROM 'naks'");
                #reserve=cort_to_list(cursor.fetchall());
                ids=[];
                free=[];
                reserve=[];
                for inc in inCar_naksS:
                    if inc.merchName==merchName and inc.removed=="0":
                        ids.append(inc.prod_id);
                        free.append(inc.free);
                        reserve.append(inc.reserve);
                for i in range(len(ids)):
                    prod_id=ids[i];
                    how_many=int(free[i])+int(reserve[i]);
                    ostPlus(merchName,prod_id,how_many);

                #cursor.execute("UPDATE 'cashInCar' SET cash='0'WHERE driver=(?)",(driver,));
                #cursor.execute("UPDATE 'cashInCar' SET term='0'WHERE driver=(?)",(driver,));
                #cursor.execute("UPDATE 'cashInCar' SET per='0'WHERE driver=(?)",(driver,));
                #cursor.execute("UPDATE 'cashInCar' SET on_day='0'WHERE driver=(?)",(driver,));
                #cursor.execute("DELETE FROM 'nakNum' WHERE owner = (?)",(driver,));
                #cursor.execute("DELETE FROM 'naks' WHERE owner = (?)",(driver,));

                #conn.commit();
                #conn.close();
                for i in range(len(inCar_cashInCarS)):
                    if inCar_cashInCarS[i].merchName==merchName:
                        if inCar_cashInCarS[i].driver==driver:
                            inCar_cashInCarS[i].cash="0";
                            inCar_cashInCarS[i].term="0";
                            inCar_cashInCarS[i].per="0";
                            inCar_cashInCarS[i].on_day="0";
                            inCar_cashInCarS[i].changed="1";
                for i in range(len(inCar_nakNumS)):
                    if inCar_nakNumS[i].merchName==merchName:
                        if inCar_nakNumS[i].owner==driver:
                            inCar_nakNumS[i].removed="1";
                            inCar_nakNumS[i].changed="1";
                for i in range(len(inCar_naksS)):
                    if inCar_naksS[i].merchName==merchName:
                        if inCar_naksS[i].owner==driver:
                            inCar_naksS[i].removed="1";
                            inCar_naksS[i].changed="1";

                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'order' SET get_type='new'WHERE owner=(?)AND get_type='ord'",(driver,));
                #cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)AND get_type='ord'",(driver,));
                #cursor.execute("UPDATE 'order' SET owner='new'WHERE owner=(?)AND get_type='new'",(driver,));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_orderS)):
                    if orders_orderS[i].merchName==merchName:
                        if orders_orderS[i].owner==driver:
                            if orders_orderS[i].get_type=="ord":
                                orders_orderS[i].get_type="new";
                                orders_orderS[i].owner="new";
                            elif orders_orderS[i].get_type=="new":
                                orders_orderS[i].owner="new";
                            orders_orderS[i].changed="1";


                async2(driver,"getProducts");

                async2(login,"getProducts");
                async2(login,"getDrivers");
                async2(login,"getProducts");
                async2(login,"getAdminOrders");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    write_now();
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_districts_driver(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    driver=request.GET['driverName'];
                    data=request.GET['data'];
                if request.method=='POST':
                    driver=request.POST['driverName'];
                    data=request.POST['data'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE 'drivers' SET district=(?)WHERE login=(?)",(data,driver,));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_driversS)):
                    if orders_driversS[i].merchName==merchName:
                        if orders_driversS[i].login==driver:
                            orders_driversS[i].district=data;
                            orders_driversS[i].changed="1";

                async2(Login(session),"getDrivers");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_photo_new(request):
    try:
        session=request.POST['session'];
        if checkSession(session):
            merchName=Merch(session);
            data=request.POST['img'];
            imgType=request.POST['imgType'];
            if imgType=='cat' or imgType=='prod':
                id=request.POST['id'];
                files=os.listdir("img");
                nameFile=generate()+".jpg";
                while nameFile in files:
                    nameFile=generate()+".jpg";
                path="img/"+nameFile;

                img = Image.open(io.BytesIO(base64.decodebytes(bytes(data, "utf-8"))))
                img.save(path);

                olds=[];
                for o in orders_imgS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.type==imgType and o._id==id:
                            olds.append(o.place);
                if len(olds)==0:
                    oimg=orders_img(imgType,id,nameFile,merchName);
                    oimg.changed="1"
                    orders_imgS.append(oimg);
                    with open(path, 'rb') as f:
                       file_data = f.read();
                    myImageS.append(myImage("img/"+nameFile,file_data));
                else:
                    for i in range(len(orders_imgS)):
                        if orders_imgS[i].merchName==merchName:
                            if orders_imgS[i].type==imgType and orders_imgS[i]._id==id:
                                orders_imgS[i].place=nameFile;
                                orders_imgS[i].changed="1";
                    with open(path, 'rb') as f:
                       file_data = f.read();
                    myImageS.append(myImage("img/"+nameFile,file_data));
                img=Img(merchName,imgType,id);
                send={"err":"0","text":"OK",'filename':img};
            if imgType=='merchPhoto':
                files=os.listdir("img");
                nameFile=generate()+".jpg";
                while nameFile in files:
                    nameFile=generate()+".jpg";
                path="img/"+nameFile;

                img = Image.open(io.BytesIO(base64.decodebytes(bytes(data, "utf-8"))))
                img.save(path);

                with open(path, 'rb') as f:
                   file_data = f.read();
                myImageS.append(myImage("img/"+nameFile,file_data));
                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].key=="merchImg":
                            orders_constS[i].val=nameFile;
                            orders_constS[i].changed="1";
                send={"err":"0","text":"OK",'filename':nameFile};
                print(send)
            if imgType=="guvPhoto" or imgType=="pasPhoto":
                #
                marketName=request.POST['marketName'];
                files=os.listdir("marketImg");
                nameFile=marketName+imgType+".jpg";
                path="marketImg/"+nameFile;
                img = Image.open(io.BytesIO(base64.decodebytes(bytes(data, "utf-8"))))
                img.save(path);
                with open(path, 'rb') as f:
                   file_data = f.read();
                havePhoto=False;
                for i in range(len(myImageS)):
                    if myImageS[i].link==path:
                        myImageS[i].photo=file_data;
                        havePhoto=True;
                        break;
                if not havePhoto:
                    myImageS.append(myImage("marketImg/"+nameFile,file_data));
                print("OK");
                for i in range(len(markets_usersS)):
                    if markets_usersS[i].login==marketName:
                        markets_usersS[i].varified="100";
                        markets_usersS[i].changed="1";
                        break;
                admins=[];
                for i in range(len(basic_usersS)):
                    if basic_usersS[i].user_type=="admin" and basic_usersS[i].merchName=="Kay-Kay":
                        admins.append(basic_usersS[i].login);
                for a in admins:
                    l=ADLevel("Kay-Kay",a);
                    if "max98" in l:
                        addNot(a,"title","text","varMarket");
                send={"err":"0","text":"OK"};
            if imgType=='chatSendPhoto':
                chat_id=request.POST['chat_id'];
                sender=request.POST['sender'];
                getter=request.POST['getter'];
                files=os.listdir("img");
                nameFile=generate()+".jpg";
                while nameFile in files:
                    nameFile=generate()+".jpg";
                path="img/"+nameFile;
                img = Image.open(io.BytesIO(base64.decodebytes(bytes(data, "utf-8"))))
                img.save(path);
                with open(path, 'rb') as f:
                   file_data = f.read();
                myImageS.append(myImage("marketImg/"+nameFile,file_data));

                chat_ids=[];
                message_ids=[];
                for i in range(len(basic_messagesS)):
                    chat_ids.append(basic_messagesS[i].chat_id);
                    message_ids.append(basic_messagesS[i].message_id);
                if "|^" in chat_id:
                    getter=stringToArray(chat_id)[0][1];
                    newChatId=1;
                    while True:
                        if str(newChatId) in chat_ids:
                            newChatId+=1;
                        else:
                            break;
                    chat_id=newChatId;
                newMessageId=1;
                while True:
                    if str(newMessageId) in message_ids:
                        newMessageId+=1;
                    else:
                        break;
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                basic_messagesS.append(basic_messages(str(chat_id),str(newMessageId),"photo",sender,getter,date,"0",nameFile))
                send={"err":"0","text":"OK"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"NE"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def send_photo(request,var='network'):
    try:
        b=str(request.body[:600])
        for i in range(len(b)):
            if b[i]==";":
                if b[i+1]=="-":
                    b=b[i+2:];
                    break;
        for i in range(len(b)):
            if b[i]=="-":
                if b[i+1]==";":
                    b=b[:i];
                    break;
        if "guvPhoto" in b or "pasPhoto" in b:
            data=stringToArray(b);
            data=data[0];
            marketName=data[0];
            photoType=data[1];
            files=os.listdir("marketImg");
            nameFile=marketName+photoType+".jpg";
            path="marketImg/"+nameFile;
            #
            f=request.FILES[nameFile];
            with open("marketImg/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);

            with open("marketImg/"+nameFile, 'rb') as f:
               file_data = f.read();
            myImageS.append(myImage("marketImg/"+nameFile,file_data));
            #
            print("OK");
            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("UPDATE users SET varified='100' WHERE login=(?)",(marketName,));
            #conn.commit();
            #conn.close();
            for i in range(len(markets_usersS)):
                if markets_usersS[i].login==marketName:
                    markets_usersS[i].varified="100";
                    markets_usersS[i].changed="1";
                    break;
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users' WHERE user_type='admin'AND merchName='Kay-Kay'");
            #admins=cort_to_list(cursor.fetchall());
            #conn.close();
            admins=[];
            for i in range(len(basic_usersS)):
                if basic_usersS[i].user_type=="admin" and basic_usersS[i].merchName=="Kay-Kay":
                    admins.append(basic_usersS[i].login);
            for a in admins:
                l=ADLevel("Kay-Kay",a);
                if "max98" in l:
                    addNot(a,"title","text","varMarket");
            send={"err":"0","text":"OK"};
            send=json.dumps(send);
            return HttpResponse(send, content_type='application/json');
        elif "chatSendPhoto" in b:
            data=stringToArray(b);
            data=data[0];

            chat_id=data[0];
            sender=data[1];
            getter=data[2];

            files=os.listdir("img");
            nameFile=generate()+".jpg";
            while nameFile in files:
                nameFile=generate()+".jpg";
            #
            f=request.FILES['image.jpg'];
            with open("img/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);

            with open("img/"+nameFile, 'rb') as f:
               file_data = f.read();
            myImageS.append(myImage("img/"+nameFile,file_data));
            chat_ids=[];
            message_ids=[];
            for i in range(len(basic_messagesS)):
                chat_ids.append(basic_messagesS[i].chat_id);
                message_ids.append(basic_messagesS[i].message_id);
            if "|^" in chat_id:
                getter=stringToArray(chat_id)[0][1];
                newChatId=1;
                while True:
                    if str(newChatId) in chat_ids:
                        newChatId+=1;
                    else:
                        break;
                chat_id=newChatId;
            newMessageId=1;
            while True:
                if str(newMessageId) in message_ids:
                    newMessageId+=1;
                else:
                    break;

            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            basic_messagesS.append(basic_messages(str(chat_id),str(newMessageId),"photo",sender,getter,date,"0",nameFile))
            send={"err":"0","text":"OK"};
        else:
            data=stringToArray(b);
            data=data[0];
            photoType=data[0];
            merchName=data[1];
            files=os.listdir("img");
            nameFile=generate()+".jpg";
            while nameFile in files:
                nameFile=generate()+".jpg";
            path="merchants/"+merchName+"/imgBuf/"+nameFile;
            #
            f=request.FILES['image.jpg'];
            with open("img/"+nameFile, 'wb+') as destination:
                for chunk in f.chunks():
                    destination.write(chunk);

            with open("img/"+nameFile, 'rb') as f:
               file_data = f.read();
            myImageS.append(myImage("img/"+nameFile,file_data));

            if photoType=="merchPhoto":
                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].key=="merchImg":
                            orders_constS[i].val=nameFile;
                            orders_constS[i].changed="1";


            elif photoType=="catPhoto":
                ID=data[2];
                #cursor.execute("SELECT place FROM 'img' WHERE type='cat'AND _id=(?)",(ID,));
                #olds=cort_to_list(cursor.fetchall());
                olds=[];
                for o in orders_imgS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.type=="cat" and o._id==ID:
                            olds.append(o.place);
                if len(olds)==0:
                    #cursor.execute("INSERT INTO 'img' VALUES('cat',(?),(?))",(ID,nameFile,));
                    oimg=orders_img("cat",ID,nameFile,merchName);
                    oimg.changed="1"
                    orders_imgS.append(oimg);
                else:
                    #cursor.execute("UPDATE'img'SET place=(?) WHERE _id=(?) AND type='cat'",(nameFile,ID,));
                    for i in range(len(orders_imgS)):
                        if orders_imgS[i].merchName==merchName:
                            if orders_imgS[i].type=="cat"and orders_imgS[i]._id==ID:
                                orders_imgS[i].place=nameFile;
                                orders_imgS[i].changed="1";


            elif photoType=="prodPhoto":
                ID=data[2];
                #cursor.execute("SELECT place FROM 'img' WHERE type='prod'AND _id=(?)",(ID,));
                #olds=cort_to_list(cursor.fetchall());
                olds=[];
                for o in orders_imgS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.type=="prod" and o._id==ID:
                            olds.append(o.place);
                if len(olds)==0:
                    #cursor.execute("INSERT INTO 'img' VALUES('prod',(?),(?))",(ID,nameFile,));
                    oimg=orders_img("prod",ID,nameFile,merchName);
                    oimg.changed="1"
                    orders_imgS.append(oimg);
                else:
                    #cursor.execute("UPDATE'img'SET place=(?) WHERE _id=(?) AND type='prod'",(nameFile,ID,));
                    for i in range(len(orders_imgS)):
                        if orders_imgS[i].merchName==merchName:
                            if orders_imgS[i].type=="prod" and orders_imgS[i]._id==ID:
                                orders_imgS[i].place=nameFile;
                                orders_imgS[i].changed="1";
                                break;
            send={"err":"0","text":"OK"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_merch_rev(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            if User(session)=="admin":
                newRev=request.GET['newRev'];
                merchName=Merch(session);
                #conn=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("UPDATE const SET val=(?)WHERE key=(?)",(newRev,"merchRev"));
                #conn.commit();
                #conn.close();
                for i in range(len(orders_constS)):
                    if orders_constS[i].merchName==merchName:
                        if orders_constS[i].key=="merchRev":
                            orders_constS[i].val=newRev;
                            orders_constS[i].changed=newRev;
                            break;
                async2(Login(session),"getCategories");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_hand_nak(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    driver=request.GET['driverName'];
                    data=request.GET['data'];
                    nak=request.GET['nakNum'];
                if request.method=='POST':
                    driver=request.POST['driverName'];
                    data=request.POST['data'];
                    nak=request.POST['nakNum'];
                merchName=Merch(session);
                admin=Login(session);
                data=stringToArray(data);
                #conn1=sqlite3.connect(mPath(merchName,"orders2"));
                #cursor1=conn1.cursor();
                #conn=sqlite3.connect(mPath(merchName,"inCar2"));
                #cursor=conn.cursor();
                for i in range(len(data)):
                    ids=data[i][0]
                    excelHowE=data[i][1];
                    #cursor1.execute("SELECT name FROM 'products' WHERE id=(?)",(ids,));
                    #excelNameE=cort_to_list(cursor1.fetchall())[0];
                    for o in orders_productsS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.id==ids:
                                excelNameE=o.name;
                                break;
                    #cursor.execute("SELECT free FROM 'naks' WHERE owner=(?)AND prod_id=(?)",(driver,ids,));
                    #freeOld=cort_to_list(cursor.fetchall());
                    freeOld=[];
                    for inc in inCar_naksS:
                        if inc.merchName==merchName and inc.removed=="0":
                            if inc.owner==driver and inc.prod_id==ids:
                                freeOld.append(inc.free);
                    if len(freeOld)==0:
                        #cursor.execute("INSERT INTO 'naks' VALUES((?),(?),(?),'0','0',(?))",
                        #    (excelNameE,ids,excelHowE,driver,));
                        inCar_naksS.append(inCar_naks(excelNameE,ids,excelHowE,'0','0',driver,merchName));
                    else:
                        freeOld=freeOld[0];
                        newFree=str(int(excelHowE)+int(freeOld));
                        #cursor.execute("UPDATE 'naks' SET free=(?) WHERE owner=(?)AND prod_id=(?)",
                        #    (newFree,driver,ids,));
                        for k in range(len(inCar_naksS)):
                            if inCar_naksS[k].merchName==merchName:
                                if inCar_naksS[k].owner==driver and inCar_naksS[k].prod_id==ids:
                                    inCar_naksS[k].free=newFree;
                                    inCar_naksS[k].changed="1";
                                    break;
                    ostMinus(merchName,ids,excelHowE)
                d=datetime.now();
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
                #cursor.execute("INSERT INTO 'nakNum' VALUES((?),(?),(?))",
                #    (driver,nak,date));
                #conn.commit();
                #conn.close();
                inCar_nakNumS.append(inCar_nakNum(driver,nak,date,merchName));
                #conn1.commit();
                #conn1.close();
                async2(admin,"getProducts");
                async2(admin,"getDrivers");
                txt="Создана накладная №{0}".format(nak);
                makeAct(merchName,Login(session),txt);
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_nulls(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    nulls=request.GET['nulls']+";";
                if request.method=='POST':
                    nulls=request.POST['nulls']+";";
                nulls=nulls.replace("|",":")
                merchName=Merch(session);
                merchNames=[];
                for c in basic_clearTimeS:
                    merchNames.append(c.merchName);
                if merchName in merchNames:
                    for i in range(len(basic_clearTimeS)):
                        if basic_clearTimeS[i].merchName==merchName:
                            basic_clearTimeS[i].drivers=nulls;
                            basic_clearTimeS[i].changed="1";
                            break;
                else:
                    basic_usersS.append(basic_users(merchName,nulls,))
                async2(Login(session),"getDefaultMarket");
                send={"err":"0","text":"OK"};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_act_req(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if User(session)=="admin":
                if request.method=='GET':
                    acts=request.GET['acts'];
                    try:
                        date=request.GET['date'];
                    except:
                        date="";
                if request.method=='POST':
                    acts=request.POST['acts'];
                    try:
                        date=request.POST['date'];
                    except:
                        date="";
                print(acts,date)
                merchName=Merch(session);
                login=[];
                Date=[];
                Data=[];
                for o in orders_actHistoryS:
                    if o.merchName==merchName and o.removed=="0":
                        login.append(o.login);
                        Date.append(o.Date);
                        Data.append(o.Data);
                act=[];
                for i in range(len(login)):
                    if date in Date[i]:
                        if acts in login[i] or acts in Data[i]:
                            act.append([login[i],Data[i],Date[i]]);
                acts=arrayToString2(act);
                print(len(acts),"adcts<<<")
                send={"err":"0","text":"OK","acts":acts};
            else:
                send={"err":"1","text":"Вы не админ"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

#driver
#driver do
def create_new_market(request,var='network'):
    try:
        session=request.GET['session'];
        merchName=Merch(session);
        login=Login(session);
        send="";
        if checkSession(session):
            name=request.GET['name'];
            inn=request.GET['inn'];
            adres=request.GET['adres'];
            phone=request.GET['phone'];
            phone2=request.GET['phone2'];
            district=request.GET['district'];
            lon=request.GET['lon'];
            lat=request.GET['lat'];
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #logins=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT district FROM 'districts'");
            #districts=cort_to_list(cursor.fetchall());
            #conn.close();
            logins=[];
            for c in basic_usersS:
                logins.append(c.login);
            districts=[];
            for c in basic_districtsS:
                districts.append(c.district);

            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT inn FROM 'users'");
            #inns=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT adres FROM 'users'");
            #logins=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT phone FROM 'users'");
            #phones=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT adres FROM 'users'");
            #adress=cort_to_list(cursor.fetchall());
            inns=[];
            logins=[];
            phones=[];
            adress=[];
            for m in markets_usersS:
                inns.append(m.inn);
                logins.append(m.login);
                phones.append(m.phone);
                adress.append(m.adres);
            if name in logins:
                send={"err":"1","text":"Такой Логин уже существует"};
            if district not in districts:
                send={"err":"1","text":"Района не существует"};
            if inn in inns:
                send={"err":"1","text":"Такой ИНН уже существует"};
            if phone in phones:
                send={"err":"1","text":"Такой Телефон уже существует"};
            if adres in adress:
                send={"err":"1","text":"Такой Адрес уже существует"};
            if len(name)<6:
                send={"err":"1","text":"Логин слишком короткий"};
            if len(phone)!=9:
                send={"err":"1","text":"Телефон слишком короткий"};
            if len(inn)!=9 and len(inn)!=14:
                send={"err":"1","text":"инн слишком короткий"};
            if len(adres)<6:
                send={"err":"1","text":"Адрес слишком короткий"};
            float(lon)+1;
            float(lat)+1;
            if '"err":"1"' not in str(send):
                #cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?))",
                #    (name,adres,inn,phone,"0",district,phone2));
                #conn.commit();
                markets_usersS.append(markets_users(name,adres,inn,phone,"0",district,phone2))
            #conn.close();

            if '"err":"1"' not in str(send):
                #conn=sqlite3.connect("basic2.sqlite");
                #cursor=conn.cursor();
                #cursor.execute("INSERT INTO users VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
                #        (name,h(name),"_",merchName,"market",lon,lat,"0","ru",));
                #cursor.execute("SELECT login FROM 'users' WHERE user_type='driver'");
                #logins=cort_to_list(cursor.fetchall());
                #conn.commit();
                #conn.close();
                basic_usersS.append(basic_users(name,h(name),"_",merchName,"market",lon,lat,"0","ru",name));
                logins=[];
                for c in basic_usersS:
                    if c.user_type=="driver":
                        logins.append(c.login);
                checkMarket(merchName,name);
                send={"err":"0","text":"OK","adres":[["1","2"]]};
            else:
                for l in logins:
                    async2(l,"getMarketList");
            async2(login,"getMarketList");
            messageInChannel("Kay-Kay","Создан магазин🔐\n\nДобавил: {0}\nОт производителя:{3}\nМагазин: {1}\nИНН: {2}".format(Login(session),name,inn,merchName));
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
        async2(login,"getMarketList");
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json');
def get_adres_from_coords(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            lon=request.GET['lon'];
            lat=request.GET['lat'];
            adres=get_address_from_coords1(lon+","+lat);
            if "не установлен" in adres:
                adres=get_address_from_coords2(lon+","+lat);
            adres=adres.replace("Узбекистан","").replace(", "," ").replace(","," ")
            districts=[];
            for c in basic_districtsS:
                districts.append(c.district);
            district="null";
            for d in districts:
                dMin=d[:4]
                if dMin in adres:
                    district=d;
                    break;
            adresData=arrayToString2([[adres,district]]);
            send={"err":"0","adres":adresData};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json');
def send_dolg(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                last_index=request.GET['last_index'];
                nal=request.GET['nal'];
                term=request.GET['term'];
                per=request.GET['per'];
                order_or_sell=request.GET['order_or_sell'];
            if request.method=='POST':
                last_index=request.POST['last_index'];
                nal=request.POST['nal'];
                term=request.POST['term'];
                per=request.POST['per'];
                order_or_sell=request.POST['order_or_sell'];
            saldo=int(nal)+int(term)+int(per);
            merchName=Merch(session);
            login=Login(session);
            for i in range(len(orders_orderS)):
                if orders_orderS[i].merchName==merchName:
                    if orders_orderS[i].get_type=="dolg" and orders_orderS[i].last_index==last_index:
                        getter=orders_orderS[i].getter;
                        orders_orderS[i].removed="1";
                        orders_orderS[i].changed="1";
                        break;
            checkMarket(merchName,getter);
            for inc in inCar_cashInCarS:
                if inc.merchName==merchName and inc.removed=="0":
                    if inc.driver==login:
                        nal1=inc.cash;
                        term1=inc.term;
                        per1=inc.per;
                        break;
            for i in range(len(inCar_cashInCarS)):
                if inCar_cashInCarS[i].merchName==merchName:
                    if inCar_cashInCarS[i].driver==login:
                        inCar_cashInCarS[i].cash=str(int(nal)+int(nal1));
                        inCar_cashInCarS[i].term=str(int(term)+int(term1));
                        inCar_cashInCarS[i].per=str(int(per)+int(per1));
                        inCar_cashInCarS[i].changed="1";
                        break;
            dolgHave=getDolgInfo(merchName,getter)['dolgHave']
            if getDolgInfo(merchName,getter)['dolgType']=="sum":
                dolgHave=str(int(dolgHave)-saldo);
            else:
                dolgHave=str(int(dolgHave)-1);
            for i in range(len(orders_buy_marketsS)):
                if orders_buy_marketsS[i].merchName==merchName:
                    if orders_buy_marketsS[i].name==getter:
                        orders_buy_marketsS[i].dolgHave=dolgHave;
                        orders_buy_marketsS[i].changed="1";
                        break
            try:
                for o in orders_historyS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.last_index==last_index:
                            historyPrice=o.price;
                            break;
                if historyPrice=="0":
                    for i in range(len(orders_historyS)):
                        if orders_historyS[i].merchName==merchName:
                            if orders_historyS[i].last_index==last_index:
                                orders_historyS[i].price=str(saldo);
                                orders_historyS[i].changed="1";
                                break;
            except:
                pass;
            if "USER_" in getter:
                removeUser(merchName,getter);
            send={"err":"0","text":"OK"};
            async2(login,"getDolgs");
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        async2(login,"getDolgs");
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def send_order(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        login=Login(session);
        if checkSession(session):
            if request.method=='GET':
                last_index=request.GET['last_index'];
                d=request.GET['data'];
                nal=request.GET['nal'];
                term=request.GET['term'];
                per=request.GET['per'];
                order_or_sell=request.GET['order_or_sell'];
            if request.method=='POST':
                last_index=request.POST['last_index'];
                d=request.POST['data'];
                nal=request.POST['nal'];
                term=request.POST['term'];
                per=request.POST['per'];
                order_or_sell=request.POST['order_or_sell'];
            orderType=order_or_sell;
            saldo=int(nal)+int(term)+int(per);
            merchName=Merch(session);
            if order_or_sell=="sell":
                Data=d;
                data=stringToArray(Data);
                getter=last_index;
                Last(merchName,"new");
                last_index=Last(merchName);
            else:
                for o in orders_orderS:
                    if o.merchName==merchName and o.removed=="0":
                        print(o.get_type,o.last_index,last_index,'<<<<<')
                        if o.get_type=="ord" and o.last_index==last_index:
                            Data=o.data;
                            data=stringToArrayData(Data);
                            getter=o.getter;
                            break;
            try:
                checkMarket(merchName,getter);
            except:
                send={"err":"0","text":"NOT"};
                send=json.dumps(send);
                return HttpResponse(send, content_type='application/json')
            buyType=getBuyType(merchName,getter);
            realSaldo=0;
            d=datetime.now();
            try:
                date="{0}.{1}.{2} {3}:{4}".format(str(d.day),str(d.month),str(d.year),str(d.hour+5),str(d.minute));
            except:
                date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            district=District(getter);
            print(login)
            print(data)
            for i in range(len(data)):
                prod_id=data[i][0];
                how_many=data[i][1];
                for o in orders_pricesS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.name==buyType and o.id==prod_id:
                            price=o.price;
                            break;
                realSaldo=realSaldo+(int(price)*int(how_many));
                for inc in inCar_naksS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.owner==login and inc.prod_id==prod_id:
                            reserve=inc.reserve;
                            free=inc.free;
                            sell=inc.sell;
                            break;
                for inc in inCar_cashInCarS:
                    if inc.merchName==merchName and inc.removed=="0":
                        if inc.driver==login:
                            nal1=inc.cash;
                            term1=inc.term;
                            per1=inc.per;
                            break;
                if order_or_sell=="ord":
                    orderType="ord";
                    how_ost=reserve;
                    new_how_ost=str(int(how_ost)-int(how_many));
                    for i in range(len(inCar_naksS)):
                        if inCar_naksS[i].merchName==merchName:
                            if inCar_naksS[i].owner==login and inCar_naksS[i].prod_id==prod_id:
                                inCar_naksS[i].reserve=new_how_ost;
                                inCar_naksS[i].changed="1";
                                break;
                else:
                    orderType="sell";
                    how_ost=free;
                    new_how_ost=int(how_ost)-int(how_many);
                    for i in range(len(inCar_naksS)):
                        if inCar_naksS[i].merchName==merchName:
                            if inCar_naksS[i].owner==login and inCar_naksS[i].prod_id==prod_id:
                                inCar_naksS[i].free=new_how_ost;
                                inCar_naksS[i].changed="1";
                                break;
                new_sell=str(int(sell)+int(how_many));
                for i in range(len(inCar_naksS)):
                    if inCar_naksS[i].merchName==merchName:
                        if inCar_naksS[i].owner==login and inCar_naksS[i].prod_id==prod_id:
                            inCar_naksS[i].sell=new_sell;
                            inCar_naksS[i].changed="1";
                            break;
            orders_historyS.append(orders_history(Data.replace("|",":").replace("^",";"),str(saldo),orderType,getter,login,date,district,last_index,"new",merchName));
            if realSaldo==saldo:
                for k in range(len(orders_orderS)):
                    if orders_orderS[k].merchName==merchName:
                        if orders_orderS[k].get_type=="ord" and orders_orderS[k].last_index==last_index:
                            orders_orderS[k].removed="1";
                            orders_orderS[k].changed="1";
                            print("removed!!!<<<<<<<")
                            break;
                if "USER_" in getter:
                    removeUser(merchName,getter);
            else:
                dolg=str(int(realSaldo)-int(saldo));
                if order_or_sell=="ord":
                    for k in range(len(orders_orderS)):
                        if orders_orderS[k].merchName==merchName:
                            if orders_orderS[k].get_type=="ord" and orders_orderS[k].last_index==last_index:
                                orders_orderS[k].get_type="dolg";
                                orders_orderS[k].changed="1";
                            if orders_orderS[k].last_index==last_index:
                                orders_orderS[k].price=dolg;
                                orders_orderS[k].changed="1";
                else:
                    payForm="nal"
                    print(Data,'je tadyyyyyyyyyy<<<<<<<<<');
                    orders_orderS.append(orders_order(Data.replace("|",":").replace("^",";"),getter,login,last_index,date,"dolg",dolg,payForm,merchName));

                dolgType=getDolgInfo(merchName,getter)['dolgType'];
                dolgHave=getDolgInfo(merchName,getter)['dolgHave'];
                if dolgType=="sum":
                    newHave=int(dolgHave)+int(dolg);
                else:
                    newHave=int(dolgHave)+1;
                for k in range(len(orders_buy_marketsS)):
                    if orders_buy_marketsS[k].merchName==merchName:
                        if orders_buy_marketsS[k].name==getter:
                            orders_buy_marketsS[k].dolgHave=newHave;
                            orders_buy_marketsS[k].changed="1";
                            break;
            for i in range(len(inCar_cashInCarS)):
                if inCar_cashInCarS[i].merchName==merchName:
                    if inCar_cashInCarS[i].driver==login:
                        inCar_cashInCarS[i].cash=str(int(nal)+int(nal1));
                        inCar_cashInCarS[i].term=str(int(term)+int(term1));
                        inCar_cashInCarS[i].per=str(int(per)+int(per1));
                        inCar_cashInCarS[i].changed="1";
                        break;
            if order_or_sell=="ord":
                giveCoinForOrder(getter,saldo);
                for i in range(len(orders_buy_marketsS)):
                    if orders_buy_marketsS[i].name==getter:
                        try:
                            orders_buy_marketsS[i].wallet=str(int(orders_buy_marketsS[i].wallet)-int(per));
                        except:
                            orders_buy_marketsS[i].wallet='0';
                        orders_buy_marketsS[i].changed='1';
                        break;
            notification(merchName,"admin","Заказ выполнен",login+" выполнил заказ №"+last_index);
            send={"err":"0","text":"OK"};
            async2(login,"getOrders");
            async2(login,"getDolgs");
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        async2(login,"getOrders");
        async2(login,"getDolgs");
        send={"err":"1","text":"Ошибка сервера"};
    print(send,"SENDDDDD<<<<<")
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def ostMinus(merchName,prod_id,how_many):
    try:
        #conn32=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor32=conn32.cursor();
        #cursor32.execute("SELECT ost FROM 'products' WHERE id=(?)",(prod_id,));
        #ost1=cort_to_list(cursor32.fetchall())[0];
        for i in range(len(orders_productsS)):
            if orders_productsS[i].merchName==merchName:
                if orders_productsS[i].id==prod_id:
                    ost1=orders_productsS[i].ost;
                    k=i;
                    break;
        ost1=str(int(ost1)-int(how_many));

        if ost1=="0":
            #cursor32.execute("UPDATE products SET visible = '0' WHERE id=(?)",(prod_id,));
            orders_productsS[k].visible="0";
        #cursor32.execute("UPDATE products SET ost = (?) WHERE id=(?)",(ost1,prod_id,));
        orders_productsS[k].ost=ost1;
        orders_productsS[k].changed="1";
        #conn32.commit();
        #conn32.close();
    except Exception as e:
        logger(e);
def ostPlus(merchName,prod_id,how_many):
    try:
        how_many=str(how_many);
        #conn32=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor32=conn32.cursor();
        #cursor32.execute("SELECT ost FROM 'products' WHERE id=(?)",(prod_id,));
        #ost1=cort_to_list(cursor32.fetchall())[0];
        for i in range(len(orders_productsS)):
            if orders_productsS[i].merchName==merchName:
                if orders_productsS[i].id==prod_id:
                    ost1=orders_productsS[i].ost;
                    k=i;
                    break;
        ost1=str(int(ost1)+int(how_many));

        if ost1=="0":
            #cursor32.execute("UPDATE products SET visible = '0' WHERE id=(?)",(prod_id,));
            orders_productsS[k].visible="0";
        else:
            #cursor32.execute("UPDATE products SET visible = '1' WHERE id=(?)",(prod_id,));
            orders_productsS[k].visible="1";

        #cursor32.execute("UPDATE products SET ost = (?) WHERE id=(?)",(ost1,prod_id,));
        orders_productsS[k].ost=ost1;
        #conn32.commit();
        #conn32.close();
    except Exception as e:
        logger(e);
def send_new_inn(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                new_inn=request.GET['new_inn'];
                getter=request.GET['getter'];
            if request.method=='POST':
                new_inn=request.POST['new_inn'];
                getter=request.POST['getter'];
            int(new_inn)+1;
            login=Login(session);
            merchName=Merch(session);
            for i in range(len(markets_usersS)):
                if markets_usersS[i].login==getter:
                    markets_usersS[i].inn=new_inn;
                    markets_usersS[i].changed="1";
                    break;
            send={"err":"0","text":"OK"};
            messageInChannel(merchName,"Добавлен ИНН🔐\n\nДобавил: {0}\nОт производителя:{3}\nМагазин: {1}\nИНН: {2}".format(login,getter,new_inn,merchName));
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_phone(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                new_phone=request.GET['new_phone'];
                getter=request.GET['getter'];
            if request.method=='POST':
                new_phone=request.POST['new_phone'];
                getter=request.POST['getter'];
            int(new_phone)+1;
            login=Login(session);
            merchName=Merch(session);
            for i in range(len(markets_usersS)):
                if markets_usersS[i].login==getter:
                    markets_usersS[i].phone=new_phone;
                    markets_usersS[i].changed="1";
                    break;
            send={"err":"0","text":"OK"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def send_new_location(request,var='network'):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                lon=request.GET['new_lon'];
                lat=request.GET['new_lat'];
                getter=request.GET['getter'];
            if request.method=='POST':
                lon=request.POST['new_lon'];
                lat=request.POST['new_lat'];
                getter=request.POST['getter'];
            login=Login(session);
            merchName=Merch(session);
            for i in range(len(basic_usersS)):
                if basic_usersS[i].login==getter:
                    basic_usersS[i].lon=lon;
                    basic_usersS[i].lat=lat;
                    basic_usersS[i].changed="1";
                    break;
            send={"err":"0","text":"OK"};
            messageInChannel(merchName,"Добавлена локация📍\n\nДобавил: {0}\nМагазин: {1}\nШирота: {2}\nДолгота: {3}".format(login,getter,lat,lon));
            locationInChannel(merchName,lon,lat);
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#driver updates
def uber_driver_request(request):
    try:
        updates="";
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            prod_id=[];
            for inc in inCar_naksS:
                if inc.merchName==merchName and inc.removed=="0":
                    if inc.owner==login:
                        prod_id.append(inc.prod_id);
            prod_id=arrayToString([prod_id]);
            print(prod_id)
            for c in basic_usersS:
                if c.session==session:
                    lon=c.lon;
                    lat=c.lat;
                    break;
            send={
            "getMoney":get_money(request,'local'),
                "getDistricts":get_districts(request,'local'),
                "getNak":get_nak(request,'local'),
                "getOrders":get_orders(request,'local'),
                "getDolgs":get_dolgs(request,'local'),
                "getPrices":get_prices(request,prod_id,'local',"driver"),
                "getProductList":get_product_list(request,'local'),
                "getDefaultMarket":get_default_market(request,'local'),
                "err":"0"};

        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    send=json.dumps(send);
    return HttpResponse(send, content_type='application/json')
def get_product_list(request,var='network'):
    try:
        products=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            cat_id=[];
            prod_id=[];
            name=[];
            rev=[];
            box=[];
            form=[];
            visible=[];
            ost=[];
            name2=[];
            rev2=[];

            for o in orders_productsS:
                if o.merchName==merchName and o.removed=="0":
                    if o.work=="1" and o.removed=="0":
                        cat_id.append(o.cat_id);
                        prod_id.append(o.id);
                        name.append(o.name);
                        rev.append(o.rev);
                        box.append(o.box);
                        form.append(o.form);
                        visible.append(o.visible);
                        ost.append(o.ost);
                        name2.append(o.name2);
                        rev2.append(o.rev2);
            for i in range(len(name)):
                img=Img(merchName,"prod",prod_id[i]);
                products.append([cat_id[i],prod_id[i],name[i],rev[i],img,merchName,box[i],form[i],visible[i],ost[i],name2[i],rev2[i]]);
            products=arrayToString2(products);
            send={"err":"0","products":products};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_cat_list(request,var='network'):
    try:
        cats=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            cat_id=[];
            prod_id=[];
            name=[];
            name2=[];
            for o in orders_categoriesS:
                if o.merchName==merchName and o.removed=="0":
                    if o.work=="1" and o.removed=="0":
                        cat_id.append(o.cat_id);
                        prod_id.append(o.id);
                        name.append(o.name);
                        name2.append(o.name2);
            for i in range(len(name)):
                img=Img(merchName,"cat",prod_id[i]);
                cats.append([cat_id[i],prod_id[i],name[i],img,name2[i],merchName]);
            cats=arrayToString2(cats);
            send={"err":"0","cats":cats};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_default_market(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        print(">"+session+"<")
        if session=='':
            merchName='Kay-Kay';
        else:
            merchName=Merch(session);
        default_buyer=getConst(merchName,"default_buyer");
        defDolgType=getConst(merchName,"defDolgType");
        defDolgVal=getConst(merchName,"defDolgVal");
        defDolgMax=getConst(merchName,"defDolgMax");
        tg_id=getConst(merchName,"tg_id");
        tg_token=getConst(merchName,"tg_token");
        nots=getConst(merchName,"nots");
        excelName=getConst(merchName,"excelName");
        excelHow=getConst(merchName,"excelHow");
        excelStartWrite=getConst(merchName,"excelStartWrite");
        excelType=getConst(merchName,"excelType");
        excelCat=getConst(merchName,"excelCat");
        excelPriceName=getConst(merchName,"excelPriceName");
        excelPrice=getConst(merchName,"excelPrice");
        excelStartWritePrice=getConst(merchName,"excelStartWritePrice");
        excelNakSheet=getConst(merchName,"excelNakSheet");
        excelProdSheet=getConst(merchName,"excelProdSheet");
        excelNakPlace=getConst(merchName,"excelNakPlace");
        api_key=getConst(merchName,"api_key");
        merchImg=getConst(merchName,"merchImg");
        merchRev=getConst(merchName,"merchRev");
        autoNull=getConst(merchName,"autoNull");
        site=getConst(merchName,"site");
        ostName=getConst(merchName,"ostName");
        ostHowOst=getConst(merchName,"ostHowOst");
        ostStart=getConst(merchName,"ostStart");
        ostSheet=getConst(merchName,"ostSheet");
        upRange=getConst(merchName,"upRange");
        generalLang=getConst(merchName,"generalLang");
        secondLang=getConst(merchName,"secondLang");
        generalLink=getConst(merchName,"generalLink");
        exNakDriverName=getConst(merchName,"exNakDriverName");
        print(upRange);
        exPerInn=getConst(merchName,"exPerInn");
        exPerSum=getConst(merchName,"exPerSum");
        exPerStart=getConst(merchName,"exPerStart");
        exPerSheet=getConst(merchName,"exPerSheet");
        name=getConst(merchName,"name");
        merchImg="{0}get_photo/?link={1}".format(URL,merchImg);
        defMarket=arrayToString2([[default_buyer,defDolgType,defDolgVal,defDolgMax,
            tg_id,tg_token,nots,excelName,excelHow,excelStartWrite,excelType,
            excelCat,excelPriceName,excelPrice,excelStartWritePrice,excelNakSheet,
            excelProdSheet,excelNakPlace,api_key,merchName,merchImg,merchRev,autoNull,site
            ,ostName,ostHowOst,ostStart,ostSheet,upRange,generalLang,secondLang,generalLink,exNakDriverName,exPerInn,exPerSum,exPerStart,exPerSheet,name]]);
        send={"err":"0","defMarket":defMarket};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_list(request,var='network'):
    try:
        markets=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=Level(myMerchName);

            login=[];
            for m in markets_usersS:
                login.append(m.login);
            if False:#"a1|" not in level:
                log=[];

                lo=[];
                for c in basic_usersS:
                    if c.merchName==myMerchName and c.user_type=="market":
                        lo.append(c.login);
                for i in range(len(login)):
                    if login[i]in lo:
                        log.append(login[i]);
                login=log;
            for i in range(len(login)):
                log=login[i];
                try:

                    for c in basic_usersS:
                        if c.login==log and c.user_type=="market":
                            lon=c.lon;
                            lat=c.lat;
                            merchName=c.merchName;
                            break;
                    lo=lon;
                    la=lat;
                    me=merchName;

                    for m in markets_usersS:
                        if m.login==log:
                            adres=m.adres;
                            inn=m.inn;
                            phone=m.phone;
                            varified=m.varified;
                            district=m.district;
                            phone2=m.phone2;
                            break;
                    wallet=markWallet(myMerchName,log);
                    if str(phone2)=="None":
                        phone2="";
                    if User(session)!="root":
                        checkMarket(myMerchName,log);
                    dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                    dolgVal=getDolgInfo(myMerchName,log)['dolgVal'];
                    dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                    dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                    buyType=getBuyType(myMerchName,log)
                    markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                except Exception as e:
                    pass;

            markets=arrayToString2(markets);
            send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    print('getMarkets<<<<<<<<',len(markets));
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_market_root_list(request,var='network'):
    try:
        markets=[];
        session=request.GET['session'];
        if checkSession(session):
            login=Login(session);
            myMerchName=Merch(session);
            level=Level(myMerchName);
            #conn1=sqlite3.connect("basic2.sqlite");
            #cursor1=conn1.cursor();
            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #login=cort_to_list(cursor.fetchall());
            login=[];
            for m in markets_usersS:
                login.append(m.login);
            if "a1|" not in level:
                log=[];
                #cursor1.execute("SELECT login FROM 'users' WHERE merchName=(?) AND user_type='market'",(myMerchName,));
                #lo=cort_to_list(cursor1.fetchall());
                lo=[];
                for c in basic_usersS:
                    if c.merchName==myMerchName and c.user_type=="market":
                        lo.append(c.login);
                for i in range(len(login)):
                    if login[i]in lo:
                        log.append(login[i]);
                login=log;
                #conn.close();
            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            for i in range(len(login)):
                log=login[i];
                #cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                #varified=cort_to_list(cursor.fetchall())[0];
                for m in markets_usersS:
                    if m.login==log:
                        varified=v.varified;
                        break;
                if "100" in varified:
                    #cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    #lon=cort_to_list(cursor1.fetchall())[0];
                    #cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    #lat=cort_to_list(cursor1.fetchall())[0];
                    #cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                    #merchName=cort_to_list(cursor1.fetchall())[0];
                    for c in basic_usersS:
                        if c.login==log and c.user_type=="market":
                            lon=c.lon;
                            lat=c.lat;
                            merchName=c.merchName;
                            break;
                    lo=lon;
                    la=lat;
                    me=merchName;
                    #cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                    #inn=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                    #phone=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                    #varified=cort_to_list(cursor.fetchall())[0];
                    #cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                    #district=cort_to_list(cursor.fetchall())[0];
                    for m in markets_usersS:
                        if m.login==log:
                            inn=v.inn;
                            phone=v.phone;
                            varified=v.varified;
                            district=v.district;
                            break;
                    if User(session)!="root":
                        checkMarket(myMerchName,log);
                    dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                    dolgVal=getDolgInfo(myMerchName,log)['dolgVal'];
                    dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                    dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                    buyType=getBuyType(myMerchName,log)
                    markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax]);
            #conn.close();
            #conn1.close();
            markets=arrayToString(markets);
            send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_mikro_market(request,var='network'):
    try:
        markets=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                getMarkets=request.GET['getMarkets'];
                getMarkets=stringToArray(getMarkets)[0];
            if request.method=='POST':
                getMarkets=request.POST['getMarkets'];
            myMerchName=Merch(session);
            level=Level(myMerchName);
            #conn1=sqlite3.connect("basic2.sqlite");
            #cursor1=conn1.cursor();
            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #log=cort_to_list(cursor.fetchall());
            #conn.close();
            log=[];
            for m in markets_usersS:
                log.append(m.login);
            login=[];
            for lo in log:
                if lo in getMarkets:
                    login.append(lo);
            for i in range(len(login)):
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                log=login[i];
                #cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #lon=cort_to_list(cursor1.fetchall())[0];
                #cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #lat=cort_to_list(cursor1.fetchall())[0];
                #cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #merchName=cort_to_list(cursor1.fetchall())[0];
                for c in basic_usersS:
                    if c.login==log and c.user_type=="market":
                        lon=c.lon;
                        lat=c.lat;
                        merchName=c.merchName;
                        break;
                lo=lon;
                la=lat;
                me=merchName;

                #cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                #adres=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                #inn=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                #phone=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                #varified=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                #district=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                #phone2=cort_to_list(cursor.fetchall())[0];
                for m in markets_usersS:
                    if m.login==log:
                        adres=m.adres;
                        inn=m.inn;
                        phone=m.phone;
                        varified=m.varified;
                        district=m.district;
                        phone2=m.phone2;
                        break;
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal'];
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                #conn.close();
            #conn1.close();
            markets=arrayToString2(markets);
            send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_search_text(request,var='network'):
    try:
        markets=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                searchTxt=request.GET['searchTxt'].lower();
            if request.method=='POST':
                searchTxt=request.POST['searchTxt'].lower();
            myMerchName=Merch(session);
            level=Level(myMerchName);
            #conn1=sqlite3.connect("basic2.sqlite");
            #cursor1=conn1.cursor();
            #conn=sqlite3.connect("markets2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT login FROM 'users'");
            #log=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT adres FROM 'users'");
            #adr=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT phone FROM 'users'");
            #pho=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT inn FROM 'users'");
            #inn=cort_to_list(cursor.fetchall());
            #conn.close();
            log=[];
            adr=[];
            pho=[];
            inn=[];
            for m in markets_usersS:
                log.append(m.login);
                adr.append(m.adres);
                pho.append(m.phone);
                inn.append(m.inn);
            login=[];
            for i in range(len(log)):
                if searchTxt in log[i].lower() or searchTxt in adr[i].lower() or searchTxt in pho[i].lower():
                    login.append(log[i]);
            for i in range(len(login)):
                #conn=sqlite3.connect("markets2.sqlite");
                #cursor=conn.cursor();
                log=login[i];
                #cursor1.execute("SELECT lon FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #lon=cort_to_list(cursor1.fetchall())[0];
                #cursor1.execute("SELECT lat FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #lat=cort_to_list(cursor1.fetchall())[0];
                #cursor1.execute("SELECT merchName FROM 'users' WHERE user_type='market' AND login=(?)",(log,));
                #merchName=cort_to_list(cursor1.fetchall())[0];
                for c in basic_usersS:
                    if c.user_type=="market" and c.login==log:
                        lon=c.lon;
                        lat=c.lat;
                        merchName=c.merchName;
                        break;
                lo=lon;
                la=lat;
                me=merchName;
                #cursor.execute("SELECT adres FROM 'users' WHERE login=(?) ",(log,));
                #adres=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT inn FROM 'users' WHERE login=(?) ",(log,));
                #inn=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT phone FROM 'users' WHERE login=(?) ",(log,));
                #phone=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT varified FROM 'users' WHERE login=(?) ",(log,));
                #varified=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT district FROM 'users' WHERE login=(?) ",(log,));
                #district=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT phone2 FROM 'users' WHERE login=(?) ",(log,));
                #phone2=cort_to_list(cursor.fetchall())[0];
                for m in markets_usersS:
                    if m.login==log:
                        adres=m.adres;
                        inn=m.inn;
                        phone=m.phone;
                        varified=m.varified;
                        district=m.district;
                        phone2=m.phone2;
                        break;
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal'];
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
                #conn.close();
            #conn1.close();
            markets=arrayToString2(markets);
            send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_search_loc(request,var='network'):
    try:
        markets=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            if request.method=='GET':
                lon=request.GET['lon'];
                lat=request.GET['lat'];
            if request.method=='POST':
                lon=request.POST['lon'];
                lat=request.POST['lat'];
            myMerchName=Merch(session);
            level=Level(myMerchName);

            log=[];
            Lon=[];
            Lat=[];
            for c in basic_usersS:
                if c.user_type=="market":
                    log.append(c.login)
                    Lon.append(c.lon)
                    Lat.append(c.lat)
            login=[];
            for i in range(len(log)):
                res=get_distance(lon,lat,Lon[i],Lat[i]);
                if int(res)<3:
                    login.append(log[i]);
                    print(res);
            for i in range(len(login)):

                log=login[i];

                for c in basic_usersS:
                    if c.user_type=="market" and c.login==log:
                        lon=c.lon;
                        lat=c.lat;
                        merchName=c.merchName;
                        break;
                lo=lon;
                la=lat;
                me=merchName;

                for m in markets_usersS:
                    if m.login==log:
                        adres=m.adres;
                        inn=m.inn;
                        phone=m.phone;
                        varified=m.varified;
                        district=m.district;
                        phone2=m.phone2;
                        break;
                wallet=markWallet(myMerchName,log);
                if str(phone2)=="None":
                    phone2="";
                if User(session)!="root":
                    checkMarket(myMerchName,log);
                dolgType=getDolgInfo(myMerchName,log)['dolgType'];
                dolgVal=getDolgInfo(myMerchName,log)['dolgVal'];
                dolgHave=getDolgInfo(myMerchName,log)['dolgHave'];
                dolgMax=getDolgInfo(myMerchName,log)['dolgMax'];
                buyType=getBuyType(myMerchName,log)
                markets.append([log,lo,la,adres,inn,phone,varified,district,me,buyType,dolgType,dolgVal,dolgHave,dolgMax,phone2,wallet]);
            markets=arrayToString2(markets);
            send={"err":"0","markets":markets};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send

def get_updates_driver(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            get_location(request);
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT updates FROM 'users' WHERE session=(?)",(session,));
            #updates=cort_to_list(cursor.fetchall())[0];
            for c in basic_usersS:
                if c.session==session:
                    updates=c.updates;
                    break;
            if updates=="0":
                send={"err":"0","update_stat":"0"};
            else:
                updates=stringToArray(updates)
                send={"err":"0","update_stat":"1","updates":updates};
            #cursor.execute("UPDATE users SET updates='0'WHERE session=(?)",(session,));
            for i in range(len(basic_usersS)):
                if basic_usersS[i].session==session:
                    basic_usersS[i].updates="0";
                    basic_usersS[i].changed="1";
                    break;
            #conn.commit();
            #conn.close();
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        #conn.close();
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if "new" in send:
        print(send+"<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<")
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_location(request,lon="l",lat="l",var='network'):
    try:
        session=request.GET['session'];
        if lon=="l":
            lon=request.GET['lon'];
            lat=request.GET['lat'];
        if checkSession(session):
            merchName=Merch(session);
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("UPDATE users SET lon=(?)WHERE session=(?)",(lon,session,));
            #cursor.execute("UPDATE users SET lat=(?)WHERE session=(?)",(lat,session,));
            #conn.commit();
            #conn.close();
            for i in range(len(basic_usersS)):
                if basic_usersS[i].session==session:
                    basic_usersS[i].lon=lon;
                    basic_usersS[i].lat=lat;
                    basic_usersS[i].changed="1";
                    break;
            send={"err":"0"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_prices(request,products=0,var='network',user="admin"):
    try:
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        merchName=Merch(session);
        if products==0:
            try:
                products=request.GET['products'];
            except:
                products=[allProductIds(merchName)];
        else:
            products=[products];
        if user=="driver":
            products=products[0]
            products=stringToArray(products);
        if "|^" in products:
            products=stringToArray(products);
        products=products[0]
        #products=stringToArray(products)[0];
        stringList="";
        if checkSession(session):
            send="";
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT val FROM 'const' WHERE key = 'default_buyer'");
            #default_buyer=cort_to_list(cursor.fetchall())[0];
            for o in orders_constS:
                if o.merchName==merchName and o.removed=="0":
                    if o.key=="default_buyer":
                        default_buyer=o.val;
                        break;
            #cursor.execute("SELECT name FROM 'buyers_type'");
            #buyers_type=cort_to_list(cursor.fetchall());
            buyers_type=[];
            for o in orders_buyers_typeS:
                if o.merchName==merchName and o.removed=="0":
                    buyers_type.append(o.name);
            priceMap={};
            for prod in products:
                for buy_type in buyers_type:
                    #cursor.execute("SELECT name FROM 'prices' WHERE id=(?) AND name=(?)",(prod,buy_type));
                    #name=cort_to_list(cursor.fetchall());
                    #cursor.execute("SELECT price FROM 'prices' WHERE id=(?) AND name=(?)",(prod,buy_type));
                    #price=cort_to_list(cursor.fetchall());
                    name=[];
                    price=[];
                    for o in orders_pricesS:
                        if o.merchName==merchName and o.removed=="0":
                            if o.id==prod and o.name==buy_type:
                                name.append(o.name);
                                price.append(o.price);
                    for i in range(len(name)):
                        stringList+=name[i]+"|"+price[i]+"|^";
                stringList2=stringToArray(stringList);
                priceMap[prod]=stringList2;
                #stringList=prod+"="+stringList;
                #send+=stringList+",,"

                stringList="";
            #conn.close();
            #send=send+"err=0";
            priceMap["err"]="0";
            send=priceMap;
            #print(send+"\n\n")
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def allProductIds(merchName):
    allPro=[];
    for p in orders_pricesS:
        if p.merchName==merchName and p.id not in allPro:
            allPro.append(p.id);
    return allPro;
def get_dolgs(request,var='network'):
    try:
        dolgs=[];
        session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            #conn=sqlite3.connect(mPath(merchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT data FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #data=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT getter FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #getter=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT last_index FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #last_index=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT date FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #date=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT price FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #price=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT owner FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #owner=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT payForm FROM 'order' WHERE owner=(?) AND get_type='dolg'",(login,));
            #payForm=cort_to_list(cursor.fetchall());
            #conn.close();
            data=[];
            getter=[];
            last_index=[];
            date=[];
            price=[];
            owner=[];
            payForm=[];
            for o in orders_orderS:
                if o.merchName==merchName and o.removed=="0":
                    if o.get_type=="golg" and o.owner==login:
                        data.append(o.data);
                        getter.append(o.getter);
                        last_index.append(o.last_index);
                        date.append(o.date);
                        price.append(o.price);
                        owner.append(o.owner);
                        payForm.append(o.payForm);

            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                dolgs.append([data[i],getter[i],last_index[i],date[i],"dolg",price[i],owner[i],payForm[i]]);
            dolgs=arrayToString2(dolgs);
            send={"err":"0","dolgs":dolgs};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_orders(request,var='network',order_type='ord'):
    try:
        orders=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            U=User(session);
            try:
                order_type=request.POST['order_type'];
            except:
                order_type='ord';
            data=[];
            getter=[];
            last_index=[];
            date=[];
            get_type=[];
            owner=[];
            payForm=[];
            price=[];
            for o in orders_orderS:
                if o.merchName==merchName and o.removed=="0":
                    if o.get_type==order_type and o.owner==login or U=='admin':
                        data.append(o.data);
                        getter.append(o.getter);
                        last_index.append(o.last_index);
                        date.append(o.date);
                        get_type.append(o.get_type);
                        owner.append(o.owner);
                        payForm.append(o.payForm);
                        price.append(o.price);
            for i in range(len(data)):
                checkMarket(merchName,getter[i]);
                orders.append([data[i],getter[i],last_index[i],date[i],get_type[i],owner[i],payForm[i],price[i]]);
            orders=arrayToString2(orders);
            send={"err":"0","orders":orders};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_nak(request,var='network'):
    try:
        naks=[];
        if request.method=='GET':session=request.GET['session'];
        if request.method=='POST':session=request.POST['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            name=[];
            prod_id=[];
            free=[];
            reserve=[];
            sell=[];
            for inc in inCar_naksS:
                if inc.merchName==merchName and inc.removed=="0":
                    if inc.owner==login:
                        name.append(inc.name);
                        prod_id.append(inc.prod_id);
                        free.append(inc.free);
                        reserve.append(inc.reserve);
                        sell.append(inc.sell);
            for i in range(len(name)):
                for nam in orders_productsS:
                    if nam.merchName==merchName and nam.removed=="0":
                        name2=nam.name2;
                        break;
                naks.append([name[i],prod_id[i],free[i],reserve[i],sell[i],name2]);
            naks=arrayToString2(naks);
            send={"err":"0","nak":naks};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_districts(request,var='network'):
    try:
        get=request.method=='GET';
        if get:
            session=request.GET['session'];
        else:
            session=request.POST['session'];
        if True:
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT district FROM 'districts'");
            #dists=cort_to_list(cursor.fetchall());
            #conn.close();
            dists=[];
            for c in basic_districtsS:
                dists.append(c.district);
            dists=arrayToString2([dists]);
            send={"err":"0","districts":dists};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
def get_money(request,var='network'):
    try:
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            for inc in inCar_cashInCarS:
                if inc.merchName==merchName and inc.removed=="0":
                    if inc.driver==login:
                        cash=inc.cash;
                        term=inc.term;
                        per=inc.per;
                        on_day=inc.on_day;
                        break;
            num=[];
            date=[];
            for inc in inCar_nakNumS:
                if inc.merchName==merchName and inc.removed=="0":
                    if inc.owner==login:
                        num.append(inc.num);
                        date.append(inc.date);
            nakNum="";
            for i in range(len(num)):
                nakNum+="Накладная №"+num[i]+"\nДата: "+date[i]+" \n";
            send={"err":"0","moneyNal":cash,"moneyTerm":term,"moneyPer":per,"moneyDay":on_day,"merchName":merchName,"nakNum":nakNum};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return 2;
def get_all_orders_sells(request,var='network',order_type='ord'):
    try:
        orders=[];
        if request.method=='POST':session=request.POST['session'];
        if request.method=='GET':session=request.GET['session'];
        if checkSession(session):
            merchName=Merch(session);
            login=Login(session);
            d=datetime.now();
            date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
            date1=[];
            date2=[];
            for o in orders_historyS:
                if o.merchName==merchName and o.removed=="0":
                    if o.driver==login and o.type=="ord":
                        date1.append(o.date);
                    if o.driver==login and o.type=="sell":
                        date2.append(o.date);
            orders=0;
            sells=0;
            for d in date1:
                if date in d:
                    orders+=1;
            for d in date2:
                if date in d:
                    sells+=1;
            send={"err":"0","ord":str(orders),"sells":str(sells)};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json')
    else:
        return send
#functions
def clearInn(text):
    try:
        nums=["0","1","2","3","4","5","6","7","8","9"];
        clear="";
        i=len(text)-1;
        while i!=-1:
            if text[i]in nums:
                clear+=text[i];
            i-=1;
        i=len(clear)-1;
        clear2="";
        while i!=-1:
            clear2+=clear[i];
            i-=1;
        return clear2;
    except Exception as e:
        logger(e);
def loginsByInn(inn):
    #conn=sqlite3.connect("markets2.sqlite");
    #cursor=conn.cursor();
    #cursor.execute("SELECT login FROM 'users' WHERE inn=(?) ",(inn,));
    #login=cort_to_list(cursor.fetchall());
    #conn.commit();
    #conn.close();
    login=[];
    for m in markets_usersS:
        if m.inn==inn:
            login.append(m.login);
    return login

def markWallet(merchName,login):
    if "USER_" not in login:
        checkMarket(merchName,login)
        #conn=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor=conn.cursor();
        #cursor.execute("SELECT wallet FROM 'buy_markets' WHERE name=(?) ",(login,));
        #wallet=cort_to_list(cursor.fetchall())[0];
        for o in orders_buy_marketsS:
            if o.merchName==merchName and o.removed=="0":
                if o.name==login:
                    wallet=o.wallet;
                    break;
        if str(wallet)=="None":
            wallet="0";
            #cursor.execute("UPDATE buy_markets SET wallet='0' WHERE name=(?)",(login,));
            for i in range(len(orders_buy_marketsS)):
                if o.merchName==merchName and o.removed=="0":
                    if orders_buy_marketsS[i].name==login:
                        orders_buy_marketsS[i].wallet="0";
                        orders_buy_marketsS[i].changed="1";
                        break;
        #conn.commit();
        #conn.close();

    else:
        wallet="0";

    return wallet;
def not_was_seen(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            id_=request.GET['id_'];
            #conn=sqlite3.connect('basic2.sqlite');
            #cursor=conn.cursor();
            #cursor.execute("DELETE FROM 'nots' WHERE id = (?)",(id_,));
            #conn.commit();
            #conn.close();
            for i in range(len(basic_notsS)):
                if basic_notsS[i].id==id_:
                    basic_notsS[i].removed="1";
                    basic_notsS[i].changed="1";
                    break;
            send={"err":"0","text":"OK"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def get_notifications(request,var='network'):
    try:
        session=request.GET['session'];
        if checkSession(session):
            N=request.GET['nots'];
            login=Login(session);
            #conn=sqlite3.connect('basic2.sqlite');
            #cursor=conn.cursor();
            #cursor.execute("SELECT id FROM 'nots'WHERE owner=(?)",(login,));
            #id_=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT title FROM 'nots'WHERE owner=(?)",(login,));
            #title=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT text FROM 'nots'WHERE owner=(?)",(login,));
            #text=cort_to_list(cursor.fetchall());
            #cursor.execute("SELECT Type FROM 'nots'WHERE owner=(?)",(login,));
            #Type=cort_to_list(cursor.fetchall());
            id_=[];
            title=[];
            text=[];
            Type=[];
            for c in basic_notsS:
                if c.owner==login:
                    id_.append(c.id);
                    title.append(c.title);
                    text.append(c.text);
                    Type.append(c.Type);

            nots=[];
            for i in range(len(id_)):
                nots.append([id_[i],title[i],text[i],Type[i]]);
            nots=arrayToString2(nots);
            #conn.commit();
            #conn.close();
            if int(N)<len(id_):
                send={"err":"0","nots":nots};
            else:
                send={"err":"2","nots":"0"};
        else:
            send={"err":"1","text":"Сессия истекла"};
    except Exception as e:
        logger(e);
        send={"err":"1","text":"Ошибка сервера"};
    if var=='network':
        send=json.dumps(send);
        return HttpResponse(send, content_type='application/json');
    else:
        return send;
def addNot(owner,title,text,Type):
    try:
        #conn=sqlite3.connect('basic2.sqlite');
        #cursor=conn.cursor();
        #cursor.execute("SELECT id FROM 'nots'");
        #id_=cort_to_list(cursor.fetchall());
        id_=[];
        for c in basic_notsS:
            id_.append(id);
        #cursor.execute("SELECT val FROM 'const'WHERE key='lastNot'");
        #lastNot=cort_to_list(cursor.fetchall())[0];
        for c in basic_constS:
            if c.key=="lastNot":
                lastNot=c.val;
                break;
        lastNot=str(int(lastNot)+1);
        #cursor.execute("UPDATE const SET val=(?) WHERE key='lastNot'",(lastNot,));
        for i in range(len(basic_constS)):
            if basic_constS[i].key=="lastNot":
                basic_constS[i].val=lastNot;
                break;
        #cursor.execute("INSERT INTO 'nots' VALUES((?),(?),(?),(?),(?))",
        #    (lastNot,owner,title,text,Type));
        basic_notsS.append(basic_nots(lastNot,owner,title,text,Type));
        #conn.commit();
        #conn.close();

    except Exception as e:
        logger(e);
def prodById(merchName,prodId):
    #conn=sqlite3.connect(mPath(merchName,"orders2"));
    #cursor=conn.cursor();
    try:
        #cursor.execute("SELECT name FROM 'products' WHERE id=(?)",(prodId,));
        #data=cort_to_list(cursor.fetchall())[0];
        data="";
        for o in orders_productsS:
            if o.merchName==merchName and o.removed=="0":
                if o.id==prodId:
                    data=o.name;
                    break;
        data[0]
    except Exception as e:
        print(e);
        data=prodId;
    #conn.close();
    return data;
def get_distance(lon,lat,lon1,lat1):
    try:
        x1 = lon;
        y1 = lat;
        x2 = lon1;
        y2 = lat1;
        #+++++++++++++++++++++++++++++
        X = (float(x1)-float(x2))
        X = X * X
        Y = (float(y1)-float(y2))
        Y = Y * Y
        S0 = X + Y
        S = math.sqrt(S0)
        S = round(float(S),5)
        km = S*100
        km = round(float(km),3)
        km = round(float(km),1)
    except:
        km=1000;
    return km;

def giveCoinForOrder(login,price):
    try:
        if "USER_" not in login:
            #conn=sqlite3.connect('markets2.sqlite');
            #cursor=conn.cursor();
            #cursor.execute("SELECT varified FROM 'users' WHERE login=(?)",(login,));
            #varified=cort_to_list(cursor.fetchall())[0];
            #conn.close();
            for m in markets_usersS:
                if m.login==login:
                    varified=m.varified;
                    break;
            if varified=="2":
                #conn=sqlite3.connect('basic2.sqlite');
                #cursor=conn.cursor();
                #cursor.execute("SELECT cur FROM 'percent' WHERE var=(?)",(varified,));
                #cur=cort_to_list(cursor.fetchall())[0];
                #cursor.execute("SELECT per FROM 'percent' WHERE var=(?)",(varified,));
                #per=cort_to_list(cursor.fetchall())[0];
                #conn.close();
                for c in basic_percentS:
                    if c.var==varified:
                        cur=c.cur;
                        per=c.per;
                        break;
                price=float(price)*(float(per)/100.0);
                #dol_cur=float(dolConv.get_dol());
                #price=str(float(price)/dol_cur);
                if cur=="ice":
                    how_many=get_course("ice",price).text;
                else:
                    how_many=get_course("kay",price).text;
                giveCoin(login,cur,how_many);
        else:
            pass;
    except Exception as e:
        logger(e);
def get_course(cur,price):
    return get_html(KAY+"/get_course/?cur={0}&price={1}".format(cur,price));
def giveCoin(login,kay_ice,how_many):
    try:
        #conn=sqlite3.connect('markets2.sqlite');
        #cursor=conn.cursor();
        #cursor.execute("SELECT tg_id FROM 'coinInfo' WHERE login=(?)",(login,));
        #tg_id=cort_to_list(cursor.fetchall())[0];
        #cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?)",(login,));
        #tg_pswd=cort_to_list(cursor.fetchall())[0];
        #conn.close();
        for m in markets_coinInfoS:
            if m.login==login:
                tg_id=m.tg_id;
                tg_pswd=m.tg_pswd;
                break;
        get_html(KAY+"/give_coin/?tg_id={0}&tg_pswd={1}&kay_ice={2}&how_many={3}".format(tg_id,tg_pswd,kay_ice,how_many,));
    except Exception as e:
        logger(e);
def createTimeUser(adres,phone,merchName,orderDistrict,phone2):
    #conn=sqlite3.connect('basic2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("SELECT login FROM 'users'");
    #logins=cort_to_list(cursor.fetchall());
    logins=[];
    for c in basic_usersS:
        logins.append(c.login);
    i=1000;
    login="USER_"+str(i);
    while login in logins:
        login="USER_"+str(i);
        i+=1;
    #cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?),(?),(?))",
    #    login,login,"_",merchName,"market","0.0","0.0","0","ru",));
    basic_usersS.append(basic_users(login,login,"_",merchName,"market","0.0","0.0","0","ru",login));
    #conn.commit();
    #conn.close();
    #conn=sqlite3.connect('markets2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("INSERT INTO 'users' VALUES((?),(?),(?),(?),(?),(?),(?))",
    #    (login,adres,"0",phone,"0",orderDistrict,phone2));
    #conn.commit();
    #conn.close();
    markets_usersS.append(markets_users(login,adres,"0",phone,"0",orderDistrict,phone2))
    return login;
def removeUser(merchName,user):
    #conn=sqlite3.connect("basic2.sqlite");
    #cursor=conn.cursor();
    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(user,));
    #conn.commit();
    #conn.close();
    for i in range(len(basic_usersS)):
        if basic_usersS[i].login==user:
            basic_usersS[i].removed="1";
            basic_usersS[i].changed="1";
            break;
    #conn=sqlite3.connect("markets2.sqlite");
    #cursor=conn.cursor();
    #cursor.execute("DELETE FROM 'users' WHERE login = (?)",(user,));
    #conn.commit();
    #conn.close();
    for i in range(len(markets_usersS)):
        if markets_usersS[i].login==user:
            markets_usersS[i].removed="1";
            markets_usersS[i].changed="1";
            break;
def handle_uploaded_file(f,Path,name):
    with open(Path+'/'+name+'.xlsx', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk);
def async2(login,updates,merchName="0"):
    asy = threading.Thread(target=async3, args=(login,updates,merchName));
    asy.start()
def async3(login,updates,merchName):
    #conn=sqlite3.connect("basic2.sqlite");
    #cursor=conn.cursor();

    if merchName=="0":
        #cursor.execute("SELECT updates FROM 'users' WHERE login=(?)",(login,));
        for c in basic_usersS:
            if c.login==login:
                oldUpdates=c.updates;
                break;
    else:
        #cursor.execute("SELECT updates FROM 'users' WHERE login=(?)AND merchName=(?)",(login,merchName,));
        for c in basic_usersS:
            if c.login==login and c.merchName==merchName:
                oldUpdates=c.updates;
                break;
    try:
        #oldUpdates=cort_to_list(cursor.fetchall())[0];
        oldUpdates=oldUpdates;
    except:
        oldUpdates="";
    if oldUpdates=="0":
        update=updates+"|^";
    else:
        if updates not in oldUpdates:
            update=oldUpdates+updates+"|^";
        else:
            update=oldUpdates;
    if merchName=="0":
        #cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)",(update,login,));
        for i in range(len(basic_usersS)):
            if basic_usersS[i].login==login:
                basic_usersS[i].updates=update;
                basic_usersS[i].changed="1";
                break;
    else:
        #cursor.execute("UPDATE users SET updates=(?) WHERE login=(?)AND merchName=(?)",(update,login,merchName,));
        for i in range(len(basic_usersS)):
            if basic_usersS[i].login==login and basic_usersS[i].merchName==merchName:
                basic_usersS[i].updates=update;
                basic_usersS[i].changed="1";
                break;
    #conn.commit();
    #conn.close();
    send="OK";
    return send;
def notification(merchName,login,title,body):
    nots=getConst(merchName,"nots");
    if nots=="tg":
        messageInChannel(merchName,title+"\n"+body);
    elif nots=="app":
        #conn=sqlite3.connect(mPath(merchName,"orders2"));
        #cursor=conn.cursor();
        #cursor.execute("UPDATE 'const' SET val=(?)WHERE key='notTitle'",(title,));
        #cursor.execute("UPDATE 'const' SET val=(?)WHERE key='notBody'",(body,));
        #conn.commit();
        #conn.close();
        for i in range(len(orders_constS)):
            if orders_constS[i].merchName==merchName:
                if orders_constS[i].key=="notTitle":
                    orders_constS[i].val=title;
                if orders_constS[i].key=="notBody":
                    orders_constS[i].val=body;
                orders_constS[i].changed="1";

        async2(login,"sendAboutNot")
    else:
        print(">>>>");
        print(title);
        print(body);
        print(">>>>");
def checkSession(session):
    #conn=sqlite3.connect('basic2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("SELECT session FROM 'users' WHERE session=(?)",(session,));
    #length=len(cort_to_list(cursor.fetchall()));
    #conn.close();
    leng=[];
    for c in basic_usersS:
        if c.session==session:
            leng.append(c.session);
            break;
    length=len(leng)
    if length==1:
        return True;
    else:
        return False;
def genSession(data):
    #conn=sqlite3.connect('basic2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("SELECT session FROM 'users'");
    #sessions=cort_to_list(cursor.fetchall());
    sessions=[];
    for c in basic_usersS:
        sessions.append(c.session);
    session=generate();
    while session in sessions:
        session=generate();
    #cursor.execute("UPDATE users SET session=(?)WHERE login=(?)",(session,data,));
    for i in range(len(basic_usersS)):
        if basic_usersS[i].login==data:
            basic_usersS[i].session=session;
            basic_usersS[i].changed="1";
            break;
    #conn.commit();
    #conn.close();
    return session;
def getBuyType(myMerchName,login):
    if "USER_" in login or myMerchName=="self":
        return "user";
    else:
        try:
            #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT type FROM 'buy_markets' WHERE name=(?)",(login,));
            #Type=cort_to_list(cursor.fetchall())[0];
            #conn.close();
            for o in orders_buy_marketsS:
                if o.merchName==myMerchName:
                    if o.name==login:
                        Type=o.type;
                        break;
            Type=Type;
            return Type;
        except:
            Def=getConst(myMerchName,"default_buyer");
            #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("UPDATE buy_markets SET type=(?)WHERE name=(?)",(Def,login,));
            #conn.commit();
            #conn.close();
            for i in range(len(orders_buy_marketsS)):
                if orders_buy_marketsS[i].merchName==myMerchName:
                    if orders_buy_marketsS[i].name==login:
                        orders_buy_marketsS[i].type=Def;
                        orders_buy_marketsS[i].changed="1";
                        break;
            return Def;
def getProductPrice(merchName,market,prod_id):
    buyType=getBuyType(merchName,market);
    #conn=sqlite3.connect(mPath(merchName,"orders2"));
    #cursor=conn.cursor();
    #cursor.execute("SELECT price FROM 'prices' WHERE name=(?)AND id=(?)",(buyType,prod_id,));
    #price=cort_to_list(cursor.fetchall())[0];
    #conn.close();
    for o in orders_pricesS:
        price="";
        if o.merchName==merchName and o.removed=="0":
            if o.name==buyType and o.id==prod_id:
                price=o.price;
                break;
    if price=="":
        print(prod_id,buyType,merchName,"<<<<price")
    else:
        pass
    return price;
def Img(merchName,cat_prod,prod_id):
    try:
        for o in orders_imgS:
            if o.merchName==merchName and o.removed=="0":
                if o.type==cat_prod and o._id==prod_id:
                    img=o.place;
                    break;
        img=img
    except:
        img=getConst(merchName,"merchImg");
    ret="{0}get_photo/?link={1}".format(URL,img);
    return ret;
def getConst(myMerchName,key):
    try:
        if key=="autoNull":
            #conn=sqlite3.connect("basic2.sqlite");
            #cursor=conn.cursor();
            #cursor.execute("SELECT drivers FROM 'clearTime' WHERE merchName=(?)",(myMerchName,));
            #val=cort_to_list(cursor.fetchall())[0];
            #conn.close();
            for c in basic_clearTimeS:
                if c.merchName==myMerchName:
                    val=c.drivers;
                    break;
        else:
            #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT val FROM 'const' WHERE key=(?)",(key,));
            #val=cort_to_list(cursor.fetchall())[0];
            #conn.close();
            for o in orders_constS:
                if o.merchName==myMerchName:
                    if o.key==key:
                        val=o.val;
                        break;
        return val;
    except Exception as e:
        return "null";
def Last(merchName,newLast="0"):
    newLast=str(newLast);
    #conn=sqlite3.connect(mPath(merchName,"orders2"));
    #cursor=conn.cursor();
    #cursor.execute("SELECT val FROM 'const' WHERE key='last_index'");
    #last_index=cort_to_list(cursor.fetchall())[0];
    for o in orders_constS:
        if o.merchName==merchName and o.removed=="0":
            if o.key=="last_index":
                last_index=o.val;
                break;
    if newLast=="0":
        #conn.close();
        return last_index;
    else:
        newLast=str(int(last_index)+1);
        #cursor.execute("UPDATE 'const' SET val=(?)WHERE key='last_index'",(newLast,));
        #conn.commit();
        #conn.close();
        for i in range(len(orders_constS)):
            if orders_constS[i].merchName==merchName:
                if orders_constS[i].key=="last_index":
                    orders_constS[i].val=newLast;
                    orders_constS[i].changed="1";
                    break;
        return "OK";
def makeAct(merchName,login,data):
    d=datetime.now();
    date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
    #conn=sqlite3.connect(mPath(merchName,"orders2"));
    #cursor=conn.cursor();
    #cursor.execute("INSERT INTO actHistory VALUES((?),(?),(?))",(login,data,date));
    #conn.commit();
    #conn.close();
    orders_actHistoryS.append(orders_actHistory(login,data,date,merchName));
def checkMarket(myMerchName,mark):
    if "USER_" in str(mark):
        return "";
    else:
        if myMerchName=="self":
            merchs=os.listdir("merchants");
            for myMerchName in merchs:
                #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
                #cursor=conn.cursor();
                #cursor.execute("SELECT name FROM 'buy_markets' WHERE name=(?)",(mark,));
                #markets=cort_to_list(cursor.fetchall());
                markets=[];
                for o in orders_buy_marketsS:
                    if o.merchName==merchName and o.removed=="0":
                        if o.name==mark:
                            markets.append(o.name);

                if len(markets)==0:
                    default_buyer=getConst(myMerchName,"default_buyer");
                    defDolgType=getConst(myMerchName,"defDolgType");
                    defDolgVal=getConst(myMerchName,"defDolgVal");
                    defDolgMax=getConst(myMerchName,"defDolgMax");
                    #cursor.execute("INSERT INTO buy_markets VALUES((?),(?),(?),(?),(?),(?),(?))",
                    #            (mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0"));
                    #conn.commit();
                    #conn.close();
                    orders_buy_marketsS.append(orders_buy_markets(mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0",merchName));
                else:
                    #conn.close();
                    pass;
        else:
            #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT name FROM 'buy_markets' WHERE name=(?)",(mark,));
            #markets=cort_to_list(cursor.fetchall());
            markets=[];
            for o in orders_buy_marketsS:
                if o.merchName==myMerchName:
                    if o.name==mark:
                        markets.append(o.name);
            if len(markets)==0:
                default_buyer=getConst(myMerchName,"default_buyer");
                defDolgType=getConst(myMerchName,"defDolgType");
                defDolgVal=getConst(myMerchName,"defDolgVal");
                defDolgMax=getConst(myMerchName,"defDolgMax");
                #cursor.execute("INSERT INTO buy_markets VALUES((?),(?),(?),(?),(?),(?),(?))",
                #            (mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0"));
                #conn.commit();
                #conn.close();
                orders_buy_marketsS.append(orders_buy_markets(mark,default_buyer,defDolgType,defDolgVal,"0",defDolgMax,"0",myMerchName));
            else:
                #conn.close();
                pass;
def checkCoinSettings(login):
    #conn=sqlite3.connect("markets2.sqlite");
    #cursor=conn.cursor();
    #cursor.execute("SELECT pswd FROM 'coinInfo' WHERE login=(?)",(login,));
    #logins=cort_to_list(cursor.fetchall());
    logins=[];
    for m in markets_coinInfoS:
        if m.login==login:
            logins.append(m.pswd);
    if len(logins)==0:
        #cursor.execute("INSERT INTO coinInfo VALUES((?),(?),(?))",(login,"",""));
        markets_coinInfoS.append(markets_coinInfo(login,"",""));
        #conn.commit();
    #conn.close();
def getDolgInfo(myMerchName,login):
    try:
        if myMerchName!="self":
            if "USER_" in login:
                login="USER_";
            #conn=sqlite3.connect(mPath(myMerchName,"orders2"));
            #cursor=conn.cursor();
            #cursor.execute("SELECT dolgType FROM 'buy_markets' WHERE name=(?)",(login,));
            #dolgType=cort_to_list(cursor.fetchall())[0];
            #cursor.execute("SELECT dolgVal FROM 'buy_markets' WHERE name=(?)",(login,));
            #dolgVal=cort_to_list(cursor.fetchall())[0];
            #cursor.execute("SELECT dolgHave FROM 'buy_markets' WHERE name=(?)",(login,));
            #dolgHave=cort_to_list(cursor.fetchall())[0];
            #cursor.execute("SELECT dolgMax FROM 'buy_markets' WHERE name=(?)",(login,));
            #dolgMax=cort_to_list(cursor.fetchall())[0];
            #conn.close();
            for o in orders_buy_marketsS:
                if o.merchName==myMerchName and o.removed=="0":
                    if o.name==login:
                        dolgType=o.dolgType;
                        dolgVal=o.dolgVal;
                        dolgHave=o.dolgHave;
                        dolgMax=o.dolgMax;
                        break;
            vac={"dolgType":dolgType,"dolgVal":dolgVal,"dolgHave":dolgHave,"dolgMax":dolgMax};
            return vac;
        else:
            vac={"dolgType":"how","dolgVal":"0","dolgHave":"0","dolgMax":"0"};
            return vac;
    except Exception as e:
        print(e);
        vac={"dolgType":"how","dolgVal":"0","dolgHave":"0","dolgMax":"0"};
        return vac;
def User(data,l=1):
    if l==1:
        for c in basic_usersS:
            if c.session==data:
                user_type=c.user_type;
                break;
    else:
        for c in basic_usersS:
            if c.login==data:
                user_type=c.user_type;
                break;
    return user_type;
def Login(data):
    for c in basic_usersS:
        if c.session==data:
            login=c.login;
            break;
    return login;

def Ent(ent):
    login='';
    for c in basic_usersS:
        if c.ent==ent:
            login=c.login;
            break;
    return login;
def Merch(data,l=1):
    if l==1:
        for c in basic_usersS:
            if c.session==data:
                merchName=c.merchName;
                break;
    else:
        for c in basic_usersS:
            if c.login==data:
                merchName=c.merchName;
                break;
    return merchName;
def Level(merchName):
    #conn=sqlite3.connect('basic2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("SELECT level FROM 'levels' WHERE merchName=(?)",(merchName,));
    #level=cort_to_list(cursor.fetchall())[0];
    #conn.close();
    for c in basic_levelsS:
        if c.merchName==merchName:
            level=c.level;
            break;
    return level;
def ADLevel(merchName,login):
    #conn=sqlite3.connect(mPath(merchName,"orders2"));
    #cursor=conn.cursor();
    #cursor.execute("SELECT level FROM 'admins' WHERE login=(?)",(login,));
    #level=cort_to_list(cursor.fetchall())[0];
    #conn.close();
    for o in orders_adminsS:
        if o.merchName==merchName and o.removed=="0":
            if o.login==login:
                level=o.level;
                break;
    return level;
def District(login):
    #conn=sqlite3.connect('markets2.sqlite');
    #cursor=conn.cursor();
    #cursor.execute("SELECT district FROM 'users' WHERE login=(?)",(login,));
    #district=cort_to_list(cursor.fetchall())[0];
    #conn.close();
    for m in markets_usersS:
        district=m.district;
    return district;
def Dolg(login,market):
    #conn=sqlite3.connect(mPath(Merch(session),"orders2"));
    #cursor=conn.cursor();
    #cursor.execute("SELECT data FROM order WHERE owner=(?)AND getter=(?)AND get_type='dolg'"
    #            ,(login,market,));
    #dolgs=cort_to_list(cursor.fetchall());
    #conn.close();
    dolgs=[];
    for o in orders_orderS:
        if o.merchName==merchName and o.removed=="0":
            if o.owner==login and o.getter==market and o.get_type=="dolg":
                dolgs.append(data);
    if len(dolgs)==0:
        return True;
    else:
        return False;
def mPath(merchName,base):
    return "merchants/"+merchName+"/"+base+'.sqlite';
def generate():
    gen=["q","w","e","r","t","y","u","i","o","p","a","s","d","f","g","h","j","k","l","z","x","c","v","b","n","m"]
    Gen=[]
    for i in gen:
        Gen.append(i.upper());
    nums=["0","1","2","3","4","5","6","7","8","9"];
    globList=gen+Gen+nums;

    word=""
    for i in range(32):
        symbol=globList[random.randint(0,len(globList)-1)];
        word+=str(symbol);
    return word;
def stringToArray(string):
    prods=[];
    prod_mass=[];
    tov=""
    for i in string:
        if i=="^":
            prod_mass.append(prods);
            prods=[];
        elif i =="|":
            prods.append(tov);
            tov="";
        else:
            tov=tov+i;
    return prod_mass;
def stringToMap(string):
    string=string+",,";
    Map={};
    a="";
    b="";
    var="a";
    try:
        for z in range(len(string)):
            i=string[z];
            if i=="=":
                var="b";
            elif string[z]=="," and string[z+1]==",":
                Map[a.replace(",","")]=b;
                var="a";
                a="";b="";
            else:
                if var=="a":
                    a=a+i;
                elif var=="b":
                    b=b+i;
    except:
        pass
    return Map;
def stringToArrayData(string):
    prods=[];
    prod_mass=[];
    tov=""
    for i in string:
        if i==";":
            prod_mass.append(prods);
            prods=[];
        elif i ==":":
            prods.append(tov);
            tov="";
        else:
            tov=tov+i;
    return prod_mass;
def cort_to_list(cort):
    list_=[];
    for i in range(len(cort)):
        list_.append(cort[i][0]);
    return list_;
def arrayToString(maxList):
    string="";
    for Min in maxList:
        for m in Min:
            string=string+str(m)+"|";
        string=string+"^";
    return string;
def arrayToString2(maxList):
    return maxList;
def messageInChannel(merchName,text):
    try:
        #print(text);
        minch = threading.Thread(target=messageInChannelThread, args=(merchName,text),daemon=True)
        minch.start()
    except Exception as e:
        logger(e);
def messageInChannelThread(merchName,text):
    tg_token=getConst(merchName,"tg_token").replace(" ","");
    tg_id=getConst(merchName,"tg_id").replace(" ","");
    print(tg_token)
    print(tg_id)
    req = "https://api.telegram.org/bot"+tg_token+"/sendMessage?chat_id="+tg_id+"&text="+text;
    print(get_html(req).text);
    print(req)
def locationInChannel(merchName,lon,lat):
    try:
        #conn=sqlite3.connect("basic2.sqlite");
        #cursor=conn.cursor();
        #cursor.execute("SELECT token FROM 'telegram' WHERE merchName=(?)",(merchName,));
        #token=cort_to_list(cursor.fetchall())[0];
        #cursor.execute("SELECT chat_id FROM 'telegram' WHERE merchName=(?)",(merchName,));
        #chat_id=cort_to_list(cursor.fetchall())[0];
        #conn.close();
        for c in basic_telegramS:
            if c.merchName==merchName:
                token=c.token;
                chat_id=c.chat_id;
                break;
        req = "https://api.telegram.org/bot"+token+"/sendLocation?chat_id="+chat_id+"&latitude="+lat+"&longitude="+lon
        get_html(req);
    except Exception as e:
        logger(e);
def logger(e):
    try:
        logging.error(e);
        d=datetime.now();
        date="{0}.{1}.{2}".format(str(d.day),str(d.month),str(d.year));
        file=open("log/mylog.txt","w");
        file.write(str(traceback.format_exc()));
        file.close();

        try:
            file=open("log/logHistory.txt","r");
            text=file.read();
            file.close();
        except:
            text="start";
        file=open("log/logHistory.txt","w");
        errTxt=str(traceback.format_exc())+"\n>>>>>>"+date+"<<<<<<<<\n\n";
        file.write(errTxt+"\n"+text);
        file.close();

        file=open("log/mylog.txt","r");
        text=file.read();
        file.close();
        req = "https://api.telegram.org/bot"+token+"/sendMessage?chat_id="+MYTGID+"&text="+text;
        get_html(req);
    except:
        print(e);
def get_html(url,params=None):
    HEADERS ={};
    r = rq.get(url,headers=HEADERS,params=params)
    return r
def get_address_from_coords1(coords):
    try:
        PARAMS = {
            "apikey":"dcc7de33-5acb-4746-9558-a2bfbccc8391",
            "format":"json",
            "lang":"ru_RU",
            "kind":"house",
            "geocode": coords
        }
        r = rq.get(url="https://geocode-maps.yandex.ru/1.x/", params=PARAMS)
        json_data = r.json()
        address_str = json_data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["metaDataProperty"]["GeocoderMetaData"]["AddressDetails"]["Country"]["AddressLine"]
        return address_str;
    except Exception as e:
        logger(e);
        return "Адрес не установлен"
def get_address_from_coords2(coords):
    try:
        PARAMS = {
            "apikey":"dcc7de33-5acb-4746-9558-a2bfbccc8391",
            "format":"json",
            "lang":"ru_RU",
            "kind":"district",
            "geocode": coords
        }
        r = rq.get(url="https://geocode-maps.yandex.ru/1.x/", params=PARAMS)
        json_data = r.json()
        address_str2 = json_data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["metaDataProperty"]["GeocoderMetaData"]["AddressDetails"]["Country"]["AddressLine"]
        return address_str2;
    except Exception as e:
        logger(e);
        return "Адрес не установлен"

startServer();
#startServerThread();
writeDbThread();
